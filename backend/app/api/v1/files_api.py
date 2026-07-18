from __future__ import annotations

import hashlib
import mimetypes
from datetime import UTC, datetime
from pathlib import Path
from tempfile import SpooledTemporaryFile
from urllib.parse import quote
from uuid import uuid4

from fastapi import APIRouter, Depends, File, Header, Query, Request, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, exists, or_, select
from sqlalchemy.orm import Session
from sqlalchemy.sql.elements import ColumnElement

from app.api.deps import (
    CurrentUser,
    get_db,
    get_project_membership,
    has_global_project_access,
)
from app.core.constants import (
    ALLOWED_UPLOAD_EXTENSIONS,
    JOB_PENDING,
    JOB_QUEUED,
    JOB_RUNNING,
    JOB_VALIDATING,
    JOB_WAITING_CAD_WORKER,
    TASK_DWG_TO_DXF,
    TASK_DXF_TO_DWG,
)
from app.core.exceptions import AppHTTPException, forbidden, not_found
from app.core.validators import validate_sort_by
from app.db.pagination import paginate_scalars
from app.models.drawing import Drawing, DrawingVersion
from app.models.file import StoredFile
from app.models.job import Job
from app.models.project import Project, ProjectMember
from app.models.result import AnalysisResult
from app.schemas.common import ok
from app.schemas.common import page as page_response
from app.schemas.file_schema import (
    BatchBulkDeleteRequest,
    BatchBulkDeleteResult,
    BulkDeleteRequest,
    DxfPreviewBoundsRead,
    DxfPreviewRead,
    FileRead,
    ZipDownloadRequest,
    ZipUploadResult,
)
from app.services.audit_service import write_audit_log
from app.services.dxf_preview_service import (
    MAX_DXF_SIZE_BYTES,
    get_or_create_dxf_preview,
    invalidate_dxf_previews_for_source,
    preview_batch_name,
    validate_dxf_source_size,
)
from app.services.file_service import (
    build_signed_download_url,
    build_zip_to_path,
    download_headers,
    validate_download_signature,
)
from app.services.file_transfer_service import (
    ACTIVE_TRANSFER_STATUSES,
    TransferSpec,
    complete_transfer_in_transaction,
    prepare_transfer_in_transaction,
    session_factory_for,
    settle_stream,
    settle_transfer,
)
from app.services.job_service import cancel_job as transition_job_to_cancelled
from app.services.storage_service import (
    get_storage_backend,
    sanitize_filename,
    save_bytes_as_file,
    save_upload_file,
    validate_dwg_header,
    validate_upload_name,
)
from app.storage.base import AbstractStorageBackend, StorageError, StorageObjectNotFound

router = APIRouter()


def _file_project_association_exists(
    *,
    active_only: bool,
    member_user_id: int | None = None,
) -> ColumnElement[bool]:
    """Build a correlated project-association predicate for the outer file row."""
    drawing_conditions = [
        DrawingVersion.file_id == StoredFile.id,
        Drawing.status != "deleted",
    ]
    result_conditions = [
        AnalysisResult.result_file_id == StoredFile.id,
        Job.project_id.is_not(None),
    ]
    if active_only:
        drawing_conditions.append(Project.status != "deleted")
        result_conditions.append(Project.status != "deleted")

    drawing_stmt = (
        select(1)
        .select_from(DrawingVersion)
        .join(Drawing, Drawing.id == DrawingVersion.drawing_id)
        .join(Project, Project.id == Drawing.project_id)
    )
    result_stmt = (
        select(1)
        .select_from(AnalysisResult)
        .join(Job, Job.id == AnalysisResult.job_id)
        .join(Project, Project.id == Job.project_id)
    )
    if member_user_id is not None:
        drawing_stmt = drawing_stmt.join(ProjectMember, ProjectMember.project_id == Project.id)
        result_stmt = result_stmt.join(ProjectMember, ProjectMember.project_id == Project.id)
        drawing_conditions.append(ProjectMember.user_id == member_user_id)
        result_conditions.append(ProjectMember.user_id == member_user_id)

    return or_(
        exists(drawing_stmt.where(*drawing_conditions)),
        exists(result_stmt.where(*result_conditions)),
    )


def _file_list_access_filter(current_user: CurrentUser) -> ColumnElement[bool]:
    """Mirror single-file access rules as one SQL predicate for list endpoints."""
    any_project = _file_project_association_exists(active_only=False)
    active_project = _file_project_association_exists(active_only=True)
    not_orphaned_by_project_deletion = or_(~any_project, active_project)

    if has_global_project_access(current_user):
        return not_orphaned_by_project_deletion

    active_membership = _file_project_association_exists(
        active_only=True,
        member_user_id=current_user.id,
    )
    return and_(
        not_orphaned_by_project_deletion,
        or_(StoredFile.uploaded_by == current_user.id, active_membership),
    )


def _file_project_ids(db: Session, file_id: int, *, include_deleted: bool = False) -> set[int]:
    drawing_stmt = (
        select(Drawing.project_id)
        .join(DrawingVersion, DrawingVersion.drawing_id == Drawing.id)
        .join(Project, Project.id == Drawing.project_id)
        .where(
            DrawingVersion.file_id == file_id,
            Drawing.status != "deleted",
        )
    )
    if not include_deleted:
        drawing_stmt = drawing_stmt.where(Project.status != "deleted")
    drawing_project_ids = db.scalars(drawing_stmt).all()

    result_stmt = (
        select(Job.project_id)
        .join(AnalysisResult, AnalysisResult.job_id == Job.id)
        .join(Project, Project.id == Job.project_id)
        .where(
            AnalysisResult.result_file_id == file_id,
            Job.project_id.is_not(None),
        )
    )
    if not include_deleted:
        result_stmt = result_stmt.where(Project.status != "deleted")
    result_project_ids = db.scalars(result_stmt).all()

    return {
        project_id
        for project_id in (*drawing_project_ids, *result_project_ids)
        if project_id is not None
    }


def _can_read_file(db: Session, current_user: CurrentUser, stored: StoredFile) -> bool:
    active_project_ids = _file_project_ids(db, stored.id, include_deleted=False)

    # If the file is attached to projects but all of them have been soft-deleted,
    # treat it as inaccessible regardless of global role / uploader status.
    if not active_project_ids and _file_project_ids(db, stored.id, include_deleted=True):
        return False

    if has_global_project_access(current_user) or stored.uploaded_by == current_user.id:
        return True
    return any(
        get_project_membership(db, current_user, project_id) for project_id in active_project_ids
    )


def _require_file_read_access(db: Session, current_user: CurrentUser, stored: StoredFile) -> None:
    # If the file is attached only to soft-deleted projects, treat as not found
    # so that soft-deleting a project cascades to its file metadata (BUG-7).
    active_ids = _file_project_ids(db, stored.id, include_deleted=False)
    all_ids = _file_project_ids(db, stored.id, include_deleted=True)
    if not active_ids and all_ids:
        raise not_found("File")
    if not _can_read_file(db, current_user, stored):
        raise forbidden("File access is restricted.")


def _require_file_delete_access(db: Session, current_user: CurrentUser, stored: StoredFile) -> None:
    # If all associated projects are soft-deleted, treat the file as not found (BUG-7).
    active_ids = _file_project_ids(db, stored.id, include_deleted=False)
    all_ids = _file_project_ids(db, stored.id, include_deleted=True)
    if not active_ids and all_ids:
        raise not_found("File")
    if has_global_project_access(current_user) or stored.uploaded_by == current_user.id:
        return
    raise forbidden("Only the uploader or an administrator can delete this file.")


def _soft_delete_file_in_transaction(
    db: Session,
    stored: StoredFile,
    *,
    actor_user_id: int,
    request_id: str,
    batch_ref: str | None = None,
) -> None:
    invalidate_dxf_previews_for_source(
        db,
        stored,
        actor_user_id=actor_user_id,
        request_id=request_id,
    )
    stored.status = "deleted"
    stored.deleted_at = datetime.now(UTC)
    transfer = prepare_transfer_in_transaction(
        db,
        TransferSpec(
            direction="internal",
            operation="soft_delete",
            actor_user_id=actor_user_id,
            request_id=request_id,
            file_id=stored.id,
            batch_ref=batch_ref,
            bucket=stored.bucket,
            storage_key=stored.storage_key,
            original_name=stored.original_name,
            expected_bytes=stored.size_bytes,
        ),
    )
    complete_transfer_in_transaction(
        db,
        transfer.transfer_uid,
        file_id=stored.id,
        bucket=stored.bucket,
        storage_key=stored.storage_key,
        original_name=stored.original_name,
        transferred_bytes=stored.size_bytes,
    )


@router.post("", status_code=status.HTTP_201_CREATED)
async def upload_file(
    request: Request,
    current_user: CurrentUser,
    upload: UploadFile = File(...),
    batch_name: str = Query(""),
    idempotency_key: str | None = Header(
        default=None,
        alias="Idempotency-Key",
        max_length=128,
    ),
    db: Session = Depends(get_db),
):
    transfer = prepare_transfer_in_transaction(
        db,
        TransferSpec(
            direction="inbound",
            operation="upload",
            actor_user_id=current_user.id,
            request_id=request.state.request_id,
            idempotency_key=idempotency_key,
            original_name=sanitize_filename(upload.filename or "unnamed.dwg"),
        ),
    )
    db.commit()

    if transfer.status == "succeeded" and transfer.file_id is not None:
        stored = db.get(StoredFile, transfer.file_id)
        if stored is None:
            raise AppHTTPException(
                409,
                "IDEMPOTENT_RESULT_MISSING",
                "The previous upload result is no longer available.",
            )
        return ok(FileRead.model_validate(stored), request.state.request_id)
    if transfer.status not in ACTIVE_TRANSFER_STATUSES:
        raise AppHTTPException(
            409,
            "IDEMPOTENT_OPERATION_FAILED",
            "The previous operation with this idempotency key did not succeed.",
            {"transfer_uid": transfer.transfer_uid},
        )

    try:
        stored = await save_upload_file(
            db,
            upload,
            uploaded_by=current_user.id,
            batch_name=sanitize_filename(batch_name.strip()) if batch_name.strip() else None,
            transfer_uid=transfer.transfer_uid,
        )
        complete_transfer_in_transaction(
            db,
            transfer.transfer_uid,
            file_id=stored.id,
            bucket=stored.bucket,
            storage_key=stored.storage_key,
            original_name=stored.original_name,
            transferred_bytes=stored.size_bytes,
        )
        write_audit_log(
            db,
            actor_user_id=current_user.id,
            action="files.upload",
            resource_type="file",
            resource_id=stored.id,
            after_json={"original_name": stored.original_name, "sha256": stored.sha256},
            request=request,
        )
        db.commit()
    except Exception as exc:
        db.rollback()
        detail = exc.detail if isinstance(exc, AppHTTPException) else None
        settle_transfer(
            session_factory_for(db),
            transfer.transfer_uid,
            status="failed",
            transferred_bytes=0,
            error_code=detail["code"] if isinstance(detail, dict) else "UPLOAD_TRANSACTION_FAILED",
            error_message=(
                detail["message"]
                if isinstance(detail, dict)
                else "Upload transaction failed before completion."
            ),
        )
        raise
    return ok(FileRead.model_validate(stored), request.state.request_id)


@router.post("/upload-zip", status_code=status.HTTP_201_CREATED)
async def upload_zip(
    request: Request,
    current_user: CurrentUser,
    upload: UploadFile = File(...),
    file_ext: str = Query(".dwg", description="只提取此扩展名的文件 (.dwg 或 .dxf)"),
    db: Session = Depends(get_db),
):
    """Upload a .zip archive, extract matching files, and auto-create StoredFile records.

    The ZIP filename (minus .zip) becomes the batch_name for all extracted files.
    Only files matching ``file_ext`` are imported; others are counted as skipped.
    """
    import io
    import zipfile

    from app.core.config import settings

    # ── validate the upload is a .zip ──────────────────────────────────────
    zip_original = sanitize_filename(upload.filename or "unnamed.zip")
    if not zip_original.lower().endswith(".zip"):
        raise AppHTTPException(
            415,
            "FILE_TYPE_NOT_ALLOWED",
            "Only .zip archives are accepted by this endpoint.",
        )
    validate_upload_name(zip_original)  # 校验 .zip 在 ALLOWED_UPLOAD_EXTENSIONS 白名单内
    if not file_ext.strip():
        raise AppHTTPException(422, "INVALID_PARAMS", "file_ext query parameter is required.")
    target_ext = file_ext.strip().lower()
    if target_ext not in (".dwg", ".dxf"):
        raise AppHTTPException(
            422,
            "INVALID_PARAMS",
            "file_ext must be .dwg or .dxf.",
        )
    if target_ext not in ALLOWED_UPLOAD_EXTENSIONS:
        raise AppHTTPException(
            422,
            "INVALID_PARAMS",
            f"Target extension {target_ext} is not allowed.",
        )

    # Derive batch_name from the ZIP filename
    batch_name = sanitize_filename(
        zip_original[: -len(".zip")]
        if zip_original.lower().endswith(".zip")
        else zip_original.rsplit(".", 1)[0]
    )
    if not batch_name or batch_name == "unnamed":
        batch_name = f"导入_{datetime.now(UTC).strftime('%Y%m%d_%H%M%S')}"

    # ── buffer the upload ──────────────────────────────────────────────────
    max_upload = settings.max_upload_size_mb * 1024 * 1024
    zip_bytes_total = 0
    with SpooledTemporaryFile(max_size=16 * 1024 * 1024, mode="w+b") as tmp:
        while chunk := await upload.read(1024 * 1024):
            zip_bytes_total += len(chunk)
            if zip_bytes_total > max_upload:
                raise AppHTTPException(
                    413,
                    "FILE_TOO_LARGE",
                    f"ZIP archive exceeds max upload size ({settings.max_upload_size_mb} MB).",
                )
            tmp.write(chunk)
        tmp.seek(0)

        # ── extract ────────────────────────────────────────────────────────
        try:
            zf = zipfile.ZipFile(tmp, "r")
        except (zipfile.BadZipFile, io.UnsupportedOperation) as e:
            raise AppHTTPException(
                415,
                "FILE_NOT_ZIP",
                "The uploaded file is not a valid ZIP archive.",
            ) from e

        bad_names = zf.testzip()
        if bad_names is not None:
            raise AppHTTPException(
                415,
                "ZIP_CORRUPTED",
                f"ZIP archive contains corrupted entry: {bad_names}",
            )

        extracted: list[StoredFile] = []
        skipped = 0
        total_extracted_bytes = 0
        max_extract = settings.max_zip_extract_mb * 1024 * 1024
        entry_count = 0

        for info in zf.infolist():
            if info.is_dir():
                continue
            entry_count += 1
            if entry_count > settings.max_zip_entry_count:
                raise AppHTTPException(
                    413,
                    "ZIP_TOO_MANY_FILES",
                    f"ZIP archive contains more than {settings.max_zip_entry_count} files.",
                )

            # Path traversal defence: sanitize + strip any directory components
            entry_name = sanitize_filename(
                info.filename.rsplit("/", 1)[-1] if "/" in info.filename else info.filename
            )
            entry_ext = Path(entry_name).suffix.lower()
            if entry_ext != target_ext:
                skipped += 1
                continue

            # Read entry
            entry_bytes = zf.read(info)
            total_extracted_bytes += len(entry_bytes)
            if total_extracted_bytes > max_extract:
                raise AppHTTPException(
                    413,
                    "ZIP_TOO_LARGE",
                    f"Extracted content exceeds {settings.max_zip_extract_mb} MB limit.",
                )

            # DWG header validation
            if entry_ext == ".dwg":
                try:
                    validate_dwg_header(entry_bytes[:6])
                except AppHTTPException:
                    skipped += 1
                    continue

            # Persist via save_bytes_as_file (reuses existing storage logic)
            content_type = mimetypes.guess_type(entry_name)[0]
            bucket = (
                settings.minio_bucket_original
                if entry_ext == ".dwg"
                else settings.minio_bucket_dxf_original
            )
            storage_key = f"uploads/{uuid4().hex}{entry_ext}"
            stored = save_bytes_as_file(
                db,
                bucket=bucket,
                storage_key=storage_key,
                original_name=entry_name,
                file_ext=entry_ext,
                content_type=content_type or "application/octet-stream",
                payload=entry_bytes,
                uploaded_by=current_user.id,
                batch_name=batch_name,
                transfer_direction="inbound",
                transfer_operation="upload_zip",
                request_id=request.state.request_id,
            )
            extracted.append(stored)

    if not extracted and skipped == 0:
        raise AppHTTPException(
            422,
            "ZIP_EMPTY",
            "The ZIP archive contains no files.",
        )

    # ── audit log ──────────────────────────────────────────────────────────
    write_audit_log(
        db,
        actor_user_id=current_user.id,
        action="files.upload_zip",
        resource_type="file",
        resource_id=0,
        after_json={
            "batch_name": batch_name,
            "target_ext": target_ext,
            "success_count": len(extracted),
            "skipped_count": skipped,
            "zip_original": zip_original,
            "zip_size": zip_bytes_total,
        },
        request=request,
    )
    db.commit()

    return ok(
        ZipUploadResult(
            batch_name=batch_name,
            files=[FileRead.model_validate(f) for f in extracted],
            success_count=len(extracted),
            skipped_count=skipped,
        ).model_dump(),
        request.state.request_id,
    )


@router.get("")
def list_files(
    request: Request,
    current_user: CurrentUser,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    sort_by: str = Query("created_at"),
    sort_dir: str = Query("desc", pattern=r"^(asc|desc)$"),
    batch_name: str = Query(""),
    file_ext: str = Query("", description="Filter by file extension, e.g. '.dwg' or '.dxf'"),
    db: Session = Depends(get_db),
):
    sort_column = validate_sort_by("files", sort_by)
    sort_dir_value = sort_dir.strip().lower()
    order_clause = getattr(StoredFile, sort_column)
    if sort_dir_value == "asc":
        order_clause = order_clause.asc()
    else:
        order_clause = order_clause.desc()
    stmt = select(StoredFile).where(StoredFile.status != "deleted")
    if batch_name.strip():
        stmt = stmt.where(StoredFile.batch_name == batch_name.strip())
    if file_ext.strip():
        stmt = stmt.where(StoredFile.file_ext == file_ext.strip())
    tie_breaker = StoredFile.id.asc() if sort_dir_value == "asc" else StoredFile.id.desc()
    stmt = stmt.where(_file_list_access_filter(current_user)).order_by(order_clause, tie_breaker)
    files, total = paginate_scalars(db, stmt, page_no=page, page_size=page_size)
    return page_response(
        [FileRead.model_validate(f) for f in files],
        page,
        page_size,
        total,
        request.state.request_id,
    )


# ── batches ─────────────────────────────────────────────────────────────────
# NOTE: must be registered BEFORE /{file_id} to avoid route shadowing.


@router.get("/batches")
def list_batches(
    request: Request,
    current_user: CurrentUser,
    file_ext: str = Query(
        "", description="Filter batches by file extension, e.g. '.dwg' or '.dxf'"
    ),
    db: Session = Depends(get_db),
):
    """Query MySQL for distinct batch names, file counts and latest creation times."""
    from sqlalchemy import func as sa_func

    where_clauses = [
        StoredFile.batch_name.is_not(None),
        StoredFile.batch_name != "",
        StoredFile.status != "deleted",
    ]
    if file_ext.strip():
        where_clauses.append(StoredFile.file_ext == file_ext.strip())
    where_clauses.append(_file_list_access_filter(current_user))

    rows = list(
        db.execute(
            select(
                StoredFile.batch_name,
                sa_func.count(StoredFile.id).label("file_count"),
                sa_func.max(StoredFile.created_at).label("latest_created_at"),
            )
            .where(*where_clauses)
            .group_by(StoredFile.batch_name)
            .order_by(sa_func.max(StoredFile.created_at).desc())
        ).all()
    )
    batches = [
        {
            "name": r.batch_name,
            "file_count": r.file_count,
            "latest_created_at": r.latest_created_at.isoformat(),
        }
        for r in rows
    ]
    return ok(batches, request.state.request_id)


@router.post("/batches/bulk-delete")
def bulk_delete_batches(
    payload: BatchBulkDeleteRequest,
    request: Request,
    current_user: CurrentUser,
    db: Session = Depends(get_db),
):
    """Atomically soft-delete one or more complete file batches."""
    batch_names = list(dict.fromkeys(payload.batch_names))
    stored_list = list(
        db.scalars(
            select(StoredFile).where(
                StoredFile.batch_name.in_(batch_names),
                StoredFile.status != "deleted",
            )
        ).all()
    )
    found_names = {stored.batch_name for stored in stored_list}
    if found_names != set(batch_names):
        raise not_found("Batch")

    for stored in stored_list:
        _require_file_delete_access(db, current_user, stored)

    source_ids = [stored.id for stored in stored_list]
    active_jobs = list(
        db.scalars(
            select(Job).where(
                Job.task_type.in_((TASK_DWG_TO_DXF, TASK_DXF_TO_DWG)),
                Job.status.in_(
                    (
                        JOB_PENDING,
                        JOB_QUEUED,
                        JOB_RUNNING,
                        JOB_VALIDATING,
                        JOB_WAITING_CAD_WORKER,
                    )
                ),
                Job.params_json["file_id"].as_integer().in_(source_ids),
            )
        ).all()
    )

    try:
        for job in active_jobs:
            transition_job_to_cancelled(db, job)
        for stored in stored_list:
            _soft_delete_file_in_transaction(
                db,
                stored,
                actor_user_id=current_user.id,
                request_id=request.state.request_id,
                batch_ref=stored.batch_name,
            )
            write_audit_log(
                db,
                actor_user_id=current_user.id,
                action="files.batch_delete",
                resource_type="file",
                resource_id=stored.id,
                after_json={"batch_name": stored.batch_name, "bulk": True},
                request=request,
            )
        result = BatchBulkDeleteResult(
            deleted_batch_count=len(batch_names),
            deleted_file_count=len(stored_list),
            cancelled_job_count=len(active_jobs),
        )
        db.commit()
    except Exception:
        db.rollback()
        raise
    return ok(result, request.state.request_id)


@router.delete("/batches/{batch_name}", status_code=status.HTTP_204_NO_CONTENT)
def delete_batch(
    batch_name: str,
    request: Request,
    current_user: CurrentUser,
    db: Session = Depends(get_db),
):
    """Soft-delete all non-deleted files in a batch.

    Uses a bulk UPDATE for efficiency on large batches. Access control is
    enforced per-file — if any file in the batch is not deletable by the
    current user the whole operation is rejected.
    """
    stored_list = list(
        db.scalars(
            select(StoredFile).where(
                StoredFile.batch_name == batch_name,
                StoredFile.status != "deleted",
            )
        ).all()
    )
    if not stored_list:
        raise not_found("Batch")

    # Verify access for every file before mutating
    for s in stored_list:
        _require_file_delete_access(db, current_user, s)

    # Bulk soft-delete
    deleted_count = 0
    for s in stored_list:
        _soft_delete_file_in_transaction(
            db,
            s,
            actor_user_id=current_user.id,
            request_id=request.state.request_id,
            batch_ref=batch_name,
        )
        deleted_count += 1
        write_audit_log(
            db,
            actor_user_id=current_user.id,
            action="files.batch_delete",
            resource_type="file",
            resource_id=s.id,
            after_json={"batch_name": batch_name},
            request=request,
        )
    db.commit()
    return None


@router.get("/batches/{batch_name}/download-zip")
def download_batch_zip(
    batch_name: str,
    request: Request,
    current_user: CurrentUser,
    db: Session = Depends(get_db),
):
    """Download all non-deleted files in a batch as a ZIP archive.

    Streams the zip directly — no intermediate storage. Only the original
    file format is included (not conversion results).
    """
    stored_list = list(
        db.scalars(
            select(StoredFile).where(
                StoredFile.batch_name == batch_name,
                StoredFile.status != "deleted",
            )
        ).all()
    )
    if not stored_list:
        raise not_found("Batch")

    for s in stored_list:
        _require_file_read_access(db, current_user, s)

    file_ids = [s.id for s in stored_list]
    # Determine format from the batch's file extension
    first_ext = stored_list[0].file_ext.lstrip(".") if stored_list else "dxf"
    formats = [first_ext] if first_ext in ("dwg", "dxf") else ["dxf"]

    clean_name = sanitize_filename(batch_name)
    prepared = build_zip_to_path(db, file_ids, formats, clean_name)
    try:
        transfer = prepare_transfer_in_transaction(
            db,
            TransferSpec(
                direction="outbound",
                operation="download_zip",
                actor_user_id=current_user.id,
                request_id=request.state.request_id,
                idempotency_key=request.state.request_id,
                batch_ref=batch_name,
                original_name=prepared.filename,
                expected_bytes=prepared.size_bytes,
            ),
        )
        write_audit_log(
            db,
            actor_user_id=current_user.id,
            action="files.batch_download_zip",
            resource_type="file",
            resource_id=0,
            after_json={"batch_name": batch_name, "file_count": len(file_ids)},
            request=request,
        )
        db.commit()
    except Exception:
        db.rollback()
        prepared.path.unlink(missing_ok=True)
        raise

    def _stream_and_cleanup():
        try:
            with prepared.path.open("rb") as f:
                while chunk := f.read(1024 * 1024):
                    yield chunk
        finally:
            prepared.path.unlink(missing_ok=True)

    encoded_filename = quote(f"{clean_name}.zip")
    return StreamingResponse(
        settle_stream(
            session_factory_for(db),
            transfer.transfer_uid,
            _stream_and_cleanup(),
        ),
        media_type="application/zip",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            "Content-Length": str(prepared.size_bytes),
        },
    )


def _column_letter(index: int) -> str:
    """Convert 0-based column index to Excel column letter(s): 0→A, 25→Z, 26→AA."""
    letters: list[str] = []
    n = index
    while True:
        n, rem = divmod(n, 26)
        letters.append(chr(ord("A") + rem))
        if n == 0:
            break
        n -= 1
    return "".join(reversed(letters))


@router.get("/{file_id}/excel-preview")
def get_excel_preview(
    file_id: int,
    request: Request,
    current_user: CurrentUser,
    sheet: str = Query("", description="Sheet name to preview (empty = first sheet)"),
    db: Session = Depends(get_db),
):
    """Read an Excel file from authoritative storage and return preview JSON."""
    stored = db.get(StoredFile, file_id)
    if not stored or stored.status == "deleted":
        raise not_found("File")
    _require_file_read_access(db, current_user, stored)

    if not stored.file_ext or stored.file_ext.lower() not in (".xlsx", ".xls"):
        raise AppHTTPException(
            415,
            "NOT_EXCEL",
            "Only .xlsx / .xls files can be previewed.",
        )

    # Read Excel bytes from storage
    storage = get_storage_backend()
    try:
        local_path = storage.local_path(stored.bucket, stored.storage_key)
        if local_path is not None:
            if not local_path.exists():
                raise StorageObjectNotFound(f"{stored.bucket}/{stored.storage_key}")
            excel_bytes = local_path.read_bytes()
        else:
            chunks: list[bytes] = []
            for chunk in storage.iter_file(stored.bucket, stored.storage_key):
                chunks.append(chunk)
            excel_bytes = b"".join(chunks)
    except StorageObjectNotFound:
        raise not_found("StoredFileObject") from None
    except StorageError as exc:
        raise AppHTTPException(
            503, "STORAGE_READ_FAILED", "Failed to read stored file object."
        ) from exc

    # Parse with openpyxl
    try:
        import io

        import openpyxl
    except ImportError as exc:
        raise AppHTTPException(
            503,
            "OPENPYXL_UNAVAILABLE",
            "openpyxl is not installed — cannot preview Excel files.",
        ) from exc

    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise AppHTTPException(
            415,
            "EXCEL_PARSE_ERROR",
            f"Failed to parse Excel file: {exc}",
        ) from exc

    sheet_names = wb.sheetnames
    if not sheet_names:
        raise AppHTTPException(415, "EXCEL_EMPTY", "Excel file has no sheets.")

    target_sheet = sheet.strip() if sheet.strip() else sheet_names[0]
    if target_sheet not in sheet_names:
        raise AppHTTPException(
            422,
            "SHEET_NOT_FOUND",
            f"Sheet '{target_sheet}' not found. Available: {', '.join(sheet_names)}",
        )

    ws = wb[target_sheet]
    rows_iter = ws.iter_rows(values_only=True)

    # First row is always the header, remaining rows are data.
    # Simple and predictable — no heuristic scoring that can accidentally
    # discard real data rows as "metadata".
    try:
        headers_raw: tuple[object, ...] = next(rows_iter)
    except StopIteration:
        headers_raw = ()

    # Build clean, unique headers. Empty cells become "Col A", "Col B", ...
    # Duplicates get a numeric suffix ("Name", "Name_2", "Name_3").
    seen: dict[str, int] = {}
    headers: list[str] = []
    for idx, h in enumerate(headers_raw or []):
        col_letter = _column_letter(idx)
        base = str(h).strip() if h is not None and str(h).strip() else f"Col {col_letter}"
        if base in seen:
            seen[base] += 1
            headers.append(f"{base}_{seen[base]}")
        else:
            seen[base] = 0
            headers.append(base)

    # If there are more data columns than headers, pad with column letters
    # (openpyxl iter_rows may return rows wider than the header row)
    _max_data_cols = 0

    data_rows: list[dict[str, object]] = []

    def _extract_row(row: tuple[object, ...]) -> dict[str, object]:
        """Convert an openpyxl row tuple into a column-keyed dict."""
        nonlocal _max_data_cols
        if len(row) > _max_data_cols:
            _max_data_cols = len(row)
        row_dict: dict[str, object] = {}
        for idx, val in enumerate(row):
            while idx >= len(headers):
                headers.append(f"Col {_column_letter(len(headers))}")
            col_name = headers[idx]
            if val is None:
                row_dict[col_name] = None
            elif isinstance(val, (int, float)):
                row_dict[col_name] = val
            else:
                row_dict[col_name] = str(val)
        return row_dict

    for row in rows_iter:
        data_rows.append(_extract_row(row))

    wb.close()

    result: dict = {
        "file": stored.original_name,
        "file_id": file_id,
        "sheets": sheet_names,
        "sheet": target_sheet,
        "headers": headers,
        "rows": data_rows,
        "total_rows": len(data_rows),
    }

    return ok(result, request.state.request_id)


def _read_dxf_preview_source(
    stored: StoredFile,
) -> tuple[bytes, AbstractStorageBackend]:
    """Read one DXF through the shared adapter with size and digest guards."""
    validate_dxf_source_size(stored.size_bytes)
    storage = get_storage_backend()
    payload = bytearray()
    digest = hashlib.sha256()
    try:
        for chunk in storage.iter_file(stored.bucket, stored.storage_key):
            payload.extend(chunk)
            if len(payload) > MAX_DXF_SIZE_BYTES:
                raise AppHTTPException(
                    413,
                    "DXF_TOO_LARGE",
                    f"DXF 文件超过在线预览上限 {MAX_DXF_SIZE_BYTES // (1024 * 1024)} MB。",
                )
            digest.update(chunk)
    except AppHTTPException:
        raise
    except StorageObjectNotFound:
        raise not_found("StoredFileObject") from None
    except StorageError as exc:
        raise AppHTTPException(
            503,
            "STORAGE_READ_FAILED",
            "Failed to read stored file object.",
        ) from exc

    if len(payload) != stored.size_bytes:
        raise AppHTTPException(
            409,
            "STORAGE_SIZE_MISMATCH",
            "DXF 对象大小与 MySQL 登记不一致，请先执行存储一致性扫描。",
        )
    if digest.hexdigest() != stored.sha256:
        raise AppHTTPException(
            409,
            "STORAGE_CHECKSUM_MISMATCH",
            "DXF 对象校验值与 MySQL 登记不一致，请先执行存储一致性扫描。",
        )
    return bytes(payload), storage


@router.get("/{file_id}/dxf-preview")
def get_dxf_preview(
    file_id: int,
    request: Request,
    current_user: CurrentUser,
    db: Session = Depends(get_db),
):
    """Generate or reuse a registered SVG preview for an accessible DXF."""
    stored = db.get(StoredFile, file_id)
    if not stored or stored.status == "deleted":
        raise not_found("File")
    _require_file_read_access(db, current_user, stored)
    if stored.file_ext.lower() != ".dxf":
        raise AppHTTPException(
            415,
            "NOT_DXF",
            "Only .dxf files can be previewed with this endpoint.",
        )

    payload, storage = _read_dxf_preview_source(stored)
    preview = get_or_create_dxf_preview(
        db,
        stored,
        payload,
        storage=storage,
        request_id=request.state.request_id,
    )
    preview_id = preview.preview_file.id
    assert preview_id is not None
    action = "files.dxf_preview_cache_hit" if preview.cached else "files.dxf_preview_generate"
    write_audit_log(
        db,
        actor_user_id=current_user.id,
        action=action,
        resource_type="file",
        resource_id=stored.id,
        after_json={"preview_file_id": preview_id},
        request=request,
    )
    db.commit()
    response = DxfPreviewRead(
        file_id=stored.id,
        file_name=stored.original_name,
        preview_file_id=preview_id,
        content_url=(f"/api/v1/files/{stored.id}/dxf-preview/content?preview_file_id={preview_id}"),
        content_type=preview.preview_file.content_type or "image/svg+xml",
        document_entities=preview.document_entities,
        modelspace_entities=preview.modelspace_entities,
        entity_counts=preview.entity_counts,
        layers=list(preview.layers),
        layer_colors=preview.layer_colors,
        bounds=DxfPreviewBoundsRead(
            min_x=preview.bounds.min_x,
            min_y=preview.bounds.min_y,
            max_x=preview.bounds.max_x,
            max_y=preview.bounds.max_y,
        ),
        cached=preview.cached,
    )
    return ok(response.model_dump(), request.state.request_id)


@router.get("/{file_id}/dxf-preview/content")
def get_dxf_preview_content(
    file_id: int,
    request: Request,
    current_user: CurrentUser,
    preview_file_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
):
    """Stream registered preview content after rechecking source-file access."""
    source = db.get(StoredFile, file_id)
    if not source or source.status == "deleted":
        raise not_found("File")
    _require_file_read_access(db, current_user, source)
    preview = db.get(StoredFile, preview_file_id)
    if (
        preview is None
        or preview.status == "deleted"
        or preview.file_ext != ".svg"
        or preview.batch_name != preview_batch_name(source)
    ):
        raise not_found("DxfPreview")

    storage = get_storage_backend()
    try:
        object_info = storage.stat_object(preview.bucket, preview.storage_key)
    except StorageObjectNotFound:
        raise not_found("DxfPreviewObject") from None
    except StorageError as exc:
        raise AppHTTPException(
            503,
            "STORAGE_READ_FAILED",
            "Failed to read preview object.",
        ) from exc
    if object_info.size_bytes != preview.size_bytes:
        raise AppHTTPException(
            409,
            "STORAGE_SIZE_MISMATCH",
            "DXF 预览对象大小与 MySQL 登记不一致。",
        )
    transfer = prepare_transfer_in_transaction(
        db,
        TransferSpec(
            direction="outbound",
            operation="preview",
            actor_user_id=current_user.id,
            request_id=request.state.request_id,
            idempotency_key=request.state.request_id,
            file_id=preview.id,
            batch_ref=preview.batch_name,
            bucket=preview.bucket,
            storage_key=preview.storage_key,
            original_name=preview.original_name,
            expected_bytes=object_info.size_bytes,
        ),
    )
    write_audit_log(
        db,
        actor_user_id=current_user.id,
        action="files.dxf_preview_view",
        resource_type="file",
        resource_id=source.id,
        after_json={"preview_file_id": preview.id},
        request=request,
    )
    db.commit()
    factory = session_factory_for(db)
    return StreamingResponse(
        settle_stream(
            factory,
            transfer.transfer_uid,
            storage.iter_file(preview.bucket, preview.storage_key),
        ),
        media_type="image/svg+xml",
        headers={
            "Content-Length": str(object_info.size_bytes),
            "Content-Disposition": (f"inline; filename*=UTF-8''{quote(preview.original_name)}"),
            "Cache-Control": "private, max-age=300",
            "X-Content-Type-Options": "nosniff",
            "Content-Security-Policy": ("sandbox; default-src 'none'; style-src 'unsafe-inline'"),
        },
    )


@router.get("/{file_id}")
def get_file(
    file_id: int, request: Request, current_user: CurrentUser, db: Session = Depends(get_db)
):
    stored = db.get(StoredFile, file_id)
    if not stored or stored.status == "deleted":
        raise not_found("File")
    _require_file_read_access(db, current_user, stored)
    return ok(FileRead.model_validate(stored), request.state.request_id)


@router.delete("/{file_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_file(
    file_id: int, request: Request, current_user: CurrentUser, db: Session = Depends(get_db)
):
    stored = db.get(StoredFile, file_id)
    if not stored or stored.status == "deleted":
        raise not_found("File")
    _require_file_delete_access(db, current_user, stored)
    _soft_delete_file_in_transaction(
        db,
        stored,
        actor_user_id=current_user.id,
        request_id=request.state.request_id,
    )
    write_audit_log(
        db,
        actor_user_id=current_user.id,
        action="files.delete",
        resource_type="file",
        resource_id=stored.id,
        request=request,
    )
    db.commit()
    return None


@router.get("/{file_id}/download-url")
def get_download_url(
    file_id: int, request: Request, current_user: CurrentUser, db: Session = Depends(get_db)
):
    stored = db.get(StoredFile, file_id)
    if not stored or stored.status == "deleted":
        raise not_found("File")
    _require_file_read_access(db, current_user, stored)
    write_audit_log(
        db,
        actor_user_id=current_user.id,
        action="files.download_url",
        resource_type="file",
        resource_id=stored.id,
        request=request,
    )
    db.commit()
    return ok(build_signed_download_url(file_id), request.state.request_id)


@router.get("/{file_id}/download")
def download_file(
    file_id: int,
    request: Request,
    current_user: CurrentUser,
    expires: int | None = Query(default=None),
    signature: str | None = Query(default=None),
    db: Session = Depends(get_db),
):
    stored = db.get(StoredFile, file_id)
    if not stored or stored.status == "deleted":
        raise not_found("File")
    _require_file_read_access(db, current_user, stored)
    if expires is None or signature is None:
        raise AppHTTPException(
            403, "INVALID_DOWNLOAD_SIGNATURE", "Download URL signature is required."
        )
    validate_download_signature(file_id, expires, signature)
    storage = get_storage_backend()
    try:
        object_info = storage.stat_object(stored.bucket, stored.storage_key)
    except StorageObjectNotFound:
        raise not_found("StoredFileObject") from None
    except StorageError as exc:
        raise AppHTTPException(
            503,
            "STORAGE_READ_FAILED",
            "Failed to read stored file object.",
        ) from exc

    transfer = prepare_transfer_in_transaction(
        db,
        TransferSpec(
            direction="outbound",
            operation="download",
            actor_user_id=current_user.id,
            request_id=request.state.request_id,
            idempotency_key=request.state.request_id,
            file_id=stored.id,
            bucket=stored.bucket,
            storage_key=stored.storage_key,
            original_name=stored.original_name,
            expected_bytes=object_info.size_bytes,
        ),
    )
    write_audit_log(
        db,
        actor_user_id=current_user.id,
        action="files.download",
        resource_type="file",
        resource_id=stored.id,
        request=request,
    )
    db.commit()
    factory = session_factory_for(db)
    return StreamingResponse(
        settle_stream(
            factory,
            transfer.transfer_uid,
            storage.iter_file(stored.bucket, stored.storage_key),
        ),
        media_type=stored.content_type or "application/octet-stream",
        headers={
            **download_headers(stored.original_name),
            "Content-Length": str(object_info.size_bytes),
        },
    )


# ── bulk operations ──────────────────────────────────────────────────────────


@router.post("/bulk-delete", status_code=status.HTTP_204_NO_CONTENT)
def bulk_delete_files(
    request: Request,
    payload: BulkDeleteRequest,
    current_user: CurrentUser,
    db: Session = Depends(get_db),
):
    if not payload.file_ids:
        raise AppHTTPException(422, "INVALID_PARAMS", "file_ids must not be empty.")
    stored_list = list(
        db.scalars(
            select(StoredFile).where(
                StoredFile.id.in_(payload.file_ids), StoredFile.status != "deleted"
            )
        ).all()
    )
    for s in stored_list:
        _require_file_delete_access(db, current_user, s)
        _soft_delete_file_in_transaction(
            db,
            s,
            actor_user_id=current_user.id,
            request_id=request.state.request_id,
        )
        write_audit_log(
            db,
            actor_user_id=current_user.id,
            action="files.bulk_delete",
            resource_type="file",
            resource_id=s.id,
            request=request,
        )
    db.commit()
    return None


@router.post("/download-zip")
def download_zip_endpoint(
    request: Request,
    payload: ZipDownloadRequest,
    current_user: CurrentUser,
    db: Session = Depends(get_db),
):
    """Build a zip archive with selected files' DWG and/or DXF versions,
    stream it directly, and clean up the temp file on completion."""
    if not payload.file_ids:
        raise AppHTTPException(422, "INVALID_PARAMS", "file_ids must not be empty.")
    if not payload.formats:
        raise AppHTTPException(
            422, "INVALID_PARAMS", "formats must not be empty — choose at least dwg or dxf."
        )

    # Verify access
    stored_list = list(
        db.scalars(
            select(StoredFile).where(
                StoredFile.id.in_(payload.file_ids), StoredFile.status != "deleted"
            )
        ).all()
    )
    if len(stored_list) != len(payload.file_ids):
        raise not_found("File")
    for s in stored_list:
        _require_file_read_access(db, current_user, s)

    clean_name = sanitize_filename(payload.folder_name) or "图纸导出"
    prepared = build_zip_to_path(db, payload.file_ids, payload.formats, clean_name)
    try:
        transfer = prepare_transfer_in_transaction(
            db,
            TransferSpec(
                direction="outbound",
                operation="download_zip",
                actor_user_id=current_user.id,
                request_id=request.state.request_id,
                idempotency_key=request.state.request_id,
                batch_ref=request.state.request_id,
                original_name=prepared.filename,
                expected_bytes=prepared.size_bytes,
            ),
        )
        write_audit_log(
            db,
            actor_user_id=current_user.id,
            action="files.download_zip",
            resource_type="file",
            resource_id=0,
            after_json={
                "file_ids": payload.file_ids,
                "formats": payload.formats,
                "folder": clean_name,
            },
            request=request,
        )
        db.commit()
    except Exception:
        db.rollback()
        prepared.path.unlink(missing_ok=True)
        raise

    def _stream_and_cleanup():
        try:
            with prepared.path.open("rb") as f:
                while chunk := f.read(1024 * 1024):  # 1 MiB chunks
                    yield chunk
        finally:
            prepared.path.unlink(missing_ok=True)

    encoded_filename = quote(f"{clean_name}.zip")
    return StreamingResponse(
        settle_stream(
            session_factory_for(db),
            transfer.transfer_uid,
            _stream_and_cleanup(),
        ),
        media_type="application/zip",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            "Content-Length": str(prepared.size_bytes),
        },
    )
