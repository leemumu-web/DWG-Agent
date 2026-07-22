"""Streaming workbook-to-MySQL projection for Excel Final results."""

from __future__ import annotations

import re
from collections import Counter
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy.orm import Session

from app.modules.excel_processing.models import ExcelFinalComponent, ExcelFinalPart
from app.modules.excel_processing.schemas import (
    ComponentsImportStats,
    ExcelFinalPartType,
    PartsImportStats,
    QualityImportStats,
    WeightValidationStatus,
)

_PART_TYPE_ALIASES = {
    "零件": ExcelFinalPartType.PART,
    "板材": ExcelFinalPartType.PLATE,
    "扁钢": ExcelFinalPartType.FLAT_BAR,
    "BH": ExcelFinalPartType.BH,
    "BH腹": ExcelFinalPartType.BH_WEB,
    "BH翼": ExcelFinalPartType.BH_FLANGE,
    "BOX": ExcelFinalPartType.BOX,
    "BOX腹": ExcelFinalPartType.BOX_WEB,
    "BOX翼": ExcelFinalPartType.BOX_FLANGE,
    "BT": ExcelFinalPartType.BT,
    "BT腹": ExcelFinalPartType.BT_WEB,
    "BT翼": ExcelFinalPartType.BT_FLANGE,
    "工字钢": ExcelFinalPartType.I_BEAM,
    "H型钢": ExcelFinalPartType.H_BEAM,
    "T型钢": ExcelFinalPartType.T_BEAM,
    "槽钢": ExcelFinalPartType.CHANNEL,
    "角钢": ExcelFinalPartType.ANGLE,
    "方管": ExcelFinalPartType.SQUARE_TUBE,
    "钢管": ExcelFinalPartType.STEEL_PIPE,
    "方钢": ExcelFinalPartType.SQUARE_BAR,
    "高频焊": ExcelFinalPartType.HFW_PIPE,
    "W型钢": ExcelFinalPartType.W_BEAM,
    "圆钢": ExcelFinalPartType.ROUND_BAR,
    "螺纹钢": ExcelFinalPartType.REBAR,
    "螺栓": ExcelFinalPartType.BOLT,
    "螺母": ExcelFinalPartType.NUT,
    "螺套": ExcelFinalPartType.THREADED_SLEEVE,
    "TT": ExcelFinalPartType.TT,
    "未分类": ExcelFinalPartType.UNCLASSIFIED,
}
_PART_TYPE_VALUES = {item.value for item in ExcelFinalPartType}
_WEIGHT_STATUS_ALIASES = {
    "通过": WeightValidationStatus.OK,
    "警告": WeightValidationStatus.WARNING,
    "严重": WeightValidationStatus.SEVERE_WARNING,
}
_WEIGHT_STATUS_VALUES = {item.value for item in WeightValidationStatus}
_REPORT_LEVELS = {"信息", "警告", "严重", "致命"}
_MAX_REPORT_CATEGORIES = 50
_MAX_REPRESENTATIVE_MESSAGES = 10
_MAX_REPRESENTATIVE_MESSAGE_LENGTH = 500


def _canonical_header(value: object) -> str:
    text = "" if value is None else str(value).strip().replace(" ", "")
    return re.split(r"[（(]", text, maxsplit=1)[0]


def _number(value: object) -> Decimal | None:
    if value is None:
        return None
    try:
        result = Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return None
    if not result.is_finite():
        raise ValueError("Excel Final output contains non-finite numeric value")
    return result


def _has_negative(record: dict[str, Any], fields: tuple[str, ...]) -> bool:
    return any(record.get(field) is not None and record[field] < 0 for field in fields)


def _text(value: object) -> str | None:
    if value is None:
        return None
    normalized = str(value).strip()
    return normalized or None


def _part_type(value: object) -> str | None:
    normalized = _text(value)
    if normalized is None:
        return None
    if normalized in _PART_TYPE_VALUES:
        return normalized
    mapped = _PART_TYPE_ALIASES.get(normalized)
    if mapped is None:
        raise ValueError(f"Excel Final output contains unknown part type: {normalized}")
    return mapped.value


def _weight_status(value: object) -> str | None:
    normalized = _text(value)
    if normalized is None:
        return None
    if normalized in _WEIGHT_STATUS_VALUES:
        return normalized
    mapped = _WEIGHT_STATUS_ALIASES.get(normalized)
    if mapped is None:
        raise ValueError(f"Excel Final output contains unknown weight status: {normalized}")
    return mapped.value


def _value(row: list[object], column: int | None) -> object:
    return row[column - 1] if column is not None and column <= len(row) else None


def import_parts_to_db(
    db: Session,
    batch_id: int,
    output_path: Path,
) -> PartsImportStats:
    """Stream the canonical part-list sheet into `excel_final_parts`."""
    workbook = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    try:
        sheet_name = "整理表" if "整理表" in workbook.sheetnames else None
        if sheet_name is None:
            return {"parts_imported": 0, "error": "No 整理表 sheet found"}

        rows = workbook[sheet_name].iter_rows(values_only=True)
        columns: dict[str, int] = {}
        for column, value in enumerate(next(rows, ()), start=1):
            columns.setdefault(_canonical_header(value), column)

        def column(*names: str) -> int | None:
            return next((columns[name] for name in names if name in columns), None)

        seq_col = column("序号")
        component_no_col = column("构件编号")
        import_component_no_col = column("导入构件编号")
        component_qty_col = column("构件数")
        part_type_col = column("类型")
        team_col = column("班组")
        source_batch_col = column("批次")
        part_no_col = column("零件号", "零件编号")
        import_part_no_col = column("导入零件号")
        profile_spec_col = column("截面型材")
        spec_col = column("规格")
        width_col = column("宽度")
        length_col = column("长度")
        left_inset_col = column("左进")
        right_inset_col = column("右进")
        cut_length_col = column("下料长度")
        material_col = column("材质")
        original_qty_col = column("原数量")
        qty_col = column("数量")
        total_qty_col = column("总数")
        total_length_col = column("总长")
        density_col = column("比重")
        density_source_col = column("比重来源")
        theo_unit_weight_col = column("理单重")
        theo_total_weight_col = column("理总重")
        material_utilization_col = column("净材利用率")
        weight_validation_col = column("重量核验")
        net_unit_weight_col = column("单净重")
        net_total_weight_col = column("总净重")
        table_net_weight_col = column("表净重")
        gross_unit_weight_col = column("单毛重")
        gross_total_weight_col = column("总毛重")
        table_gross_weight_col = column("表毛重")
        surface_area_col = column("单表面积")
        total_surface_area_col = column("总表面积")

        required_columns = {
            "序号": seq_col,
            "构件编号": component_no_col,
            "零件号": part_no_col,
            "规格": spec_col,
            "长度": length_col,
            "材质": material_col,
            "数量": qty_col,
        }
        missing = [name for name, position in required_columns.items() if position is None]
        if missing:
            raise ValueError(
                "Excel Final output is missing required columns: " + ", ".join(missing)
            )

        parts: list[dict[str, Any]] = []
        for row_number, values in enumerate(rows, start=2):
            row = list(values)
            if all(value is None for value in row):
                continue
            part_no = _text(_value(row, part_no_col))
            component_no = _text(_value(row, component_no_col))
            summary_values = [
                _text(_value(row, position))
                for position in (component_no_col, part_no_col, profile_spec_col, spec_col)
                if position is not None
            ]
            if (
                not part_no
                or not component_no
                or any(value and value.startswith("合计") for value in summary_values)
            ):
                continue

            seq = _number(_value(row, seq_col))
            component_qty = _number(_value(row, component_qty_col))
            record = {
                    "batch_id": batch_id,
                    "seq": int(seq or 0) if seq_col else row_number - 1,
                    "import_component_no": _text(_value(row, import_component_no_col)),
                    "import_part_no": _text(_value(row, import_part_no_col)),
                    "source_batch": _text(_value(row, source_batch_col)),
                    "team": _text(_value(row, team_col)),
                    "original_qty": _number(_value(row, original_qty_col)),
                    "component_no": component_no,
                    "component_qty": (
                        int(component_qty) if component_qty is not None else None
                    ),
                    "part_type": _part_type(_value(row, part_type_col)),
                    "part_no": part_no,
                    "profile_spec": _text(_value(row, profile_spec_col)),
                    "spec": _text(_value(row, spec_col)),
                    "width": _number(_value(row, width_col)),
                    "length": _number(_value(row, length_col)),
                    "left_inset": _number(_value(row, left_inset_col)),
                    "right_inset": _number(_value(row, right_inset_col)),
                    "cut_length": _number(_value(row, cut_length_col)),
                    "material": _text(_value(row, material_col)),
                    "qty": _number(_value(row, qty_col)),
                    "total_qty": _number(_value(row, total_qty_col)),
                    "total_length": _number(_value(row, total_length_col)),
                    "density": _number(_value(row, density_col)),
                    "density_source": _text(_value(row, density_source_col)),
                    "theo_unit_weight": _number(_value(row, theo_unit_weight_col)),
                    "theo_total_weight": _number(_value(row, theo_total_weight_col)),
                    "material_utilization": _number(
                        _value(row, material_utilization_col)
                    ),
                    "weight_validation": _weight_status(
                        _value(row, weight_validation_col)
                    ),
                    "net_unit_weight": _number(_value(row, net_unit_weight_col)),
                    "net_total_weight": _number(_value(row, net_total_weight_col)),
                    "table_net_weight": _number(_value(row, table_net_weight_col)),
                    "gross_unit_weight": _number(_value(row, gross_unit_weight_col)),
                    "gross_total_weight": _number(_value(row, gross_total_weight_col)),
                    "table_gross_weight": _number(_value(row, table_gross_weight_col)),
                    "surface_area": _number(_value(row, surface_area_col)),
                    "total_surface_area": _number(_value(row, total_surface_area_col)),
                }
            if _has_negative(
                record,
                (
                    "component_qty",
                    "original_qty",
                    "width",
                    "length",
                    "left_inset",
                    "right_inset",
                    "cut_length",
                    "qty",
                    "total_qty",
                    "total_length",
                    "density",
                    "theo_unit_weight",
                    "theo_total_weight",
                    "material_utilization",
                    "net_unit_weight",
                    "net_total_weight",
                    "table_net_weight",
                    "gross_unit_weight",
                    "gross_total_weight",
                    "table_gross_weight",
                    "surface_area",
                    "total_surface_area",
                ),
            ):
                continue
            parts.append(record)
    finally:
        workbook.close()

    if parts:
        db.bulk_insert_mappings(ExcelFinalPart, parts)
        db.flush()
    return {"parts_imported": len(parts)}


def import_quality_report(output_path: Path) -> QualityImportStats:
    """Read the canonical quality ledger and return a bounded aggregate."""
    workbook = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    try:
        if "处理报告" not in workbook.sheetnames:
            return {
                "quality_status": "ok",
                "warning_count": 0,
                "severe_warning_count": 0,
                "report_summary": None,
            }
        rows = workbook["处理报告"].iter_rows(values_only=True)
        columns = {
            _canonical_header(value): index
            for index, value in enumerate(next(rows, ()))
        }
        missing = [name for name in ("级别", "类别", "说明") if name not in columns]
        if missing:
            raise ValueError(
                "Excel Final quality report is missing required columns: "
                + ", ".join(missing)
            )

        level_counts: Counter[str] = Counter()
        category_counts: Counter[str] = Counter()
        representative_messages: list[str] = []
        for values in rows:
            level = _text(values[columns["级别"]])
            category = _text(values[columns["类别"]])
            description = _text(values[columns["说明"]])
            if level is None and category is None and description is None:
                continue
            if level not in _REPORT_LEVELS:
                raise ValueError(f"Excel Final quality report has unknown level: {level}")
            level_counts[level] += 1
            category_counts[category or "未分类"] += 1
            if description and len(representative_messages) < _MAX_REPRESENTATIVE_MESSAGES:
                representative_messages.append(
                    description[:_MAX_REPRESENTATIVE_MESSAGE_LENGTH]
                )

        warning_count = level_counts["警告"]
        severe_warning_count = level_counts["严重"]
        quality_status = "ok"
        if severe_warning_count or level_counts["致命"]:
            quality_status = "severe_warning"
        elif warning_count:
            quality_status = "warning"
        return {
            "quality_status": quality_status,
            "warning_count": warning_count,
            "severe_warning_count": severe_warning_count,
            "report_summary": {
                "info_count": level_counts["信息"],
                "warning_count": warning_count,
                "severe_warning_count": severe_warning_count,
                "category_counts": dict(
                    category_counts.most_common(_MAX_REPORT_CATEGORIES)
                ),
                "representative_messages": representative_messages,
            },
        }
    finally:
        workbook.close()


def import_components_to_db(
    db: Session,
    batch_id: int,
    output_path: Path,
) -> ComponentsImportStats:
    """Stream the component-summary sheet into `excel_final_components`."""
    workbook = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    try:
        if "构件表" not in workbook.sheetnames:
            return {"components_imported": 0}
        rows = workbook["构件表"].iter_rows(values_only=True)
        headers = [str(value or "") for value in next(rows, ())]
        component_no_col = next(
            (index for index, header in enumerate(headers) if "构件编号" in header),
            None,
        )
        component_qty_col = next(
            (index for index, header in enumerate(headers) if "构件数" in header),
            None,
        )
        weight_col = next(
            (
                index
                for index, header in enumerate(headers)
                if "总净重" in header or "总重" in header
            ),
            None,
        )
        if component_no_col is None:
            raise ValueError("Excel Final component sheet is missing required column: 构件编号")

        components: list[dict[str, Any]] = []
        seen_component_numbers: set[str] = set()
        for values in rows:
            component_no = str(values[component_no_col] or "").strip()
            if not component_no or "合计" in component_no:
                continue
            qty = _number(values[component_qty_col]) if component_qty_col is not None else None
            weight = _number(values[weight_col]) if weight_col is not None else None
            if (qty is not None and qty < 0) or (weight is not None and weight < 0):
                continue
            if component_no in seen_component_numbers:
                raise ValueError(f"duplicate component identity: {component_no}")
            seen_component_numbers.add(component_no)
            components.append(
                {
                    "batch_id": batch_id,
                    "component_no": component_no,
                    "component_qty": int(qty) if qty is not None else None,
                    "total_weight": weight,
                }
            )
    finally:
        workbook.close()

    if components:
        db.bulk_insert_mappings(ExcelFinalComponent, components)
        db.flush()
    return {"components_imported": len(components)}


__all__ = ["import_components_to_db", "import_parts_to_db", "import_quality_report"]
