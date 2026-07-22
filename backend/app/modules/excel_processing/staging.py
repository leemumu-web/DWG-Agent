"""Source-file resolution, download and format detection for Excel Final."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from sqlalchemy.orm import Session

from app.modules.files.interface import (
    StoredFile,
    get_storage_backend,
    sanitize_filename,
)
from app.modules.jobs.interface import Job
from app.platform.config.constants import EXCEL_FILE_EXTENSIONS
from app.platform.storage.base import StorageObjectNotFound

_INIT_TABLE_SIGNATURE = [
    "零件号",
    "截面型材",
    "长度(mm)",
    "材质",
    "数量",
    "单重(kg)",
    "总重(kg)",
    "总面积(m2)",
    "备注",
]


def resolve_file_id(job: Job) -> int | None:
    """Read a positive file id from a Job's persisted parameters."""
    raw = (job.params_json or {}).get("file_id")
    if isinstance(raw, int) and raw > 0:
        return raw
    if isinstance(raw, str) and raw.strip().isdigit():
        return int(raw.strip())
    return None


def detect_source_format(filepath: Path) -> str:
    """Detect the Stage's `init` workbook input versus Tekla `tsv` fallback."""
    if filepath.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            workbook = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except Exception:
            # Malformed workbooks are handed to the Stage so its parser can
            # produce the canonical safe error mapping.
            return "tsv"
        try:
            if len(workbook.sheetnames) != 1:
                raise ValueError("Excel Final input must contain exactly one worksheet")
            if "初始表" in workbook.sheetnames:
                return "init"
            worksheet = workbook.worksheets[0]
            row2_cells = [
                str(worksheet.cell(row=2, column=column).value or "")
                for column in range(1, 10)
            ]
            match_count = sum(
                1
                for keyword in _INIT_TABLE_SIGNATURE
                if any(keyword in cell for cell in row2_cells)
            )
            if match_count >= 7:
                return "init"
        finally:
            workbook.close()
    return "tsv"


def stage_excel_source(
    db: Session,
    file_id: int,
    work_dir: Path,
) -> tuple[Path, StoredFile]:
    """Download one registered source object into the attempt work directory."""
    stored = db.get(StoredFile, file_id)
    if not stored or stored.status == "deleted":
        raise FileNotFoundError(f"File {file_id} not found or deleted")
    if stored.file_ext and stored.file_ext.lower() not in EXCEL_FILE_EXTENSIONS:
        raise ValueError(f"File {file_id} is not an Excel file (ext={stored.file_ext})")

    storage = get_storage_backend()
    destination = work_dir / sanitize_filename(stored.original_name)
    local = storage.local_path(stored.bucket, stored.storage_key)
    if local is not None:
        if not local.exists() or not local.is_file():
            raise StorageObjectNotFound(f"{stored.bucket}/{stored.storage_key}")
        destination.write_bytes(local.read_bytes())
    else:
        with destination.open("wb") as output:
            for chunk in storage.iter_file(stored.bucket, stored.storage_key):
                output.write(chunk)
    return destination, stored


__all__ = ["detect_source_format", "resolve_file_id", "stage_excel_source"]
