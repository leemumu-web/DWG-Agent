from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from tempfile import NamedTemporaryFile

from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.modules.files.interface import StoredFile
from app.modules.identity.interface import User
from app.modules.remnant_inventory.models import (
    Remnant,
    RemnantMaterial,
    RemnantPart,
)
from app.platform.time import as_business_time, business_now

EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
STATUS_LABELS = {
    "available": "可用",
    "reserved": "已预留",
    "used": "已领用",
    "archived": "已归档",
}
HEADERS = (
    "余料编号",
    "材质",
    "厚度(mm)",
    "项目编号一",
    "项目编号二",
    "库存位置",
    "备注一",
    "备注二",
    "零件编号",
    "库存状态",
    "原始图纸文件名",
    "导入人",
    "导入时间",
    "当前预留人",
    "预留时间",
    "领用人",
    "领用时间",
    "最后更新时间",
)
COLUMN_WIDTHS = (12, 14, 12, 32, 32, 20, 28, 28, 42, 12, 36, 16, 20, 16, 20, 16, 20, 20)


@dataclass(frozen=True)
class PreparedRemnantExport:
    path: Path
    filename: str
    row_count: int


class CleanupFileResponse(FileResponse):
    """File response that removes its temporary export on every exit path."""

    async def __call__(self, scope, receive, send) -> None:
        try:
            await super().__call__(scope, receive, send)
        finally:
            Path(self.path).unlink(missing_ok=True)


def _excel_datetime(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    return as_business_time(value).replace(tzinfo=None)


def _excel_safe_value(value: object) -> object:
    """把可能被 Excel 当作公式执行的字符串转义为纯文本。

    威胁模型（公式注入）：导出内容来自工人手工填写的项目编号/备注等字段，
    若以 ``=``、``+``、``-``、``@`` 开头，Excel/CSV 会把整格当作公式执行，
    可能读取/修改外部数据。这里强制前缀单引号转义为文本；用
    ``value.lstrip()[:1]`` 判断是为了防止前导空格绕过检查。这是安全措施，
    不要当成无意义的字符串变换删除。
    """
    if not isinstance(value, str):
        return value
    if value.lstrip()[:1] in {"=", "+", "-", "@"}:
        return f"'{value}"
    return value


def _indexed(rows) -> dict[int, object]:
    return {row.id: row for row in rows}


def build_remnant_export(db: Session) -> PreparedRemnantExport:
    remnants = list(db.scalars(select(Remnant).order_by(Remnant.id)).all())
    remnant_ids = [row.id for row in remnants]
    material_ids = {row.material_id for row in remnants}
    file_ids = {row.source_file_id for row in remnants}
    user_ids = {
        user_id
        for row in remnants
        for user_id in (row.imported_by, row.reserved_by, row.used_by)
        if user_id is not None
    }
    materials = (
        _indexed(
            db.scalars(select(RemnantMaterial).where(RemnantMaterial.id.in_(material_ids))).all()
        )
        if material_ids
        else {}
    )
    files = (
        _indexed(db.scalars(select(StoredFile).where(StoredFile.id.in_(file_ids))).all())
        if file_ids
        else {}
    )
    users = (
        _indexed(db.scalars(select(User).where(User.id.in_(user_ids))).all()) if user_ids else {}
    )
    parts: dict[int, list[str]] = {remnant_id: [] for remnant_id in remnant_ids}
    if remnant_ids:
        for remnant_id, part_no in db.execute(
            select(RemnantPart.remnant_id, RemnantPart.part_no)
            .where(RemnantPart.remnant_id.in_(remnant_ids))
            .order_by(RemnantPart.id)
        ):
            parts[remnant_id].append(part_no)

    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("全部余料")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:R{max(1, len(remnants) + 1)}"
    for index, width in enumerate(COLUMN_WIDTHS, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    header_cells = []
    for value in HEADERS:
        cell = WriteOnlyCell(sheet, value=value)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_cells.append(cell)
    sheet.append(header_cells)

    for row in remnants:
        material = materials.get(row.material_id)
        source = files.get(row.source_file_id)
        importer = users.get(row.imported_by)
        reserver = users.get(row.reserved_by) if row.status == "reserved" else None
        used_by = users.get(row.used_by)
        values = [
            row.id,
            material.code if material else None,
            float(row.thickness_mm),
            row.project_no,
            row.project_no_secondary,
            row.storage_location,
            row.remark_1,
            row.remark_2,
            "、".join(parts[row.id]),
            STATUS_LABELS.get(row.status, row.status),
            source.original_name if source else None,
            importer.real_name if importer else None,
            _excel_datetime(row.confirmed_at),
            reserver.real_name if reserver else None,
            _excel_datetime(row.reserved_at) if reserver else None,
            used_by.real_name if used_by else None,
            _excel_datetime(row.used_at),
            _excel_datetime(row.updated_at),
        ]
        cells = []
        for index, value in enumerate(values):
            cell = WriteOnlyCell(sheet, value=_excel_safe_value(value))
            cell.alignment = Alignment(vertical="top", wrap_text=index in (3, 4, 5, 6, 7, 8))
            if isinstance(value, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            cells.append(cell)
        sheet.append(cells)

    temporary = NamedTemporaryFile(suffix=".xlsx", delete=False)
    path = Path(temporary.name)
    temporary.close()
    try:
        workbook.save(path)
    except BaseException:
        path.unlink(missing_ok=True)
        raise
    filename = f"余料库_{business_now():%Y%m%d_%H%M%S}.xlsx"
    return PreparedRemnantExport(path=path, filename=filename, row_count=len(remnants))
