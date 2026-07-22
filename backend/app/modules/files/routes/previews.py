from __future__ import annotations

import hashlib
from urllib.parse import quote

from fastapi import APIRouter, Depends, Query, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.modules.cad_processing.interface import (
    MAX_DXF_SIZE_BYTES,
    get_or_create_dxf_preview,
    preview_batch_name,
    validate_dxf_source_size,
)
from app.modules.files.access import require_file_read_access
from app.modules.files.models import StoredFile
from app.modules.files.schemas import DxfPreviewBoundsRead, DxfPreviewRead
from app.modules.files.storage_transactions import (
    TransferSpec,
    prepare_transfer_in_transaction,
    session_factory_for,
    settle_stream,
)
from app.modules.identity.interface import CurrentUser
from app.modules.operations.audit.interface import write_audit_log
from app.platform.http.dependencies import get_db
from app.platform.http.envelopes import ok
from app.platform.http.exceptions import AppHTTPException, not_found
from app.platform.storage import factory as storage_factory
from app.platform.storage.base import (
    AbstractStorageBackend,
    StorageError,
    StorageObjectNotFound,
)

router = APIRouter()


def _column_letter(index: int) -> str:
    """Convert 0-based column index to Excel column letter(s): 0→A, 25→Z, 26→AA."""
    letters: list[str] = []
    n = index
    while True:
        n, rem = divmod(n, 26)
        letters.append(chr(ord("A") + rem))
        if n == 0:
            break
        n -= 1
    return "".join(reversed(letters))


@router.get("/{file_id}/excel-preview")
def get_excel_preview(
    file_id: int,
    request: Request,
    current_user: CurrentUser,
    sheet: str = Query("", description="Sheet name to preview (empty = first sheet)"),
    db: Session = Depends(get_db),
):
    """Read an Excel file from authoritative storage and return preview JSON."""
    stored = db.get(StoredFile, file_id)
    if not stored or stored.status == "deleted":
        raise not_found("File")
    require_file_read_access(db, current_user, stored, purpose="preview")

    if not stored.file_ext or stored.file_ext.lower() not in (".xlsx", ".xls"):
        raise AppHTTPException(
            415,
            "NOT_EXCEL",
            "Only .xlsx / .xls files can be previewed.",
        )

    # Read Excel bytes from storage
    storage = storage_factory.get_storage_backend()
    try:
        local_path = storage.local_path(stored.bucket, stored.storage_key)
        if local_path is not None:
            if not local_path.exists():
                raise StorageObjectNotFound(f"{stored.bucket}/{stored.storage_key}")
            excel_bytes = local_path.read_bytes()
        else:
            chunks: list[bytes] = []
            for chunk in storage.iter_file(stored.bucket, stored.storage_key):
                chunks.append(chunk)
            excel_bytes = b"".join(chunks)
    except StorageObjectNotFound:
        raise not_found("StoredFileObject") from None
    except StorageError as exc:
        raise AppHTTPException(
            503, "STORAGE_READ_FAILED", "Failed to read stored file object."
        ) from exc

    # Parse with openpyxl
    try:
        import io

        import openpyxl
    except ImportError as exc:
        raise AppHTTPException(
            503,
            "OPENPYXL_UNAVAILABLE",
            "openpyxl is not installed — cannot preview Excel files.",
        ) from exc

    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise AppHTTPException(
            415,
            "EXCEL_PARSE_ERROR",
            f"Failed to parse Excel file: {exc}",
        ) from exc

    sheet_names = wb.sheetnames
    if not sheet_names:
        raise AppHTTPException(415, "EXCEL_EMPTY", "Excel file has no sheets.")

    target_sheet = sheet.strip() if sheet.strip() else sheet_names[0]
    if target_sheet not in sheet_names:
        raise AppHTTPException(
            422,
            "SHEET_NOT_FOUND",
            f"Sheet '{target_sheet}' not found. Available: {', '.join(sheet_names)}",
        )

    ws = wb[target_sheet]
    rows_iter = ws.iter_rows(values_only=True)

    # First row is always the header, remaining rows are data.
    # Simple and predictable — no heuristic scoring that can accidentally
    # discard real data rows as "metadata".
    try:
        headers_raw: tuple[object, ...] = next(rows_iter)
    except StopIteration:
        headers_raw = ()

    # Build clean, unique headers. Empty cells become "Col A", "Col B", ...
    # Duplicates get a numeric suffix ("Name", "Name_2", "Name_3").
    seen: dict[str, int] = {}
    headers: list[str] = []
    for idx, h in enumerate(headers_raw or []):
        col_letter = _column_letter(idx)
        base = str(h).strip() if h is not None and str(h).strip() else f"Col {col_letter}"
        if base in seen:
            seen[base] += 1
            headers.append(f"{base}_{seen[base]}")
        else:
            seen[base] = 0
            headers.append(base)

    # If there are more data columns than headers, pad with column letters
    # (openpyxl iter_rows may return rows wider than the header row)
    _max_data_cols = 0

    data_rows: list[dict[str, object]] = []

    def _extract_row(row: tuple[object, ...]) -> dict[str, object]:
        """Convert an openpyxl row tuple into a column-keyed dict."""
        nonlocal _max_data_cols
        if len(row) > _max_data_cols:
            _max_data_cols = len(row)
        row_dict: dict[str, object] = {}
        for idx, val in enumerate(row):
            while idx >= len(headers):
                headers.append(f"Col {_column_letter(len(headers))}")
            col_name = headers[idx]
            if val is None:
                row_dict[col_name] = None
            elif isinstance(val, (int, float)):
                row_dict[col_name] = val
            else:
                row_dict[col_name] = str(val)
        return row_dict

    for row in rows_iter:
        data_rows.append(_extract_row(row))

    wb.close()

    result: dict = {
        "file": stored.original_name,
        "file_id": file_id,
        "sheets": sheet_names,
        "sheet": target_sheet,
        "headers": headers,
        "rows": data_rows,
        "total_rows": len(data_rows),
    }

    return ok(result, request.state.request_id)


def _read_dxf_preview_source(
    stored: StoredFile,
) -> tuple[bytes, AbstractStorageBackend]:
    """Read one DXF through the shared adapter with size and digest guards."""
    validate_dxf_source_size(stored.size_bytes)
    storage = storage_factory.get_storage_backend()
    payload = bytearray()
    digest = hashlib.sha256()
    try:
        for chunk in storage.iter_file(stored.bucket, stored.storage_key):
            payload.extend(chunk)
            if len(payload) > MAX_DXF_SIZE_BYTES:
                raise AppHTTPException(
                    413,
                    "DXF_TOO_LARGE",
                    f"DXF 文件超过在线预览上限 {MAX_DXF_SIZE_BYTES // (1024 * 1024)} MB。",
                )
            digest.update(chunk)
    except AppHTTPException:
        raise
    except StorageObjectNotFound:
        raise not_found("StoredFileObject") from None
    except StorageError as exc:
        raise AppHTTPException(
            503,
            "STORAGE_READ_FAILED",
            "Failed to read stored file object.",
        ) from exc

    if len(payload) != stored.size_bytes:
        raise AppHTTPException(
            409,
            "STORAGE_SIZE_MISMATCH",
            "DXF 对象大小与 MySQL 登记不一致，请先执行存储一致性扫描。",
        )
    if digest.hexdigest() != stored.sha256:
        raise AppHTTPException(
            409,
            "STORAGE_CHECKSUM_MISMATCH",
            "DXF 对象校验值与 MySQL 登记不一致，请先执行存储一致性扫描。",
        )
    return bytes(payload), storage


@router.get("/{file_id}/dxf-preview")
def get_dxf_preview(
    file_id: int,
    request: Request,
    current_user: CurrentUser,
    db: Session = Depends(get_db),
):
    """Generate or reuse a registered SVG preview for an accessible DXF."""
    stored = db.get(StoredFile, file_id)
    if not stored or stored.status == "deleted":
        raise not_found("File")
    require_file_read_access(db, current_user, stored, purpose="preview")
    if stored.file_ext.lower() != ".dxf":
        raise AppHTTPException(
            415,
            "NOT_DXF",
            "Only .dxf files can be previewed with this endpoint.",
        )

    payload, storage = _read_dxf_preview_source(stored)
    preview = get_or_create_dxf_preview(
        db,
        stored,
        payload,
        storage=storage,
        request_id=request.state.request_id,
    )
    preview_id = preview.preview_file.id
    assert preview_id is not None
    action = "files.dxf_preview_cache_hit" if preview.cached else "files.dxf_preview_generate"
    write_audit_log(
        db,
        actor_user_id=current_user.id,
        action=action,
        resource_type="file",
        resource_id=stored.id,
        after_json={"preview_file_id": preview_id},
        request=request,
    )
    db.commit()
    response = DxfPreviewRead(
        file_id=stored.id,
        file_name=stored.original_name,
        preview_file_id=preview_id,
        content_url=(f"/api/v1/files/{stored.id}/dxf-preview/content?preview_file_id={preview_id}"),
        content_type=preview.preview_file.content_type or "image/svg+xml",
        document_entities=preview.document_entities,
        modelspace_entities=preview.modelspace_entities,
        entity_counts=preview.entity_counts,
        layers=list(preview.layers),
        layer_colors=preview.layer_colors,
        bounds=DxfPreviewBoundsRead(
            min_x=preview.bounds.min_x,
            min_y=preview.bounds.min_y,
            max_x=preview.bounds.max_x,
            max_y=preview.bounds.max_y,
        ),
        cached=preview.cached,
    )
    return ok(response.model_dump(), request.state.request_id)


@router.get("/{file_id}/dxf-preview/content")
def get_dxf_preview_content(
    file_id: int,
    request: Request,
    current_user: CurrentUser,
    preview_file_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
):
    """Stream registered preview content after rechecking source-file access."""
    source = db.get(StoredFile, file_id)
    if not source or source.status == "deleted":
        raise not_found("File")
    require_file_read_access(db, current_user, source, purpose="preview")
    preview = db.get(StoredFile, preview_file_id)
    if (
        preview is None
        or preview.status == "deleted"
        or preview.file_ext != ".svg"
        or preview.batch_name != preview_batch_name(source)
    ):
        raise not_found("DxfPreview")

    storage = storage_factory.get_storage_backend()
    try:
        object_info = storage.stat_object(preview.bucket, preview.storage_key)
    except StorageObjectNotFound:
        raise not_found("DxfPreviewObject") from None
    except StorageError as exc:
        raise AppHTTPException(
            503,
            "STORAGE_READ_FAILED",
            "Failed to read preview object.",
        ) from exc
    if object_info.size_bytes != preview.size_bytes:
        raise AppHTTPException(
            409,
            "STORAGE_SIZE_MISMATCH",
            "DXF 预览对象大小与 MySQL 登记不一致。",
        )
    transfer = prepare_transfer_in_transaction(
        db,
        TransferSpec(
            direction="outbound",
            operation="preview",
            actor_user_id=current_user.id,
            request_id=request.state.request_id,
            idempotency_key=request.state.request_id,
            file_id=preview.id,
            batch_ref=preview.batch_name,
            bucket=preview.bucket,
            storage_key=preview.storage_key,
            original_name=preview.original_name,
            expected_bytes=object_info.size_bytes,
        ),
    )
    write_audit_log(
        db,
        actor_user_id=current_user.id,
        action="files.dxf_preview_view",
        resource_type="file",
        resource_id=source.id,
        after_json={"preview_file_id": preview.id},
        request=request,
    )
    db.commit()
    factory = session_factory_for(db)
    return StreamingResponse(
        settle_stream(
            factory,
            transfer.transfer_uid,
            storage.iter_file(preview.bucket, preview.storage_key),
        ),
        media_type="image/svg+xml",
        headers={
            "Content-Length": str(object_info.size_bytes),
            "Content-Disposition": (f"inline; filename*=UTF-8''{quote(preview.original_name)}"),
            "Cache-Control": "private, max-age=300",
            "X-Content-Type-Options": "nosniff",
            "Content-Security-Policy": ("sandbox; default-src 'none'; style-src 'unsafe-inline'"),
        },
    )
