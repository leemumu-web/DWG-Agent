"""Excel→Final Part-List 处理编排服务。

链路: 下载源 Excel 文件 (file_id) → 自动检测格式 (Tekla TSV / 初始表)
      → excel_final pipeline 处理 → 持久化输出 Excel → MySQL 入库
      → AnalysisResult 登记 → SSE 推送。

设计要点:
- 仿 dxf2excel_service.py 的状态机结构。
- 输入: 单个 Excel 文件 (file_id), 非批次。
- 输出: 处理后的 Excel + MySQL 结构化数据。
- 每步写 job_steps + publish_job_event。
"""

from __future__ import annotations

import logging
import re
import tempfile
from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path
from uuid import uuid4

import openpyxl
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.integrations.excel_final import (
    ExcelFinalUnavailableError,
    run_excel_final_pipeline,
)
from app.models.excel_final import ExcelFinalBatch, ExcelFinalComponent, ExcelFinalPart
from app.models.job import Job, JobStep
from app.models.result import AnalysisResult
from app.modules.files.interface import (
    StoredFile,
    get_storage_backend,
    sanitize_filename,
    save_bytes_as_file,
)
from app.platform.config.constants import (
    JOB_RUNNING,
    PIPELINE_EXCEL_FINAL,
    STEP_DOWNLOAD_EXCEL_SOURCE,
    STEP_IMPORT_PARTS_DB,
    STEP_PERSIST_EXCEL_FINAL,
    STEP_RUN_EXCEL_FINAL,
    TASK_EXCEL_FINAL,
)
from app.platform.config.settings import settings
from app.platform.database.session import SessionLocal
from app.platform.storage.base import StorageObjectNotFound
from app.services.job_events import make_event
from app.services.job_service import (
    claim_queued_job,
    commit_job_progress,
    complete_job_attempt,
    fail_job_attempt,
)

logger = logging.getLogger(__name__)

ERROR_CODE_EMPTY_INPUT = "EXCEL_FINAL_EMPTY_INPUT"
ERROR_CODE_PIPELINE_FAILED = "EXCEL_FINAL_PIPELINE_FAILED"
ERROR_CODE_NO_OUTPUT = "EXCEL_FINAL_NO_OUTPUT"
ERROR_CODE_UNAVAILABLE = "EXCEL_FINAL_UNAVAILABLE"
ERROR_CODE_STORAGE_FAILED = "EXCEL_FINAL_STORAGE_FAILED"
ERROR_CODE_NOT_EXCEL = "EXCEL_FINAL_NOT_EXCEL"
ERROR_CODE_DB_IMPORT_FAILED = "EXCEL_FINAL_DB_IMPORT_FAILED"

_EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_EXCEL_EXT = ".xlsx"
_ALGO_VERSION = "excel_final"

# 初始表 format signature headers (for auto-detection)
_INIT_TABLE_SIGNATURE = [
    "零件号",
    "截面型材",
    "长度(mm)",
    "材质",
    "数量",
    "单重(kg)",
    "总重(kg)",
    "总面积(m2)",
    "备注",
]


def _exception_message(exc: Exception) -> str:
    detail = getattr(exc, "detail", None)
    if isinstance(detail, dict):
        message = detail.get("message")
        if isinstance(message, str) and message:
            return message
    message = str(exc)
    return message or exc.__class__.__name__


def _mark_job_failed(
    db: Session,
    job_id: int,
    attempt: int,
    exc: Exception,
    error_code: str = ERROR_CODE_PIPELINE_FAILED,
) -> bool:
    """在 worker 当前事务内原子提交失败状态与待写步骤。"""
    try:
        db.execute(delete(ExcelFinalBatch).where(ExcelFinalBatch.job_id == job_id))
        return (
            fail_job_attempt(
                db,
                job_id,
                attempt=attempt,
                error_code=error_code,
                error_message=_exception_message(exc),
            )
            is not None
        )
    except Exception:
        db.rollback()
        logger.exception("Failed to mark excel_final job %s as failed", job_id)
        return False


def _add_step(
    db: Session,
    job_id: int,
    attempt: int,
    step_name: str,
    worker_name: str,
    status: str,
    *,
    input_json: dict | None = None,
    output_json: dict | None = None,
    error_message: str | None = None,
    started_at: datetime | None = None,
) -> None:
    db.add(
        JobStep(
            job_id=job_id,
            attempt=attempt,
            step_name=step_name,
            worker_name=worker_name,
            status=status,
            input_json=input_json,
            output_json=output_json,
            error_message=error_message,
            started_at=started_at,
            finished_at=datetime.now(UTC),
        )
    )


def _resolve_file_id(job: Job) -> int | None:
    """从 job.params_json 取 file_id。"""
    params = job.params_json or {}
    raw = params.get("file_id")
    if isinstance(raw, int) and raw > 0:
        return raw
    if isinstance(raw, str) and raw.strip().isdigit():
        return int(raw.strip())
    return None


def _detect_format(filepath: Path) -> str:
    """Auto-detect input format: 'init' for 初始表, 'tsv' for Tekla TSV."""
    if filepath.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            if "初始表" in wb.sheetnames:
                wb.close()
                return "init"
            ws = wb.worksheets[0]
            row2_cells = [str(ws.cell(row=2, column=c).value or "") for c in range(1, 10)]
            match_count = sum(
                1 for kw in _INIT_TABLE_SIGNATURE if any(kw in cell for cell in row2_cells)
            )
            wb.close()
            if match_count >= 7:
                return "init"
        except Exception:
            pass
    return "tsv"


def _stage_excel_source(
    db: Session,
    file_id: int,
    work_dir: Path,
) -> tuple[Path, StoredFile]:
    """下载源 Excel 文件到 work_dir。

    Returns (local_path, stored_file).
    """
    sfile = db.get(StoredFile, file_id)
    if not sfile or sfile.status == "deleted":
        raise FileNotFoundError(f"File {file_id} not found or deleted")

    if sfile.file_ext and sfile.file_ext.lower() not in (".xlsx", ".xls"):
        raise ValueError(f"File {file_id} is not an Excel file (ext={sfile.file_ext})")

    storage = get_storage_backend()
    dest = work_dir / sanitize_filename(sfile.original_name)

    local = storage.local_path(sfile.bucket, sfile.storage_key)
    if local is not None:
        if not local.exists() or not local.is_file():
            raise StorageObjectNotFound(f"{sfile.bucket}/{sfile.storage_key}")
        dest.write_bytes(local.read_bytes())
    else:
        with dest.open("wb") as out:
            for chunk in storage.iter_file(sfile.bucket, sfile.storage_key):
                out.write(chunk)

    return dest, sfile


def _import_parts_to_db(
    db: Session,
    batch_id: int,
    output_path: Path,
) -> dict:
    """从输出 Excel 的 整理表 sheet 读取零件数据，批量写入 excel_final_parts。

    兼容两种 sheet 命名：整理表（最终）和 整理表_拆板后（中间产物）。
    """
    wb = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    sheet_name = None
    for name in ("整理表", "整理表_拆板后"):
        if name in wb.sheetnames:
            sheet_name = name
            break
    if sheet_name is None:
        wb.close()
        return {"parts_imported": 0, "error": "No 整理表 sheet found"}

    ws = wb[sheet_name]
    # Match canonical names exactly after stripping unit suffixes. Substring
    # matching is unsafe here (for example, 长度 also matches 下料长度).
    def _header_name(value: object) -> str:
        text = "" if value is None else str(value).strip().replace(" ", "")
        return re.split(r"[（(]", text, maxsplit=1)[0]

    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, ())
    columns: dict[str, int] = {}
    for column, value in enumerate(header_row, start=1):
        columns.setdefault(_header_name(value), column)

    def _col(*names: str) -> int | None:
        return next((columns[name] for name in names if name in columns), None)

    seq_col = _col("序号")
    comp_no_col = _col("构件编号")
    comp_qty_col = _col("构件数")
    type_col = _col("类型")
    part_no_col = _col("零件号", "零件编号")
    profile_spec_col = _col("截面型材")
    spec_col = _col("规格")
    width_col = _col("宽度")
    len_col = _col("长度")
    left_col = _col("左进")
    right_col = _col("右进")
    cut_len_col = _col("下料长度")
    mat_col = _col("材质")
    qty_col = _col("数量")
    total_qty_col = _col("总数")
    total_len_col = _col("总长")
    density_col = _col("比重")
    theo_unit_col = _col("理单重")
    theo_total_col = _col("理总重")
    net_unit_col = _col("单净重")
    net_total_col = _col("总净重")
    table_net_col = _col("表净重")
    gross_unit_col = _col("单毛重")
    gross_total_col = _col("总毛重")
    table_gross_col = _col("表毛重")
    surface_col = _col("单表面积")
    total_surface_col = _col("总表面积")

    required_columns = {
        "序号": seq_col,
        "构件编号": comp_no_col,
        "零件号": part_no_col,
        "规格": spec_col,
        "长度": len_col,
        "材质": mat_col,
        "数量": qty_col,
    }
    missing_columns = [name for name, column in required_columns.items() if column is None]
    if missing_columns:
        wb.close()
        raise ValueError(
            "Excel Final output is missing required columns: " + ", ".join(missing_columns)
        )

    def _f(val) -> float | None:
        """Safe float conversion."""
        if val is None:
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    def _s(val) -> str | None:
        """Safe string conversion."""
        if val is None:
            return None
        s = str(val).strip()
        return s if s else None

    parts = []
    for r, row_vals_tuple in enumerate(rows, start=2):
        row_vals = list(row_vals_tuple)

        # Skip completely empty rows
        if all(v is None for v in row_vals):
            continue

        part_no = _s(row_vals[part_no_col - 1]) if part_no_col else None
        summary_values = [
            _s(row_vals[column - 1])
            for column in (comp_no_col, part_no_col, profile_spec_col, spec_col)
            if column is not None
        ]
        if not part_no or any(value and value.startswith("合计") for value in summary_values):
            continue

        parts.append(
            {
                "batch_id": batch_id,
                "seq": int(_f(row_vals[seq_col - 1]) or 0) if seq_col else r - 1,
                "component_no": _s(row_vals[comp_no_col - 1]) if comp_no_col else None,
                "component_qty": int(_f(row_vals[comp_qty_col - 1]) or 0)
                if comp_qty_col and row_vals[comp_qty_col - 1] is not None
                else None,
                "part_type": _s(row_vals[type_col - 1]) if type_col else None,
                "part_no": part_no,
                "profile_spec": _s(row_vals[profile_spec_col - 1]) if profile_spec_col else None,
                "spec": _s(row_vals[spec_col - 1]) if spec_col else None,
                "width": _f(row_vals[width_col - 1]) if width_col else None,
                "length": _f(row_vals[len_col - 1]) if len_col else None,
                "left_inset": _f(row_vals[left_col - 1]) if left_col else None,
                "right_inset": _f(row_vals[right_col - 1]) if right_col else None,
                "cut_length": _f(row_vals[cut_len_col - 1]) if cut_len_col else None,
                "material": _s(row_vals[mat_col - 1]) if mat_col else None,
                "qty": _f(row_vals[qty_col - 1]) if qty_col else None,
                "total_qty": _f(row_vals[total_qty_col - 1]) if total_qty_col else None,
                "total_length": _f(row_vals[total_len_col - 1]) if total_len_col else None,
                "density": _f(row_vals[density_col - 1]) if density_col else None,
                "theo_unit_weight": _f(row_vals[theo_unit_col - 1]) if theo_unit_col else None,
                "theo_total_weight": _f(row_vals[theo_total_col - 1]) if theo_total_col else None,
                "net_unit_weight": _f(row_vals[net_unit_col - 1]) if net_unit_col else None,
                "net_total_weight": _f(row_vals[net_total_col - 1]) if net_total_col else None,
                "table_net_weight": _f(row_vals[table_net_col - 1]) if table_net_col else None,
                "gross_unit_weight": _f(row_vals[gross_unit_col - 1]) if gross_unit_col else None,
                "gross_total_weight": _f(row_vals[gross_total_col - 1])
                if gross_total_col
                else None,
                "table_gross_weight": _f(row_vals[table_gross_col - 1])
                if table_gross_col
                else None,
                "surface_area": _f(row_vals[surface_col - 1]) if surface_col else None,
                "total_surface_area": _f(row_vals[total_surface_col - 1])
                if total_surface_col
                else None,
            }
        )

    wb.close()

    if parts:
        db.bulk_insert_mappings(ExcelFinalPart, parts)
        db.flush()

    return {"parts_imported": len(parts)}


def _import_components_to_db(
    db: Session,
    batch_id: int,
    output_path: Path,
) -> dict:
    """从输出 Excel 的 构件表 sheet 读取构件汇总，批量写入 excel_final_components。"""
    wb = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    if "构件表" not in wb.sheetnames:
        wb.close()
        return {"components_imported": 0}

    ws = wb["构件表"]
    rows = ws.iter_rows(values_only=True)
    headers = [str(value or "") for value in next(rows, ())]
    comp_no_col = next((i for i, h in enumerate(headers) if "构件编号" in h), None)
    comp_qty_col = next((i for i, h in enumerate(headers) if "构件数" in h), None)
    weight_col = next(
        (i for i, h in enumerate(headers) if "总净重" in h or "总重" in h),
        None,
    )
    if comp_no_col is None:
        wb.close()
        raise ValueError("Excel Final component sheet is missing required column: 构件编号")

    def _f(val) -> float | None:
        try:
            return float(val) if val is not None else None
        except (ValueError, TypeError):
            return None

    components = []
    for row_vals in rows:
        comp_no = str(row_vals[comp_no_col] or "").strip()
        if not comp_no or "合计" in comp_no:
            continue

        qty = _f(row_vals[comp_qty_col]) if comp_qty_col is not None else None
        weight = _f(row_vals[weight_col]) if weight_col is not None else None

        components.append(
            {
                "batch_id": batch_id,
                "component_no": comp_no,
                "component_qty": int(qty) if qty is not None else None,
                "total_weight": weight,
            }
        )

    wb.close()

    if components:
        db.bulk_insert_mappings(ExcelFinalComponent, components)
        db.flush()

    return {"components_imported": len(components)}


def _replace_batch_for_job(
    db: Session,
    *,
    job_id: int,
    file_id: int,
    source_type: str,
    source_name: str,
) -> ExcelFinalBatch:
    """Replace a batch left by an earlier failed/cancelled attempt."""
    existing = db.scalar(select(ExcelFinalBatch).where(ExcelFinalBatch.job_id == job_id))
    if existing is not None:
        db.delete(existing)
        db.flush()

    batch = ExcelFinalBatch(
        job_id=job_id,
        file_id=file_id,
        source_type=source_type,
        source_name=source_name,
    )
    db.add(batch)
    db.flush()
    return batch


def run_excel_final_processing(
    job_id: int,
    worker_name: str = "celery_excel_final",
    expected_attempt: int = 1,
) -> None:
    """Celery excel_final 队列任务体: Excel → 零件清单全链路。

    失败不抛（除导入错误外），通过 job.status/error_code 体现。
    """
    db = SessionLocal()
    try:
        job = claim_queued_job(
            db,
            job_id,
            expected_attempt=expected_attempt,
            pipeline=PIPELINE_EXCEL_FINAL,
            progress=5,
            message="开始处理 Excel",
        )
        if job is None:
            logger.info("ExcelFinal job %s was not claimable", job_id)
            return
        attempt = job.attempt

        file_id = _resolve_file_id(job)
        if file_id is None:
            _mark_job_failed(
                db,
                job_id,
                attempt,
                AppError("ExcelFinal job 缺少 params.file_id"),
                error_code=ERROR_CODE_EMPTY_INPUT,
            )
            return

        with tempfile.TemporaryDirectory(prefix=f"excel_final_job_{job_id}_") as work_dir_str:
            work_dir = Path(work_dir_str)

            # ---- 2. 下载源 Excel 文件 ----
            download_started = datetime.now(UTC)
            try:
                source_path, sfile = _stage_excel_source(db, file_id, work_dir)
            except FileNotFoundError:
                _mark_job_failed(
                    db,
                    job_id,
                    attempt,
                    AppError(f"文件 {file_id} 不存在"),
                    error_code=ERROR_CODE_EMPTY_INPUT,
                )
                return
            except ValueError as exc:
                _mark_job_failed(
                    db,
                    job_id,
                    attempt,
                    AppError(str(exc)),
                    error_code=ERROR_CODE_NOT_EXCEL,
                )
                return
            except StorageObjectNotFound:
                _mark_job_failed(
                    db,
                    job_id,
                    attempt,
                    AppError(f"文件 {file_id} 存储对象缺失"),
                    error_code=ERROR_CODE_STORAGE_FAILED,
                )
                return

            source_stats = {
                "file_id": file_id,
                "original_name": sfile.original_name,
                "size_bytes": sfile.size_bytes,
            }
            _add_step(
                db,
                job_id,
                attempt,
                STEP_DOWNLOAD_EXCEL_SOURCE,
                worker_name,
                "succeeded",
                input_json={"file_id": file_id},
                output_json=source_stats,
                started_at=download_started,
            )
            job = commit_job_progress(
                db,
                job_id,
                attempt=attempt,
                progress=15,
                event=make_event(
                    type_="progress",
                    progress=15,
                    step_name=STEP_DOWNLOAD_EXCEL_SOURCE,
                    status=JOB_RUNNING,
                    message=f"已下载: {sfile.original_name}",
                ),
            )
            if job is None:
                return

            # ---- 3. 自动检测格式 ----
            fmt = _detect_format(source_path)
            logger.info("Detected format for file_id=%s: %s", file_id, fmt)

            # ---- 4. 运行 excel_final pipeline ----
            pipeline_started = datetime.now(UTC)
            output_path = (
                work_dir / f"{sanitize_filename(sfile.original_name.rsplit('.', 1)[0])}_处理后.xlsx"
            )

            try:
                run_excel_final_pipeline(
                    source_path,
                    output_path,
                    source_format=fmt,
                )
            except ExcelFinalUnavailableError as exc:
                _add_step(
                    db,
                    job_id,
                    attempt,
                    STEP_RUN_EXCEL_FINAL,
                    worker_name,
                    "failed",
                    input_json={"file_id": file_id, "format": fmt},
                    error_message=f"Excel Final Stage 不可用: {exc}",
                    started_at=pipeline_started,
                )
                _mark_job_failed(
                    db,
                    job_id,
                    attempt,
                    AppError(f"Excel Final Stage 不可用: {exc}"),
                    error_code=ERROR_CODE_UNAVAILABLE,
                )
                return
            except Exception as exc:
                logger.exception("excel_final pipeline failed for job %s", job_id)
                _add_step(
                    db,
                    job_id,
                    attempt,
                    STEP_RUN_EXCEL_FINAL,
                    worker_name,
                    "failed",
                    input_json={"file_id": file_id, "format": fmt},
                    error_message=_exception_message(exc),
                    started_at=pipeline_started,
                )
                _mark_job_failed(
                    db,
                    job_id,
                    attempt,
                    AppError(f"流水线处理失败: {exc}"),
                    error_code=ERROR_CODE_PIPELINE_FAILED,
                )
                return

            _add_step(
                db,
                job_id,
                attempt,
                STEP_RUN_EXCEL_FINAL,
                worker_name,
                "succeeded",
                input_json={"file_id": file_id, "format": fmt},
                output_json={"output": str(output_path)},
                started_at=pipeline_started,
            )
            job = commit_job_progress(
                db,
                job_id,
                attempt=attempt,
                progress=60,
                event=make_event(
                    type_="progress",
                    progress=60,
                    step_name=STEP_RUN_EXCEL_FINAL,
                    status=JOB_RUNNING,
                    message=f"流水线完成 (format={fmt})",
                ),
            )
            if job is None:
                return

            # Check output
            if not output_path.is_file():
                _mark_job_failed(
                    db,
                    job_id,
                    attempt,
                    AppError("Excel 输出文件未生成"),
                    error_code=ERROR_CODE_NO_OUTPUT,
                )
                return

            # ---- 5. MySQL 入库 ----
            import_started = datetime.now(UTC)
            try:
                batch = _replace_batch_for_job(
                    db,
                    job_id=job.id,
                    file_id=file_id,
                    source_type=fmt,
                    source_name=sfile.original_name,
                )

                parts_stats = _import_parts_to_db(db, batch.id, output_path)
                comps_stats = _import_components_to_db(db, batch.id, output_path)

                batch.part_count = parts_stats.get("parts_imported", 0)
                batch.component_count = comps_stats.get("components_imported", 0)

                # Sum total net weight from parts
                if batch.part_count > 0:
                    from sqlalchemy import func as sa_func

                    total_net = db.scalar(
                        select(sa_func.sum(ExcelFinalPart.net_total_weight)).where(
                            ExcelFinalPart.batch_id == batch.id
                        )
                    )
                    batch.total_net_weight = float(total_net) if total_net else None

                db_stats = {
                    "batch_id": batch.id,
                    **parts_stats,
                    **comps_stats,
                }
            except Exception as exc:
                logger.exception("DB import failed for excel_final job %s", job_id)
                db.rollback()
                _mark_job_failed(
                    db,
                    job_id,
                    attempt,
                    AppError(f"MySQL 入库失败: {exc}"),
                    error_code=ERROR_CODE_DB_IMPORT_FAILED,
                )
                return

            _add_step(
                db,
                job_id,
                attempt,
                STEP_IMPORT_PARTS_DB,
                worker_name,
                "succeeded",
                input_json={"output_path": str(output_path)},
                output_json=db_stats,
                started_at=import_started,
            )
            job = commit_job_progress(
                db,
                job_id,
                attempt=attempt,
                progress=75,
                event=make_event(
                    type_="progress",
                    progress=75,
                    step_name=STEP_IMPORT_PARTS_DB,
                    status=JOB_RUNNING,
                    message=f"MySQL 入库完成: {batch.part_count} 个零件, {batch.component_count} 个构件",
                    stats=db_stats,
                ),
            )
            if job is None:
                return

            # ---- 6. 持久化输出 Excel ----
            persist_started = datetime.now(UTC)
            excel_bytes = output_path.read_bytes()
            storage_key = f"jobs/{job.id}/{uuid4().hex}{_EXCEL_EXT}"
            output_basename = sanitize_filename(sfile.original_name.rsplit(".", 1)[0])

            excel_file = save_bytes_as_file(
                db,
                bucket=settings.minio_bucket_reports,
                storage_key=storage_key,
                original_name=f"{output_basename}_处理后{_EXCEL_EXT}",
                file_ext=_EXCEL_EXT,
                content_type=_EXCEL_CONTENT_TYPE,
                payload=excel_bytes,
                uploaded_by=job.created_by,
            )

            result_payload = {
                "source": "excel_final",
                "job_id": job.id,
                "task_type": TASK_EXCEL_FINAL,
                "file_id": file_id,
                "format": fmt,
                "source_name": sfile.original_name,
                **db_stats,
                "excel_file_id": excel_file.id,
            }
            analysis = AnalysisResult(
                job_id=job.id,
                drawing_id=job.drawing_id,
                result_type=TASK_EXCEL_FINAL,
                result_json=result_payload,
                confidence=Decimal("1.0000"),
                result_file_id=excel_file.id,
                algorithm_version=_ALGO_VERSION,
                tool_version="excel_final",
                status="succeeded",
            )
            db.add(analysis)
            db.flush()

            _add_step(
                db,
                job_id,
                attempt,
                STEP_PERSIST_EXCEL_FINAL,
                worker_name,
                "succeeded",
                input_json={"excel_size": len(excel_bytes)},
                output_json={
                    "excel_file_id": excel_file.id,
                    "analysis_result_id": analysis.id,
                    **db_stats,
                },
                started_at=persist_started,
            )
            completed_job = complete_job_attempt(
                db,
                job_id,
                attempt=attempt,
                event=make_event(
                    type_="done",
                    status="succeeded",
                    progress=100,
                    step_name=STEP_PERSIST_EXCEL_FINAL,
                    message=f"处理完成: {batch.part_count} 个零件, {batch.component_count} 个构件",
                    excel_file_id=excel_file.id,
                    excel_name=f"{output_basename}_处理后{_EXCEL_EXT}",
                    part_count=batch.part_count,
                    component_count=batch.component_count,
                    **db_stats,
                ),
            )
            if completed_job is None:
                return

    except Exception as exc:
        db.rollback()
        if "attempt" in locals():
            _mark_job_failed(
                db,
                job_id,
                attempt,
                exc,
                error_code=ERROR_CODE_PIPELINE_FAILED,
            )
        logger.exception("ExcelFinal processing failed for job %s", job_id)
    finally:
        db.close()


class AppError(Exception):
    """excel_final_service 内部业务错误（消息友好）。"""
