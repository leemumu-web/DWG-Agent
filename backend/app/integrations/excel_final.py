from __future__ import annotations

import importlib.util
import json
import logging
import os
import re
import subprocess
import sys
from pathlib import Path
from typing import Literal

import pymysql
from openpyxl import load_workbook

from app.core.config import settings

logger = logging.getLogger(__name__)

SourceFormat = Literal["init", "tsv"]
_RESULT_PREFIX = "DWG_EXCEL_FINAL_RESULT="
_REQUIRED_STAGE_FILES = ("main.py", "pipeline.py", "handbook.py")


class ExcelFinalIntegrationError(RuntimeError):
    """Base error for the platform-to-Excel-Final process boundary."""


class ExcelFinalUnavailableError(ExcelFinalIntegrationError):
    """Raised when the standalone Stage or one of its dependencies is missing."""


class ExcelFinalProcessError(ExcelFinalIntegrationError):
    """Raised when the isolated Stage process exits unsuccessfully."""


def get_excel_final_stage_root() -> Path:
    """Resolve the tracked standalone Stage in source and container layouts."""
    integration_file = Path(__file__).resolve()
    candidates: list[Path] = []
    if settings.excel_final_stage_root is not None:
        candidates.append(settings.excel_final_stage_root.expanduser())
    candidates.extend(
        (
            integration_file.parents[3] / "Stages" / "excel_final",
            integration_file.parents[2] / "Stages" / "excel_final",
        )
    )

    checked: list[str] = []
    for candidate in candidates:
        resolved = candidate.resolve()
        if str(resolved) in checked:
            continue
        checked.append(str(resolved))
        if all((resolved / filename).is_file() for filename in _REQUIRED_STAGE_FILES):
            return resolved

    raise ExcelFinalUnavailableError(
        "Excel Final Stage is unavailable; checked: " + ", ".join(checked)
    )


def excel_final_dependencies_available() -> bool:
    """Return whether all third-party modules required by the Stage are installed."""
    return all(
        importlib.util.find_spec(module_name) is not None
        for module_name in ("numpy", "openpyxl", "pandas", "pymysql", "xlrd")
    )


def handbook_database_available() -> bool:
    """Probe the handbook database without importing the legacy Stage modules."""
    connection = None
    try:
        connection = pymysql.connect(**settings.handbook_database_config)
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
            return cursor.fetchone() == (1,)
    except pymysql.MySQLError as exc:
        logger.warning("Hardware handbook database health check failed: %s", exc)
        return False
    finally:
        if connection is not None:
            connection.close()


def run_excel_final_pipeline(
    source_path: Path,
    output_path: Path,
    *,
    source_format: SourceFormat,
) -> Path:
    """Run the legacy Stage in an isolated Python process and return its output."""
    if source_format not in ("init", "tsv"):
        raise ValueError(f"Unsupported Excel Final source format: {source_format}")
    if not source_path.is_file():
        raise ExcelFinalProcessError(f"Excel Final source file does not exist: {source_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    completed = _run_stage(
        "process",
        "--format",
        source_format,
        "--input",
        str(source_path.resolve()),
        "--output",
        str(output_path.resolve()),
    )
    _raise_for_failed_stage(completed)
    if not output_path.is_file():
        raise ExcelFinalProcessError("Excel Final Stage exited successfully without an output file")
    normalize_excel_final_output(output_path)
    return output_path


def normalize_excel_final_output(output_path: Path) -> int:
    """Repair deterministic column shifts emitted for fixed-width bolt rows.

    The legacy whitespace reader collapses an empty profile column. A row such
    as ``M20, C, 90, 2, 0, 0`` then arrives one column early. The adapter owns
    this compatibility rule so the standalone Stage remains isolated and both
    the downloadable workbook and MySQL projection expose the canonical fields.
    """
    workbook = load_workbook(output_path)
    if "整理表" not in workbook.sheetnames:
        workbook.close()
        return 0
    sheet = workbook["整理表"]
    columns = {
        _canonical_header(sheet.cell(row=1, column=column).value): column
        for column in range(1, sheet.max_column + 1)
    }
    required = {"零件号", "截面型材", "规格", "长度", "材质", "数量"}
    if not required.issubset(columns):
        workbook.close()
        return 0

    changed = 0
    for row in range(2, sheet.max_row + 1):
        part_no = _cell_text(sheet.cell(row=row, column=columns["零件号"]).value)
        profile = _cell_text(sheet.cell(row=row, column=columns["截面型材"]).value)
        spec = _cell_text(sheet.cell(row=row, column=columns["规格"]).value)
        material = _cell_text(sheet.cell(row=row, column=columns["材质"]).value)
        collapsed_qty = _cell_number(sheet.cell(row=row, column=columns["长度"]).value)
        zero_weight = _cell_number(sheet.cell(row=row, column=columns["数量"]).value)
        actual_length = _cell_number(material)

        is_collapsed_bolt = (
            bool(re.fullmatch(r"M\d+(?:\.\d+)?", part_no, flags=re.IGNORECASE))
            and bool(profile)
            and profile == spec
            and actual_length is not None
            and collapsed_qty is not None
            and zero_weight == 0
        )
        if not is_collapsed_bolt:
            continue

        sheet.cell(row=row, column=columns["截面型材"], value=part_no)
        sheet.cell(row=row, column=columns["规格"], value=part_no)
        sheet.cell(row=row, column=columns["长度"], value=_compact_number(actual_length))
        sheet.cell(row=row, column=columns["材质"], value=profile)
        sheet.cell(row=row, column=columns["数量"], value=_compact_number(collapsed_qty))
        if "类型" in columns:
            sheet.cell(row=row, column=columns["类型"], value="紧固件")
        if "总数" in columns:
            sheet.cell(row=row, column=columns["总数"], value=_compact_number(collapsed_qty))
        if "总长" in columns:
            sheet.cell(
                row=row,
                column=columns["总长"],
                value=_compact_number(actual_length * collapsed_qty),
            )
        changed += 1

    if changed:
        workbook.save(output_path)
    workbook.close()
    return changed


def lookup_excel_final_weight(spec: str) -> tuple[float | None, str]:
    """Look up a steel profile weight through the isolated legacy implementation."""
    completed = _run_stage("lookup", "--spec", spec)
    _raise_for_failed_stage(completed)
    result_line = next(
        (line for line in reversed(completed.stdout.splitlines()) if line.startswith(_RESULT_PREFIX)),
        None,
    )
    if result_line is None:
        raise ExcelFinalProcessError("Excel Final lookup returned no structured result")
    try:
        payload = json.loads(result_line.removeprefix(_RESULT_PREFIX))
        weight = payload["weight_kg_per_m"]
        source = str(payload["source"])
    except (KeyError, TypeError, ValueError, json.JSONDecodeError) as exc:
        raise ExcelFinalProcessError("Excel Final lookup returned an invalid result") from exc
    return (float(weight) if weight is not None else None, source)


def _run_stage(*arguments: str) -> subprocess.CompletedProcess[str]:
    stage_root = get_excel_final_stage_root()
    if not excel_final_dependencies_available():
        raise ExcelFinalUnavailableError("Excel Final Python dependencies are unavailable")

    command = [
        sys.executable,
        "-m",
        "app.integrations.excel_final_runner",
        *arguments,
        "--stage-root",
        str(stage_root),
    ]
    try:
        return subprocess.run(
            command,
            cwd=stage_root,
            env=_stage_environment(),
            capture_output=True,
            text=True,
            timeout=settings.excel_final_timeout_seconds,
            check=False,
        )
    except subprocess.TimeoutExpired as exc:
        raise ExcelFinalProcessError(
            f"Excel Final Stage exceeded {settings.excel_final_timeout_seconds} seconds"
        ) from exc
    except OSError as exc:
        raise ExcelFinalUnavailableError(f"Unable to start Excel Final Stage: {exc}") from exc


def _stage_environment() -> dict[str, str]:
    environment = os.environ.copy()
    database_config = settings.handbook_database_config
    environment.update(
        {
            "DWG_HANDBOOK_MYSQL_HOST": str(database_config["host"]),
            "DWG_HANDBOOK_MYSQL_PORT": str(database_config["port"]),
            "DWG_HANDBOOK_MYSQL_DATABASE": str(database_config["database"]),
            "DWG_HANDBOOK_MYSQL_USER": str(database_config["user"]),
            "DWG_HANDBOOK_MYSQL_PASSWORD": str(database_config["password"]),
        }
    )
    return environment


def _canonical_header(value: object) -> str:
    text = _cell_text(value).replace(" ", "")
    return re.split(r"[（(]", text, maxsplit=1)[0]


def _cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _cell_number(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _compact_number(value: float) -> int | float:
    return int(value) if value.is_integer() else value


def _raise_for_failed_stage(completed: subprocess.CompletedProcess[str]) -> None:
    if completed.returncode == 0:
        return
    details = (completed.stderr or completed.stdout or "unknown error").strip()
    logger.error(
        "Excel Final Stage exited with status %s: %s",
        completed.returncode,
        details[-16000:],
    )
    if any(
        marker in details
        for marker in (
            "ParserError",
            "BadZipFile",
            "Excel file format cannot be determined",
            "Unsupported format",
        )
    ):
        message = "The input file could not be parsed as a supported Excel Final format."
    else:
        message = "Excel Final Stage failed while processing the input."
    raise ExcelFinalProcessError(message)
