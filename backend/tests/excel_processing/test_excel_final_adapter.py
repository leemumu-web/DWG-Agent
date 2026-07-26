from __future__ import annotations

import hashlib
import json
import runpy
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.modules.excel_processing import stage_adapter as excel_final
from app.modules.excel_processing import stage_runner
from app.modules.excel_processing.staging import stage_excel_source
from tests.support.paths import BACKEND_ROOT


def _protocol_line(payload: dict[str, object]) -> str:
    return "DWG_EXCEL_FINAL_RESULT=" + json.dumps(payload, ensure_ascii=False)


def _error_protocol_line(failure: dict[str, object]) -> str:
    return "DWG_EXCEL_FINAL_ERROR=" + json.dumps(
        {
            "protocol_version": 1,
            "operation": "inspect",
            "failure": failure,
        },
        ensure_ascii=False,
    )


def _inspection_payload() -> dict[str, object]:
    return {
        "protocol_version": 1,
        "operation": "inspect",
        "input_contract_version": 1,
        "source_format": "standard_workbook",
        "sheet_name": "原表",
        "header_row": 6,
        "part_count": 12,
        "component_count": 3,
        "warnings": [],
        "ignored_sheets": [],
    }


def _input_failure_payload() -> dict[str, object]:
    return {
        "code": "EXCEL_INPUT_REQUIRED_COLUMNS_MISSING",
        "message": "表格缺少 Excel 第一阶段所需列。",
        "action": "请在正式标题行中补充：零件号。",
        "contract_version": 1,
        "issues": [
            {
                "sheet": "原表",
                "row": 6,
                "column": None,
                "field": "零件号",
                "value": None,
                "reason": "required_column_missing",
            }
        ],
        "sheets": ["原表"],
        "meta": {
            "missing_fields": ["零件号"],
            "issue_count": 1,
            "issues_truncated": False,
            "sheet_count": 1,
            "sheets_truncated": False,
        },
    }


def _process_payload(output_path: Path) -> dict[str, object]:
    return {
        "protocol_version": 1,
        "operation": "process",
        "output_path": str(output_path.resolve()),
        "quality_status": "warning",
        "warning_count": 1,
        "severe_warning_count": 0,
        "report_summary": {
            "info_count": 0,
            "warning_count": 1,
            "severe_warning_count": 0,
            "category_counts": {"手册查无": 1},
            "representative_messages": ["规格查无"],
        },
    }


def _inspect_command(source_path: Path) -> list[str]:
    return [
        sys.executable,
        "-m",
        "app.modules.excel_processing.stage_runner",
        "inspect",
        "--input",
        str(source_path),
        "--stage-root",
        str(excel_final.get_excel_final_stage_root()),
    ]


def _runner_environment_without_handbook() -> dict[str, str]:
    return {
        key: value
        for key, value in excel_final._stage_environment().items()
        if not key.startswith("DWG_HANDBOOK_MYSQL_")
    }


def _sentinel_payload(stdout: str, prefix: str) -> dict[str, object]:
    lines = [line for line in stdout.splitlines() if line.startswith(prefix)]
    assert len(lines) == 1
    payload = json.loads(lines[0].removeprefix(prefix))
    assert isinstance(payload, dict)
    return payload


def test_stage_runner_inspects_canonical_input_without_handbook_database(
    tmp_path: Path,
):
    source = tmp_path / "canonical.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原表"
    sheet.append(["构件编号", "零件号", "规格", "长度(mm)", "材质", "数量"])
    sheet.append(["C1", None, "BH500*300*12*20", 1000, "Q355B", 1])
    sheet.append([None, "P1", "PL10*200", 100, "Q355B", 2])
    workbook.save(source)
    workbook.close()

    completed = subprocess.run(
        _inspect_command(source),
        cwd=excel_final.get_excel_final_stage_root(),
        env=_runner_environment_without_handbook(),
        text=True,
        capture_output=True,
        check=False,
    )

    assert completed.returncode == 0, completed.stderr
    payload = _sentinel_payload(completed.stdout, "DWG_EXCEL_FINAL_RESULT=")
    assert payload == {
        "protocol_version": 1,
        "operation": "inspect",
        "input_contract_version": 1,
        "source_format": "standard_workbook",
        "sheet_name": "原表",
        "header_row": 1,
        "part_count": 1,
        "component_count": 1,
        "warnings": [],
        "ignored_sheets": [],
    }
    assert completed.stderr == ""


def test_stage_runner_emits_one_safe_structured_input_error(tmp_path: Path):
    source = tmp_path / "private-invalid-source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "不合格原表"
    sheet.append(["构件编号", "规格", "长度(mm)", "材质", "数量"])
    workbook.save(source)
    workbook.close()

    completed = subprocess.run(
        _inspect_command(source),
        cwd=excel_final.get_excel_final_stage_root(),
        env=_runner_environment_without_handbook(),
        text=True,
        capture_output=True,
        check=False,
    )

    assert completed.returncode == 2
    payload = _sentinel_payload(completed.stdout, "DWG_EXCEL_FINAL_ERROR=")
    assert payload["protocol_version"] == 1
    assert payload["operation"] == "inspect"
    failure = payload["failure"]
    assert failure["code"] == "EXCEL_INPUT_COMPONENT_ONLY"
    assert failure["message"]
    assert failure["action"]
    assert failure["contract_version"] == 1
    assert len(failure["issues"]) <= 20
    assert str(source.resolve()) not in completed.stdout
    assert "Traceback" not in completed.stdout
    assert "Traceback" not in completed.stderr


def test_adapter_parses_strict_inspection_result(monkeypatch, tmp_path: Path):
    source = tmp_path / "source.xlsx"
    source.write_bytes(b"source")
    captured: tuple[str, ...] | None = None

    def fake_run(*arguments: str):
        nonlocal captured
        captured = arguments
        return subprocess.CompletedProcess(
            arguments,
            0,
            stdout=_protocol_line(_inspection_payload()),
            stderr="",
        )

    monkeypatch.setattr(excel_final, "_run_stage", fake_run)

    result = excel_final.inspect_excel_stage1_path(source)

    assert captured == ("inspect", "--input", str(source.resolve()))
    assert result.protocol_version == 1
    assert result.input_contract_version == 1
    assert result.source_format == "standard_workbook"
    assert result.sheet_name == "原表"
    assert result.header_row == 6
    assert result.part_count == 12
    assert result.component_count == 3


def test_adapter_restores_structured_input_failure(monkeypatch, tmp_path: Path):
    source = tmp_path / "source.xlsx"
    source.write_bytes(b"source")
    failure_payload = _input_failure_payload()

    monkeypatch.setattr(
        excel_final,
        "_run_stage",
        lambda *arguments: subprocess.CompletedProcess(
            arguments,
            2,
            stdout=_error_protocol_line(failure_payload),
            stderr="",
        ),
    )

    with pytest.raises(excel_final.ExcelFinalInputError) as caught:
        excel_final.inspect_excel_stage1_path(source)

    assert caught.value.failure.code == "EXCEL_INPUT_REQUIRED_COLUMNS_MISSING"
    assert caught.value.failure.message == failure_payload["message"]
    assert caught.value.failure.action == failure_payload["action"]
    assert caught.value.failure.issues[0].sheet == "原表"
    assert caught.value.failure.issues[0].row == 6
    assert caught.value.failure.as_dict() == failure_payload


@pytest.mark.parametrize(
    "mutation",
    [
        lambda payload: payload.update(protocol_version=2),
        lambda payload: payload.update(operation="process"),
        lambda payload: payload["failure"].update(traceback="/private/server.py"),
        lambda payload: payload["failure"].update(issues=[{}] * 21),
        lambda payload: payload["failure"].update(sheets=["sheet"] * 11),
    ],
)
def test_adapter_rejects_malformed_or_unbounded_error_protocol(
    monkeypatch,
    tmp_path: Path,
    mutation,
):
    source = tmp_path / "source.xlsx"
    source.write_bytes(b"source")
    payload = {
        "protocol_version": 1,
        "operation": "inspect",
        "failure": _input_failure_payload(),
    }
    mutation(payload)
    monkeypatch.setattr(
        excel_final,
        "_run_stage",
        lambda *arguments: subprocess.CompletedProcess(
            arguments,
            2,
            stdout=(
                "DWG_EXCEL_FINAL_ERROR="
                + json.dumps(payload, ensure_ascii=False)
            ),
            stderr="",
        ),
    )

    with pytest.raises(
        excel_final.ExcelFinalProcessError,
        match="invalid inspect error",
    ):
        excel_final.inspect_excel_stage1_path(source)


def test_byte_inspection_rejects_changed_object_without_starting_stage(
    monkeypatch,
):
    payload = b"current object bytes"
    expected_sha256 = hashlib.sha256(b"previous object bytes").hexdigest()

    monkeypatch.setattr(
        excel_final,
        "_run_stage",
        lambda *_arguments: pytest.fail("checksum mismatch must not start the Stage"),
    )

    with pytest.raises(excel_final.ExcelFinalInputError) as caught:
        excel_final.inspect_excel_stage1_bytes(
            file_name="source.xlsx",
            payload=payload,
            expected_sha256=expected_sha256,
        )

    assert caught.value.failure.code == "EXCEL_INPUT_OBJECT_CHANGED"
    assert caught.value.failure.issues == ()
    assert caught.value.failure.meta["expected_sha256"] == expected_sha256
    assert caught.value.failure.meta["actual_sha256"] == hashlib.sha256(payload).hexdigest()


def test_byte_inspection_removes_private_temporary_file(monkeypatch):
    payload = b"valid object bytes"
    observed_path: Path | None = None

    def fake_run(*arguments: str):
        nonlocal observed_path
        observed_path = Path(arguments[arguments.index("--input") + 1])
        assert observed_path.is_file()
        assert observed_path.read_bytes() == payload
        return subprocess.CompletedProcess(
            arguments,
            0,
            stdout=_protocol_line(_inspection_payload()),
            stderr="",
        )

    monkeypatch.setattr(excel_final, "_run_stage", fake_run)

    result = excel_final.inspect_excel_stage1_bytes(
        file_name="../../source.XLSX",
        payload=payload,
    )

    assert result.part_count == 12
    assert observed_path is not None
    assert observed_path.suffix == ".xlsx"
    assert not observed_path.exists()


def test_excel_final_stage_root_resolves_tracked_standalone_layout():
    stage_root = excel_final.get_excel_final_stage_root()

    assert "config.py" in excel_final._REQUIRED_STAGE_FILES
    assert "material_routing.py" in excel_final._REQUIRED_STAGE_FILES
    assert stage_runner._REQUIRED_STAGE_FILES == excel_final._REQUIRED_STAGE_FILES
    assert (stage_root / "main.py").is_file()
    assert (stage_root / "pipeline.py").is_file()
    assert (stage_root / "handbook.py").is_file()
    assert (stage_root / "material_routing.py").is_file()


def test_excel_final_stage_root_accepts_protected_bytecode_layout(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
):
    stage_root = tmp_path / "protected-excel-final"
    stage_root.mkdir()
    for source_name in excel_final._REQUIRED_STAGE_FILES:
        (stage_root / Path(source_name).with_suffix(".pyc")).write_bytes(b"protected")

    monkeypatch.setattr(excel_final.settings, "excel_final_stage_root", stage_root)
    monkeypatch.setattr(sys, "path", list(sys.path))

    assert excel_final.get_excel_final_stage_root() == stage_root.resolve()
    assert stage_runner._configure_stage_imports(stage_root) == stage_root.resolve()


def test_excel_final_stage_root_failure_does_not_disclose_checked_paths(
    monkeypatch,
    tmp_path: Path,
):
    secret_root = tmp_path / "private-deployment-root"
    monkeypatch.setattr(excel_final.settings, "excel_final_stage_root", secret_root)
    monkeypatch.setattr(excel_final, "_REQUIRED_STAGE_FILES", ("never-present.codex",))

    with pytest.raises(excel_final.ExcelFinalUnavailableError) as raised:
        excel_final.get_excel_final_stage_root()

    assert str(secret_root) not in str(raised.value)
    assert "checked:" not in str(raised.value)


def test_handbook_health_log_does_not_disclose_mysql_host(monkeypatch, caplog):
    def fail_connect(**_kwargs):
        raise excel_final.pymysql.OperationalError(
            2003, "Can't connect to MySQL server on 'secret-db.internal'"
        )

    monkeypatch.setattr(excel_final.pymysql, "connect", fail_connect)

    assert excel_final.handbook_database_available() is False
    assert "secret-db.internal" not in caplog.text


def test_excel_final_dependency_probe_includes_legacy_xls_reader(monkeypatch):
    checked: list[str] = []

    def available(module_name: str):
        checked.append(module_name)
        return object()

    monkeypatch.setattr(excel_final.importlib.util, "find_spec", available)

    assert excel_final.excel_final_dependencies_available() is True
    assert "xlrd" in checked


def test_backend_and_stage_share_one_d_series_material_routing_contract():
    stage_rules = runpy.run_path(
        str(excel_final.get_excel_final_stage_root() / "material_routing.py")
    )

    assert excel_final._D_MATERIAL_CATEGORY_BY_PREFIX == {
        "HRB": "rebar",
        "HPB": "round_bar",
        "Q235B": "round_bar",
        "Q355B": "round_bar",
    }
    assert (
        excel_final._D_MATERIAL_CATEGORY_BY_PREFIX
        == stage_rules["D_MATERIAL_CATEGORY_BY_PREFIX"]
    )


def test_excel_final_fixed_width_probe_rejects_unrecognized_text():
    script = """
from pathlib import Path
from tempfile import TemporaryDirectory
import reader
from input_contract import InputContractError

with TemporaryDirectory() as directory:
    source = Path(directory) / 'not-tekla.xls'
    source.write_text('ordinary text without a production header', encoding='utf-8')
    try:
        reader._decode_fixed_text(source)
    except InputContractError as exc:
        assert exc.failure.code == 'EXCEL_INPUT_TEXT_UNRECOGNIZED'
        assert '重新导出' in exc.failure.action
        assert str(source.resolve()) not in str(exc.failure.as_dict())
    else:
        raise AssertionError('unrecognized text was accepted as fixed-width Tekla')
"""
    completed = subprocess.run(
        [sys.executable, "-c", script],
        cwd=excel_final.get_excel_final_stage_root(),
        text=True,
        capture_output=True,
        check=False,
    )

    assert completed.returncode == 0, completed.stderr


def test_excel_final_pipeline_runs_in_isolated_subprocess(monkeypatch, tmp_path: Path):
    source_path = tmp_path / "source.xls"
    output_path = tmp_path / "result.xlsx"
    source_path.write_bytes(b"input")
    captured: dict[str, object] = {}

    def fake_run(command, **kwargs):
        captured["command"] = command
        captured.update(kwargs)
        workbook = Workbook()
        workbook.active.title = "整理表"
        stage_output_path = Path(command[command.index("--output") + 1])
        workbook.save(stage_output_path)
        internal_output_path = Path(
            command[command.index("--internal-output") + 1]
        )
        workbook.save(internal_output_path)
        return subprocess.CompletedProcess(
            command,
            0,
            stdout=_protocol_line(_process_payload(stage_output_path)),
            stderr="",
        )

    monkeypatch.setattr(excel_final.subprocess, "run", fake_run)
    monkeypatch.setattr(excel_final.settings, "handbook_mysql_password", "not-on-command-line")

    result = excel_final.run_excel_final_pipeline(
        source_path,
        output_path,
    )

    command = captured["command"]
    assert command[:3] == [
        sys.executable,
        "-m",
        "app.modules.excel_processing.stage_runner",
    ]
    assert "not-on-command-line" not in command
    assert "--format" not in command
    assert captured["cwd"] == excel_final.get_excel_final_stage_root()
    assert captured["env"]["DWG_HANDBOOK_MYSQL_PASSWORD"] == "not-on-command-line"
    assert result.output_path == output_path.resolve()
    assert result.internal_output_path == (
        output_path.with_name(f".{output_path.stem}.internal.xlsx").resolve()
    )
    assert result.internal_output_path.is_file()
    assert result.protocol_version == 1
    assert result.quality_status == "warning"
    assert result.warning_count == 1
    assert result.report_summary["category_counts"] == {
        "手册查无": 1,
    }


def test_excel_final_pipeline_does_not_accept_preexisting_output_as_new_result(
    monkeypatch,
    tmp_path: Path,
):
    source_path = tmp_path / "source.xls"
    output_path = tmp_path / "result.xlsx"
    source_path.write_bytes(b"input")
    output_path.write_bytes(b"previous successful result")

    def fake_run(command, **kwargs):
        stage_output_path = Path(command[command.index("--output") + 1])
        internal_output_path = Path(
            command[command.index("--internal-output") + 1]
        )
        internal_output_path.touch()
        return subprocess.CompletedProcess(
            command,
            0,
            stdout=_protocol_line(_process_payload(stage_output_path)),
            stderr="",
        )

    monkeypatch.setattr(excel_final.subprocess, "run", fake_run)

    with pytest.raises(
        excel_final.ExcelFinalProcessError,
        match="without an output file",
    ):
        excel_final.run_excel_final_pipeline(source_path, output_path)

    assert output_path.read_bytes() == b"previous successful result"


def test_excel_final_pipeline_rejects_non_xlsx_output_before_stage(
    monkeypatch, tmp_path: Path
):
    source_path = tmp_path / "source.xlsm"
    output_path = tmp_path / "result.xlsm"
    source_path.write_bytes(b"input")

    def unexpected_run(*_args, **_kwargs):
        pytest.fail("invalid output suffix must not start the Stage")

    monkeypatch.setattr(excel_final.subprocess, "run", unexpected_run)

    with pytest.raises(ValueError, match=r"\.xlsx"):
        excel_final.run_excel_final_pipeline(
            source_path,
            output_path,
        )


def test_excel_final_runner_is_importable_from_stage_working_directory():
    completed = subprocess.run(
        [
            sys.executable,
            "-m",
            "app.modules.excel_processing.stage_runner",
            "--help",
        ],
        cwd=excel_final.get_excel_final_stage_root(),
        env=excel_final._stage_environment(),
        text=True,
        capture_output=True,
        check=False,
    )

    assert completed.returncode == 0, completed.stderr
    assert "Isolated Excel Final Stage runner" in completed.stdout


def test_excel_final_pipeline_logs_internal_failure_but_raises_safe_message(
    monkeypatch,
    tmp_path: Path,
    caplog,
):
    source_path = tmp_path / "source.xls"
    output_path = tmp_path / "result.xlsx"
    source_path.write_bytes(b"input")
    internal_output_path: Path | None = None

    def fake_run(command, **kwargs):
        nonlocal internal_output_path
        internal_output_path = Path(
            command[command.index("--internal-output") + 1]
        )
        internal_output_path.touch()
        return subprocess.CompletedProcess(command, 1, stdout="", stderr="pipeline exploded")

    monkeypatch.setattr(excel_final.subprocess, "run", fake_run)

    with pytest.raises(excel_final.ExcelFinalProcessError) as failure:
        excel_final.run_excel_final_pipeline(
            source_path,
            output_path,
        )

    assert str(failure.value) == "Excel Final Stage failed while processing the input."
    assert "pipeline exploded" not in str(failure.value)
    assert "pipeline exploded" not in caplog.text
    assert "processing failure" in caplog.text
    assert internal_output_path is not None
    assert not internal_output_path.exists()


@pytest.mark.parametrize(
    ("mutation", "match"),
    [
        (lambda payload: payload.pop("quality_status"), "invalid process result"),
        (lambda payload: payload.update(protocol_version=2), "invalid process result"),
        (lambda payload: payload.update(quality_status="mystery"), "invalid process result"),
        (lambda payload: payload.update(warning_count=-1), "invalid process result"),
        (
            lambda payload: payload.update(traceback="mysql://user:secret@internal/db"),
            "invalid process result",
        ),
    ],
)
def test_excel_final_process_protocol_rejects_malformed_or_extra_fields(
    monkeypatch,
    tmp_path: Path,
    mutation,
    match: str,
):
    source_path = tmp_path / "source.xlsx"
    output_path = tmp_path / "result.xlsx"
    source_path.write_bytes(b"input")
    payload = _process_payload(output_path)
    mutation(payload)

    def fake_run(command, **kwargs):
        output_path.touch()
        return subprocess.CompletedProcess(
            command,
            0,
            stdout=_protocol_line(payload),
            stderr="",
        )

    monkeypatch.setattr(excel_final.subprocess, "run", fake_run)

    with pytest.raises(excel_final.ExcelFinalProcessError, match=match):
        excel_final.run_excel_final_pipeline(
            source_path,
            output_path,
        )


def test_excel_final_lookup_protocol_passes_category_spec_and_material(monkeypatch):
    captured: dict[str, object] = {}

    def fake_run(command, **kwargs):
        captured["command"] = command
        return subprocess.CompletedProcess(
            command,
            0,
            stdout=_protocol_line(
                {
                    "protocol_version": 1,
                    "operation": "lookup",
                    "category": "round_bar",
                    "normalized_spec": "24",
                    "material": "Q355B",
                    "weight_kg_per_m": 3.55,
                    "source": "round_square_bar:round_bar",
                    "status": "hit",
                }
            ),
            stderr="",
        )

    monkeypatch.setattr(excel_final.subprocess, "run", fake_run)

    result = excel_final.lookup_excel_final_weight(
        category="round_bar",
        spec="D24",
        material="Q355B",
    )

    assert result.weight_kg_per_m == 3.55
    assert result.normalized_spec == "24"
    command = captured["command"]
    assert command[command.index("--category") + 1] == "round_bar"
    assert command[command.index("--spec") + 1] == "24"
    assert command[command.index("--material") + 1] == "Q355B"


def test_excel_final_lookup_protocol_accepts_q235b_round_bar(monkeypatch):
    captured: dict[str, object] = {}

    def fake_run(command, **kwargs):
        captured["command"] = command
        return subprocess.CompletedProcess(
            command,
            0,
            stdout=_protocol_line(
                {
                    "protocol_version": 1,
                    "operation": "lookup",
                    "category": "round_bar",
                    "normalized_spec": "8",
                    "material": "Q235B",
                    "weight_kg_per_m": 0.395,
                    "source": "round_square_bar:round_bar",
                    "status": "hit",
                }
            ),
            stderr="",
        )

    monkeypatch.setattr(excel_final.subprocess, "run", fake_run)

    result = excel_final.lookup_excel_final_weight(
        category="round_bar",
        spec="D8",
        material="Q235B",
    )

    assert result.weight_kg_per_m == 0.395
    assert result.normalized_spec == "8"
    command = captured["command"]
    assert command[command.index("--category") + 1] == "round_bar"
    assert command[command.index("--spec") + 1] == "8"
    assert command[command.index("--material") + 1] == "Q235B"


def test_excel_final_lookup_normalizes_full_width_material_whitespace():
    assert excel_final._normalize_lookup_request(
        "round_bar",
        "D8",
        "Q235　B",
    ) == ("round_bar", "8", "Q235B")


@pytest.mark.parametrize(
    ("category", "spec", "expected"),
    [
        ("steel_pipe", "PIP219*8", 41.62608),
        ("square_tube", "PD100×4", 9.46944),
    ],
)
def test_excel_final_lookup_uses_pip_pd_formula_without_handbook(
    category: str,
    spec: str,
    expected: float,
):
    result = excel_final.lookup_excel_final_weight(
        category=category,
        spec=spec,
        material="Q355B",
    )

    assert result.status == "hit"
    assert result.weight_kg_per_m == pytest.approx(expected)
    assert result.source == "circular_hollow_formula:0.02466"


@pytest.mark.parametrize(
    ("category", "spec"),
    [
        ("square_tube", "PIP219*8"),
        ("steel_pipe", "PD100*4"),
        ("steel_pipe", "PIP60*30"),
    ],
)
def test_excel_final_lookup_rejects_invalid_pip_pd_request(
    category: str,
    spec: str,
):
    with pytest.raises(ValueError):
        excel_final.lookup_excel_final_weight(
            category=category,
            spec=spec,
            material="Q355B",
        )


def test_excel_final_lookup_protocol_preserves_authoritative_source_conflict(monkeypatch):
    def fake_run(command, **kwargs):
        return subprocess.CompletedProcess(
            command,
            0,
            stdout=_protocol_line(
                {
                    "protocol_version": 1,
                    "operation": "lookup",
                    "category": "hfw_pipe",
                    "normalized_spec": "LH200*100*3.2*6",
                    "material": None,
                    "weight_kg_per_m": None,
                    "source": "hfw_pipe:conflict",
                    "status": "conflict",
                }
            ),
            stderr="",
        )

    monkeypatch.setattr(excel_final.subprocess, "run", fake_run)

    result = excel_final.lookup_excel_final_weight(
        category="hfw_pipe",
        spec="LH200*100*3.2*6",
    )

    assert result.status == "conflict"
    assert result.weight_kg_per_m is None
    assert result.source == "hfw_pipe:conflict"


def test_retired_post_output_repair_is_absent():
    assert not hasattr(excel_final, "normalize_excel_final_output")


def test_excel_final_staging_accepts_macro_enabled_workbook(
    db,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
):
    from app.modules.files.interface import StoredFile

    source = tmp_path / "stored-source.xlsm"
    workbook = Workbook()
    workbook.active.title = "原表"
    workbook.save(source)
    stored = StoredFile(
        bucket="dwg-reports",
        storage_key="uploads/stored-source.xlsm",
        original_name="source.xlsm",
        file_ext=".xlsm",
        content_type="application/vnd.ms-excel.sheet.macroEnabled.12",
        size_bytes=source.stat().st_size,
        sha256="3" * 64,
        status="available",
    )
    db.add(stored)
    db.flush()

    class LocalStorage:
        def local_path(self, _bucket, _storage_key):
            return source

    monkeypatch.setattr(
        "app.modules.excel_processing.staging.get_storage_backend",
        lambda: LocalStorage(),
    )

    work_dir = tmp_path / "attempt"
    work_dir.mkdir()
    staged, returned = stage_excel_source(db, stored.id, work_dir)

    assert returned.id == stored.id
    assert staged.suffix == ".xlsm"
    assert staged.is_file()


def test_excel_final_completion_event_does_not_pass_duplicate_batch_id():
    source = (
        BACKEND_ROOT
        / "app/modules/excel_processing/execution.py"
    ).read_text(encoding="utf-8")

    assert "batch_id=batch.id" not in source
