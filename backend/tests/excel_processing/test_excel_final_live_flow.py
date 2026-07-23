from __future__ import annotations

import os
from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.bootstrap.seed import init_db
from app.main import app
from app.modules.excel_processing.execution import run_excel_final_processing
from app.modules.excel_processing.models import (
    ExcelFinalBatch,
    ExcelFinalComponent,
    ExcelFinalPart,
)
from app.modules.files.interface import FileTransfer, StoredFile
from app.modules.identity.interface import User
from app.modules.jobs.interface import AnalysisResult, Job
from app.platform.storage.local import LocalFileStorage
from tests.support.paths import STAGES_ROOT

LIVE_SOURCE = (
    STAGES_ROOT
    / "excel_final/data/preprocessed/"
    "20260320-首都体育学院B7#地下部分-构件零件清单(毛净重)去gyb(3)_原表.xlsx"
)

pytestmark = [
    pytest.mark.live_excel_final,
    pytest.mark.skipif(
        os.environ.get("DWG_RUN_LIVE_EXCEL_FINAL") != "1",
        reason="set DWG_RUN_LIVE_EXCEL_FINAL=1 to run the real Stage/MySQL flow",
    ),
]


def _admin_client(db: Session) -> tuple[TestClient, dict[str, str], User]:
    init_db()
    client = TestClient(app)
    response = client.post(
        "/api/v1/auth/sessions",
        json={"username": "admin", "password": "SuperAdminPass1"},
    )
    assert response.status_code == 201, response.text
    admin = db.scalar(select(User).where(User.username == "admin"))
    assert admin is not None
    headers = {"Authorization": f"Bearer {response.json()['data']['access_token']}"}
    return client, headers, admin


def test_live_upload_worker_catalog_and_download_flow(
    db: Session,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    assert LIVE_SOURCE.is_file(), LIVE_SOURCE
    storage = LocalFileStorage(tmp_path / "storage")
    monkeypatch.setattr(
        "app.platform.storage.factory.get_storage_backend",
        lambda: storage,
    )
    monkeypatch.setattr(
        "app.modules.excel_processing.availability.settings.excel_final_pipeline_enabled",
        True,
    )
    dispatched: list[int] = []
    monkeypatch.setattr(
        "app.modules.excel_processing.routes.processing.dispatch_committed_job",
        lambda _db, job: dispatched.append(job.id),
    )
    client, headers, _admin = _admin_client(db)

    submitted = client.post(
        "/api/v1/excel-final/upload-and-process",
        headers={**headers, "Idempotency-Key": "live-excel-final-flow"},
        files={
            "upload": (
                LIVE_SOURCE.name,
                LIVE_SOURCE.read_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert submitted.status_code == 202, submitted.text
    job_id = submitted.json()["data"]["job_id"]
    assert dispatched == [job_id]

    run_excel_final_processing(
        job_id,
        worker_name="live-excel-final-test",
        expected_attempt=1,
    )
    db.expire_all()

    status = client.get(f"/api/v1/excel-final/process/{job_id}", headers=headers)
    assert status.status_code == 200, status.text
    payload = status.json()["data"]
    assert payload["status"] == "succeeded"
    assert payload["batch"]["part_count"] == 527
    assert payload["batch"]["component_count"] == 46
    assert payload["batch"]["quality_status"] == "ok"
    assert payload["batch"]["warning_count"] == 0
    assert payload["batch"]["severe_warning_count"] == 0
    assert payload["batch"]["total_net_weight"] == 122013.557
    assert payload["batch"]["total_gross_weight"] == 124831.881

    batch = db.scalar(select(ExcelFinalBatch).where(ExcelFinalBatch.job_id == job_id))
    assert batch is not None
    assert batch.source_type == "canonical"
    assert db.scalar(
        select(func.count()).select_from(ExcelFinalPart).where(
            ExcelFinalPart.batch_id == batch.id
        )
    ) == 527
    assert db.scalar(
        select(func.count()).select_from(ExcelFinalComponent).where(
            ExcelFinalComponent.batch_id == batch.id
        )
    ) == 46
    for field in (
        ExcelFinalPart.density_source,
        ExcelFinalPart.material_utilization,
        ExcelFinalPart.weight_validation,
    ):
        assert db.scalar(
            select(func.count()).select_from(ExcelFinalPart).where(
                ExcelFinalPart.batch_id == batch.id,
                field.is_not(None),
            )
        ) > 0
    d_parts = list(
        db.scalars(
            select(ExcelFinalPart).where(
                ExcelFinalPart.batch_id == batch.id,
                ExcelFinalPart.profile_spec.in_(("D24", "D30")),
            )
        )
    )
    assert len(d_parts) == 4
    assert all(
        part.density_source == "round_square_bar:round_bar"
        for part in d_parts
    )
    plate_catalog = client.get(
        f"/api/v1/excel-final/batches/{batch.id}/parts?part_type=plate&page_size=500",
        headers=headers,
    )
    assert plate_catalog.status_code == 200, plate_catalog.text
    assert plate_catalog.json()["pagination"]["total"] == 394
    assert {item["part_type"] for item in plate_catalog.json()["data"]} == {"plate"}

    analysis = db.scalar(select(AnalysisResult).where(AnalysisResult.job_id == job_id))
    assert analysis is not None and analysis.result_file_id is not None
    result_file = db.get(StoredFile, analysis.result_file_id)
    assert result_file is not None
    assert result_file.file_ext == ".xlsx"
    assert result_file.original_name.endswith("_处理后.xlsx")

    download_url = client.get(
        f"/api/v1/excel-final/process/{job_id}/download",
        headers=headers,
    )
    assert download_url.status_code == 200, download_url.text
    downloaded = client.get(download_url.json()["data"]["url"], headers=headers)
    assert downloaded.status_code == 200, downloaded.text
    assert downloaded.content == b"".join(
        storage.iter_file(result_file.bucket, result_file.storage_key)
    )
    workbook = load_workbook(BytesIO(downloaded.content), data_only=True)
    try:
        assert workbook.sheetnames == [
            "原表",
            "清洗表",
            "构件表",
            "整理表",
            "part",
            "处理报告",
        ]
        assert workbook["整理表"].max_row == 528
        part_headers = [cell.value for cell in workbook["part"][1]]
        part_rows = [
            dict(zip(part_headers, values, strict=True))
            for values in workbook["part"].iter_rows(min_row=2, values_only=True)
        ]
        component_types = {
            "BH腹", "BH翼", "BOX腹", "BOX翼", "BT腹", "BT翼",
        }
        component_scoped = [
            row for row in part_rows if row["类型"] in component_types
        ]
        global_scoped = [
            row for row in part_rows if row["类型"] not in component_types
        ]
        assert len(part_rows) == 122
        assert len(component_scoped) == 84
        assert len(global_scoped) == 38
        assert all(row["导入构件编号"] for row in component_scoped)
        assert all(row["导入构件编号"] is None for row in global_scoped)
        assert sum(row["汇总"] for row in global_scoped) == 1216
        assert all(row["文件"] is None for row in part_rows)
        assert workbook["处理报告"]["A2"].value == "无"
        assert workbook["处理报告"].max_row == 2
        for sheet_name in ("清洗表", "构件表", "整理表", "part", "处理报告"):
            worksheet = workbook[sheet_name]
            for column in range(1, worksheet.max_column + 1):
                letter = get_column_letter(column)
                width = worksheet.column_dimensions[letter].width
                if sheet_name == "处理报告" and column in (7, 8):
                    assert 16 <= width <= 48
                else:
                    assert 8 <= width <= 32
        for coordinate in ("G2", "H2"):
            assert workbook["处理报告"][coordinate].alignment.wrap_text is True
            assert workbook["处理报告"][coordinate].alignment.vertical == "top"
        assert workbook["构件表"].auto_filter.ref == "A1:O1"
        assert workbook["整理表"].auto_filter.ref == "A1:AF1"
        for sheet_name, removed_headers in (
            ("整理表", ("比重来源", "净材利用率", "重量核验")),
            ("构件表", ("来源sheet", "行类型", "小计来源行")),
        ):
            worksheet = workbook[sheet_name]
            header_values = {cell.value for cell in worksheet[1]}
            assert not (set(removed_headers) & header_values)
    finally:
        workbook.close()

    assert db.scalar(select(func.count()).select_from(Job)) == 1
    assert db.scalar(select(func.count()).select_from(AnalysisResult)) == 1
    assert db.scalar(select(func.count()).select_from(FileTransfer)) >= 3
