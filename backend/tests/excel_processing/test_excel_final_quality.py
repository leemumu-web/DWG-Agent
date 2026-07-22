from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.modules.excel_processing.models import ExcelFinalPart
from app.modules.excel_processing.persistence import import_workbook_for_job
from app.modules.excel_processing.presentation import batch_summary, part_detail
from app.modules.files.interface import StoredFile
from app.modules.jobs.interface import Job
from app.platform.config.constants import TASK_EXCEL_FINAL


def _job(db: Session) -> tuple[Job, StoredFile]:
    source = StoredFile(
        bucket="dwg-reports",
        storage_key="uploads/canonical.xlsx",
        original_name="canonical.xlsx",
        file_ext=".xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        size_bytes=10,
        sha256="c" * 64,
        status="available",
    )
    db.add(source)
    db.flush()
    job = Job(
        task_type=TASK_EXCEL_FINAL,
        precision_level="normal",
        pipeline="excel_final",
        status="running",
        priority=0,
        progress=60,
        params_json={"file_id": source.id},
    )
    db.add(job)
    db.flush()
    return job, source


def _canonical_workbook(path: Path) -> None:
    workbook = Workbook()
    organized = workbook.active
    organized.title = "整理表"
    organized.append(
        [
            "序号",
            "构件编号",
            "导入构件编号",
            "构件数",
            "类型",
            "批次",
            "零件号",
            "导入零件号",
            "规格",
            "长度(mm)",
            "材质",
            "原数量",
            "数量",
            "总净重(kg)",
            "表净重(kg)",
            "总毛重(kg)",
            "表毛重(kg)",
            "比重来源",
            "净材利用率",
            "重量核验",
        ]
    )
    organized.append(
        [
            1,
            "C-1",
            "C-1",
            2,
            "BOX腹",
            "B-1",
            "P-1",
            "P-1_腹",
            8,
            1000,
            "Q355B",
            1,
            1,
            100,
            200,
            110,
            220,
            "plate_constant:7.85",
            0.9,
            "严重",
        ]
    )
    organized.append(
        [
            1,
            "C-1",
            "C-1",
            2,
            "BOX翼",
            "B-1",
            "P-1",
            "P-1_翼",
            10,
            1000,
            "Q355B",
            1,
            2,
            50,
            None,
            60,
            None,
            None,
            None,
            None,
        ]
    )

    report = workbook.create_sheet("处理报告")
    report.append(["级别", "类别", "说明"])
    report.append(["信息", "RECT未证明", "仅保留候选，不填写文件"])
    report.append(["警告", "手册查无", "规格 X10 在指定类别中查无"])
    report.append(["严重", "重量矛盾", "表单重与理单重严重不一致"])
    workbook.save(path)


def test_workbook_import_persists_quality_and_table_weight_totals(
    db: Session,
    tmp_path: Path,
):
    job, source = _job(db)
    output_path = tmp_path / "canonical.xlsx"
    _canonical_workbook(output_path)

    batch, stats = import_workbook_for_job(
        db,
        job_id=job.id,
        file_id=source.id,
        source_type="init_table",
        source_name=source.original_name,
        output_path=output_path,
    )

    assert batch.part_count == 2
    assert batch.total_net_weight == 200
    assert batch.total_gross_weight == 220
    assert batch.quality_status == "severe_warning"
    assert batch.warning_count == 1
    assert batch.severe_warning_count == 1
    assert batch.report_summary == {
        "info_count": 1,
        "warning_count": 1,
        "severe_warning_count": 1,
        "category_counts": {"RECT未证明": 1, "手册查无": 1, "重量矛盾": 1},
        "representative_messages": [
            "仅保留候选，不填写文件",
            "规格 X10 在指定类别中查无",
            "表单重与理单重严重不一致",
        ],
    }
    assert stats["quality_status"] == "severe_warning"
    assert stats["total_net_weight"] == 200
    assert stats["total_gross_weight"] == 220

    first_part = db.query(ExcelFinalPart).order_by(ExcelFinalPart.seq).first()
    assert first_part is not None
    assert batch_summary(batch)["quality_status"] == "severe_warning"
    assert part_detail(first_part)["density_source"] == "plate_constant:7.85"
    assert part_detail(first_part)["weight_validation"] == "severe_warning"


def test_workbook_import_sums_physical_weights_as_exact_decimals(
    db: Session,
    tmp_path: Path,
):
    job, source = _job(db)
    output_path = tmp_path / "exact-weights.xlsx"
    workbook = Workbook()
    organized = workbook.active
    organized.title = "整理表"
    organized.append(
        [
            "序号",
            "构件编号",
            "零件号",
            "规格",
            "长度",
            "材质",
            "数量",
            "表净重",
            "表毛重",
        ]
    )
    organized.append([1, "C-1", "P-1", 10, 1000, "Q355B", 1, 122000.5, 124800.88])
    organized.append([2, "C-2", "P-2", 10, 1000, "Q355B", 1, 13.057, 31.001])
    workbook.save(output_path)
    workbook.close()

    batch, stats = import_workbook_for_job(
        db,
        job_id=job.id,
        file_id=source.id,
        source_type="init_table",
        source_name=source.original_name,
        output_path=output_path,
    )

    assert batch.total_net_weight == Decimal("122013.557")
    assert batch.total_gross_weight == Decimal("124831.881")
    assert stats["total_net_weight"] == 122013.557
    assert stats["total_gross_weight"] == 124831.881
    projected = batch_summary(batch)
    assert projected["total_net_weight"] == 122013.557
    assert projected["total_gross_weight"] == 124831.881
    assert isinstance(projected["total_gross_weight"], float)


def test_workbook_quality_must_match_independent_stage_summary(
    db: Session,
    tmp_path: Path,
):
    job, source = _job(db)
    output_path = tmp_path / "quality-mismatch.xlsx"
    _canonical_workbook(output_path)

    with pytest.raises(ValueError, match="quality summary mismatch"):
        import_workbook_for_job(
            db,
            job_id=job.id,
            file_id=source.id,
            source_type="init_table",
            source_name=source.original_name,
            output_path=output_path,
            expected_quality={
                "quality_status": "ok",
                "warning_count": 0,
                "severe_warning_count": 0,
            },
        )


def test_legacy_workbook_without_report_gets_safe_quality_defaults(
    db: Session,
    tmp_path: Path,
):
    job, source = _job(db)
    output_path = tmp_path / "legacy.xlsx"
    workbook = Workbook()
    organized = workbook.active
    organized.title = "整理表"
    organized.append(["序号", "构件编号", "零件号", "规格", "长度", "材质", "数量"])
    organized.append([1, "C-1", "P-1", "PL10*100", 200, "Q355B", 1])
    workbook.save(output_path)

    batch, _ = import_workbook_for_job(
        db,
        job_id=job.id,
        file_id=source.id,
        source_type="init_table",
        source_name=source.original_name,
        output_path=output_path,
    )

    assert batch.quality_status == "ok"
    assert batch.warning_count == 0
    assert batch.severe_warning_count == 0
    assert batch.report_summary is None
    part = db.query(ExcelFinalPart).one()
    assert part.import_component_no is None
    assert part.import_part_no is None
    assert part.weight_validation is None


def test_zero_table_totals_are_preserved_instead_of_becoming_null(
    db: Session,
    tmp_path: Path,
):
    job, source = _job(db)
    output_path = tmp_path / "zeros.xlsx"
    workbook = Workbook()
    organized = workbook.active
    organized.title = "整理表"
    organized.append(
        [
            "序号",
            "构件编号",
            "零件号",
            "规格",
            "长度",
            "材质",
            "数量",
            "表净重",
            "表毛重",
        ]
    )
    organized.append([1, "C-0", "P-0", "PL10*100", 0, "Q355B", 0, 0, 0])
    workbook.save(output_path)

    batch, _ = import_workbook_for_job(
        db,
        job_id=job.id,
        file_id=source.id,
        source_type="init_table",
        source_name=source.original_name,
        output_path=output_path,
    )

    assert batch.total_net_weight == 0
    assert batch.total_gross_weight == 0
