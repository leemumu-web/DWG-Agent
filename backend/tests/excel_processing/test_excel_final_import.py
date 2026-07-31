from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from unittest.mock import Mock

import pytest
from openpyxl import Workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.modules.excel_processing import importers
from app.modules.excel_processing.importers import (
    import_components_to_db,
    import_parts_to_db,
    import_quality_report,
)
from app.modules.excel_processing.models import (
    ExcelFinalBatch,
    ExcelFinalComponent,
    ExcelFinalPart,
)
from app.modules.files.interface import StoredFile
from app.modules.jobs.interface import Job
from app.platform.config.constants import TASK_EXCEL_FINAL


def _batch(db: Session) -> ExcelFinalBatch:
    source = StoredFile(
        bucket="dwg-reports",
        storage_key="uploads/source.xls",
        original_name="source.xls",
        file_ext=".xls",
        content_type="application/vnd.ms-excel",
        size_bytes=10,
        sha256="b" * 64,
        status="available",
    )
    db.add(source)
    db.flush()
    job = Job(
        task_type=TASK_EXCEL_FINAL,
        precision_level="normal",
        pipeline="excel_final",
        status="running",
        priority=0,
        progress=60,
        params_json={"file_id": source.id},
    )
    db.add(job)
    db.flush()
    batch = ExcelFinalBatch(
        job_id=job.id,
        file_id=source.id,
        source_type="tsv",
        source_name="source.xls",
    )
    db.add(batch)
    db.flush()
    return batch


def _workbook(
    path: Path,
    headers: list[str],
    rows: list[list[object]],
    *,
    sheet_name: str = "整理表",
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def test_excel_final_combined_import_opens_workbook_once(
    db: Session,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    batch = _batch(db)
    output_path = tmp_path / "combined-import.xlsx"
    workbook = Workbook()
    organized = workbook.active
    organized.title = "整理表"
    organized.append(
        ["序号", "构件编号", "零件号", "规格", "长度(mm)", "材质", "数量"]
    )
    organized.append([1, "C-1", "P-1", "PL10*100", 200, "Q355B", 2])
    components = workbook.create_sheet("构件表")
    components.append(["构件编号", "构件数", "总净重"])
    components.append(["C-1", 1, 15.7])
    report = workbook.create_sheet("处理报告")
    report.append(["级别", "类别", "说明"])
    report.append(["无", None, None])
    workbook.save(output_path)

    real_load_workbook = importers.openpyxl.load_workbook
    opened = 0

    def counted_load_workbook(*args, **kwargs):
        nonlocal opened
        opened += 1
        return real_load_workbook(*args, **kwargs)

    monkeypatch.setattr(
        importers.openpyxl,
        "load_workbook",
        counted_load_workbook,
    )

    stats = importers.import_workbook_to_db(db, batch.id, output_path)

    assert opened == 1
    assert stats["parts_imported"] == 1
    assert stats["components_imported"] == 1
    assert stats["quality_status"] == "ok"


def test_excel_final_quality_import_treats_no_report_sentinel_as_empty(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "no-report.xlsx"
    _workbook(
        output_path,
        [
            "级别",
            "类别",
            "来源位置",
            "构件编号",
            "零件号",
            "涉及字段",
            "说明",
            "建议操作",
        ],
        [["无", None, None, None, None, None, None, None]],
        sheet_name="处理报告",
    )

    assert import_quality_report(output_path) == {
        "quality_status": "ok",
        "warning_count": 0,
        "severe_warning_count": 0,
        "report_summary": {
            "info_count": 0,
            "warning_count": 0,
            "severe_warning_count": 0,
            "category_counts": {},
            "representative_messages": [],
        },
    }


def test_excel_final_quality_import_rejects_mixed_no_report_sentinel(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "mixed-no-report.xlsx"
    _workbook(
        output_path,
        ["级别", "类别", "说明"],
        [["无", "重量异常", "不能与无混用"]],
        sheet_name="处理报告",
    )

    with pytest.raises(ValueError, match="sentinel"):
        import_quality_report(output_path)


def test_excel_final_quality_import_accepts_legacy_report_columns(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "legacy-report.xlsx"
    _workbook(
        output_path,
        [
            "级别",
            "类别",
            "来源sheet",
            "来源行",
            "构件编号",
            "零件号",
            "规格",
            "字段",
            "实际值",
            "期望值",
            "绝对误差",
            "相对误差",
            "是否影响part",
            "比重来源",
            "说明",
        ],
        [["警告", "手册查无", "原表", 8, "C1", "P1", "L999", "比重",
          "查无", "手册命中", None, None, "否", "angle:not_found", "规格查无"]],
        sheet_name="处理报告",
    )

    stats = import_quality_report(output_path)

    assert stats["quality_status"] == "warning"
    assert stats["warning_count"] == 1
    assert stats["report_summary"]["category_counts"] == {"手册查无": 1}


def test_excel_final_import_skips_totals_row(db: Session, tmp_path: Path):
    batch = _batch(db)
    output_path = tmp_path / "result.xlsx"
    headers = ["序号", "构件编号", "零件号", "规格", "长度(mm)", "材质", "数量"]
    _workbook(
        output_path,
        headers,
        [
            [1, "C-1", "P-1", "PL10*100", 200, "Q355B", 2],
            [2, "C-1", "1.00", "合计：", 200, "123.45", None],
        ],
    )

    stats = import_parts_to_db(db, batch.id, output_path)

    assert stats == {"parts_imported": 1}
    assert db.scalar(select(func.count()).select_from(ExcelFinalPart)) == 1
    part = db.scalar(select(ExcelFinalPart))
    assert part is not None
    assert part.part_no == "P-1"
    assert part.material == "Q355B"


def test_excel_final_import_accepts_separated_unit_suffixes(
    db: Session,
    tmp_path: Path,
) -> None:
    batch = _batch(db)
    output_path = tmp_path / "result-with-unit-separators.xlsx"
    _workbook(
        output_path,
        [
            "序号",
            "构件编号",
            "零件号",
            "规格",
            "长度/mm",
            "材质",
            "数量-件",
            "单毛重[kg]",
            "总毛重(kg)",
        ],
        [[1, "C-1", "P-1", "PL10*100", 200, "Q355B", 2, 1.5, 3.0]],
    )

    stats = import_parts_to_db(db, batch.id, output_path)

    assert stats == {"parts_imported": 1}


def test_excel_final_import_rejects_missing_core_headers(db: Session, tmp_path: Path):
    batch = _batch(db)
    output_path = tmp_path / "invalid.xlsx"
    _workbook(output_path, ["序号", "零件号"], [[1, "P-1"]])

    with pytest.raises(ValueError, match="missing required columns"):
        import_parts_to_db(db, batch.id, output_path)


def test_excel_final_parts_imports_sparse_rows_and_optional_fields(db: Session, tmp_path: Path):
    batch = _batch(db)
    output_path = tmp_path / "sparse.xlsx"
    _workbook(
        output_path,
        ["序号", "构件编号", "零件号", "截面型材", "规格", "长度(mm)", "材质", "数量"],
        [["2", " C-2 ", " P-2 ", None, "PL8*80", "120.5", " Q235B ", "0"]],
    )

    stats = import_parts_to_db(db, batch.id, output_path)

    assert stats == {"parts_imported": 1}
    part = db.scalar(select(ExcelFinalPart))
    assert part is not None
    assert part.seq == 2
    assert part.component_no == "C-2"
    assert part.part_no == "P-2"
    assert part.profile_spec is None
    assert part.length == 120.5
    assert part.material == "Q235B"
    assert part.qty == 0


def test_excel_final_import_preserves_d_series_display_spec(
    db: Session,
    tmp_path: Path,
) -> None:
    batch = _batch(db)
    output_path = tmp_path / "d-series.xlsx"
    _workbook(
        output_path,
        ["序号", "构件编号", "零件号", "截面型材", "规格", "长度", "材质", "数量"],
        [[1, "C-1", "P-D24", "D24", "D24", 1000, "Q355B", 1]],
    )

    assert import_parts_to_db(db, batch.id, output_path) == {"parts_imported": 1}
    part = db.scalar(select(ExcelFinalPart))
    assert part is not None
    assert part.profile_spec == "D24"
    assert part.spec == "D24"


def test_excel_final_parts_never_persists_negative_physical_values(
    db: Session,
    tmp_path: Path,
):
    batch = _batch(db)
    output_path = tmp_path / "negative-parts.xlsx"
    _workbook(
        output_path,
        [
            "序号", "构件编号", "零件号", "规格", "长度", "材质", "数量",
            "表净重", "表毛重", "单表面积", "总表面积",
        ],
        [
            [1, "C-1", "P-valid", "PL10*100", 1000, "Q355B", 1, 7, 8, 1, 1],
            [2, "C-1", "P-length", "PL10*100", -1, "Q355B", 1, 7, 8, 1, 1],
            [3, "C-1", "P-qty", "PL10*100", 1000, "Q355B", -1, 7, 8, 1, 1],
            [4, "C-1", "P-weight", "PL10*100", 1000, "Q355B", 1, -1, -1, 1, 1],
            [5, "C-1", "P-area", "PL10*100", 1000, "Q355B", 1, 7, 8, -1, -1],
        ],
    )

    assert import_parts_to_db(db, batch.id, output_path) == {"parts_imported": 1}
    assert [part.part_no for part in db.scalars(select(ExcelFinalPart))] == ["P-valid"]


@pytest.mark.parametrize("value", ["NaN", "Infinity", "-Infinity"])
def test_excel_final_parts_rejects_nonfinite_numeric_values(
    db: Session,
    tmp_path: Path,
    value: str,
):
    batch = _batch(db)
    output_path = tmp_path / "nonfinite-parts.xlsx"
    _workbook(
        output_path,
        ["序号", "构件编号", "零件号", "规格", "长度", "材质", "数量"],
        [[1, "C-1", "P-1", "PL10*100", value, "Q355B", 1]],
    )

    with pytest.raises(ValueError, match="non-finite numeric value"):
        import_parts_to_db(db, batch.id, output_path)


def test_excel_final_parts_skips_rows_without_component_identity(
    db: Session,
    tmp_path: Path,
):
    batch = _batch(db)
    output_path = tmp_path / "missing-component-identity.xlsx"
    _workbook(
        output_path,
        ["序号", "构件编号", "零件号", "规格", "长度", "材质", "数量"],
        [[1, None, "P-1", "PL10*100", 1000, "Q355B", 1]],
    )

    assert import_parts_to_db(db, batch.id, output_path) == {"parts_imported": 0}
    assert db.scalar(select(func.count()).select_from(ExcelFinalPart)) == 0


def test_excel_final_parts_rejects_retired_intermediate_sheet(db: Session, tmp_path: Path):
    batch = _batch(db)
    output_path = tmp_path / "intermediate.xlsx"
    _workbook(
        output_path,
        ["序号", "构件编号", "零件号", "规格", "长度", "材质", "数量"],
        [[1, "C-1", "P-1", "PL10*100", 200, "Q355B", 2]],
        sheet_name="整理表_拆板后",
    )

    assert import_parts_to_db(db, batch.id, output_path) == {
        "parts_imported": 0,
        "error": "No 整理表 sheet found",
    }


def test_excel_final_parts_imports_canonical_provenance_and_quality_fields(
    db: Session,
    tmp_path: Path,
):
    batch = _batch(db)
    output_path = tmp_path / "canonical.xlsx"
    _workbook(
        output_path,
        [
            "序号",
            "构件编号",
            "导入构件编号",
            "构件数",
            "类型",
            "班组",
            "批次",
            "零件号",
            "导入零件号",
            "规格",
            "长度(mm)",
            "下料长度(mm)",
            "材质",
            "原数量",
            "数量",
            "比重",
            "比重来源",
            "净材利用率",
            "重量核验",
        ],
        [
            [
                7,
                "C-SPLIT",
                "C-RAW",
                3,
                "BOX腹",
                "",
                "B-01",
                "P-DISPLAY",
                "P-RAW_腹",
                8,
                1200,
                1180,
                "Q355B",
                2,
                1,
                7.85,
                "plate_constant:7.85",
                0.8125,
                "警告",
            ]
        ],
    )

    assert import_parts_to_db(db, batch.id, output_path) == {"parts_imported": 1}
    part = db.scalar(select(ExcelFinalPart))
    assert part is not None
    assert part.import_component_no == "C-RAW"
    assert part.import_part_no == "P-RAW_腹"
    assert part.source_batch == "B-01"
    assert part.team is None
    assert part.original_qty == 2
    assert part.cut_length == 1180
    assert part.density_source == "plate_constant:7.85"
    assert part.material_utilization == 0.8125
    assert part.weight_validation == "warning"
    assert part.part_type == "box_web"


def test_excel_final_parts_rejects_unknown_part_type(db: Session, tmp_path: Path):
    batch = _batch(db)
    output_path = tmp_path / "unknown-type.xlsx"
    _workbook(
        output_path,
        ["序号", "构件编号", "类型", "零件号", "规格", "长度", "材质", "数量"],
        [[1, "C-1", "神秘型材", "P-1", "X10", 200, "Q355B", 1]],
    )

    with pytest.raises(ValueError, match="unknown part type"):
        import_parts_to_db(db, batch.id, output_path)


@pytest.mark.parametrize(
    ("stage_type", "stored_type"),
    [
        ("H型钢", "h_beam"),
        ("T型钢", "t_beam"),
        ("槽钢", "channel"),
        ("角钢", "angle"),
        ("方管", "square_tube"),
        ("钢管", "steel_pipe"),
        ("圆管", "steel_pipe"),
        ("方钢", "square_bar"),
        ("高频焊", "hfw_pipe"),
        ("W型钢", "w_beam"),
    ],
)
def test_excel_final_part_type_projection_covers_every_handbook_profile(
    stage_type: str,
    stored_type: str,
):
    assert importers._part_type(stage_type) == stored_type


def test_excel_final_parts_returns_error_when_sheet_is_absent(db: Session, tmp_path: Path):
    batch = _batch(db)
    output_path = tmp_path / "missing-sheet.xlsx"
    _workbook(output_path, ["other"], [], sheet_name="Sheet1")

    assert import_parts_to_db(db, batch.id, output_path) == {
        "parts_imported": 0,
        "error": "No 整理表 sheet found",
    }


def test_excel_final_components_imports_rows_and_preserves_zero_qty(
    db: Session,
    tmp_path: Path,
):
    batch = _batch(db)
    output_path = tmp_path / "components.xlsx"
    _workbook(
        output_path,
        ["序号", "批次", "构件编号", "构件数", "总净重"],
        [
            [1, "B-1", "C-1", 2, 123.5],
            [2, "B-2", "C-2", 0, 0],
            [3, "", "合计", 2, 123.5],
            [4, "", " ", 1, 20],
        ],
        sheet_name="构件表",
    )

    stats = import_components_to_db(db, batch.id, output_path)

    assert stats == {"components_imported": 2}
    components = list(db.scalars(select(ExcelFinalComponent).order_by(ExcelFinalComponent.id)))
    assert [(item.component_no, item.component_qty, item.total_weight) for item in components] == [
        ("C-1", 2, 123.5),
        ("C-2", 0, 0),
    ]


def test_excel_final_components_never_persists_negative_physical_values(
    db: Session,
    tmp_path: Path,
):
    batch = _batch(db)
    output_path = tmp_path / "negative-components.xlsx"
    _workbook(
        output_path,
        ["构件编号", "构件数", "总净重"],
        [["C-valid", 1, 10], ["C-qty", -1, 10], ["C-weight", 1, -10]],
        sheet_name="构件表",
    )

    assert import_components_to_db(db, batch.id, output_path) == {
        "components_imported": 1
    }
    component = db.scalar(select(ExcelFinalComponent))
    assert component is not None
    assert component.component_no == "C-valid"


@pytest.mark.parametrize("value", ["NaN", "Infinity", "-Infinity"])
def test_excel_final_components_rejects_nonfinite_numeric_values(
    db: Session,
    tmp_path: Path,
    value: str,
):
    batch = _batch(db)
    output_path = tmp_path / "nonfinite-components.xlsx"
    _workbook(
        output_path,
        ["构件编号", "构件数", "总净重"],
        [["C-1", 1, value]],
        sheet_name="构件表",
    )

    with pytest.raises(ValueError, match="non-finite numeric value"):
        import_components_to_db(db, batch.id, output_path)


def test_excel_final_components_allows_missing_optional_columns(db: Session, tmp_path: Path):
    batch = _batch(db)
    output_path = tmp_path / "component-number-only.xlsx"
    _workbook(
        output_path,
        ["构件编号"],
        [["C-1"]],
        sheet_name="构件表",
    )

    stats = import_components_to_db(db, batch.id, output_path)

    assert stats == {"components_imported": 1}
    component = db.scalar(select(ExcelFinalComponent))
    assert component is not None
    assert component.component_no == "C-1"
    assert component.component_qty is None
    assert component.total_weight is None


def test_excel_final_components_rejects_missing_component_number(
    db: Session,
    tmp_path: Path,
):
    batch = _batch(db)
    output_path = tmp_path / "invalid-components.xlsx"
    _workbook(
        output_path,
        ["构件数", "总重"],
        [[2, 123.5]],
        sheet_name="构件表",
    )

    with pytest.raises(ValueError, match="missing required column: 构件编号"):
        import_components_to_db(db, batch.id, output_path)


def test_excel_final_components_rejects_duplicate_component_identity(
    db: Session,
    tmp_path: Path,
):
    batch = _batch(db)
    output_path = tmp_path / "duplicate-components.xlsx"
    _workbook(
        output_path,
        ["构件编号", "构件数", "总净重"],
        [["C-1", 2, 123.5], ["C-1", 2, 123.5]],
        sheet_name="构件表",
    )

    with pytest.raises(ValueError, match="duplicate component identity"):
        import_components_to_db(db, batch.id, output_path)


def test_excel_final_components_returns_zero_when_sheet_is_absent(
    db: Session,
    tmp_path: Path,
):
    batch = _batch(db)
    output_path = tmp_path / "no-components.xlsx"
    _workbook(output_path, ["序号"], [], sheet_name="整理表")

    assert import_components_to_db(db, batch.id, output_path) == {"components_imported": 0}


def test_excel_final_importers_never_use_random_cell_access(monkeypatch, tmp_path: Path):
    output_path = tmp_path / "streaming.xlsx"
    workbook = Workbook()
    parts_sheet = workbook.active
    parts_sheet.title = "整理表"
    parts_sheet.append(["序号", "构件编号", "零件号", "规格", "长度", "材质", "数量"])
    parts_sheet.append([1, "C-1", "P-1", "PL10*100", 200, "Q355B", 2])
    component_sheet = workbook.create_sheet("构件表")
    component_sheet.append(["构件编号", "构件数", "总净重"])
    component_sheet.append(["C-1", 1, 123.5])
    workbook.save(output_path)

    real_load_workbook = importers.openpyxl.load_workbook

    def guarded_load_workbook(*args, **kwargs):
        loaded = real_load_workbook(*args, **kwargs)
        for sheet in loaded.worksheets:
            sheet.cell = Mock(side_effect=AssertionError("random worksheet access is forbidden"))
        return loaded

    monkeypatch.setattr(importers.openpyxl, "load_workbook", guarded_load_workbook)
    db = Mock()

    assert import_parts_to_db(db, 1, output_path) == {"parts_imported": 1}
    assert import_components_to_db(db, 1, output_path) == {"components_imported": 1}


def test_excel_final_number_parser_preserves_decimal_text_exactly():
    assert importers._number(124831.881) == Decimal("124831.881")
    assert importers._number("0.100000001") == Decimal("0.100000001")
