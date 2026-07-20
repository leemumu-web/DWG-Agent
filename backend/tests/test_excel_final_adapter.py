from __future__ import annotations

import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from app.modules.excel_processing import stage_adapter as excel_final


def test_excel_final_stage_root_resolves_tracked_standalone_layout():
    stage_root = excel_final.get_excel_final_stage_root()

    assert (stage_root / "main.py").is_file()
    assert (stage_root / "pipeline.py").is_file()
    assert (stage_root / "handbook.py").is_file()


def test_excel_final_dependency_probe_includes_legacy_xls_reader(monkeypatch):
    checked: list[str] = []

    def available(module_name: str):
        checked.append(module_name)
        return object()

    monkeypatch.setattr(excel_final.importlib.util, "find_spec", available)

    assert excel_final.excel_final_dependencies_available() is True
    assert "xlrd" in checked


def test_excel_final_text_probe_falls_through_after_parser_error():
    script = """
from pathlib import Path
import pandas as pd
import reader

def fail_parse(*args, **kwargs):
    raise pd.errors.ParserError('not delimited text')

reader.pd.read_csv = fail_parse
assert reader._try_read(Path('binary.xls'), '\\t', 'latin-1') is None
"""
    completed = subprocess.run(
        [sys.executable, "-c", script],
        cwd=excel_final.get_excel_final_stage_root(),
        text=True,
        capture_output=True,
        check=False,
    )

    assert completed.returncode == 0, completed.stderr


def test_excel_final_pipeline_runs_in_isolated_subprocess(monkeypatch, tmp_path: Path):
    source_path = tmp_path / "source.xls"
    output_path = tmp_path / "result.xlsx"
    source_path.write_bytes(b"input")
    captured: dict[str, object] = {}

    def fake_run(command, **kwargs):
        captured["command"] = command
        captured.update(kwargs)
        workbook = Workbook()
        workbook.active.title = "整理表"
        workbook.save(output_path)
        return subprocess.CompletedProcess(command, 0, stdout="", stderr="")

    monkeypatch.setattr(excel_final.subprocess, "run", fake_run)
    monkeypatch.setattr(excel_final.settings, "handbook_mysql_password", "not-on-command-line")

    result = excel_final.run_excel_final_pipeline(
        source_path,
        output_path,
        source_format="tsv",
    )

    command = captured["command"]
    assert command[:3] == [
        sys.executable,
        "-m",
        "app.modules.excel_processing.stage_runner",
    ]
    assert "not-on-command-line" not in command
    assert captured["cwd"] == excel_final.get_excel_final_stage_root()
    assert captured["env"]["DWG_HANDBOOK_MYSQL_PASSWORD"] == "not-on-command-line"
    assert result == output_path


def test_excel_final_runner_is_importable_from_stage_working_directory():
    completed = subprocess.run(
        [
            sys.executable,
            "-m",
            "app.modules.excel_processing.stage_runner",
            "--help",
        ],
        cwd=excel_final.get_excel_final_stage_root(),
        env=excel_final._stage_environment(),
        text=True,
        capture_output=True,
        check=False,
    )

    assert completed.returncode == 0, completed.stderr
    assert "Isolated Excel Final Stage runner" in completed.stdout


def test_excel_final_pipeline_logs_internal_failure_but_raises_safe_message(
    monkeypatch,
    tmp_path: Path,
    caplog,
):
    source_path = tmp_path / "source.xls"
    output_path = tmp_path / "result.xlsx"
    source_path.write_bytes(b"input")

    def fake_run(command, **kwargs):
        return subprocess.CompletedProcess(command, 1, stdout="", stderr="pipeline exploded")

    monkeypatch.setattr(excel_final.subprocess, "run", fake_run)

    with pytest.raises(excel_final.ExcelFinalProcessError) as failure:
        excel_final.run_excel_final_pipeline(
            source_path,
            output_path,
            source_format="init",
        )

    assert str(failure.value) == "Excel Final Stage failed while processing the input."
    assert "pipeline exploded" not in str(failure.value)
    assert "pipeline exploded" in caplog.text


def test_excel_final_adapter_normalizes_legacy_fixed_width_bolt_row(tmp_path: Path):
    output_path = tmp_path / "result.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "整理表"
    sheet.append(
        [
            "序号",
            "构件编号",
            "构件数",
            "类型",
            "零件号",
            "截面型材",
            "规格",
            "宽度(mm)",
            "长度(mm)",
            "材质",
            "数量",
            "总数",
            "总长(mm)",
            "单净重(kg)",
            "总净重(kg)",
        ]
    )
    sheet.append([8, "C-1", 1, None, "M20", "C", "C", None, 2, "90", 0, 0, 0, 0, 0])
    workbook.save(output_path)

    normalized_count = excel_final.normalize_excel_final_output(output_path)

    result = load_workbook(output_path, read_only=True, data_only=True)
    row = [cell.value for cell in result["整理表"][2]]
    result.close()
    assert normalized_count == 1
    assert row[3:15] == [
        "紧固件",
        "M20",
        "M20",
        "M20",
        None,
        90,
        "C",
        2,
        2,
        180,
        0,
        0,
    ]


def test_excel_final_completion_event_does_not_pass_duplicate_batch_id():
    source = (
        Path(__file__).resolve().parents[1]
        / "app/modules/excel_processing/execution.py"
    ).read_text(encoding="utf-8")

    assert "batch_id=batch.id" not in source
