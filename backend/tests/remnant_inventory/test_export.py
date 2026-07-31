from __future__ import annotations

import asyncio
from datetime import datetime
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import select

from app.bootstrap.seed import init_db
from app.main import app
from app.modules.files.interface import StoredFile
from app.modules.identity.interface import Role, User
from app.modules.operations.audit.models import AuditLog
from app.modules.remnant_inventory.export import CleanupFileResponse
from app.modules.remnant_inventory.models import (
    Remnant,
    RemnantImportBatch,
    RemnantImportItem,
    RemnantMaterial,
    RemnantPart,
)
from app.platform.config.settings import settings
from app.platform.security.tokens import hash_password
from app.platform.time import BUSINESS_TIMEZONE
from tests.support.database import get_test_session_factory

HEADERS = [
    "余料编号",
    "材质",
    "厚度(mm)",
    "项目编号一",
    "项目编号二",
    "库存位置",
    "备注一",
    "备注二",
    "零件编号",
    "库存状态",
    "原始图纸文件名",
    "导入人",
    "导入时间",
    "当前预留人",
    "预留时间",
    "领用人",
    "领用时间",
    "最后更新时间",
]


@pytest.fixture
def client(monkeypatch) -> TestClient:
    monkeypatch.setattr(settings, "remnant_inventory_enabled", True)
    init_db()
    return TestClient(app, raise_server_exceptions=False)


@pytest.fixture
def worker_headers(client: TestClient) -> dict[str, str]:
    with get_test_session_factory()() as db:
        role = db.scalar(select(Role).where(Role.code == "operator"))
        assert role is not None
        worker = User(
            username="export-worker",
            real_name="导出工人",
            password_hash=hash_password("WorkerPass123"),
            roles=[role],
        )
        db.add(worker)
        db.commit()
    response = client.post(
        "/api/v1/auth/sessions",
        json={"username": "export-worker", "password": "WorkerPass123"},
    )
    assert response.status_code == 201, response.text
    return {"Authorization": f"Bearer {response.json()['data']['access_token']}"}


def _seed_remnant() -> int:
    with get_test_session_factory()() as db:
        worker = db.scalar(select(User).where(User.username == "export-worker"))
        assert worker is not None
        material = RemnantMaterial(code="Q355B", family_code="Q355", enabled=True)
        db.add(material)
        db.flush()
        source = StoredFile(
            bucket="test",
            storage_key="remnants/export-source.dxf",
            original_name="精武路余料图.dxf",
            file_ext=".dxf",
            content_type="application/dxf",
            size_bytes=100,
            sha256="1" * 64,
            uploaded_by=worker.id,
        )
        db.add(source)
        db.flush()
        batch = RemnantImportBatch(
            created_by=worker.id,
            status="confirmed",
            total_count=1,
            confirmed_count=1,
        )
        db.add(batch)
        db.flush()
        item = RemnantImportItem(
            batch_id=batch.id,
            source_file_id=source.id,
            dxf_file_id=source.id,
            source_sha256="1" * 64,
            source_ext=".dxf",
            status="confirmed",
        )
        db.add(item)
        db.flush()
        timestamp = datetime(2026, 7, 23, 9, 2, 3, tzinfo=BUSINESS_TIMEZONE)
        remnant = Remnant(
            import_item_id=item.id,
            source_file_id=source.id,
            dxf_file_id=source.id,
            source_sha256="1" * 64,
            thickness_mm="28.000",
            material_id=material.id,
            project_no="精武路外框项目",
            project_no_secondary="合同-02",
            storage_location="A区-03架",
            remark_1="待复核",
            remark_2="优先使用",
            status="used",
            imported_by=worker.id,
            confirmed_by=worker.id,
            confirmed_at=timestamp,
            used_by=worker.id,
            used_at=timestamp,
        )
        db.add(remnant)
        db.flush()
        db.add_all(
            [
                RemnantPart(remnant_id=remnant.id, part_no="JWL-1014-B-4"),
                RemnantPart(remnant_id=remnant.id, part_no="ND-1053-3"),
            ]
        )
        db.commit()
        return remnant.id


def test_worker_exports_all_remnants_as_one_styled_row_per_remnant(client, worker_headers) -> None:
    remnant_id = _seed_remnant()

    response = client.get("/api/v1/remnants/export.xlsx", headers=worker_headers)

    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "filename*=utf-8''" in response.headers["content-disposition"].lower()
    assert response.content.startswith(b"PK")
    workbook = openpyxl.load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["全部余料"]
    sheet = workbook["全部余料"]
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:R2"
    assert [cell.value for cell in sheet[1]] == HEADERS
    assert sheet["A1"].font.bold is True
    assert sheet["A1"].fill.fgColor.rgb == "001F4E78"
    assert sheet["A1"].alignment.wrap_text is True
    assert sheet.column_dimensions["I"].width == 42
    values = [cell.value for cell in sheet[2]]
    assert values[:12] == [
        remnant_id,
        "Q355B",
        28,
        "精武路外框项目",
        "合同-02",
        "A区-03架",
        "待复核",
        "优先使用",
        "JWL-1014-B-4、ND-1053-3",
        "已领用",
        "精武路余料图.dxf",
        "导出工人",
    ]
    assert values[12] == datetime(2026, 7, 23, 9, 2, 3)
    assert values[13:16] == [None, None, "导出工人"]
    assert values[16] == datetime(2026, 7, 23, 9, 2, 3)
    assert isinstance(values[17], datetime)
    assert sheet["D2"].alignment.wrap_text is True
    assert sheet["I2"].alignment.wrap_text is True
    with get_test_session_factory()() as db:
        assert db.scalar(select(AuditLog).where(AuditLog.action == "remnants.export")) is not None


def test_export_neutralizes_formula_like_worker_text(client, worker_headers) -> None:
    remnant_id = _seed_remnant()
    with get_test_session_factory()() as db:
        row = db.get(Remnant, remnant_id)
        assert row is not None
        row.project_no = "=1+1"
        row.project_no_secondary = " +SUM(A1:A2)"
        row.storage_location = "-2+3"
        row.remark_1 = "@危险公式"
        db.commit()

    response = client.get("/api/v1/remnants/export.xlsx", headers=worker_headers)

    assert response.status_code == 200, response.text
    sheet = openpyxl.load_workbook(BytesIO(response.content))["全部余料"]
    assert [sheet.cell(2, column).value for column in range(4, 8)] == [
        "'=1+1",
        "' +SUM(A1:A2)",
        "'-2+3",
        "'@危险公式",
    ]
    assert all(sheet.cell(2, column).data_type == "s" for column in range(4, 8))


def test_empty_inventory_export_still_contains_the_header(client, worker_headers) -> None:
    response = client.get("/api/v1/remnants/export.xlsx", headers=worker_headers)

    assert response.status_code == 200, response.text
    workbook = openpyxl.load_workbook(BytesIO(response.content))
    sheet = workbook["全部余料"]
    assert sheet.max_row == 1
    assert [cell.value for cell in sheet[1]] == HEADERS


def test_export_requires_authentication(client) -> None:
    response = client.get("/api/v1/remnants/export.xlsx")

    assert response.status_code == 401


def _run_file_response(response: CleanupFileResponse, range_header: bytes | None = None) -> None:
    headers = [] if range_header is None else [(b"range", range_header)]
    scope = {
        "type": "http",
        "method": "GET",
        "path": "/export.xlsx",
        "headers": headers,
    }

    async def receive():
        return {"type": "http.disconnect"}

    async def send(_message):
        return None

    asyncio.run(response(scope, receive, send))


def test_export_temp_file_is_removed_after_invalid_range(tmp_path: Path) -> None:
    export_path = tmp_path / "invalid-range.xlsx"
    export_path.write_bytes(b"xlsx")

    response = CleanupFileResponse(export_path, filename="余料库.xlsx")
    _run_file_response(response, b"not-a-valid-range")

    assert not export_path.exists()


def test_export_temp_file_is_removed_when_sending_fails(tmp_path: Path) -> None:
    export_path = tmp_path / "send-failure.xlsx"
    export_path.write_bytes(b"xlsx")
    response = CleanupFileResponse(export_path, filename="余料库.xlsx")
    scope = {"type": "http", "method": "GET", "path": "/", "headers": []}

    async def receive():
        return {"type": "http.disconnect"}

    async def failing_send(_message):
        raise OSError("client disconnected")

    with pytest.raises(OSError, match="client disconnected"):
        asyncio.run(response(scope, receive, failing_send))

    assert not export_path.exists()
