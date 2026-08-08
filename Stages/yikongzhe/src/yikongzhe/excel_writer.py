"""Excel 输出模块。

将分类结果写入 .xlsx 文件，包含"分类结果"和"统计"两个 Sheet。
"""

from __future__ import annotations

import logging
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from yikongzhe.models import DxfResult

logger = logging.getLogger(__name__)

_HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
_CELL_ALIGNMENT = Alignment(horizontal="center", vertical="center")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def write_excel(results: list[DxfResult], output_path: str) -> None:
    """将分类结果写入 Excel 文件。

    输出两个 Sheet：
    - "分类结果"：板件名称 | 图形类别
    - "统计"：各类别计数

    Args:
        results: 所有 DXF 文件的分类结果。
        output_path: 输出 .xlsx 文件路径。
    """
    wb = Workbook()

    # Sheet 1: 分类结果
    ws_main = wb.active
    ws_main.title = "分类结果"
    _write_result_sheet(ws_main, results)

    # Sheet 2: 统计
    ws_stats = wb.create_sheet("统计")
    _write_stats_sheet(ws_stats, results)

    wb.save(output_path)
    logger.info("Excel 已保存: %s", output_path)


def _write_result_sheet(ws, results: list[DxfResult]) -> None:
    """填写分类结果 Sheet。"""
    headers = ["板件名称", "图形类别"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER

    row = 2
    for dxf_result in results:
        for pc in dxf_result.parts:
            ws.cell(row=row, column=1, value=pc.part_name)
            ws.cell(row=row, column=2, value=pc.category)
            for col in (1, 2):
                c = ws.cell(row=row, column=col)
                c.alignment = _CELL_ALIGNMENT
                c.border = _THIN_BORDER
            row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15


def _write_stats_sheet(ws, results: list[DxfResult]) -> None:
    """填写统计 Sheet。"""
    headers = ["图形类别", "数量"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER

    counter: Counter[str] = Counter()
    for dxf_result in results:
        for pc in dxf_result.parts:
            counter[pc.category] += 1

    # 按定义的顺序输出
    category_order = ["方", "异", "方孔", "异孔", "方折", "异折", "方孔折", "异孔折"]
    row = 2
    for cat in category_order:
        if cat in counter:
            ws.cell(row=row, column=1, value=cat)
            ws.cell(row=row, column=2, value=counter[cat])
            for col in (1, 2):
                c = ws.cell(row=row, column=col)
                c.alignment = _CELL_ALIGNMENT
                c.border = _THIN_BORDER
            row += 1

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 10