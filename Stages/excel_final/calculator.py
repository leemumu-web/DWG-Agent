"""Steps 15-19: Insert calculation columns and compute weights.

Step 15: Insert 总数, 总长, 比重, 理单重, 理总重 between 数量 and 单净重.
Step 16: 总数 = 数量 × 构件数;  总长 = 长度 × 总数.
Step 17: Tag unnamed plates as 类型="零件".
Step 18: Plates (non-blank 类型): 理单重 = spec×width×len×0.00000785.
Step 19: Profiles (blank 类型):   lookup 比重 from 五金手册.
"""

from __future__ import annotations

import logging

from openpyxl.styles import Font

import config as cfg
from utils import safe_str, safe_float, get_headers, find_col_by_keyword, insert_column
from handbook import lookup_steel_weight

log = logging.getLogger(__name__)


def steps_15_19_calculations(wb, ws_name):
    """Run steps 15-19 on sheet *ws_name*.

    Returns (total_col, total_len_col, density_col, theo_wt_col, theo_total_wt_col).
    """
    ws = wb[ws_name]
    headers = get_headers(ws)

    qty_col = find_col_by_keyword(headers, cfg.KW_数量)
    len_col = find_col_by_keyword(headers, cfg.KW_长度)
    comp_qty_col = find_col_by_keyword(headers, "构件数")
    spec_col = find_col_by_keyword(headers, "规格")
    width_col = find_col_by_keyword(headers, "宽度")
    type_col = find_col_by_keyword(headers, "类型")

    # Find 截面型材 for step-19 lookup
    sec_col = None
    for i, h in enumerate(headers):
        if "截面型材" in safe_str(h):
            sec_col = i + 1
            break

    # ---- Step 15: Insert 5 columns after 数量 ----
    insert_names = ["总数", "总长", "比重", "理单重", "理总重"]
    col_positions = {}
    insert_pos = qty_col
    for col_name in insert_names:
        insert_column(ws, insert_pos, col_name)
        insert_pos += 1
        col_positions[col_name] = insert_pos

    # Refresh after inserts
    headers = get_headers(ws)
    total_col = col_positions["总数"]
    total_len_col = col_positions["总长"]
    density_col = col_positions["比重"]
    theo_wt_col = col_positions["理单重"]
    theo_total_wt_col = col_positions["理总重"]
    qty_col = find_col_by_keyword(headers, cfg.KW_数量)
    comp_qty_col = find_col_by_keyword(headers, "构件数")
    len_col = find_col_by_keyword(headers, cfg.KW_长度)
    spec_col = find_col_by_keyword(headers, "规格")
    width_col = find_col_by_keyword(headers, "宽度")
    type_col = find_col_by_keyword(headers, "类型")

    log.info("Step 15: Inserted %s between 数量 and 单净重.", insert_names)

    # ---- Step 16: 总数 & 总长 ----
    for r in range(2, ws.max_row + 1):
        qty = safe_float(ws.cell(row=r, column=qty_col).value)
        comp_qty = safe_float(ws.cell(row=r, column=comp_qty_col).value)
        length = safe_float(ws.cell(row=r, column=len_col).value)

        if qty is not None and comp_qty is not None:
            total_val = qty * comp_qty
            ws.cell(row=r, column=total_col).value = total_val
        else:
            total_val = None

        if length is not None and total_val is not None:
            ws.cell(row=r, column=total_len_col).value = length * total_val
    log.info("Step 16: Calculated 总数 and 总长.")

    # ---- Step 17: Tag unnamed plates as 类型="零件" ----
    if spec_col and type_col:
        fixed = 0
        for r in range(2, ws.max_row + 1):
            spec_val = safe_str(ws.cell(row=r, column=spec_col).value)
            type_val = safe_str(ws.cell(row=r, column=type_col).value)
            if spec_val and not type_val:
                # Tag numeric specs as PL parts (including "6*30" merged plates)
                has_digit = spec_val[0].isdigit() or spec_val[0] == "."
                # Also tag PL/- prefixed specs that weren't fully parsed
                is_pl = spec_val.upper().startswith(("PL", "-"))
                if has_digit or is_pl:
                    ws.cell(row=r, column=type_col).value = "零件"
                    fixed += 1
        log.info("Step 17: Set 类型='零件' for %d unnamed plate rows.", fixed)

    # ---- Step 18: Plates (non-blank 类型) → 理单重/理总重 ----
    if all(c is not None for c in [spec_col, width_col, len_col, type_col]):
        fixed = 0
        for r in range(2, ws.max_row + 1):
            type_val = safe_str(ws.cell(row=r, column=type_col).value)
            if not type_val:
                continue
            spec = safe_float(ws.cell(row=r, column=spec_col).value)
            width = safe_float(ws.cell(row=r, column=width_col).value)
            length = safe_float(ws.cell(row=r, column=len_col).value)
            total = safe_float(ws.cell(row=r, column=total_col).value)

            if spec and width and length:
                theo_unit = spec * width * length * 7.85 / 1_000_000
                ws.cell(row=r, column=theo_wt_col).value = round(theo_unit, 3)
                if total:
                    ws.cell(row=r, column=theo_total_wt_col).value = round(theo_unit * total, 2)
                fixed += 1
        log.info("Step 18: Calculated 理单重/理总重 for %d plate rows.", fixed)

    # ---- Step 19: Profiles (blank 类型) → 五金手册 lookup ----
    if all(c is not None for c in [spec_col, len_col, type_col]):
        fixed = 0
        missed: list[str] = []
        red_font = Font(color="FF0000")
        import re as _re_step19

        for r in range(2, ws.max_row + 1):
            type_val = safe_str(ws.cell(row=r, column=type_col).value)
            if type_val:
                continue  # already handled in step 18

            # Prefer 截面型材 for handbook lookup
            lookup_spec = ""
            if sec_col:
                lookup_spec = safe_str(ws.cell(row=r, column=sec_col).value)
            if not lookup_spec and spec_col:
                lookup_spec = safe_str(ws.cell(row=r, column=spec_col).value)
            if not lookup_spec:
                continue

            # Skip non-profile specs: bolts, studs, summary text, single chars
            if _re_step19.match(r"^(TS|HS)\s*\d+", lookup_spec, _re_step19.I):
                continue
            if _re_step19.match(r"^M\d+", lookup_spec, _re_step19.I):
                continue
            if _re_step19.match(r"^(STUD|stud|栓钉)$", lookup_spec.strip()):
                continue
            if any(kw in lookup_spec for kw in ["合计", "总计"]):
                continue
            if len(lookup_spec) <= 2 and not lookup_spec.isdigit():
                continue

            density, source = lookup_steel_weight(lookup_spec)
            length = safe_float(ws.cell(row=r, column=len_col).value)
            total = safe_float(ws.cell(row=r, column=total_col).value)

            if density is not None:
                ws.cell(row=r, column=density_col).value = density
                if length:
                    if source and source.startswith("杂项"):
                        # MISC items (nuts, threaded rods): weight is per PIECE,
                        # not per meter.  Use the piece weight directly.
                        theo_unit = density
                    else:
                        theo_unit = length * density / 1000
                    ws.cell(row=r, column=theo_wt_col).value = round(theo_unit, 3)
                    if total:
                        ws.cell(row=r, column=theo_total_wt_col).value = round(theo_unit * total, 2)
                fixed += 1
            elif lookup_spec:
                cell = ws.cell(row=r, column=density_col)
                cell.value = "查无"
                cell.font = red_font
                missed.append(lookup_spec)

        if missed:
            ws.cell(row=1, column=density_col).font = red_font
            unique = list(dict.fromkeys(missed))
            log.warning("Step 19: %d rows with UNKNOWN profiles. Unique (%d): %s",
                        len(missed), len(unique), unique[:30])
        log.info("Step 19: Looked up 比重 for %d profile rows (missed %d).", fixed, len(missed))

    return total_col, total_len_col, density_col, theo_wt_col, theo_total_wt_col
