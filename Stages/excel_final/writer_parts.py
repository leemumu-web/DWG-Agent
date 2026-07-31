"""Write and verify the canonical six-sheet Excel Final workbook."""

from __future__ import annotations

from decimal import Decimal
from enum import Enum, StrEnum
import logging
import math
import os
from pathlib import Path
import shutil
import tempfile
from typing import Iterable, Mapping, Sequence
import unicodedata

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from domain import ComponentSourceRow, PipelineOutcome, SourcePart
from ooxml_formula import FormulaCache, patch_formula_caches
from part_builder import PartRow
from quality import IssueLevel, QualityIssue, QualityLedger
from weights import (
    CIRCULAR_HOLLOW_DENSITY_SOURCE,
    CIRCULAR_HOLLOW_LINEAR_WEIGHT_FACTOR,
)


log = logging.getLogger(__name__)

CANONICAL_SHEET_NAMES = ["原表", "清洗表", "构件表", "整理表", "part", "处理报告"]

CLEAN_HEADERS = [
    "来源sheet", "来源行", "序号", "批次", "构件编号", "构件数", "零件号",
    "原规格", "材质", "长度(mm)", "原数量", "单净重(kg)", "总净重(kg)",
    "单毛重(kg)", "总毛重(kg)", "单表面积(㎡)", "总表面积(㎡)", "分类",
]

COMPONENT_HEADERS = [
    "来源sheet", "来源行", "行类型", "小计来源行", "批次", "构件编号", "构件数", "原规格",
    "材质", "单净重(kg)", "总净重(kg)", "单毛重(kg)", "总毛重(kg)",
    "单表面积(㎡)", "总表面积(㎡)", "构件长度(mm)", "构件宽度(mm)", "构件高度(mm)",
]

ORGANIZED_HEADERS = [
    "序号", "构件编号", "导入构件编号", "构件数", "类型", "班组", "批次", "零件号",
    "导入零件号", "截面型材", "规格", "宽度", "长度(mm)", "左进(mm)", "右进(mm)",
    "下料长度(mm)", "材质", "原数量", "数量", "总数", "总长(mm)", "比重", "比重来源",
    "理单重(kg)", "理总重(kg)", "单净重(kg)", "总净重(kg)", "表净重(kg)",
    "单毛重(kg)", "总毛重(kg)", "表毛重(kg)", "净材利用率", "重量核验",
    "单表面积(㎡)", "总表面积(㎡)",
]

PART_HEADERS = [
    "导入构件编号", "导入零件号", "规格", "宽度", "下料长度", "材质",
    "汇总", "班组", "图形", "类型", "备注", "文件",
]

REPORT_HEADERS = [
    "级别", "类别", "来源位置", "构件编号",
    "零件号", "涉及字段", "说明", "建议操作",
]

_AUTO_WIDTH_SHEETS = ("清洗表", "构件表", "整理表", "part", "处理报告")
_DEFAULT_WIDTH_BOUNDS = (8, 32)
_WIDTH_BOUNDS_BY_HEADER = {
    ("处理报告", "说明"): (16, 48),
    ("处理报告", "建议操作"): (16, 48),
}
_WRAPPED_REPORT_HEADERS = ("说明", "建议操作")
_REMOVED_OUTPUT_COLUMNS = {
    "构件表": ("来源sheet", "行类型", "小计来源行"),
    "整理表": ("比重来源", "净材利用率", "重量核验"),
}

_LIGHT_RED_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")
_WARNING_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
_RED_FONT = Font(color="FF0000")
_SEVERE_FONT = Font(color="9C0006")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
_COMPONENT_SCOPED_TYPES = frozenset({
    "BH腹",
    "BH翼",
    "BOX腹",
    "BOX翼",
    "BT腹",
    "BT翼",
})


class FormulaLengthBasis(StrEnum):
    MODEL_LENGTH = "model_length"
    CUT_LENGTH = "cut_length"


def _thin_border() -> Border:
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


def _excel_value(value: object) -> object:
    if isinstance(value, Enum):
        return value.value
    return value


def _display_width(value: object) -> int:
    if value is None:
        return 0
    return max(
        (
            sum(
                2 if unicodedata.east_asian_width(character) in {"W", "F"} else 1
                for character in line
            )
            for line in str(value).splitlines()
        ),
        default=0,
    )


def _canonical_headers(ws, headers: Sequence[str]) -> None:
    border = _thin_border()
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _canonical_row(ws, row_number: int, values: Sequence[object]) -> None:
    border = _thin_border()
    for column, value in enumerate(values, start=1):
        cell = ws.cell(row=row_number, column=column, value=_excel_value(value))
        cell.border = border
        cell.alignment = Alignment(vertical="center")


def _write_clean_sheet(ws, parts: Iterable[SourcePart]) -> None:
    _canonical_headers(ws, CLEAN_HEADERS)
    for row_number, part in enumerate(parts, start=2):
        missing = set(part.invalid_fields)
        _canonical_row(ws, row_number, [
            part.source_sheet, part.source_row, part.source_seq, part.batch,
            part.component_no, None if "构件数" in missing else part.component_qty,
            None if "零件号" in missing else part.part_no,
            None if "规格" in missing else part.original_spec,
            None if "材质" in missing else part.material,
            None if "长度" in missing else part.length,
            None if "数量" in missing else part.original_qty, part.source_unit_net,
            part.source_total_net, part.source_unit_gross, part.source_total_gross,
            part.source_unit_area, part.source_total_area, part.classification,
        ])


def _write_component_sheet(ws, rows: Iterable[ComponentSourceRow]) -> None:
    _canonical_headers(ws, COMPONENT_HEADERS)
    for row_number, item in enumerate(rows, start=2):
        _canonical_row(ws, row_number, [
            item.source_sheet, item.source_row, item.kind, item.subtotal_source_row,
            item.batch, item.component_no,
            item.component_qty, item.original_spec, item.material, item.source_unit_net,
            item.source_total_net, item.source_unit_gross, item.source_total_gross,
            item.source_unit_area, item.source_total_area, item.component_length,
            item.component_width, item.component_height,
        ])


def _decimal_or_zero(value: object) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    return Decimal(str(value))


def _formula_cache_for_row(row: Mapping[str, object]) -> Decimal | None:
    if row.get("长度(mm)") in (None, ""):
        return None
    return (
        _decimal_or_zero(row.get("长度(mm)"))
        - _decimal_or_zero(row.get("左进(mm)"))
        - _decimal_or_zero(row.get("右进(mm)"))
    )


def _write_organized_sheet(
    ws,
    rows: Sequence[Mapping[str, object]],
) -> None:
    _canonical_headers(ws, ORGANIZED_HEADERS)
    weight_columns = {
        index for index, header in enumerate(ORGANIZED_HEADERS, start=1)
        if "重(kg)" in header
    }
    for row_number, item in enumerate(rows, start=2):
        cached_cut_length = _formula_cache_for_row(item)
        supplied = item.get("下料长度(mm)")
        if cached_cut_length is None and supplied not in (None, ""):
            raise ValueError(
                f"整理表来源行 {item.get('_source_row')!r} 缺少长度但提供了下料长度"
            )
        if (
            cached_cut_length is not None
            and supplied not in (None, "")
            and Decimal(str(supplied)) != cached_cut_length
        ):
            raise ValueError(
                f"整理表来源行 {item.get('_source_row')!r} 的下料长度与长度-左进-右进不一致"
            )
        values = [item.get(header) for header in ORGANIZED_HEADERS]
        _canonical_row(ws, row_number, values)
        for column in weight_columns:
            ws.cell(row=row_number, column=column).number_format = "0.000"
        ws.cell(row=row_number, column=32).number_format = "0.0000%"
        ws.cell(row=row_number, column=34).number_format = "0.00"
        ws.cell(row=row_number, column=35).number_format = "0.00"
        if item.get("比重") in {"查无", "冲突"}:
            ws.cell(row=row_number, column=22).font = _RED_FONT


def _formula_number(value: Decimal) -> str:
    return format(value, "f")


def _theory_basis_formula(
    item: Mapping[str, object],
    *,
    row_number: int,
    columns: Mapping[str, str],
    formula_length_basis: FormulaLengthBasis,
) -> str | None:
    if (
        item.get("理单重(kg)") in (None, "")
        and item.get("_stage2_status") != "manual"
    ):
        return None
    length_header = (
        "下料长度(mm)"
        if formula_length_basis is FormulaLengthBasis.CUT_LENGTH
        else "长度(mm)"
    )
    length = f"{columns[length_header]}{row_number}"
    density = f"{columns['比重']}{row_number}"
    source = str(item.get("比重来源") or "")
    if source == "plate_constant:7.85":
        spec = f"{columns['规格']}{row_number}"
        width = f"{columns['宽度']}{row_number}"
        return f"{spec}*{width}*{length}*{density}/1000000"
    return f"{density}*{length}/1000"


def _density_formula(
    item: Mapping[str, object],
    *,
    row_number: int,
    columns: Mapping[str, str],
) -> str | None:
    if item.get("比重来源") != CIRCULAR_HOLLOW_DENSITY_SOURCE:
        return None
    outer_diameter = f"{columns['规格']}{row_number}"
    wall_thickness = f"{columns['宽度']}{row_number}"
    return (
        f"=({outer_diameter}-{wall_thickness})"
        f"*{wall_thickness}*{_formula_number(CIRCULAR_HOLLOW_LINEAR_WEIGHT_FACTOR)}"
    )


def _apply_organized_formulas(
    ws,
    rows: Sequence[Mapping[str, object]],
    *,
    formula_length_basis: FormulaLengthBasis = FormulaLengthBasis.MODEL_LENGTH,
) -> dict[str, FormulaCache]:
    columns = {
        cell.value: get_column_letter(cell.column)
        for cell in ws[1]
    }
    length = columns["长度(mm)"]
    left_inset = columns["左进(mm)"]
    right_inset = columns["右进(mm)"]
    cut_length = columns["下料长度(mm)"]
    formula_length = (
        cut_length
        if formula_length_basis is FormulaLengthBasis.CUT_LENGTH
        else length
    )
    caches: dict[str, FormulaCache] = {}
    for row_number, item in enumerate(rows, start=2):
        stage2_status = item.get("_stage2_status")
        manual_stage2 = stage2_status == "manual"
        values = {
            "长度(mm)": ws[f"{length}{row_number}"].value,
            "左进(mm)": ws[f"{left_inset}{row_number}"].value,
            "右进(mm)": ws[f"{right_inset}{row_number}"].value,
        }
        cached_value = _formula_cache_for_row(values)
        cell = ws[f"{cut_length}{row_number}"]
        if manual_stage2:
            if values["长度(mm)"] in (None, ""):
                raise ValueError("BH人工补录行缺少模型长度")
            formula = (
                f'=IF(OR({left_inset}{row_number}="",'
                f'{right_inset}{row_number}=""),"",'
                f'{length}{row_number}-{left_inset}{row_number}-'
                f'{right_inset}{row_number})'
            )
            cell.value = formula
            caches[cell.coordinate] = FormulaCache(formula, None)
        elif cached_value is None:
            cell.value = None
            continue
        else:
            formula = (
                f"={length}{row_number}"
                if stage2_status == "missing"
                else (
                    f"={length}{row_number}-{left_inset}{row_number}-"
                    f"{right_inset}{row_number}"
                )
            )
            cell.value = formula
            caches[cell.coordinate] = FormulaCache(formula, cached_value)

        formula_specs: list[tuple[str, str, object]] = []
        density_formula = _density_formula(
            item,
            row_number=row_number,
            columns=columns,
        )
        if density_formula is not None:
            formula_specs.append(("比重", density_formula, item["比重"]))
        component_qty = f"{columns['构件数']}{row_number}"
        quantity = f"{columns['数量']}{row_number}"
        total_count = f"{columns['总数']}{row_number}"
        if item.get("总数") not in (None, ""):
            formula_specs.append(("总数", f"={component_qty}*{quantity}", item["总数"]))
        if item.get("总长(mm)") not in (None, "") or manual_stage2:
            formula_specs.append((
                "总长(mm)",
                f"={formula_length}{row_number}*{total_count}",
                item["总长(mm)"],
            ))

        theory_basis = _theory_basis_formula(
            item,
            row_number=row_number,
            columns=columns,
            formula_length_basis=formula_length_basis,
        )
        if theory_basis is not None:
            formula_specs.append((
                "理单重(kg)",
                f"=ROUND({theory_basis},3)",
                item["理单重(kg)"],
            ))
            formula_specs.append((
                "理总重(kg)",
                f"=ROUND({theory_basis}*{total_count},3)",
                item["理总重(kg)"],
            ))
            if (
                "净材利用率" in columns
                and item.get("净材利用率") not in (None, "")
            ):
                unit_net = f"{columns['单净重(kg)']}{row_number}"
                utilization_basis = theory_basis
                parent_theory_unit = item.get(
                    "_material_utilization_theory_unit"
                )
                if parent_theory_unit not in (None, ""):
                    utilization_basis = _formula_number(
                        Decimal(str(parent_theory_unit))
                    )
                formula_specs.append((
                    "净材利用率",
                    f"={unit_net}/({utilization_basis})",
                    item["净材利用率"],
                ))

        for target, source_total in (
            ("表净重(kg)", "总净重(kg)"),
            ("表毛重(kg)", "总毛重(kg)"),
        ):
            if item.get(target) not in (None, ""):
                source_cell = f"{columns[source_total]}{row_number}"
                formula_specs.append((
                    target,
                    f"=ROUND({source_cell}*{component_qty},3)",
                    item[target],
                ))

        for header, formula, value in formula_specs:
            if manual_stage2 and header in {
                "总长(mm)",
                "理单重(kg)",
                "理总重(kg)",
            }:
                formula = (
                    f'=IF({cut_length}{row_number}="","",'
                    f'{formula.removeprefix("=")})'
                )
            coordinate = f"{columns[header]}{row_number}"
            ws[coordinate] = formula
            caches[coordinate] = FormulaCache(formula, value)
    return caches


def _part_matches_organized(part: PartRow, item: Mapping[str, object]) -> bool:
    if item.get("重量核验") == "严重" or item.get("类型") != part.part_type:
        return False
    if item.get("导入零件号") != part.import_part_no:
        return False
    if (
        item.get("规格") != part.spec
        or item.get("宽度") != part.width
        or item.get("下料长度(mm)") != part.cut_length
        or item.get("材质") != part.material
        or (item.get("班组") or "") != part.team
    ):
        return False
    stage2_identity = (
        part.model_length,
        part.left_setback,
        part.right_setback,
    )
    if stage2_identity != (None, None, None) and stage2_identity != (
        item.get("长度(mm)"),
        item.get("左进(mm)"),
        item.get("右进(mm)"),
    ):
        return False
    if part.part_type in _COMPONENT_SCOPED_TYPES:
        return item.get("导入构件编号") == part.import_component_no
    return not part.import_component_no


def _part_match_key(
    *,
    import_component_no: object,
    import_part_no: object,
    spec: object,
    width: object,
    cut_length: object,
    material: object,
    part_type: object,
    team: object,
    model_length: object,
    left_setback: object,
    right_setback: object,
) -> tuple[object, ...] | None:
    """Return the exact identity used by ``_part_matches_organized``.

    All canonical values are scalar.  The hashability guard keeps malformed
    inputs on the old fail-closed matcher path instead of raising a surprising
    ``TypeError`` while building the index.
    """
    key = (
        import_component_no,
        import_part_no,
        spec,
        width,
        cut_length,
        material,
        part_type,
        team or "",
        model_length,
        left_setback,
        right_setback,
    )
    try:
        hash(key)
    except TypeError:
        return None
    return key


def _organized_part_match_keys(
    item: Mapping[str, object],
) -> tuple[tuple[object, ...], ...]:
    if item.get("重量核验") == "严重":
        return ()
    part_type = item.get("类型")
    if part_type in _COMPONENT_SCOPED_TYPES:
        import_component_no = item.get("导入构件编号")
    else:
        import_component_no = ""
    common = {
        "import_component_no": import_component_no,
        "import_part_no": item.get("导入零件号"),
        "spec": item.get("规格"),
        "width": item.get("宽度"),
        "cut_length": item.get("下料长度(mm)"),
        "material": item.get("材质"),
        "part_type": part_type,
        "team": item.get("班组"),
    }
    keys = [
        _part_match_key(
            **common,
            model_length=None,
            left_setback=None,
            right_setback=None,
        ),
    ]
    stage2_key = _part_match_key(
        **common,
        model_length=item.get("长度(mm)"),
        left_setback=item.get("左进(mm)"),
        right_setback=item.get("右进(mm)"),
    )
    if stage2_key is not None and stage2_key != keys[0]:
        keys.append(stage2_key)
    return tuple(key for key in keys if key is not None)


def _part_row_match_key(part: PartRow) -> tuple[object, ...] | None:
    if part.part_type not in _COMPONENT_SCOPED_TYPES and part.import_component_no:
        return None
    return _part_match_key(
        import_component_no=(
            part.import_component_no
            if part.part_type in _COMPONENT_SCOPED_TYPES
            else ""
        ),
        import_part_no=part.import_part_no,
        spec=part.spec,
        width=part.width,
        cut_length=part.cut_length,
        material=part.material,
        part_type=part.part_type,
        team=part.team,
        model_length=part.model_length,
        left_setback=part.left_setback,
        right_setback=part.right_setback,
    )


def _apply_part_formulas(
    ws,
    parts: Sequence[PartRow],
    organized_rows: Sequence[Mapping[str, object]],
    organized_ws,
) -> dict[str, FormulaCache]:
    part_columns = {
        cell.value: get_column_letter(cell.column)
        for cell in ws[1]
    }
    organized_columns = {
        cell.value: get_column_letter(cell.column)
        for cell in organized_ws[1]
    }
    total_count_column = organized_columns["总数"]
    cut_length_column = organized_columns["下料长度(mm)"]
    organized_index: dict[tuple[object, ...], list[int]] = {}
    for source_row_number, item in enumerate(organized_rows, start=2):
        for key in _organized_part_match_keys(item):
            organized_index.setdefault(key, []).append(source_row_number)
    caches: dict[str, FormulaCache] = {}
    for part_row_number, part in enumerate(parts, start=2):
        part_key = _part_row_match_key(part)
        if part_key is None:
            source_rows = [
                source_row_number
                for source_row_number, item in enumerate(organized_rows, start=2)
                if _part_matches_organized(part, item)
            ]
        else:
            source_rows = organized_index.get(part_key, [])
        contribution = sum(
            (
                _decimal_or_zero(organized_rows[source_row_number - 2].get("总数"))
                for source_row_number in source_rows
            ),
            Decimal("0"),
        )
        if contribution != part.summary:
            raise ValueError(
                f"part {part.import_part_no!r} 的汇总无法与整理表贡献行核对: "
                f"{contribution} != {part.summary}"
            )
        references = ",".join(
            f"'整理表'!{total_count_column}{source_row_number}"
            for source_row_number in source_rows
        )
        formula = f"=SUM({references})"
        coordinate = f"{part_columns['汇总']}{part_row_number}"
        ws[coordinate] = formula
        caches[coordinate] = FormulaCache(formula, part.summary)
        if part.cut_length is None and part.model_length is not None:
            cut_references = [
                f"'整理表'!{cut_length_column}{source_row_number}"
                for source_row_number in source_rows
            ]
            if not cut_references:
                raise ValueError(
                    f"part {part.import_part_no!r} 的人工下料长度没有整理表来源行"
                )
            if any(
                organized_rows[source_row_number - 2].get("_stage2_status")
                != "manual"
                for source_row_number in source_rows
            ):
                raise ValueError(
                    f"part {part.import_part_no!r} 仅人工补录行可使用空下料长度"
                )
            first_reference = cut_references[0]
            blank_checks = [
                f'{reference}=""' for reference in cut_references
            ]
            equality_checks = [
                f"{reference}<>{first_reference}"
                for reference in cut_references[1:]
            ]
            conditions = blank_checks + equality_checks
            condition = (
                conditions[0]
                if len(conditions) == 1
                else f"OR({','.join(conditions)})"
            )
            cut_formula = f'=IF({condition},"",{first_reference})'
            cut_coordinate = (
                f"{part_columns['下料长度']}{part_row_number}"
            )
            ws[cut_coordinate] = cut_formula
            caches[cut_coordinate] = FormulaCache(cut_formula, None)
    return caches


def _write_part_sheet(ws, rows: Iterable[PartRow]) -> None:
    _canonical_headers(ws, PART_HEADERS)
    for row_number, item in enumerate(rows, start=2):
        _canonical_row(ws, row_number, [
            item.import_component_no, item.import_part_no, item.spec, item.width,
            item.cut_length, item.material, item.summary, item.team, item.graphic,
            item.part_type, None, None,
        ])


def _write_report_sheet(
    ws,
    report_rows: Sequence[Mapping[str, object]],
) -> None:
    _canonical_headers(ws, REPORT_HEADERS)
    if not report_rows:
        ws["A2"] = "无"
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        return
    for row_number, report in enumerate(report_rows, start=2):
        _canonical_row(ws, row_number, [report[header] for header in REPORT_HEADERS])
        if report["级别"] in (IssueLevel.SEVERE.value, IssueLevel.FATAL.value):
            for cell in ws[row_number]:
                cell.fill = _LIGHT_RED_FILL
                cell.font = _SEVERE_FONT
        elif report["级别"] == IssueLevel.WARNING.value:
            for cell in ws[row_number]:
                cell.fill = _WARNING_FILL


def _apply_quality_styles(
    ws,
    rows: Sequence[Mapping[str, object]],
    issues: Sequence[QualityIssue],
) -> None:
    output_rows_by_source: dict[int, list[int]] = {}
    for output_row, item in enumerate(rows, start=2):
        source_row = item.get("_source_row")
        if source_row is not None:
            output_rows_by_source.setdefault(int(source_row), []).append(output_row)

    column_by_header = {
        header: index for index, header in enumerate(ORGANIZED_HEADERS, start=1)
    }
    for output_row, item in enumerate(rows, start=2):
        stage2_status = item.get("_stage2_status")
        if stage2_status not in {"missing", "manual"}:
            continue
        headers = ["左进(mm)", "右进(mm)", "下料长度(mm)"]
        if stage2_status == "manual":
            headers.extend(("总长(mm)", "理单重(kg)", "理总重(kg)"))
        for header in headers:
            cell = ws.cell(output_row, column_by_header[header])
            cell.fill = _LIGHT_RED_FILL
            cell.font = _RED_FONT
    issue_field_to_headers = {
        "构件编号": ("构件编号", "导入构件编号"),
        "零件号": ("零件号",),
        "规格": ("截面型材",),
        "长度": ("长度(mm)",),
        "材质": ("材质",),
        "数量": ("原数量", "数量"),
        "构件数": ("构件数",),
        "比重": ("比重",),
        "理单重": ("理单重(kg)",),
        "理总重": ("理总重(kg)",),
        "单净重": ("单净重(kg)",),
        "总净重": ("总净重(kg)",),
        "表净重": ("表净重(kg)",),
        "单毛重": ("单毛重(kg)",),
        "总毛重": ("总毛重(kg)",),
        "表毛重": ("表毛重(kg)",),
        "净材利用率": ("净材利用率",),
        "单表面积": ("单表面积(㎡)",),
        "总表面积": ("总表面积(㎡)",),
    }
    for issue in issues:
        for output_row in output_rows_by_source.get(issue.source_row, []):
            if issue.level is IssueLevel.WARNING and issue.field == "比重":
                ws.cell(output_row, column_by_header["比重"]).font = _RED_FONT
            if issue.level is not IssueLevel.SEVERE:
                continue
            headers = issue_field_to_headers.get(issue.field or "", ())
            for header in headers:
                cell = ws.cell(output_row, column_by_header[header])
                cell.fill = _LIGHT_RED_FILL
                cell.font = _SEVERE_FONT
            status_cell = ws.cell(output_row, column_by_header["重量核验"])
            status_cell.fill = _LIGHT_RED_FILL
            status_cell.font = _SEVERE_FONT


def _apply_part_stage2_styles(ws, parts: Sequence[PartRow]) -> None:
    cut_length_column = PART_HEADERS.index("下料长度") + 1
    for output_row, part in enumerate(parts, start=2):
        if part.cut_length is not None or part.model_length is None:
            continue
        cell = ws.cell(output_row, cut_length_column)
        cell.fill = _LIGHT_RED_FILL
        cell.font = _RED_FONT


def _apply_clean_quality_styles(
    ws,
    parts: Sequence[SourcePart],
    issues: Sequence[QualityIssue],
) -> None:
    output_rows_by_source = {
        (part.source_sheet, part.source_row): output_row
        for output_row, part in enumerate(parts, start=2)
    }
    column_by_header = {header: index for index, header in enumerate(CLEAN_HEADERS, start=1)}
    issue_field_to_header = {
        "构件编号": "构件编号",
        "零件号": "零件号",
        "规格": "原规格",
        "长度": "长度(mm)",
        "材质": "材质",
        "数量": "原数量",
        "构件数": "构件数",
        "单净重": "单净重(kg)",
        "总净重": "总净重(kg)",
        "单毛重": "单毛重(kg)",
        "总毛重": "总毛重(kg)",
        "单表面积": "单表面积(㎡)",
        "总表面积": "总表面积(㎡)",
    }
    for issue in issues:
        if issue.level is not IssueLevel.SEVERE:
            continue
        output_row = output_rows_by_source.get((issue.source_sheet, issue.source_row))
        header = issue_field_to_header.get(issue.field or "")
        if output_row is None or header is None:
            continue
        cell = ws.cell(output_row, column_by_header[header])
        cell.fill = _LIGHT_RED_FILL
        cell.font = _SEVERE_FONT


def _format_canonical_workbook(workbook) -> None:
    for sheet_name, removed_headers in _REMOVED_OUTPUT_COLUMNS.items():
        ws = workbook[sheet_name]
        headers = [cell.value for cell in ws[1]]
        columns = sorted(
            (headers.index(header) + 1 for header in removed_headers),
            reverse=True,
        )
        for column in columns:
            ws.delete_cols(column)
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}1"

    organized = workbook["整理表"]
    organized_headers = [cell.value for cell in organized[1]]
    organized_type_column = organized_headers.index("类型") + 1
    for row in range(2, organized.max_row + 1):
        cell = organized.cell(row=row, column=organized_type_column)
        if cell.value not in _COMPONENT_SCOPED_TYPES:
            cell.value = None

    part = workbook["part"]
    part_headers = [cell.value for cell in part[1]]
    type_column = part_headers.index("类型") + 1
    notes_column = part_headers.index("备注") + 1
    file_column = part_headers.index("文件") + 1
    for row in range(1, part.max_row + 1):
        part_type = part.cell(row=row, column=type_column).value
        notes = part.cell(row=row, column=notes_column).value
        file_value = part.cell(row=row, column=file_column).value
        part.cell(row=row, column=type_column).value = notes
        part.cell(row=row, column=notes_column).value = file_value
        part.cell(row=row, column=file_column).value = (
            "类型"
            if row == 1
            else part_type if part_type in _COMPONENT_SCOPED_TYPES else None
        )
    part.auto_filter.ref = f"A1:{get_column_letter(part.max_column)}1"

    for sheet_name in _AUTO_WIDTH_SHEETS:
        ws = workbook[sheet_name]
        for column in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=column).value
            minimum, maximum = _WIDTH_BOUNDS_BY_HEADER.get(
                (sheet_name, header),
                _DEFAULT_WIDTH_BOUNDS,
            )
            content_width = max(
                _display_width(ws.cell(row=row, column=column).value)
                for row in range(1, ws.max_row + 1)
            )
            width = min(max(content_width + 2, minimum), maximum)
            ws.column_dimensions[get_column_letter(column)].width = width

    report = workbook["处理报告"]
    for header in _WRAPPED_REPORT_HEADERS:
        column = REPORT_HEADERS.index(header) + 1
        report.cell(row=1, column=column).alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        for row in range(2, report.max_row + 1):
            report.cell(row=row, column=column).alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )


def _verify_formula_caches(
    workbook_path: Path,
    sheet_name: str,
    caches: Mapping[str, FormulaCache],
) -> None:
    if not caches:
        return
    # Stream both views row-by-row instead of reopening random-access cells for
    # every coordinate. This keeps the check linear in workbook size and avoids
    # the long tail that shows up on larger stage2 runs.
    formulas = load_workbook(workbook_path, data_only=False, read_only=True)
    values = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        formula_sheet = formulas[sheet_name]
        value_sheet = values[sheet_name]
        for formula_row, value_row in zip(
            formula_sheet.iter_rows(),
            value_sheet.iter_rows(),
            strict=True,
        ):
            for formula_cell, value_cell in zip(formula_row, value_row, strict=True):
                cache = caches.get(formula_cell.coordinate)
                if cache is None:
                    continue
                if formula_cell.value != cache.formula:
                    raise ValueError(
                        f"公式回读失败: {formula_cell.coordinate}={formula_cell.value!r}"
                    )
                cached_value = value_cell.value
                if cache.value is None:
                    if cached_value is not None:
                        raise ValueError(
                            f"公式缓存应为空: {formula_cell.coordinate}={cached_value!r}"
                        )
                    continue
                if cached_value is None or not math.isclose(
                    float(cached_value),
                    float(cache.value),
                    rel_tol=1e-12,
                    abs_tol=1e-12,
                ):
                    raise ValueError(
                        f"公式缓存回读失败: {formula_cell.coordinate}={cached_value!r}"
                    )
    finally:
        formulas.close()
        values.close()


def write_canonical_workbook(
    source_path: str | Path,
    output_path: str | Path,
    *,
    cleaned_parts: Iterable[SourcePart],
    component_rows: Iterable[ComponentSourceRow],
    organized_rows: Iterable[Mapping[str, object]],
    part_rows: Iterable[PartRow],
    issues: Iterable[QualityIssue],
    internal_output_path: str | Path | None = None,
    formula_length_basis: FormulaLengthBasis = FormulaLengthBasis.MODEL_LENGTH,
) -> PipelineOutcome:
    """Write and verify the fixed six-sheet normalized workbook atomically."""
    source = Path(source_path).resolve()
    output = Path(output_path).resolve()
    internal_output = (
        Path(internal_output_path).resolve()
        if internal_output_path is not None
        else None
    )
    if not source.is_file():
        raise FileNotFoundError(source)
    if output.suffix.lower() != ".xlsx":
        raise ValueError("Excel Final output must use the .xlsx extension")
    if source == output:
        raise ValueError("source_path and output_path must be different")
    if internal_output is not None:
        if internal_output.suffix.lower() != ".xlsx":
            raise ValueError("Excel Final internal output must use the .xlsx extension")
        if internal_output in {source, output}:
            raise ValueError("internal_output_path must differ from source and output")
        internal_output.parent.mkdir(parents=True, exist_ok=True)
    output.parent.mkdir(parents=True, exist_ok=True)

    cleaned = tuple(cleaned_parts)
    components = tuple(component_rows)
    organized = tuple(organized_rows)
    parts = tuple(part_rows)
    issue_list = tuple(issues)
    ledger = QualityLedger()
    for issue in issue_list:
        ledger.add(issue)

    temp_fd, temp_name = tempfile.mkstemp(
        prefix=f".{output.stem}.", suffix=output.suffix or ".xlsx", dir=output.parent
    )
    os.close(temp_fd)
    temp_path = Path(temp_name)
    internal_temp_path: Path | None = None
    if internal_output is not None:
        internal_fd, internal_temp_name = tempfile.mkstemp(
            prefix=f".{internal_output.stem}.",
            suffix=internal_output.suffix,
            dir=internal_output.parent,
        )
        os.close(internal_fd)
        internal_temp_path = Path(internal_temp_name)
    try:
        shutil.copy2(source, temp_path)
        workbook = load_workbook(temp_path)
        try:
            source_sheet = workbook.worksheets[0]
            for ignored_sheet in workbook.worksheets[1:]:
                workbook.remove(ignored_sheet)
            source_sheet.title = "原表"
            clean_sheet = workbook.create_sheet("清洗表")
            component_sheet = workbook.create_sheet("构件表")
            organized_sheet = workbook.create_sheet("整理表")
            part_sheet = workbook.create_sheet("part")
            report_sheet = workbook.create_sheet("处理报告")

            _write_clean_sheet(clean_sheet, cleaned)
            _write_component_sheet(component_sheet, components)
            _write_organized_sheet(organized_sheet, organized)
            _write_part_sheet(part_sheet, parts)
            _write_report_sheet(report_sheet, ledger.report_rows())
            _apply_clean_quality_styles(clean_sheet, cleaned, issue_list)
            _apply_quality_styles(organized_sheet, organized, issue_list)
            _apply_part_stage2_styles(part_sheet, parts)
            internal_organized_caches = _apply_organized_formulas(
                organized_sheet,
                organized,
                formula_length_basis=formula_length_basis,
            )
            internal_part_caches = _apply_part_formulas(
                part_sheet,
                parts,
                organized,
                organized_sheet,
            )
            if internal_temp_path is not None:
                workbook.save(internal_temp_path)
            _format_canonical_workbook(workbook)
            final_organized_caches = _apply_organized_formulas(
                workbook["整理表"],
                organized,
                formula_length_basis=formula_length_basis,
            )
            final_part_caches = _apply_part_formulas(
                workbook["part"],
                parts,
                organized,
                workbook["整理表"],
            )
            workbook.save(temp_path)
        finally:
            workbook.close()

        if internal_temp_path is not None:
            patch_formula_caches(
                internal_temp_path,
                "整理表",
                internal_organized_caches,
            )
            patch_formula_caches(internal_temp_path, "part", internal_part_caches)
            _verify_formula_caches(
                internal_temp_path,
                "整理表",
                internal_organized_caches,
            )
            _verify_formula_caches(
                internal_temp_path,
                "part",
                internal_part_caches,
            )
        patch_formula_caches(temp_path, "整理表", final_organized_caches)
        patch_formula_caches(temp_path, "part", final_part_caches)
        _verify_formula_caches(temp_path, "整理表", final_organized_caches)
        _verify_formula_caches(temp_path, "part", final_part_caches)
        os.replace(temp_path, output)
        if internal_temp_path is not None and internal_output is not None:
            os.replace(internal_temp_path, internal_output)
    finally:
        temp_path.unlink(missing_ok=True)
        if internal_temp_path is not None:
            internal_temp_path.unlink(missing_ok=True)

    log.info(
        "规范输出: %d 清洗行, %d 构件行, %d 整理行, %d part 行 → %s",
        len(cleaned), len(components), len(organized), len(parts), output.name,
    )
    return ledger.to_outcome(output)
