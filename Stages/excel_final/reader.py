"""Step 0-1: Load .xls (TSV or space-delimited), detect encoding, find header,
create workbook.

Produces an openpyxl Workbook with two sheets:
  - 原表 (preserved original, cleaned)
  - 整理表 (working copy)
"""

from __future__ import annotations

import logging
import re
import unicodedata
from collections import Counter
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl

from domain import ComponentRowKind, ComponentSourceRow, SourcePart
from input_contract import (
    HeaderDetection,
    InputContractError,
    InputKind,
    detect_canonical_header,
    inspect_production_input,
    is_repeated_canonical_header,
)
from input_errors import ExcelInputIssue, input_failure
from legacy_xls import open_legacy_workbook
from quality import IssueLevel, QualityIssue

log = logging.getLogger(__name__)

# Keywords for detecting steel-table content (encoding confirmation)
_CONTENT_KWS = ["构件编号", "零件", "规格", "长度", "材质", "数量", "型材", "型 材"]


@dataclass(frozen=True, slots=True)
class CanonicalWorkbookRead:
    source_path: Path
    sheet_name: str
    header: HeaderDetection
    working_values: tuple[tuple[Any, ...], ...]
    parts: tuple[SourcePart, ...]
    component_rows: tuple[ComponentSourceRow, ...]
    issues: tuple[QualityIssue, ...]


def _working_value(value: Any) -> Any:
    if isinstance(value, str):
        return value.replace(" ", "").replace("　", "")
    return value


def _row_value(row: tuple[Any, ...], columns: dict[str, int] | Any, field: str) -> Any:
    column = columns.get(field)
    if column is None or column > len(row):
        return None
    return row[column - 1]


def _text(value: Any) -> str | None:
    if value is None:
        return None
    result = str(value)
    return result if result else None


def _decimal(
    value: Any,
    *,
    field: str,
    source_row: int,
    sheet_name: str,
) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        result = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        failure = input_failure(
            "EXCEL_INPUT_ROW_VALUE_INVALID",
            "表格中存在无法读取的数值。",
            f"请检查 {sheet_name} 第 {source_row} 行“{field}”，填写有效数字。",
            issues=(
                ExcelInputIssue.create(
                    sheet=sheet_name,
                    row=source_row,
                    field=field,
                    value=value,
                    reason="not_numeric",
                ),
            ),
        )
        raise InputContractError(
            failure,
            diagnostic=f"row {source_row} field {field} is not numeric: {value!r}",
        ) from exc
    if not result.is_finite():
        failure = input_failure(
            "EXCEL_INPUT_ROW_VALUE_INVALID",
            "表格中存在无效数值。",
            f"请检查 {sheet_name} 第 {source_row} 行“{field}”，填写有限数字。",
            issues=(
                ExcelInputIssue.create(
                    sheet=sheet_name,
                    row=source_row,
                    field=field,
                    value=value,
                    reason="not_finite",
                ),
            ),
        )
        raise InputContractError(
            failure,
            diagnostic=f"row {source_row} field {field} must be finite",
        )
    return result


def _has_part_payload(row: tuple[Any, ...], columns: Any) -> bool:
    return any(
        _row_value(row, columns, field) not in (None, "")
        for field in (
            "规格", "零件长度", "材质", "数量", "单净重", "总净重",
            "单毛重", "总毛重", "单表面积", "总表面积",
        )
    )


_STACKED_PART_FIELDS = (
    "零件号",
    "规格",
    "零件长度",
    "材质",
    "数量",
    "单净重",
    "总净重",
    "单毛重",
    "总毛重",
    "单表面积",
    "总表面积",
)


def _cell_items(value: Any) -> tuple[Any, ...]:
    """Return logical items from a cell while preserving internal blank lines."""
    if value in (None, ""):
        return ()
    if not isinstance(value, str) or "\n" not in value and "\r" not in value:
        return (value,)
    lines = value.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    while lines and lines[0] == "":
        lines.pop(0)
    while lines and lines[-1] == "":
        lines.pop()
    return tuple(lines)


def _stacked_part_count_failure(
    *,
    sheet_name: str,
    source_row: int,
    reference_field: str = "零件号",
    reference_count: int,
    mismatches: tuple[tuple[str, int, Any], ...],
) -> InputContractError:
    details = "、".join(f"{field}有 {count} 条" for field, count, _ in mismatches)
    failure = input_failure(
        "EXCEL_INPUT_MULTILINE_ROW_AMBIGUOUS",
        "同一行中的多条零件无法一一对应。",
        (
            f"请检查 {sheet_name} 第 {source_row} 行："
            f"{reference_field}有 {reference_count} 条，但{details}；"
            "请让各换行字段的条目数一致，或仅保留一个明确的共用值。"
        ),
        issues=tuple(
            ExcelInputIssue.create(
                sheet=sheet_name,
                row=source_row,
                field=field,
                value=value,
                reason="multiline_item_count_mismatch",
            )
            for field, _, value in mismatches
        ),
    )
    return InputContractError(
        failure,
        diagnostic=(
            f"row {source_row} stacked part item counts do not match: "
            f"{reference_field}={reference_count}, "
            + ", ".join(f"{field}={count}" for field, count, _ in mismatches)
        ),
    )


def _expand_stacked_part_row(
    row: tuple[Any, ...],
    columns: Any,
    *,
    sheet_name: str,
    source_row: int,
) -> tuple[tuple[Any, ...], ...]:
    """Expand Tekla cells that stack several part lines in one physical row.

    A single non-empty value is an explicit shared attribute. Empty cells remain
    empty. Itemized fields must otherwise have exactly as many entries as the
    itemized part-number cell; inconsistent rows are rejected instead of paired
    by guesswork.
    """
    field_items = {
        field: _cell_items(_row_value(row, columns, field))
        for field in _STACKED_PART_FIELDS
        if columns.get(field) is not None
    }
    multiline_counts = {
        field: len(items)
        for field, items in field_items.items()
        if len(items) > 1
    }
    if not multiline_counts:
        return (row,)

    part_items = field_items.get("零件号", ())
    part_count = len(part_items)
    if part_count < 2:
        reference_field = max(multiline_counts, key=multiline_counts.get)
        reference_count = multiline_counts[reference_field]
        mismatch_value = _row_value(row, columns, "零件号")
        raise _stacked_part_count_failure(
            sheet_name=sheet_name,
            source_row=source_row,
            reference_field=reference_field,
            reference_count=reference_count,
            mismatches=(("零件号", part_count, mismatch_value),),
        )

    mismatches = tuple(
        (
            field,
            len(items),
            _row_value(row, columns, field),
        )
        for field, items in field_items.items()
        if len(items) not in (0, 1, part_count)
    )
    if mismatches:
        raise _stacked_part_count_failure(
            sheet_name=sheet_name,
            source_row=source_row,
            reference_count=part_count,
            mismatches=mismatches,
        )

    expanded: list[tuple[Any, ...]] = []
    for item_index in range(part_count):
        values = list(row)
        for field, items in field_items.items():
            column = columns.get(field)
            if column is None:
                continue
            if not items:
                value = None
            elif len(items) == 1:
                value = items[0]
            else:
                value = items[item_index]
            values[column - 1] = value
        expanded.append(tuple(values))
    return tuple(expanded)


def _iter_canonical_rows(
    *,
    working_values: tuple[tuple[Any, ...], ...],
    header: HeaderDetection,
    sheet_name: str,
):
    for source_row, row in enumerate(
        working_values[header.row_number:],
        start=header.row_number + 1,
    ):
        if is_repeated_canonical_header(row, header.columns):
            yield source_row, row
            continue
        for expanded in _expand_stacked_part_row(
            row,
            header.columns,
            sheet_name=sheet_name,
            source_row=source_row,
        ):
            yield source_row, expanded


def _component_source_row(
    row: tuple[Any, ...],
    columns: Any,
    *,
    sheet_name: str,
    source_row: int,
    kind: ComponentRowKind,
) -> ComponentSourceRow:
    component_no = _text(_row_value(row, columns, "构件编号"))
    if not component_no:
        failure = input_failure(
            "EXCEL_INPUT_ROW_VALUE_INVALID",
            "构件行缺少构件编号。",
            f"请检查 {sheet_name} 第 {source_row} 行并填写构件编号。",
            issues=(
                ExcelInputIssue.create(
                    sheet=sheet_name,
                    row=source_row,
                    field="构件编号",
                    reason="required_value_missing",
                ),
            ),
        )
        raise InputContractError(
            failure,
            diagnostic=f"row {source_row} component source row has no component number",
        )
    return ComponentSourceRow(
        source_sheet=sheet_name,
        source_row=source_row,
        kind=kind,
        batch=_text(_row_value(row, columns, "批次")),
        component_no=component_no,
        component_qty=_decimal(
            _row_value(row, columns, "数量"),
            field="数量",
            source_row=source_row,
            sheet_name=sheet_name,
        ),
        original_spec=_text(_row_value(row, columns, "规格")),
        material=_text(_row_value(row, columns, "材质")),
        source_unit_net=_decimal(
            _row_value(row, columns, "单净重"),
            field="单净重",
            source_row=source_row,
            sheet_name=sheet_name,
        ),
        source_total_net=_decimal(
            _row_value(row, columns, "总净重"),
            field="总净重",
            source_row=source_row,
            sheet_name=sheet_name,
        ),
        source_unit_gross=_decimal(
            _row_value(row, columns, "单毛重"),
            field="单毛重",
            source_row=source_row,
            sheet_name=sheet_name,
        ),
        source_total_gross=_decimal(
            _row_value(row, columns, "总毛重"),
            field="总毛重",
            source_row=source_row,
            sheet_name=sheet_name,
        ),
        source_unit_area=_decimal(
            _row_value(row, columns, "单表面积"),
            field="单表面积",
            source_row=source_row,
            sheet_name=sheet_name,
        ),
        source_total_area=_decimal(
            _row_value(row, columns, "总表面积"),
            field="总表面积",
            source_row=source_row,
            sheet_name=sheet_name,
        ),
        component_length=_decimal(
            _row_value(row, columns, "构件长度"),
            field="构件长度",
            source_row=source_row,
            sheet_name=sheet_name,
        ),
        component_width=_decimal(
            _row_value(row, columns, "构件宽度"),
            field="构件宽度",
            source_row=source_row,
            sheet_name=sheet_name,
        ),
        component_height=_decimal(
            _row_value(row, columns, "构件高度"),
            field="构件高度",
            source_row=source_row,
            sheet_name=sheet_name,
        ),
    )


_COMPONENT_IDENTITY_FIELDS = ("batch", "component_qty", "original_spec", "material")
_COMPONENT_METRIC_FIELDS = (
    "source_unit_net",
    "source_total_net",
    "source_unit_gross",
    "source_total_gross",
    "source_unit_area",
    "source_total_area",
    "component_length",
    "component_width",
    "component_height",
)


def _summarize_component_rows(
    rows: list[ComponentSourceRow],
) -> tuple[tuple[ComponentSourceRow, ...], tuple[QualityIssue, ...]]:
    """Collapse Tekla start/subtotal records into one canonical component row."""
    grouped: dict[str, list[ComponentSourceRow]] = {}
    for row in rows:
        grouped.setdefault(row.component_no, []).append(row)

    summaries: list[ComponentSourceRow] = []
    issues: list[QualityIssue] = []
    for component_no, group in grouped.items():
        starts = [row for row in group if row.kind == ComponentRowKind.START]
        subtotals = [row for row in group if row.kind == ComponentRowKind.SUBTOTAL]
        base = starts[0] if starts else group[0]
        values: dict[str, object | None] = {}
        value_sources: dict[str, ComponentSourceRow | None] = {}

        for field in (*_COMPONENT_IDENTITY_FIELDS, *_COMPONENT_METRIC_FIELDS):
            preferred = (
                [*starts, *subtotals]
                if field in _COMPONENT_IDENTITY_FIELDS
                else [*subtotals, *starts]
            )
            present = [getattr(row, field) for row in preferred if getattr(row, field) is not None]
            selected = present[0] if present else None
            values[field] = selected
            value_sources[field] = next(
                (row for row in preferred if getattr(row, field) == selected),
                None,
            ) if selected is not None else None
            conflicts = [value for value in present[1:] if value != selected]
            if conflicts:
                conflict_row = next(
                    row
                    for row in preferred
                    if getattr(row, field) is not None and getattr(row, field) != selected
                )
                issues.append(QualityIssue(
                    level=IssueLevel.SEVERE,
                    category="构件编号冲突",
                    source_sheet=conflict_row.source_sheet,
                    source_row=conflict_row.source_row,
                    component_no=component_no,
                    part_no=None,
                    spec=base.original_spec,
                    field=field,
                    actual_value=getattr(conflict_row, field),
                    expected_value=selected,
                    absolute_error=None,
                    relative_error=None,
                    affects_part=True,
                    density_source=None,
                    description=(
                        f"构件编号 {component_no} 的字段 {field} 在来源行中不一致"
                    ),
                ))

        summary = ComponentSourceRow(
            source_sheet=base.source_sheet,
            source_row=base.source_row,
            kind=ComponentRowKind.SUMMARY,
            component_no=component_no,
            subtotal_source_row=subtotals[0].source_row if subtotals else None,
            **values,
        )
        summaries.append(summary)

        physical_fields = (
            ("component_qty", "构件数", ">0", True),
            ("component_length", "构件长度", ">0", False),
            ("component_width", "构件宽度", ">0", False),
            ("component_height", "构件高度", ">0", False),
            ("source_unit_net", "单净重", ">=0", False),
            ("source_total_net", "总净重", ">=0", False),
            ("source_unit_gross", "单毛重", ">=0", False),
            ("source_total_gross", "总毛重", ">=0", False),
            ("source_unit_area", "单表面积", ">=0", False),
            ("source_total_area", "总表面积", ">=0", False),
        )
        for attribute, field, expected, required in physical_fields:
            value = getattr(summary, attribute)
            invalid = (
                value is not None
                and (not value.is_finite() or value < 0 or (expected == ">0" and value == 0))
            )
            if required and value is None:
                invalid = True
            if not invalid:
                continue
            source = value_sources.get(attribute) or base
            issues.append(QualityIssue(
                level=IssueLevel.SEVERE,
                category="构件物理量非法",
                source_sheet=source.source_sheet,
                source_row=source.source_row,
                component_no=component_no,
                part_no=None,
                spec=summary.original_spec,
                field=field,
                actual_value=value,
                expected_value=expected,
                absolute_error=None,
                relative_error=None,
                affects_part=True,
                density_source=None,
                description=f"构件 {component_no} 的{field}必须满足 {expected}",
            ))
    return tuple(summaries), tuple(issues)


def _canonicalize_values(
    *,
    source_path: Path,
    sheet_name: str,
    header: HeaderDetection,
    working_values: tuple[tuple[Any, ...], ...],
) -> CanonicalWorkbookRead:
    columns = header.columns
    parts: list[SourcePart] = []
    component_rows: list[ComponentSourceRow] = []
    issues: list[QualityIssue] = []
    current: ComponentSourceRow | None = None

    for source_row, row in _iter_canonical_rows(
        working_values=working_values,
        header=header,
        sheet_name=sheet_name,
    ):
        if is_repeated_canonical_header(row, columns):
            continue
        batch = _text(_row_value(row, columns, "批次"))
        component_no = _text(_row_value(row, columns, "构件编号"))
        part_no = _text(_row_value(row, columns, "零件号"))
        has_total_marker = any(
            "合计" in str(value)
            for value in row
            if value not in (None, "")
        )

        if has_total_marker and not part_no:
            continue
        if part_no == "构件小计":
            # Subtotal rows carry component-level aggregate weights and
            # dimensions but often have empty component_no / batch / spec
            # columns — inherit identity from the current component start.
            if current is None:
                failure = input_failure(
                    "EXCEL_INPUT_PART_WITHOUT_COMPONENT",
                    "构件小计行前缺少所属构件起始行。",
                    (
                        f"请在 {sheet_name} 第 {source_row} 行"
                        f"构件小计前补充对应的构件起始行。"
                    ),
                    issues=(
                        ExcelInputIssue.create(
                            sheet=sheet_name,
                            row=source_row,
                            field="构件编号",
                            value="构件小计",
                            reason="subtotal_without_component",
                        ),
                    ),
                )
                raise InputContractError(
                    failure,
                    diagnostic=(
                        f"row {source_row} subtotal has no preceding component row"
                    ),
                )
            subtotal = ComponentSourceRow(
                source_sheet=sheet_name,
                source_row=source_row,
                kind=ComponentRowKind.SUBTOTAL,
                batch=current.batch,
                component_no=current.component_no,
                component_qty=current.component_qty,
                original_spec=current.original_spec,
                material=current.material,
                source_unit_net=_decimal(
                    _row_value(row, columns, "单净重"),
                    field="单净重",
                    source_row=source_row,
                    sheet_name=sheet_name,
                ),
                source_total_net=_decimal(
                    _row_value(row, columns, "总净重"),
                    field="总净重",
                    source_row=source_row,
                    sheet_name=sheet_name,
                ),
                source_unit_gross=_decimal(
                    _row_value(row, columns, "单毛重"),
                    field="单毛重",
                    source_row=source_row,
                    sheet_name=sheet_name,
                ),
                source_total_gross=_decimal(
                    _row_value(row, columns, "总毛重"),
                    field="总毛重",
                    source_row=source_row,
                    sheet_name=sheet_name,
                ),
                source_unit_area=_decimal(
                    _row_value(row, columns, "单表面积"),
                    field="单表面积",
                    source_row=source_row,
                    sheet_name=sheet_name,
                ),
                source_total_area=_decimal(
                    _row_value(row, columns, "总表面积"),
                    field="总表面积",
                    source_row=source_row,
                    sheet_name=sheet_name,
                ),
                component_length=_decimal(
                    _row_value(row, columns, "构件长度"),
                    field="构件长度",
                    source_row=source_row,
                    sheet_name=sheet_name,
                ),
                component_width=_decimal(
                    _row_value(row, columns, "构件宽度"),
                    field="构件宽度",
                    source_row=source_row,
                    sheet_name=sheet_name,
                ),
                component_height=_decimal(
                    _row_value(row, columns, "构件高度"),
                    field="构件高度",
                    source_row=source_row,
                    sheet_name=sheet_name,
                ),
            )
            component_rows.append(subtotal)
            continue
        if component_no and not part_no:
            start = _component_source_row(
                row,
                columns,
                sheet_name=sheet_name,
                source_row=source_row,
                kind=ComponentRowKind.START,
            )
            component_rows.append(start)
            current = start
            continue
        if not part_no and (current is None or not _has_part_payload(row, columns)):
            continue
        if current is None:
            failure = input_failure(
                "EXCEL_INPUT_PART_WITHOUT_COMPONENT",
                "零件明细前缺少所属构件行。",
                (
                    f"请在 {sheet_name} 第 {source_row} 行零件 "
                    f"{part_no or '<空>'} 前补充对应的构件起始行。"
                ),
                issues=(
                    ExcelInputIssue.create(
                        sheet=sheet_name,
                        row=source_row,
                        field="构件编号",
                        value=part_no,
                        reason="part_without_component",
                    ),
                ),
            )
            raise InputContractError(
                failure,
                diagnostic=(
                    f"row {source_row} part {part_no!r} has no preceding component row"
                ),
            )

        length = _decimal(
            _row_value(row, columns, "零件长度"),
            field="零件长度",
            source_row=source_row,
            sheet_name=sheet_name,
        )
        quantity = _decimal(
            _row_value(row, columns, "数量"),
            field="数量",
            source_row=source_row,
            sheet_name=sheet_name,
        )
        original_spec = _text(_row_value(row, columns, "规格"))
        material = _text(_row_value(row, columns, "材质"))
        invalid_fields = tuple(
            field
            for field, missing in (
                ("零件号", not part_no),
                ("规格", not original_spec),
                ("长度", length is None),
                ("材质", not material),
                ("数量", quantity is None),
                ("构件数", current.component_qty is None),
            )
            if missing
        )
        parts.append(SourcePart(
            source_sheet=sheet_name,
            source_row=source_row,
            source_seq=source_row - header.row_number,
            batch=batch or current.batch,
            component_no=current.component_no,
            component_qty=current.component_qty or Decimal("0"),
            part_no=part_no or "",
            original_spec=original_spec or "",
            material=material or "",
            length=length or Decimal("0"),
            original_qty=quantity or Decimal("0"),
            source_unit_net=_decimal(
                _row_value(row, columns, "单净重"),
                field="单净重",
                source_row=source_row,
                sheet_name=sheet_name,
            ),
            source_total_net=_decimal(
                _row_value(row, columns, "总净重"),
                field="总净重",
                source_row=source_row,
                sheet_name=sheet_name,
            ),
            source_unit_gross=_decimal(
                _row_value(row, columns, "单毛重"),
                field="单毛重",
                source_row=source_row,
                sheet_name=sheet_name,
            ),
            source_total_gross=_decimal(
                _row_value(row, columns, "总毛重"),
                field="总毛重",
                source_row=source_row,
                sheet_name=sheet_name,
            ),
            source_unit_area=_decimal(
                _row_value(row, columns, "单表面积"),
                field="单表面积",
                source_row=source_row,
                sheet_name=sheet_name,
            ),
            source_total_area=_decimal(
                _row_value(row, columns, "总表面积"),
                field="总表面积",
                source_row=source_row,
                sheet_name=sheet_name,
            ),
            classification=None,
            invalid_fields=invalid_fields,
        ))

    component_summaries, component_issues = _summarize_component_rows(component_rows)
    issues.extend(component_issues)
    return CanonicalWorkbookRead(
        source_path=source_path,
        sheet_name=sheet_name,
        header=header,
        working_values=working_values,
        parts=tuple(parts),
        component_rows=component_summaries,
        issues=tuple(issues),
    )


def _worksheet_values(worksheet: Any) -> tuple[tuple[Any, ...], ...]:
    return tuple(
        tuple(_working_value(value) for value in row)
        for row in worksheet.iter_rows(values_only=True)
    )


def read_canonical_workbook(path: str | Path) -> CanonicalWorkbookRead:
    """Read one reviewed worksheet into immutable canonical source records."""
    inspected = inspect_production_input(Path(path))
    if inspected.kind not in {InputKind.WORKBOOK, InputKind.LEGACY_WORKBOOK}:
        raise ValueError("canonical workbook reader requires a one-sheet workbook source")
    if inspected.sheet_name is None:
        raise ValueError("canonical workbook reader requires a named worksheet")

    workbook = (
        open_legacy_workbook(inspected.path)
        if inspected.kind is InputKind.LEGACY_WORKBOOK
        else openpyxl.load_workbook(inspected.path, read_only=True, data_only=False)
    )
    try:
        worksheet = workbook[inspected.sheet_name]
        header = detect_canonical_header(worksheet)
        working_values = _worksheet_values(worksheet)
    finally:
        workbook.close()
    return _canonicalize_values(
        source_path=inspected.path,
        sheet_name=inspected.sheet_name,
        header=header,
        working_values=working_values,
    )


def _tab_text_workbook(path: Path) -> openpyxl.Workbook | None:
    for encoding in ("utf-8-sig", "gb18030", "gbk", "gb2312"):
        try:
            text = path.read_text(encoding=encoding)
        except UnicodeError:
            continue
        if "\t" not in text:
            return None
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "原表"
        for line in text.splitlines():
            worksheet.append([value if value != "" else None for value in line.split("\t")])
        return workbook
    return None


def read_canonical_source(path: str | Path) -> CanonicalWorkbookRead:
    """Dispatch a workbook or Tekla text export into the canonical reader."""
    inspected = inspect_production_input(Path(path))
    if inspected.kind in {InputKind.WORKBOOK, InputKind.LEGACY_WORKBOOK}:
        return read_canonical_workbook(inspected.path)

    workbook = _tab_text_workbook(inspected.path)
    if workbook is None:
        workbook = _space_text_workbook(inspected.path)
    try:
        worksheet = workbook["原表"]
        header = detect_canonical_header(worksheet)
        working_values = _worksheet_values(worksheet)
    finally:
        workbook.close()
    return _canonicalize_values(
        source_path=inspected.path,
        sheet_name="原表",
        header=header,
        working_values=working_values,
    )


_FIXED_HEADER_PATTERNS = (
    (re.compile(r"构件\s*编号"), "构件编号"),
    (re.compile(r"零件\s*(?:编号|号)"), "零件号"),
    (re.compile(r"型\s*材"), "规格"),
    (re.compile(r"构件\s*名称"), "构件名称"),
    (re.compile(r"材\s*质"), "材质"),
    (re.compile(r"长度(?:\([^)]*\))?"), "长度"),
    (re.compile(r"数量"), "数量"),
    (re.compile(r"单净重(?:\([^)]*\))?"), "单净重"),
    (re.compile(r"总净重(?:\([^)]*\))?"), "总净重"),
    (re.compile(r"单毛重(?:\([^)]*\))?"), "单毛重"),
    (re.compile(r"总毛重(?:\([^)]*\))?"), "总毛重"),
    (re.compile(r"单(?:表)?面积(?:\([^)]*\))?"), "单表面积"),
    (re.compile(r"总(?:表)?面积(?:\([^)]*\))?"), "总表面积"),
    (re.compile(r"备\s*注"), "备注"),
)


def _display_width(character: str) -> int:
    if unicodedata.combining(character):
        return 0
    return 2 if unicodedata.east_asian_width(character) in {"W", "F"} else 1


def _display_offset(text: str, character_index: int) -> int:
    return sum(_display_width(character) for character in text[:character_index])


def _display_slice(text: str, start: int, end: int | None) -> str:
    result: list[str] = []
    position = 0
    for character in text:
        width = _display_width(character)
        next_position = position + width
        if next_position > start and (end is None or position < end):
            result.append(character)
        if end is not None and position >= end:
            break
        position = next_position
    return "".join(result).strip()


def _fixed_header_spans(
    header: str,
) -> tuple[tuple[str, int, int | None], ...]:
    matches: list[tuple[int, str]] = []
    for pattern, canonical in _FIXED_HEADER_PATTERNS:
        match = pattern.search(header)
        if match is not None:
            matches.append((_display_offset(header, match.start()), canonical))
    matches.sort()
    if len(matches) < 8:
        raise ValueError(
            f"fixed-width Tekla header has only {len(matches)} recognized columns"
        )
    duplicate_names = {
        name for _, name in matches if sum(other == name for _, other in matches) > 1
    }
    if duplicate_names:
        raise ValueError(
            f"fixed-width Tekla header has duplicate columns: {sorted(duplicate_names)}"
        )
    return tuple(
        (name, start, matches[index + 1][0] if index + 1 < len(matches) else None)
        for index, (start, name) in enumerate(matches)
    )


def _calibrate_fixed_spans(
    lines: list[str],
    header_index: int,
    spans: tuple[tuple[str, int, int | None], ...],
) -> tuple[tuple[str, int, int | None], ...]:
    token_starts: Counter[int] = Counter()
    for line in lines[header_index + 1:header_index + 501]:
        for match in re.finditer(r"\S+", line):
            token_starts[_display_offset(line, match.start())] += 1

    calibrated_starts: list[int] = []
    for index, (_, expected, _) in enumerate(spans):
        if index == 0:
            calibrated_starts.append(expected)
            continue
        nearby = [
            position
            for position in range(max(0, expected - 2), expected + 3)
            if token_starts[position]
        ]
        if not nearby:
            calibrated_starts.append(expected)
            continue
        calibrated_starts.append(max(
            nearby,
            key=lambda position: (
                token_starts[position],
                -abs(position - expected),
                -position,
            ),
        ))

    if any(
        right <= left
        for left, right in zip(calibrated_starts, calibrated_starts[1:])
    ):
        raise ValueError(
            f"fixed-width Tekla column starts are not increasing: {calibrated_starts}"
        )
    return tuple(
        (
            spans[index][0],
            start,
            calibrated_starts[index + 1]
            if index + 1 < len(calibrated_starts)
            else None,
        )
        for index, start in enumerate(calibrated_starts)
    )


def _decode_fixed_text(input_file: Path) -> tuple[list[str], str]:
    for encoding in ("utf-8-sig", "gb18030", "gbk", "gb2312"):
        try:
            text = input_file.read_text(encoding=encoding)
        except UnicodeError:
            continue
        if any(keyword in text[:5000] for keyword in _CONTENT_KWS):
            return text.splitlines(), encoding
    failure = input_failure(
        "EXCEL_INPUT_TEXT_UNRECOGNIZED",
        "无法识别 Tekla 文本格式的 XLS 文件。",
        (
            "请从 Tekla 重新导出包含构件和零件明细的文本格式 XLS，"
            "或将有效工作簿另存为 XLSX 后上传。"
        ),
    )
    raise InputContractError(
        failure,
        diagnostic=f"fixed-width Tekla text is unrecognized: {input_file}",
    )


def _has_component_only_header(lines: list[str]) -> bool:
    for line in lines[:100]:
        compact = "".join(line.split())
        has_component = "构件编号" in compact or "构件号" in compact
        has_spec = "截面型材" in compact or "规格" in compact
        has_core_metrics = all(
            field in compact
            for field in ("材质", "长度", "数量")
        )
        has_part = "零件编号" in compact or "零件号" in compact
        if has_component and has_spec and has_core_metrics and not has_part:
            return True
    return False


def _space_text_workbook(input_file: Path) -> openpyxl.Workbook:
    """Adapt a fixed-width Tekla export without collapsing blank columns."""
    lines, encoding = _decode_fixed_text(input_file)
    if _has_component_only_header(lines):
        failure = input_failure(
            "EXCEL_INPUT_COMPONENT_ONLY",
            "输入只有构件汇总，没有零件明细。",
            "请从 Tekla 导出包含零件号的构件零件明细清单后重新上传。",
            sheets=("原表",),
        )
        raise InputContractError(
            failure,
            diagnostic="输入只有构件汇总，没有零件明细，不能生成 Excel Final part",
        )
    candidates: list[tuple[int, tuple[tuple[str, int, int | None], ...]]] = []
    for index, line in enumerate(lines[:100]):
        try:
            spans = _fixed_header_spans(line)
        except ValueError:
            continue
        candidates.append((index, spans))
    candidate_rows = [index + 1 for index, _ in candidates]
    if not candidates:
        failure = input_failure(
            "EXCEL_INPUT_HEADER_NOT_FOUND",
            "未检测到固定宽度 Tekla 明细表的标题行。",
            "请重新从 Tekla 导出构件零件明细清单，不要手工改变列间距。",
            sheets=("原表",),
        )
        raise InputContractError(
            failure,
            diagnostic="fixed-width Tekla header is not unique: candidate_rows=[]",
        )
    if len(candidates) > 1:
        failure = input_failure(
            "EXCEL_INPUT_HEADER_AMBIGUOUS",
            "固定宽度 Tekla 文本中检测到多个标题行。",
            "请只保留一行正式列标题，并删除重复标题行。",
            issues=tuple(
                ExcelInputIssue.create(
                    sheet="原表",
                    row=row,
                    reason="ambiguous_header",
                )
                for row in candidate_rows
            ),
            sheets=("原表",),
            meta={"candidate_rows": candidate_rows},
        )
        raise InputContractError(
            failure,
            diagnostic=(
                "fixed-width Tekla header is not unique: "
                f"candidate_rows={candidate_rows}"
            ),
        )
    header_index, spans = candidates[0]
    spans = _calibrate_fixed_spans(lines, header_index, spans)
    log.info(
        "Loading %s as fixed-width text: encoding=%s header_row=%d columns=%d",
        input_file.name,
        encoding,
        header_index + 1,
        len(spans),
    )

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "原表"
    worksheet.append([name for name, _, _ in spans])
    for line in lines[header_index + 1:]:
        values = [
            _display_slice(line, start, end) or None
            for _, start, end in spans
        ]
        if any(value is not None for value in values):
            worksheet.append(values)
    return workbook
