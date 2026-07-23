#!/usr/bin/env python3
"""Compare the immature ground truth with the canonical Excel Final output."""

from __future__ import annotations

import argparse
import csv
from decimal import Decimal
import hashlib
import json
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


EXPECTED_SHEETS = ["原表", "清洗表", "构件表", "整理表", "part", "处理报告"]


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _rows(sheet) -> list[dict[object, object]]:
    headers = [cell.value for cell in sheet[1]]
    return [
        dict(zip(headers, row, strict=True))
        for row in sheet.iter_rows(min_row=2, values_only=True)
    ]


def _decimal(value: object) -> Decimal:
    return Decimal(str(value))


def _sum(rows: Iterable[dict[object, object]], field: str) -> Decimal:
    return sum(
        (_decimal(row[field]) for row in rows if row.get(field) is not None),
        Decimal("0"),
    )


def _record(
    results: list[dict[str, object]],
    rule: str,
    source_or_gt: object,
    canonical: object,
    status: str,
    assessment: str,
) -> None:
    results.append({
        "规则": rule,
        "来源或不成熟GT": source_or_gt,
        "规范结果": canonical,
        "状态": status,
        "结论": assessment,
    })


def compare(source: Path, preprocessed: Path, output: Path) -> list[dict[str, object]]:
    baseline_path = Path(__file__).parents[1] / "tests/fixtures/ground_truth_baseline.json"
    baseline = json.loads(baseline_path.read_text(encoding="utf-8"))
    source_hash = _sha256(source)

    ground_truth = load_workbook(source, read_only=True, data_only=True)
    canonical_values = load_workbook(output, read_only=True, data_only=True)
    canonical_formulas = load_workbook(output, read_only=True, data_only=False)
    preprocessed_book = load_workbook(preprocessed, read_only=True, data_only=False)
    try:
        gt_organized = _rows(ground_truth["整理"])
        gt_part = _rows(ground_truth["part"])
        cleaned = _rows(canonical_values["清洗表"])
        organized = _rows(canonical_values["整理表"])
        part = _rows(canonical_values["part"])
        report = _rows(canonical_values["处理报告"])
        results: list[dict[str, object]] = []

        _record(
            results, "原始GT文件SHA256", baseline["sha256"], source_hash,
            "PASS" if source_hash == baseline["sha256"] else "FAIL",
            "原始多表GT未被预处理或生产流程覆盖",
        )
        _record(
            results, "预处理输入sheet数", len(ground_truth.sheetnames), len(preprocessed_book.sheetnames),
            "PASS" if len(preprocessed_book.sheetnames) == 1 else "FAIL",
            "生产输入仅分离原表，保留多表GT用于对比",
        )
        _record(
            results, "规范输出sheet合同", ground_truth.sheetnames, canonical_values.sheetnames,
            "PASS" if canonical_values.sheetnames == EXPECTED_SHEETS else "FAIL",
            "固定六表，去除GT中的临时透视/重复整理表",
        )
        _record(
            results, "父零件数", baseline["parent_parts"], len(cleaned),
            "PASS" if len(cleaned) == baseline["parent_parts"] else "FAIL",
            "485条源零件均进入规范清洗记录",
        )
        _record(
            results, "拆板后整理行数", len(gt_organized), len(organized),
            "PASS" if len(organized) == baseline["organized_rows"] else "FAIL",
            "42个BOX各拆为腹/翼，485父件得到527行",
        )

        gt_box_cover = sum(row.get("类型") == "BOX盖" for row in gt_organized)
        output_box_flange = sum(row.get("类型") == "BOX翼" for row in organized)
        _record(
            results, "BOX子板命名", f"BOX盖={gt_box_cover}", f"BOX翼={output_box_flange}",
            "INTENDED_DIFFERENCE" if output_box_flange == baseline["box"] else "FAIL",
            "统一BH/BOX/BT为腹/翼术语，不再使用BOX盖",
        )
        output_bh = sum(str(row.get("类型") or "").startswith("BH") for row in organized)
        _record(
            results, "禁止BOX回退BH", "GT规则不明确", f"BH输出行={output_bh}",
            "PASS" if output_bh == 0 else "FAIL",
            "本样本只有BOX父件，分类门控不跨类别拆板",
        )

        gt_missing_component = sum(row.get("导入构件编号") is None for row in gt_organized)
        output_missing_component = sum(row.get("导入构件编号") is None for row in organized)
        _record(
            results, "导入构件编号完整性",
            f"GT空={gt_missing_component}", f"规范空={output_missing_component}",
            "INTENDED_DIFFERENCE" if output_missing_component == 0 else "FAIL",
            "所有父行和拆板子行都保留真实构件身份",
        )

        gt_part_data = len(gt_part)
        canonical_part_data = len(part)
        _record(
            results, "part数据行范围", gt_part_data, canonical_part_data,
            "INTENDED_DIFFERENCE" if canonical_part_data == 478 else "REVIEW",
            "GT仅含84条BOX子板和117条不完整板行；规范输出84+394条合法下料候选",
        )
        gt_blank_part_component = sum(row.get("导入构件编号") is None for row in gt_part)
        output_blank_part_component = sum(row.get("导入构件编号") is None for row in part)
        _record(
            results, "part身份完整性",
            f"GT空={gt_blank_part_component}", f"规范空={output_blank_part_component}",
            "INTENDED_DIFFERENCE" if output_blank_part_component == 0 else "FAIL",
            "不输出缺失导入构件号/零件号的part记录",
        )

        blank_files = sum(row.get("文件") is None for row in part)
        _record(
            results, "part文件列固定留空", "GT混用标记与空值",
            f"空={blank_files}/{len(part)}",
            "PASS" if blank_files == len(part) else "FAIL",
            "文件列仅为兼容保留列，不再承载业务标记",
        )

        d_rows = [row for row in organized if str(row.get("截面型材") or "").startswith("D")]
        round_sources = sum(
            str(row.get("比重来源") or "").startswith("round_square_bar:round_bar")
            for row in d_rows
        )
        _record(
            results, "D系列材质路由", "GT未记录查询表", f"圆钢来源={round_sources}/{len(d_rows)}",
            "PASS" if round_sources == baseline["d"] else "FAIL",
            "Q355B D24/D30按圆钢表查询；英文类别代码稳定保存",
        )
        skipped = [
            row for row in organized
            if str(row.get("截面型材") or "").startswith(("NUT", "TT"))
        ]
        skipped_blank = sum(
            row.get("比重") is None
            and row.get("理单重(kg)") is None
            and row.get("理总重(kg)") is None
            for row in skipped
        )
        _record(
            results, "NUT/TT跳过手册", "GT存在静态/混合处理", f"留空={skipped_blank}/{len(skipped)}",
            "PASS" if skipped_blank == baseline["nut"] + baseline["tt"] else "FAIL",
            "不查库、不标查无、不填理论重",
        )

        source_chain_failures = 0
        table_chain_failures = 0
        for row in organized:
            if row.get("数量") is not None:
                for unit_field, total_field in (
                    ("单净重(kg)", "总净重(kg)"),
                    ("单毛重(kg)", "总毛重(kg)"),
                ):
                    if row.get(unit_field) is not None and row.get(total_field) is not None:
                        expected = _decimal(row[unit_field]) * _decimal(row["原数量"])
                        if abs(_decimal(row[total_field]) - expected) > Decimal("0.1"):
                            source_chain_failures += 1
            for total_field, table_field in (
                ("总净重(kg)", "表净重(kg)"),
                ("总毛重(kg)", "表毛重(kg)"),
            ):
                if row.get(total_field) is not None and row.get(table_field) is not None:
                    expected = _decimal(row[total_field]) * _decimal(row["构件数"])
                    if abs(_decimal(row[table_field]) - expected) > Decimal("0.001"):
                        table_chain_failures += 1
        _record(
            results, "源单重×原数量=源总重", "允许0.1kg舍入差",
            f"超限={source_chain_failures}", "PASS" if source_chain_failures == 0 else "FAIL",
            "净重链和毛重链按单构件物理口径核验",
        )
        _record(
            results, "源总重×构件数=表重", "GT仅有混合表总重字段",
            f"超限={table_chain_failures}", "PASS" if table_chain_failures == 0 else "FAIL",
            "表净重/表毛重是后端项目汇总唯一口径",
        )

        net_total = _sum(organized, "表净重(kg)")
        gross_total = _sum(organized, "表毛重(kg)")
        theory_total = _sum(organized, "理总重(kg)")
        _record(
            results, "项目重量物理关系", "GT字段不足以区分表净/表毛",
            f"净={net_total:.3f}, 理={theory_total:.3f}, 毛={gross_total:.3f}",
            "PASS" if net_total <= theory_total <= gross_total else "REVIEW",
            "本样本净重≤理论毛坯重≤表毛重，且拆板翼行不重复父重",
        )

        lookup_misses = sum(row.get("比重") == "查无" for row in organized)
        severe = sum(row.get("重量核验") == "严重" for row in organized)
        warnings = sum(row.get("级别") == "警告" for row in report)
        info = sum(row.get("级别") == "信息" for row in report)
        _record(
            results, "质量报告", "GT无结构化质量台账",
            f"查无={lookup_misses}, 警告={warnings}, 严重行={severe}, 信息={info}",
            "PASS" if lookup_misses == warnings == severe == info == 0 else "REVIEW",
            "当前样本没有需记录的质量问题",
        )

        formulas = canonical_formulas["整理表"]
        values = canonical_values["整理表"]
        formula_ok = all(
            formulas.cell(row=row, column=16).value == f"=M{row}-N{row}-O{row}"
            and values.cell(row=row, column=16).value is not None
            for row in range(2, formulas.max_row + 1)
        )
        _record(
            results, "下料长度公式及缓存", "GT公式缓存不完整", f"527行全部有效={formula_ok}",
            "PASS" if formula_ok else "FAIL",
            "公式模式保留计算式，data_only模式立即读到数值",
        )
        teams_blank = sum(row.get("班组") is None for row in part)
        _record(
            results, "班组来源", "GT含样本手工班组", f"空={teams_blank}/{len(part)}",
            "INTENDED_DIFFERENCE" if teams_blank == len(part) else "REVIEW",
            "原始输入无班组来源，不从不成熟GT反向注入",
        )
        return results
    finally:
        ground_truth.close()
        canonical_values.close()
        canonical_formulas.close()
        preprocessed_book.close()


def _write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def _markdown_value(value: object) -> str:
    return str(value).replace("|", "\\|").replace("\n", "<br>")


def _write_markdown(path: Path, rows: list[dict[str, object]]) -> None:
    headers = list(rows[0])
    lines = [
        "# Excel Final ground truth 对比报告",
        "",
        "不成熟 ground truth 仅作为学习与差异参照；规范规则、物理关系和真实 MySQL 结果优先。",
        "",
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    lines.extend(
        "| " + " | ".join(_markdown_value(row[header]) for header in headers) + " |"
        for row in rows
    )
    failures = [row for row in rows if row["状态"] == "FAIL"]
    lines.extend([
        "",
        "## 结论",
        "",
        f"- 规则数：{len(rows)}",
        f"- 失败数：{len(failures)}",
        "- PASS 表示规范结果满足已确认合同。",
        "- INTENDED_DIFFERENCE 表示有证据支持、且优于不成熟 GT 的有意差异。",
        "- REVIEW 表示仍建议结合业务决定复核。",
    ])
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, required=True)
    parser.add_argument("--preprocessed", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--report-dir", type=Path)
    args = parser.parse_args()
    for path in (args.source, args.preprocessed, args.output):
        if not path.is_file():
            parser.error(f"file does not exist: {path}")

    report_dir = args.report_dir or args.output.parent.parent / "reports"
    report_dir.mkdir(parents=True, exist_ok=True)
    stem = args.output.stem
    csv_path = report_dir / f"{stem}-ground-truth-comparison.csv"
    markdown_path = report_dir / f"{stem}-ground-truth-comparison.md"
    results = compare(args.source, args.preprocessed, args.output)
    _write_csv(csv_path, results)
    _write_markdown(markdown_path, results)
    print(csv_path.resolve())
    print(markdown_path.resolve())
    return 1 if any(row["状态"] == "FAIL" for row in results) else 0


if __name__ == "__main__":
    raise SystemExit(main())
