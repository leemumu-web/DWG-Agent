#!/usr/bin/env python3
"""Compare canonical output with a reviewed two-sheet supervision workbook.

The comparison is intentionally diagnostic.  It only compares cells populated
on both sides and never decides whether the canonical output or the reviewed
sample is authoritative.
"""

from __future__ import annotations

import argparse
import csv
from collections import Counter, defaultdict
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
import hashlib
import json
from pathlib import Path
from typing import Iterable, Mapping, Sequence

from openpyxl import load_workbook


SHEETS = ("整理表", "part")
ORGANIZED_DISCRIMINATORS = (
    ("序号", 8),
    ("零件号", 8),
    ("规格", 5),
    ("宽度", 5),
    ("下料长度(mm)", 5),
    ("长度(mm)", 4),
    ("材质", 4),
    ("截面型材", 3),
)
PART_DISCRIMINATORS = (
    ("规格", 5),
    ("宽度", 5),
    ("下料长度", 4),
    ("材质", 4),
)


@dataclass(frozen=True)
class SharedFieldComparison:
    compared: dict[str, tuple[object, object]]
    differences: dict[str, tuple[object, object]]


@dataclass(frozen=True)
class RowMatch:
    program: dict[str, object]
    ground_truth: dict[str, object]
    match_kind: str
    ambiguous: bool = False


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _nonempty(value: object) -> bool:
    return value is not None and value != ""


def _normalized_number(value: object) -> object:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float, Decimal)):
        decimal_value = Decimal(str(value))
        integral = decimal_value.to_integral_value()
        return int(integral) if decimal_value == integral else decimal_value.normalize()
    return value


def normalized_part_no(value: object) -> object:
    if not isinstance(value, str):
        return value
    return value.strip().replace("BOX盖", "BOX翼")


def _normalized_text(value: object) -> object:
    if isinstance(value, str):
        return normalized_part_no(value.strip())
    return _normalized_number(value)


def _precision(field: str) -> int | None:
    if "(kg)" in field or "(KG)" in field.upper():
        return 3
    if "面积" in field:
        return 2
    return None


def values_equal(program: object, ground_truth: object, *, field: str) -> bool:
    if isinstance(program, bool) or isinstance(ground_truth, bool):
        return program is ground_truth
    if isinstance(program, (int, float, Decimal)) and isinstance(
        ground_truth, (int, float, Decimal)
    ):
        left = Decimal(str(program))
        right = Decimal(str(ground_truth))
        precision = _precision(field)
        if precision is None:
            return left == right
        quantum = Decimal(1).scaleb(-precision)
        return left.quantize(quantum, rounding=ROUND_HALF_UP) == right.quantize(
            quantum, rounding=ROUND_HALF_UP
        )
    return _normalized_text(program) == _normalized_text(ground_truth)


def compare_shared_fields(
    program: Mapping[str, object],
    ground_truth: Mapping[str, object],
    *,
    fields: Iterable[str],
) -> SharedFieldComparison:
    compared: dict[str, tuple[object, object]] = {}
    differences: dict[str, tuple[object, object]] = {}
    for field in fields:
        program_value = program.get(field)
        ground_truth_value = ground_truth.get(field)
        if not (_nonempty(program_value) and _nonempty(ground_truth_value)):
            continue
        compared[field] = (program_value, ground_truth_value)
        if not values_equal(program_value, ground_truth_value, field=field):
            differences[field] = (program_value, ground_truth_value)
    return SharedFieldComparison(compared=compared, differences=differences)


def part_key(row: Mapping[str, object]) -> tuple[object, ...]:
    return (
        _normalized_text(row.get("导入构件编号")),
        normalized_part_no(row.get("导入零件号")),
        _normalized_number(row.get("规格")),
        _normalized_number(row.get("宽度")),
        _normalized_number(row.get("下料长度")),
        _normalized_text(row.get("材质")),
    )


def _identity_key(row: Mapping[str, object]) -> tuple[object, object]:
    component = row.get("导入构件编号")
    if not _nonempty(component):
        component = row.get("构件编号")
    return _normalized_text(component), normalized_part_no(row.get("导入零件号"))


def _shared_part_identity(row: Mapping[str, object]) -> object:
    return normalized_part_no(row.get("导入零件号"))


def _components_are_compatible(
    program: Mapping[str, object],
    ground_truth: Mapping[str, object],
) -> bool:
    program_component, _ = _identity_key(program)
    gt_component, _ = _identity_key(ground_truth)
    return not (_nonempty(program_component) and _nonempty(gt_component))


def _rows(sheet) -> list[dict[str, object]]:
    headers = [cell.value for cell in sheet[1]]
    if any(header is None for header in headers):
        raise ValueError(f"{sheet.title}: blank header is not supported")
    return [
        {
            **dict(zip(headers, values, strict=True)),
            "__row__": row_number,
        }
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        )
    ]


def _row_score(
    program: Mapping[str, object],
    ground_truth: Mapping[str, object],
    discriminators: Sequence[tuple[str, int]],
) -> int:
    score = 0
    for field, weight in discriminators:
        left = program.get(field)
        right = ground_truth.get(field)
        if not (_nonempty(left) and _nonempty(right)):
            continue
        score += weight if values_equal(left, right, field=field) else -weight
    return score


def _pair_group(
    program_rows: Sequence[dict[str, object]],
    ground_truth_rows: Sequence[dict[str, object]],
    *,
    match_kind: str,
    discriminators: Sequence[tuple[str, int]],
    require_compatible_components: bool = False,
) -> tuple[list[RowMatch], list[dict[str, object]], list[dict[str, object]]]:
    remaining_program = list(program_rows)
    remaining_gt = list(ground_truth_rows)
    matches: list[RowMatch] = []
    while remaining_program and remaining_gt:
        candidates = [
            (
                _row_score(program, gt, discriminators),
                int(program["__row__"]),
                int(gt["__row__"]),
                program,
                gt,
            )
            for program in remaining_program
            for gt in remaining_gt
            if not require_compatible_components
            or _components_are_compatible(program, gt)
        ]
        if not candidates:
            break
        candidates.sort(key=lambda item: (-item[0], item[1], item[2]))
        best = candidates[0]
        best_score = best[0]
        tied_for_program = sum(
            score == best_score and program is best[3]
            for score, _, _, program, _ in candidates
        )
        tied_for_gt = sum(
            score == best_score and gt is best[4]
            for score, _, _, _, gt in candidates
        )
        matches.append(
            RowMatch(
                program=best[3],
                ground_truth=best[4],
                match_kind=match_kind,
                ambiguous=tied_for_program > 1 or tied_for_gt > 1,
            )
        )
        remaining_program.remove(best[3])
        remaining_gt.remove(best[4])
    return matches, remaining_program, remaining_gt


def _match_rows(
    sheet_name: str,
    program_rows: Sequence[dict[str, object]],
    ground_truth_rows: Sequence[dict[str, object]],
) -> tuple[list[RowMatch], list[dict[str, object]], list[dict[str, object]]]:
    matches: list[RowMatch] = []
    remaining_program = list(program_rows)
    remaining_gt = list(ground_truth_rows)

    if sheet_name == "part":
        program_by_key: dict[tuple[object, ...], list[dict[str, object]]] = defaultdict(list)
        gt_by_key: dict[tuple[object, ...], list[dict[str, object]]] = defaultdict(list)
        for row in remaining_program:
            program_by_key[part_key(row)].append(row)
        for row in remaining_gt:
            gt_by_key[part_key(row)].append(row)
        for key in sorted(set(program_by_key) & set(gt_by_key), key=str):
            paired, _, _ = _pair_group(
                program_by_key[key],
                gt_by_key[key],
                match_kind="完整参数键",
                discriminators=PART_DISCRIMINATORS,
            )
            matches.extend(paired)
            for match in paired:
                remaining_program.remove(match.program)
                remaining_gt.remove(match.ground_truth)

    program_by_identity: dict[tuple[object, object], list[dict[str, object]]] = defaultdict(list)
    gt_by_identity: dict[tuple[object, object], list[dict[str, object]]] = defaultdict(list)
    for row in remaining_program:
        program_by_identity[_identity_key(row)].append(row)
    for row in remaining_gt:
        gt_by_identity[_identity_key(row)].append(row)

    discriminators = (
        ORGANIZED_DISCRIMINATORS if sheet_name == "整理表" else PART_DISCRIMINATORS
    )
    for key in sorted(set(program_by_identity) & set(gt_by_identity), key=str):
        paired, _, _ = _pair_group(
            program_by_identity[key],
            gt_by_identity[key],
            match_kind="身份消歧",
            discriminators=discriminators,
        )
        matches.extend(paired)
        for match in paired:
            remaining_program.remove(match.program)
            remaining_gt.remove(match.ground_truth)

    program_by_part: dict[object, list[dict[str, object]]] = defaultdict(list)
    gt_by_part: dict[object, list[dict[str, object]]] = defaultdict(list)
    for row in remaining_program:
        part_identity = _shared_part_identity(row)
        if _nonempty(part_identity):
            program_by_part[part_identity].append(row)
    for row in remaining_gt:
        part_identity = _shared_part_identity(row)
        if _nonempty(part_identity):
            gt_by_part[part_identity].append(row)
    for key in sorted(set(program_by_part) & set(gt_by_part), key=str):
        paired, _, _ = _pair_group(
            program_by_part[key],
            gt_by_part[key],
            match_kind="共享零件身份消歧",
            discriminators=discriminators,
            require_compatible_components=True,
        )
        matches.extend(paired)
        for match in paired:
            remaining_program.remove(match.program)
            remaining_gt.remove(match.ground_truth)

    return matches, remaining_program, remaining_gt


def compare_workbooks(
    canonical_path: Path,
    ground_truth_path: Path,
) -> tuple[list[dict[str, object]], dict[str, object]]:
    canonical = load_workbook(canonical_path, read_only=True, data_only=True)
    ground_truth = load_workbook(ground_truth_path, read_only=True, data_only=True)
    try:
        missing_canonical = [sheet for sheet in SHEETS if sheet not in canonical.sheetnames]
        missing_gt = [sheet for sheet in SHEETS if sheet not in ground_truth.sheetnames]
        if missing_canonical or missing_gt:
            raise ValueError(
                f"missing sheets: canonical={missing_canonical}, ground_truth={missing_gt}"
            )

        details: list[dict[str, object]] = []
        sheet_summaries: dict[str, object] = {}
        for sheet_name in SHEETS:
            program_rows = _rows(canonical[sheet_name])
            gt_rows = _rows(ground_truth[sheet_name])
            matches, unmatched_program, unmatched_gt = _match_rows(
                sheet_name, program_rows, gt_rows
            )
            common_fields = [
                header.value
                for header in canonical[sheet_name][1]
                if header.value in {cell.value for cell in ground_truth[sheet_name][1]}
            ]
            compared_cells = 0
            different_cells = 0
            differences_by_field: Counter[str] = Counter()
            for match in matches:
                comparison = compare_shared_fields(
                    match.program,
                    match.ground_truth,
                    fields=common_fields,
                )
                for field, (program_value, gt_value) in comparison.compared.items():
                    status = "DIFFERENT" if field in comparison.differences else "EQUAL"
                    compared_cells += 1
                    if status == "DIFFERENT":
                        different_cells += 1
                        differences_by_field[field] += 1
                    details.append({
                        "sheet": sheet_name,
                        "program_row": match.program["__row__"],
                        "ground_truth_row": match.ground_truth["__row__"],
                        "match_kind": match.match_kind,
                        "ambiguous": match.ambiguous,
                        "field": field,
                        "program_value": program_value,
                        "ground_truth_value": gt_value,
                        "status": status,
                    })
            for row in unmatched_program:
                details.append({
                    "sheet": sheet_name,
                    "program_row": row["__row__"],
                    "ground_truth_row": None,
                    "match_kind": "未匹配",
                    "ambiguous": False,
                    "field": None,
                    "program_value": _identity_key(row),
                    "ground_truth_value": None,
                    "status": "UNMATCHED_PROGRAM",
                })
            for row in unmatched_gt:
                details.append({
                    "sheet": sheet_name,
                    "program_row": None,
                    "ground_truth_row": row["__row__"],
                    "match_kind": "未匹配",
                    "ambiguous": False,
                    "field": None,
                    "program_value": None,
                    "ground_truth_value": _identity_key(row),
                    "status": "UNMATCHED_GT",
                })
            sheet_summaries[sheet_name] = {
                "program_rows": len(program_rows),
                "ground_truth_rows": len(gt_rows),
                "matched_rows": len(matches),
                "ambiguous_matches": sum(match.ambiguous for match in matches),
                "unmatched_program_rows": len(unmatched_program),
                "unmatched_ground_truth_rows": len(unmatched_gt),
                "compared_cells": compared_cells,
                "equal_cells": compared_cells - different_cells,
                "different_cells": different_cells,
                "differences_by_field": dict(sorted(differences_by_field.items())),
            }
        summary = {
            "canonical_path": str(canonical_path.resolve()),
            "canonical_sha256": _sha256(canonical_path),
            "ground_truth_path": str(ground_truth_path.resolve()),
            "ground_truth_sha256": _sha256(ground_truth_path),
            "sheets": sheet_summaries,
        }
        return details, summary
    finally:
        canonical.close()
        ground_truth.close()


def _write_csv(path: Path, details: Sequence[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "sheet",
        "program_row",
        "ground_truth_row",
        "match_kind",
        "ambiguous",
        "field",
        "program_value",
        "ground_truth_value",
        "status",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(details)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--canonical", type=Path, required=True)
    parser.add_argument("--ground-truth", type=Path, required=True)
    parser.add_argument("--csv", type=Path, required=True)
    parser.add_argument("--json", type=Path, required=True)
    args = parser.parse_args()
    for path in (args.canonical, args.ground_truth):
        if not path.is_file():
            parser.error(f"file does not exist: {path}")

    details, summary = compare_workbooks(args.canonical, args.ground_truth)
    _write_csv(args.csv, details)
    args.json.parent.mkdir(parents=True, exist_ok=True)
    args.json.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2, default=str) + "\n",
        encoding="utf-8",
    )
    print(args.csv.resolve())
    print(args.json.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
