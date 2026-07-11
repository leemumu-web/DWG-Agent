from pathlib import Path
import sys

import openpyxl
import pytest

STAGE_ROOT = Path(__file__).resolve().parents[2]
if str(STAGE_ROOT) not in sys.path:
    sys.path.insert(0, str(STAGE_ROOT))

from transformer import steps_7_9_modify  # noqa: E402


def _workbook(component_values: tuple[object, object, object, object]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "整理表"
    ws.append([
        "批次",
        "构件编号",
        "构件数",
        "零件号",
        "规格",
        "长度",
        "材质",
        "数量",
        "长度",
        "宽度",
        "高度",
        "版本",
    ])
    ws.append([
        component_values[0],
        "G1",
        1,
        "P1",
        "PL8*100",
        1200,
        "Q355B",
        2,
        component_values[1],
        component_values[2],
        component_values[3],
        None,
    ])
    return wb, ws


def test_component_only_columns_allow_blank_strings():
    wb, ws = _workbook(("", " ", "", "\t"))

    steps_7_9_modify(wb, ws)

    assert "批次" not in [cell.value for cell in ws[1]]
    assert "高度" not in [cell.value for cell in ws[1]]


def test_component_only_columns_reject_real_values():
    wb, ws = _workbook(("BATCH-1", "", "", ""))

    with pytest.raises(ValueError, match="'批次' row 2"):
        steps_7_9_modify(wb, ws)


def test_component_only_columns_report_later_non_empty_row():
    wb, ws = _workbook(("", "", "", ""))
    ws.append(["", "G1", 1, "P2", "PL8*100", 900, "Q355B", 1, "", "", "120", None])

    with pytest.raises(ValueError, match="'高度' row 3"):
        steps_7_9_modify(wb, ws)


def test_component_only_columns_allow_unicode_whitespace():
    wb, ws = _workbook(("　", "\n", " ", "\r\n"))

    steps_7_9_modify(wb, ws)

    headers = [cell.value for cell in ws[1]]
    assert "批次" not in headers
    assert "宽度" in headers


def test_ancillary_columns_may_contain_values():
    wb, ws = _workbook(("", "", "", ""))
    ws.insert_cols(9)
    ws.cell(row=1, column=9, value="备注")
    ws.cell(row=2, column=9, value="保留源备注")
    ws.insert_cols(10)
    ws.cell(row=1, column=10, value="构件名称")
    ws.cell(row=2, column=10, value="构件名称值")

    steps_7_9_modify(wb, ws)

    headers = [cell.value for cell in ws[1]]
    assert "备注" not in headers
    assert "构件名称" not in headers
