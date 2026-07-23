"""Excel I/O helpers for SunFire.

Reads Excel files into pandas DataFrames and writes DataFrames back to Excel.
Uses openpyxl under the hood -- no Excel installation required.
"""

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def read_excel(
    path: str | Path,
    sheet_name: str | int = 0,
    header_row: int | None = None,
) -> tuple[pd.DataFrame, int]:
    """Read an Excel sheet into a DataFrame with optional header auto-detection.

    Args:
        path: Path to .xlsx/.xls file.
        sheet_name: Sheet name or 0-based index.
        header_row: If provided, 0-based row to use as header.
                    If None, reads raw data and returns header_row=0
                    (caller can then use utils.detect_header_row).

    Returns:
        Tuple of (DataFrame, detected_header_row_index).
    """
    if header_row is not None:
        # Read with proper type inference; empty cells become NaN
        df = pd.read_excel(path, sheet_name=sheet_name, header=header_row)
        return df, header_row

    # Read raw (no header) so caller can run header detection.
    # Use dtype=str here to prevent pandas from guessing header row from data types.
    df = pd.read_excel(path, sheet_name=sheet_name, header=None, dtype=str)
    return df, 0


def write_excel(
    df: pd.DataFrame,
    path: str | Path,
    sheet_name: str = "Sheet1",
    column_styles: dict[str, dict] | None = None,
) -> None:
    """Write a DataFrame to an Excel file.

    Args:
        df: DataFrame to write.
        path: Output file path (.xlsx).
        sheet_name: Name of the output sheet.
        column_styles: Optional per-column styling.
            Key = column name or prefix pattern (e.g. '目标-*'),
            value = dict of openpyxl Font kwargs (e.g. {'color': '7F00FF'}).
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        if column_styles:
            ws = writer.sheets[sheet_name]

            # Build column index map
            col_map = {}
            for col_idx, col_name in enumerate(df.columns):
                col_map[str(col_name)] = col_idx + 1  # 1-based for openpyxl

                # Also map prefix patterns
                for pattern, style in column_styles.items():
                    if pattern.endswith("*") and str(col_name).startswith(pattern[:-1]):
                        col_letter = get_column_letter(col_idx + 1)
                        _apply_style_to_column(ws, col_letter, style)

            # Apply exact column name styles
            for col_name, style in column_styles.items():
                if col_name in col_map and not col_name.endswith("*"):
                    col_letter = get_column_letter(col_map[col_name])
                    _apply_style_to_column(ws, col_letter, style)


def _apply_style_to_column(ws, col_letter: str, style: dict) -> None:
    """Apply font style to all cells in a column."""
    font_kwargs = {}
    if "color" in style:
        font_kwargs["color"] = style["color"]
    if "bold" in style:
        font_kwargs["bold"] = style["bold"]
    if not font_kwargs:
        return

    font = Font(**font_kwargs)
    for row in range(2, ws.max_row + 1):  # Skip header (row 1)
        cell = ws[f"{col_letter}{row}"]
        cell.font = font
