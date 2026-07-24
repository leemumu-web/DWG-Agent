#!/usr/bin/env python3
"""Controlled extraction of the reviewed raw sheet from a comparison workbook."""

from __future__ import annotations

import argparse
import hashlib
import json
from dataclasses import asdict, dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class PreprocessResult:
    source_path: Path
    output_path: Path
    source_sha256: str
    output_sha256: str
    source_sheet_name: str
    output_sheet_name: str
    dimension: str


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _dimension_attributes(dimension: object) -> tuple[object, ...]:
    return tuple(
        getattr(dimension, name, None)
        for name in (
            "width",
            "height",
            "hidden",
            "outlineLevel",
            "collapsed",
            "bestFit",
            "style",
            "min",
            "max",
        )
    )


def _verify_equivalent(source_path: Path, output_path: Path) -> tuple[str, str]:
    source_book = load_workbook(source_path, data_only=False)
    output_book = load_workbook(output_path, data_only=False)
    try:
        if len(output_book.worksheets) != 1 or output_book.sheetnames != ["原表"]:
            raise ValueError("preprocessed workbook must contain only the 原表 sheet")

        source_sheet = source_book.worksheets[0]
        output_sheet = output_book["原表"]
        scalar_pairs = {
            "max_row": (source_sheet.max_row, output_sheet.max_row),
            "max_column": (source_sheet.max_column, output_sheet.max_column),
            "dimension": (
                source_sheet.calculate_dimension(),
                output_sheet.calculate_dimension(),
            ),
            "freeze_panes": (source_sheet.freeze_panes, output_sheet.freeze_panes),
            "show_grid_lines": (
                source_sheet.sheet_view.showGridLines,
                output_sheet.sheet_view.showGridLines,
            ),
        }
        for label, (source_value, output_value) in scalar_pairs.items():
            if source_value != output_value:
                raise ValueError(
                    f"preprocessed sheet {label} mismatch: {source_value!r} != {output_value!r}"
                )

        source_merges = {str(item) for item in source_sheet.merged_cells.ranges}
        output_merges = {str(item) for item in output_sheet.merged_cells.ranges}
        if source_merges != output_merges:
            raise ValueError("preprocessed sheet merged-cell ranges do not match source")

        for source_row, output_row in zip(source_sheet.iter_rows(), output_sheet.iter_rows()):
            for source_cell, output_cell in zip(source_row, output_row):
                source_state = (source_cell.value, source_cell.data_type, source_cell.style_id)
                output_state = (output_cell.value, output_cell.data_type, output_cell.style_id)
                if source_state != output_state:
                    raise ValueError(
                        f"preprocessed cell mismatch at {source_cell.coordinate}: "
                        f"{source_state!r} != {output_state!r}"
                    )

        column_keys = set(source_sheet.column_dimensions) | set(output_sheet.column_dimensions)
        for key in column_keys:
            source_state = _dimension_attributes(source_sheet.column_dimensions[key])
            output_state = _dimension_attributes(output_sheet.column_dimensions[key])
            if source_state != output_state:
                raise ValueError(f"preprocessed column dimension mismatch at {key}")

        row_keys = set(source_sheet.row_dimensions) | set(output_sheet.row_dimensions)
        for key in row_keys:
            source_state = _dimension_attributes(source_sheet.row_dimensions[key])
            output_state = _dimension_attributes(output_sheet.row_dimensions[key])
            if source_state != output_state:
                raise ValueError(f"preprocessed row dimension mismatch at {key}")

        return source_sheet.title, output_sheet.calculate_dimension()
    finally:
        source_book.close()
        output_book.close()


def preprocess_ground_truth(source: Path, output: Path) -> PreprocessResult:
    source = source.resolve()
    output = output.resolve()
    if source == output:
        raise ValueError("source and output must be different paths / 源文件与输出文件必须不同")
    if not source.is_file():
        raise FileNotFoundError(source)

    source_hash_before = _sha256(source)
    workbook = load_workbook(source, data_only=False)
    try:
        if not workbook.worksheets:
            raise ValueError("source workbook contains no worksheets")
        raw_sheet = workbook.worksheets[0]
        source_sheet_name = raw_sheet.title
        for sheet in list(workbook.worksheets[1:]):
            workbook.remove(sheet)
        raw_sheet.title = "原表"
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output)
    finally:
        workbook.close()

    if _sha256(source) != source_hash_before:
        raise RuntimeError("source workbook changed during preprocessing")
    verified_source_name, dimension = _verify_equivalent(source, output)
    if verified_source_name != source_sheet_name:
        raise RuntimeError("source sheet changed during preprocessing verification")

    return PreprocessResult(
        source_path=source,
        output_path=output,
        source_sha256=source_hash_before,
        output_sha256=_sha256(output),
        source_sheet_name=source_sheet_name,
        output_sheet_name="原表",
        dimension=dimension,
    )


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="提取经审查的首个原始清单工作表")
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    try:
        result = preprocess_ground_truth(args.source, args.output)
    except (OSError, ValueError, RuntimeError) as exc:
        raise SystemExit(str(exc)) from exc
    payload = asdict(result)
    payload["source_path"] = str(result.source_path)
    payload["output_path"] = str(result.output_path)
    print(json.dumps(payload, ensure_ascii=False, sort_keys=True))


if __name__ == "__main__":
    main()
