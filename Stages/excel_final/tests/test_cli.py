from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

import main
from domain import PipelineOutcome
from handbook import HandbookInfrastructureError


def _single_sheet(path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "原表"
    workbook.save(path)
    workbook.close()


def test_cli_rejects_multi_sheet_through_source_intake(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    path = tmp_path / "multi.xlsx"
    workbook = Workbook()
    workbook.active.title = "原表"
    workbook.create_sheet("part")
    workbook.save(path)
    workbook.close()

    exit_code = main.main([str(path)])
    captured = capsys.readouterr()

    assert exit_code == 2
    assert "exactly one worksheet" in captured.err


def test_cli_requires_explicit_input_without_sample_specific_fallback(
    capsys: pytest.CaptureFixture[str],
) -> None:
    exit_code = main.main([])
    captured = capsys.readouterr()

    assert exit_code == 2
    assert "用法" in captured.err
    assert "20260320" not in captured.err


def test_cli_database_failure_is_fatal_without_secret_or_traceback(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
) -> None:
    source = tmp_path / "source.xlsx"
    _single_sheet(source)

    def fail(*_args, **_kwargs):
        raise HandbookInfrastructureError(
            "connect failed host=10.0.0.8 password=do-not-print dsn=mysql://secret"
        )

    monkeypatch.setattr(main, "run_auto_pipeline", fail)

    exit_code = main.main([str(source)])
    captured = capsys.readouterr()

    assert exit_code == 2
    assert "五金手册数据库不可用" in captured.err
    assert "10.0.0.8" not in captured.err
    assert "do-not-print" not in captured.err
    assert "mysql://" not in captured.err
    assert "Traceback" not in captured.err


def test_cli_prints_quality_status_counts_and_actionable_lookup_warning(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "result.xlsx"
    _single_sheet(source)
    monkeypatch.setattr(
        main,
        "run_auto_pipeline",
        lambda *_args, **_kwargs: PipelineOutcome(
            output_path=output,
            quality_status="warning",
            warning_count=2,
            severe_warning_count=0,
            report_summary={
                "category_counts": {"五金手册查无": 1, "源重量缺失": 1},
                "representative_messages": ["I999 在工字钢表查无"],
            },
        ),
    )

    exit_code = main.main([str(source), "-o", str(output)])
    captured = capsys.readouterr()

    assert exit_code == 0
    assert "warning" in captured.out
    assert "警告=2" in captured.out
    assert "严重=0" in captured.out
    assert "五金手册查无" in captured.out
    assert "处理报告" in captured.out
    assert output.name in captured.out
    assert str(tmp_path) not in captured.out
