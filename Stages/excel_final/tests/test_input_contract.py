from __future__ import annotations

import importlib
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook


def _contract():
    try:
        return importlib.import_module("input_contract")
    except ModuleNotFoundError as exc:
        pytest.fail(f"input contract module is missing: {exc}")


def _workbook(path: Path, sheet_names: tuple[str, ...]) -> Path:
    workbook = Workbook()
    workbook.active.title = sheet_names[0]
    for name in sheet_names[1:]:
        workbook.create_sheet(name)
    workbook.save(path)
    return path


def test_production_workbook_uses_first_sheet_and_warns(tmp_path: Path) -> None:
    contract = _contract()
    source = _workbook(tmp_path / "multi.xlsx", ("原表", "整理", "part"))

    inspected = contract.inspect_production_input(source)

    assert inspected.sheet_name == "原表"
    assert inspected.warnings == (
        "检测到 3 张工作表，仅处理第一张“原表”；其余工作表已忽略：整理、part。",
    )
    assert inspected.ignored_sheets == ("整理", "part")


def test_production_single_sheet_and_tekla_text_have_distinct_kinds(tmp_path: Path) -> None:
    contract = _contract()
    workbook_source = _workbook(tmp_path / "single.xlsx", ("原表",))
    text_source = tmp_path / "tekla.xls"
    text_source.write_text("批次\t构件编号\t零件号\t规格\n", encoding="utf-8")

    workbook_input = contract.inspect_production_input(workbook_source)
    text_input = contract.inspect_production_input(text_source)

    assert workbook_input.kind is contract.InputKind.WORKBOOK
    assert workbook_input.sheet_name == "原表"
    assert text_input.kind is contract.InputKind.TEKLA_TEXT
    assert text_input.sheet_name is None


def test_corrupt_binary_xls_reports_unreadable(tmp_path: Path) -> None:
    contract = _contract()
    source = tmp_path / "legacy.xls"
    source.write_bytes(bytes.fromhex("D0CF11E0A1B11AE1") + b"\x00" * 64)

    with pytest.raises(contract.InputContractError) as caught:
        contract.inspect_production_input(source)

    failure = caught.value.failure
    assert failure.code == "EXCEL_INPUT_UNREADABLE"
    assert "另存为" in failure.action
    assert failure.issues == ()
    assert str(source.resolve()) not in str(failure.as_dict())


def test_real_binary_xls_is_registered_as_legacy_workbook() -> None:
    contract = _contract()
    source = (
        Path(__file__).resolve().parents[4]
        / "Data"
        / "十份排版"
        / "排版1"
        / "C区域四节钢柱（宝冶）"
        / "1.构件图"
        / "C区核心筒四节钢柱构件清单.xls"
    )
    if not source.is_file():
        pytest.skip("authoritative binary XLS fixture is absent")

    inspected = contract.inspect_production_input(source)

    assert inspected.kind is contract.InputKind.LEGACY_WORKBOOK
    assert inspected.sheet_name == "C区域四钢柱构件清单新"


def test_binary_workbook_reaches_schema_validation() -> None:
    source = (
        Path(__file__).resolve().parents[4]
        / "Data"
        / "十份排版"
        / "排版1"
        / "C区域四节钢柱（宝冶）"
        / "1.构件图"
        / "C区核心筒四节钢柱构件清单.xls"
    )
    if not source.is_file():
        pytest.skip("binary XLS fixture is absent")

    intake = importlib.import_module("source_intake")
    contract = _contract()
    with pytest.raises(contract.InputContractError) as caught:
        intake.read_production_source(source)

    assert caught.value.failure.code == "EXCEL_INPUT_COMPONENT_ONLY"


def test_abbreviated_duplicate_weight_pairs_map_first_to_net_and_second_to_gross() -> None:
    contract = _contract()

    columns, conflicts = contract._columns_for_header_row(
        ("构件编号", "零件号", "单重(kg)", "总重(kg)", "单重(kg)", "总重(kg)")
    )

    assert conflicts == ()
    assert columns["单净重"] == 3
    assert columns["总净重"] == 4
    assert columns["单毛重"] == 5
    assert columns["总毛重"] == 6


def test_incomplete_duplicate_weight_pair_stays_ambiguous() -> None:
    contract = _contract()

    _columns, conflicts = contract._columns_for_header_row(
        ("构件编号", "零件号", "单重(kg)", "总重(kg)", "单重(kg)")
    )

    assert conflicts == ("单毛重 columns=[3, 5]",)


@pytest.mark.parametrize(
    ("header", "expected"),
    [
        ("长度/mm", "长度"),
        ("数量-件", "数量"),
        ("单毛重[kg]", "单毛重"),
        ("总表面积【㎡】", "总表面积"),
        ("材质 kg", "材质"),
        ("规格-自定义", "规格-自定义"),
    ],
)
def test_header_units_after_canonical_name_are_ignored(
    header: str,
    expected: str,
) -> None:
    contract = _contract()

    assert contract._normalized_header(header) == expected


def test_header_detection_resolves_duplicate_length_by_group_semantics(tmp_path: Path) -> None:
    contract = _contract()
    source = tmp_path / "grouped-header.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原表"
    sheet.append(["零件清单"])
    sheet.append([])
    sheet.append([])
    sheet.append([])
    sheet.append(
        [None, None, None, None, None, None, None, "净重", None, "毛重", None,
         "表面积(㎡)", None, "构件形状尺寸", None, None, None]
    )
    sheet.append(
        [
            "批次", "构件编号", "零件号", "规格", "长度(mm)", "材质", "数量",
            "单净重(kg)", "总净重(kg)", "单毛重(kg)", "总毛重(kg)",
            "单表面积(㎡)", "总表面积(㎡)", "长度(mm)", "宽度(mm)",
            "高度(mm)", "版本",
        ]
    )
    workbook.save(source)
    workbook.close()
    loaded = load_workbook(source, read_only=True, data_only=False)
    try:
        detection = contract.detect_canonical_header(loaded["原表"])
    finally:
        loaded.close()

    assert detection.row_number == 6
    assert detection.columns == {
        "批次": 1,
        "构件编号": 2,
        "零件号": 3,
        "规格": 4,
        "零件长度": 5,
        "材质": 6,
        "数量": 7,
        "单净重": 8,
        "总净重": 9,
        "单毛重": 10,
        "总毛重": 11,
        "单表面积": 12,
        "总表面积": 13,
        "构件长度": 14,
        "构件宽度": 15,
        "构件高度": 16,
        "版本": 17,
    }


def test_header_detection_rejects_equally_complete_rows_with_diagnostics(tmp_path: Path) -> None:
    contract = _contract()
    source = tmp_path / "ambiguous-header.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    header = ["批次", "构件编号", "零件号", "规格", "长度(mm)", "材质", "数量"]
    sheet.append(["零件清单"])
    sheet.append(header)
    sheet.append(["说明"])
    sheet.append(header)
    workbook.save(source)
    workbook.close()

    loaded = load_workbook(source, read_only=True, data_only=False)
    try:
        with pytest.raises(contract.InputContractError) as caught:
            contract.detect_canonical_header(loaded.active)
    finally:
        loaded.close()

    failure = caught.value.failure
    assert failure.code == "EXCEL_INPUT_HEADER_AMBIGUOUS"
    assert failure.message == "表格中检测到多个同等有效的标题行。"
    assert failure.action == "请只保留一行正式列标题，并删除重复标题行。"
    assert [issue.row for issue in failure.issues] == [2, 4]
    assert failure.meta["candidate_rows"] == [2, 4]


def test_header_detection_rejects_incomplete_row_six_without_fallback(tmp_path: Path) -> None:
    contract = _contract()
    source = tmp_path / "incomplete-row-six.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    for _ in range(5):
        sheet.append([None])
    sheet.append(["批次", "构件编号", "零件号", "规格", "长度(mm)", "材质"])
    workbook.save(source)
    workbook.close()

    loaded = load_workbook(source, read_only=True, data_only=False)
    try:
        with pytest.raises(contract.InputContractError) as caught:
            contract.detect_canonical_header(loaded.active)
    finally:
        loaded.close()

    failure = caught.value.failure
    assert failure.code == "EXCEL_INPUT_REQUIRED_COLUMNS_MISSING"
    assert failure.message == "表格缺少 Excel 第一阶段所需列。"
    assert failure.action == "请在正式标题行中补充：数量。"
    assert failure.meta["missing_fields"] == ["数量"]
    assert failure.issues[0].sheet is None
    assert failure.issues[0].row == 6
    assert failure.issues[0].field == "数量"
    assert failure.issues[0].reason == "required_column_missing"


def test_header_detection_accepts_common_aliases_without_batch(tmp_path: Path) -> None:
    contract = _contract()
    source = tmp_path / "aliases-without-batch.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["构件号", "零件编号", "截面型材", "长度(mm)", "材质", "数量"])
    workbook.save(source)
    workbook.close()

    loaded = load_workbook(source, read_only=True, data_only=False)
    try:
        detection = contract.detect_canonical_header(loaded.active)
    finally:
        loaded.close()

    assert detection.row_number == 1
    assert detection.columns == {
        "构件编号": 1,
        "零件号": 2,
        "规格": 3,
        "零件长度": 4,
        "材质": 5,
        "数量": 6,
    }


def test_header_detection_accepts_section_spec_alias(tmp_path: Path) -> None:
    contract = _contract()
    source = tmp_path / "section-spec-alias.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["构件号", "零件编号", "截面规格", "长度(mm)", "材质", "数量"])
    workbook.save(source)
    workbook.close()

    loaded = load_workbook(source, read_only=True, data_only=False)
    try:
        detection = contract.detect_canonical_header(loaded.active)
    finally:
        loaded.close()

    assert detection.row_number == 1
    assert detection.columns == {
        "构件编号": 1,
        "零件号": 2,
        "规格": 3,
        "零件长度": 4,
        "材质": 5,
        "数量": 6,
    }


def test_header_detection_rejects_duplicate_alias_for_core_field(tmp_path: Path) -> None:
    contract = _contract()
    source = tmp_path / "duplicate-part-number.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "构件编号",
        "零件号",
        "零件编号",
        "规格",
        "长度(mm)",
        "材质",
        "数量",
    ])
    workbook.save(source)
    workbook.close()

    loaded = load_workbook(source, read_only=True, data_only=False)
    try:
        with pytest.raises(contract.InputContractError) as caught:
            contract.detect_canonical_header(loaded.active)
    finally:
        loaded.close()

    failure = caught.value.failure
    assert failure.code == "EXCEL_INPUT_DUPLICATE_COLUMNS"
    assert failure.message == "同一业务字段对应了多个标题列。"
    assert failure.meta["duplicate_fields"] == {"零件号": [2, 3]}
    assert [
        (issue.field, issue.column, issue.reason)
        for issue in failure.issues
    ] == [
        ("零件号", "B", "duplicate_column"),
        ("零件号", "C", "duplicate_column"),
    ]
