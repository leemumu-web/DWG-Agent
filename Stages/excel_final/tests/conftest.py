from __future__ import annotations

from pathlib import Path
import sys

import pytest
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


STAGE_ROOT = Path(__file__).resolve().parents[1]
if str(STAGE_ROOT) not in sys.path:
    sys.path.insert(0, str(STAGE_ROOT))


@pytest.fixture
def styled_multisheet_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "immature-ground-truth.xlsx"
    workbook = Workbook()
    raw = workbook.active
    raw.title = "原始清单"
    raw["A1"] = "钢结构零件清单"
    raw["A1"].font = Font(name="微软雅黑", size=16, bold=True, color="FF102030")
    raw["A1"].fill = PatternFill("solid", fgColor="FFFCE4D6")
    raw["A1"].alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="FF000000")
    raw["A1"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
    raw["A2"] = "=1+1"
    raw["A2"].number_format = "0.000"
    raw["B2"] = 12.5
    raw["C3"] = "PL10*200"
    raw.merge_cells("A1:C1")
    raw.column_dimensions["A"].width = 24.5
    raw.column_dimensions["C"].hidden = True
    raw.row_dimensions[1].height = 28.0
    raw.sheet_view.showGridLines = False
    raw.freeze_panes = "A2"

    comparison = workbook.create_sheet("整理")
    comparison["A1"] = "旧流程结果，不应进入预处理产物"
    part = workbook.create_sheet("part")
    part["A1"] = "旧 part，不应进入预处理产物"
    workbook.save(path)
    return path
