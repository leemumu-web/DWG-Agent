from __future__ import annotations

from decimal import Decimal
import logging
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from handbook import HandbookLookupResult, LookupStatus
from pipeline import run_auto_pipeline


class FakeHandbook:
    def __init__(self) -> None:
        self.requests: list[tuple[str, str, str | None]] = []

    def lookup(self, category, normalized_spec: str, *, material: str | None = None):
        category_value = getattr(category, "value", str(category))
        self.requests.append((category_value, normalized_spec, material))
        values = {
            ("flat_steel", "6*30"): Decimal("1.413"),
            ("round_bar", "24"): Decimal("3.55"),
            ("rebar", "24"): Decimal("3.55"),
        }
        value = values.get((category_value, normalized_spec))
        status = LookupStatus.HIT if value is not None else LookupStatus.NOT_FOUND
        table = {
            "flat_steel": "flat_steel",
            "round_bar": "round_square_bar",
            "rebar": "rebar",
            "i_beam": "i_beam",
        }.get(category_value, category_value)
        source = f"{table}:{category_value}" if value is not None else f"{table}:not_found"
        return HandbookLookupResult(category_value, normalized_spec, value, source, status)

    def log_stats(self) -> None:
        return None


class ConflictHandbook(FakeHandbook):
    def lookup(self, category, normalized_spec: str, *, material: str | None = None):
        category_value = getattr(category, "value", str(category))
        if (category_value, normalized_spec) in {
            ("flat_steel", "9*91"),
            ("i_beam", "I999"),
        }:
            self.requests.append((category_value, normalized_spec, material))
            return HandbookLookupResult(
                category_value,
                normalized_spec,
                None,
                f"{category_value}:conflict",
                LookupStatus.CONFLICT,
                ("测试源!10", "测试源!11"),
            )
        return super().lookup(category, normalized_spec, material=material)


PARTS = (
    ("p-plate", "PL10*100", "Q355B"),
    ("p-flat", "PL6*30", "Q355B"),
    ("p-bare", "9*91", "Q355B"),
    ("p-pip", "PIP2000*60", "Q355B"),
    ("p-pd", "PD114.3*6.3", "Q355B"),
    ("p-box", "BOX100*100*10*10", "Q355B"),
    ("p-round", "D24", "Q355B"),
    ("p-rebar", "D24", "HRB400"),
    ("p-nut", "NUT24", "Q355B"),
    ("p-tt", "TT25", "Q355B"),
    ("p-miss", "I999", "Q355B"),
)


def _tekla_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原表"
    sheet.append(["测试零件清单"])
    sheet.append([
        "批次", "构件编号", "零件号", "规格", "长度(mm)", "材质", "数量",
        "单净重(kg)", "总净重(kg)", "单毛重(kg)", "总毛重(kg)",
        "单表面积(㎡)", "总表面积(㎡)", "长度(mm)", "宽度(mm)",
        "高度(mm)", "版本",
    ])
    sheet.append(["B1", "C1", None, "BOX100*100*10*10", 1000, "Q355B", 2])
    for part_no, spec, material in PARTS:
        sheet.append([None, None, part_no, spec, 1000, material, 1])
    sheet.append(["B1", "C1", "构件小计", None, None, None, 2])
    workbook.save(path)
    workbook.close()


def _initial_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "初始表"
    sheet.append(["C1材  料  表构件数量：2构件总重：0"])
    sheet.append(["零件号", "截面型材", "长度", "材质", "数量", "单重", "总重", "总面积", "备注"])
    for part_no, spec, material in PARTS:
        sheet.append([part_no, spec, 1000, material, 1, None, None, None, None])
    sheet.append(["合计"])
    workbook.save(path)
    workbook.close()


def _two_component_shared_part_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原表"
    sheet.append([
        "批次", "构件编号", "零件号", "规格", "长度(mm)", "材质", "数量",
        "单净重(kg)", "总净重(kg)", "单毛重(kg)", "总毛重(kg)",
        "单表面积(㎡)", "总表面积(㎡)", "长度(mm)", "宽度(mm)",
        "高度(mm)", "版本",
    ])
    for component_no, component_qty in (("C1", 2), ("C2", 3)):
        sheet.append([
            "B1", component_no, None, "BOX100*100*10*10", 1000, "Q355B",
            component_qty,
        ])
        sheet.append([
            None, None, "p-box", "BOX100*100*10*10", 1000, "Q355B", 1,
        ])
        sheet.append([
            None, None, "p-shared", "PL10*100", 1000, "Q355B", 1,
        ])
        sheet.append([
            "B1", component_no, "构件小计", None, None, None, component_qty,
        ])
    workbook.save(path)
    workbook.close()


def _organized(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook["整理表"]
        headers = [cell.value for cell in sheet[1]]
        return [dict(zip(headers, row, strict=True)) for row in sheet.iter_rows(min_row=2, values_only=True)]
    finally:
        workbook.close()


def _cell_by_header(sheet, row: int, header: str):
    headers = [cell.value for cell in sheet[1]]
    return sheet.cell(row, headers.index(header) + 1)


def _semantic_rows(path: Path) -> list[tuple[object, ...]]:
    return [
        (
            row["零件号"], row["规格"], row["宽度"],
            row["比重"], row["理单重(kg)"], row["理总重(kg)"], row["导入零件号"],
        )
        for row in _organized(path)
    ]


def test_both_input_adapters_share_the_canonical_engine(tmp_path: Path) -> None:
    tekla = tmp_path / "tekla.xlsx"
    initial = tmp_path / "initial.xlsx"
    _tekla_workbook(tekla)
    _initial_workbook(initial)
    tekla_output = tmp_path / "tekla-output.xlsx"
    initial_output = tmp_path / "initial-output.xlsx"
    tekla_internal_output = tmp_path / "tekla-internal-output.xlsx"
    tekla_handbook = FakeHandbook()
    initial_handbook = FakeHandbook()

    tekla_outcome = run_auto_pipeline(
        tekla,
        tekla_output,
        handbook_repository=tekla_handbook,
        internal_output_file=tekla_internal_output,
    )
    initial_outcome = run_auto_pipeline(
        initial, initial_output, handbook_repository=initial_handbook
    )

    assert _semantic_rows(tekla_output) == _semantic_rows(initial_output)
    internal_workbook = load_workbook(tekla_internal_output, read_only=True)
    try:
        internal_headers = [cell.value for cell in internal_workbook["整理表"][1]]
        assert {"比重来源", "净材利用率", "重量核验"} <= set(internal_headers)
    finally:
        internal_workbook.close()
    assert tekla_outcome.output_path == tekla_output.resolve()
    assert initial_outcome.output_path == initial_output.resolve()
    assert Path(tekla_outcome) == tekla_output.resolve()
    assert Path(initial_outcome) == initial_output.resolve()
    assert tekla_outcome.quality_status == initial_outcome.quality_status == "warning"
    assert tekla_handbook.requests == initial_handbook.requests
    assert ("round_bar", "24", "Q355B") in tekla_handbook.requests
    assert ("rebar", "24", "HRB400") in tekla_handbook.requests
    assert not any(request[1] in {"NUT24", "TT25"} for request in tekla_handbook.requests)


def test_split_parent_utilization_remains_a_parent_formula_in_internal_output(
    tmp_path: Path,
) -> None:
    source = tmp_path / "box-with-source-weights.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原表"
    sheet.append([
        "批次", "构件编号", "零件号", "规格", "长度(mm)", "材质", "数量",
        "单净重(kg)", "总净重(kg)", "单毛重(kg)", "总毛重(kg)",
        "单表面积(㎡)", "总表面积(㎡)", "长度(mm)", "宽度(mm)",
        "高度(mm)", "版本",
    ])
    sheet.append([
        "B1", "C1", None, "BOX100*100*10*10", 1000, "Q355B", 1,
    ])
    sheet.append([
        None, None, "p-box", "BOX100*100*10*10", 1000, "Q355B", 1,
        28, 28, 28.26, 28.26,
    ])
    sheet.append(["B1", "C1", "构件小计", None, None, None, 1])
    workbook.save(source)
    workbook.close()

    output = tmp_path / "box-output.xlsx"
    internal = tmp_path / "box-internal.xlsx"
    outcome = run_auto_pipeline(
        source,
        output,
        handbook_repository=FakeHandbook(),
        internal_output_file=internal,
    )

    formulas = load_workbook(internal, data_only=False)
    values = load_workbook(internal, data_only=True)
    try:
        formula_sheet = formulas["整理表"]
        value_sheet = values["整理表"]
        headers = [cell.value for cell in formula_sheet[1]]
        columns = {header: index + 1 for index, header in enumerate(headers)}
        split_rows = {
            formula_sheet.cell(row, columns["类型"]).value: row
            for row in range(2, formula_sheet.max_row + 1)
        }
        web_row = split_rows["BOX腹"]
        flange_row = split_rows["BOX翼"]
        utilization_column = columns["净材利用率"]
        theory_column = get_column_letter(columns["理单重(kg)"])
        unit_net_column = get_column_letter(columns["单净重(kg)"])
        formula = formula_sheet.cell(web_row, utilization_column).value

        assert value_sheet.cell(web_row, columns["单净重(kg)"]).value == 28
        assert value_sheet.cell(flange_row, columns["单净重(kg)"]).value is None
        assert value_sheet.cell(web_row, columns["单毛重(kg)"]).value == 28.26
        assert value_sheet.cell(flange_row, columns["单毛重(kg)"]).value is None
        assert (
            value_sheet.cell(web_row, columns["理总重(kg)"]).value
            + value_sheet.cell(flange_row, columns["理总重(kg)"]).value
        ) == pytest.approx(28.26)
        assert formula.startswith(f"={unit_net_column}{web_row}/(")
        assert f"{theory_column}{web_row}" not in formula
        assert value_sheet.cell(web_row, utilization_column).value == pytest.approx(
            28 / 28.26
        )
        assert formula_sheet.cell(flange_row, utilization_column).value is None
        assert outcome.quality_status == "ok"
        report_rows = list(
            values["处理报告"].iter_rows(min_row=2, values_only=True)
        )
        assert report_rows == [("无", None, None, None, None, None, None, None)]
    finally:
        formulas.close()
        values.close()


def test_auto_pipeline_selects_standard_and_initial_sources(tmp_path: Path) -> None:
    tekla = tmp_path / "auto-tekla.xlsx"
    initial = tmp_path / "auto-initial.xlsx"
    _tekla_workbook(tekla)
    _initial_workbook(initial)

    tekla_outcome = run_auto_pipeline(
        tekla,
        tmp_path / "auto-tekla-result.xlsx",
        handbook_repository=FakeHandbook(),
    )
    initial_outcome = run_auto_pipeline(
        initial,
        tmp_path / "auto-initial-result.xlsx",
        handbook_repository=FakeHandbook(),
    )

    assert tekla_outcome.output_path.name == "auto-tekla-result.xlsx"
    assert initial_outcome.output_path.name == "auto-initial-result.xlsx"


def test_macro_enabled_input_is_normalized_to_new_xlsx_output(tmp_path: Path) -> None:
    source = tmp_path / "tekla.xlsm"
    output = tmp_path / "normalized.xlsx"
    _tekla_workbook(source)

    outcome = run_auto_pipeline(source, output, handbook_repository=FakeHandbook())

    assert outcome.output_path == output.resolve()
    assert output.is_file()


def test_pipeline_logs_only_file_names(
    tmp_path: Path, caplog: pytest.LogCaptureFixture
) -> None:
    source = tmp_path / "private-source.xlsx"
    output = tmp_path / "private-output.xlsx"
    _tekla_workbook(source)

    with caplog.at_level(logging.INFO):
        run_auto_pipeline(source, output, handbook_repository=FakeHandbook())

    assert source.name in caplog.text
    assert output.name in caplog.text
    assert str(tmp_path) not in caplog.text


@pytest.mark.parametrize("adapter", ["tekla", "initial"])
def test_pipeline_rejects_non_xlsx_output(tmp_path: Path, adapter: str) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "result.xlsm"
    if adapter == "tekla":
        _tekla_workbook(source)
        runner = run_auto_pipeline
    else:
        _initial_workbook(source)
        runner = run_auto_pipeline

    with pytest.raises(ValueError, match=r"\.xlsx"):
        runner(source, output, handbook_repository=FakeHandbook())

    assert not output.exists()


def test_canonical_pipeline_applies_lookup_split_skip_and_report_rules(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _tekla_workbook(source)
    handbook = FakeHandbook()

    outcome = run_auto_pipeline(source, output, handbook_repository=handbook)
    rows = _organized(output)
    by_part: dict[str, list[dict[str, object]]] = {}
    for row in rows:
        by_part.setdefault(str(row["零件号"]), []).append(row)

    assert by_part["p-plate"][0]["比重"] == 7.85
    assert by_part["p-flat"][0]["比重"] == 1.413
    assert by_part["p-flat"][0]["规格"] == "6*30"
    assert by_part["p-flat"][0]["宽度"] is None
    assert by_part["p-bare"][0]["比重"] == 7.85
    assert by_part["p-bare"][0]["规格"] == 9
    assert by_part["p-bare"][0]["宽度"] == 91
    assert by_part["p-pip"][0]["比重"] == pytest.approx(2870.424)
    assert by_part["p-pip"][0]["规格"] == 2000
    assert by_part["p-pip"][0]["宽度"] == 60
    assert by_part["p-pip"][0]["理单重(kg)"] == pytest.approx(2870.424)
    assert by_part["p-pd"][0]["比重"] == pytest.approx(16.778664)
    assert by_part["p-pd"][0]["规格"] == pytest.approx(114.3)
    assert by_part["p-pd"][0]["宽度"] == pytest.approx(6.3)
    assert by_part["p-pd"][0]["理单重(kg)"] == pytest.approx(16.779)
    assert [row["导入零件号"] for row in by_part["p-box"]] == [
        "p-box-BOX腹", "p-box-BOX翼",
    ]
    assert by_part["p-box"][0]["序号"] == by_part["p-box"][1]["序号"]
    # p-box（BOX100*100*10*10）拆板推导：腹板取宽 100-2*10=80，
    # 理单重 6.28 = 10×80×1000×7.85/1e6，翼缘 7.85 = 10×100×1000×7.85/1e6；
    # 数量 4 时 25.12 + 31.40 = 56.52 断言的是「拆板子件理论总重
    # 守恒于父理论重」（splitter 精确守恒规则）。
    assert by_part["p-box"][0]["理单重(kg)"] == 6.28
    assert by_part["p-box"][0]["理总重(kg)"] == 25.12
    assert by_part["p-box"][1]["理单重(kg)"] == 7.85
    assert by_part["p-box"][1]["理总重(kg)"] == 31.4
    assert by_part["p-box"][1]["比重"] == 7.85
    assert sum(row["理总重(kg)"] for row in by_part["p-box"]) == pytest.approx(56.52)
    assert by_part["p-round"][0]["规格"] == "D24"
    assert by_part["p-rebar"][0]["规格"] == "D24"
    assert ("round_bar", "24", "Q355B") in handbook.requests
    assert ("rebar", "24", "HRB400") in handbook.requests
    assert by_part["p-nut"][0]["比重"] is None
    assert by_part["p-nut"][0]["理单重(kg)"] is None
    assert by_part["p-tt"][0]["比重"] is None
    assert by_part["p-miss"][0]["比重"] == "查无"
    assert outcome.warning_count > 0
    assert not any(
        request[0] == "steel_pipe"
        or request[1] in {"PIP2000*60", "PD114.3*6.3"}
        for request in handbook.requests
    )

    workbook = load_workbook(output, data_only=False, read_only=True)
    try:
        organized_sheet = workbook["整理表"]
        headers = [cell.value for cell in organized_sheet[1]]
        rows_by_part = {
            row[headers.index("零件号")]: row
            for row in organized_sheet.iter_rows(min_row=2, values_only=True)
        }
        for part_no in ("p-pip", "p-pd"):
            density_formula = rows_by_part[part_no][headers.index("比重")]
            assert isinstance(density_formula, str)
            assert density_formula.startswith("=")
            assert "*0.02466" in density_formula
        report_rows = list(workbook["处理报告"].iter_rows(min_row=2, values_only=True))
        assert any(row[1] == "五金手册查无" and row[4] == "p-miss" for row in report_rows)
        assert not any(
            row[1] == "五金手册查无" and row[4] in {"p-pip", "p-pd"}
            for row in report_rows
        )
        assert not any(row[1] == "五金手册查无" and row[4] in {"p-nut", "p-tt"} for row in report_rows)
        part_headers = [cell.value for cell in workbook["part"][1]]
        file_index = part_headers.index("文件")
        file_cells = [row[file_index] for row in workbook["part"].iter_rows(
            min_row=2, values_only=True
        )]
        assert file_cells
        assert all(value is None for value in file_cells)
    finally:
        workbook.close()


def test_nonphysical_circular_hollow_dimensions_are_reported_without_crashing(
    tmp_path: Path,
) -> None:
    source = tmp_path / "invalid-pipe.xlsx"
    output = tmp_path / "invalid-pipe-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原表"
    sheet.append([
        "批次", "构件编号", "零件号", "规格", "长度(mm)", "材质", "数量",
    ])
    sheet.append(["B1", "C1", None, "BOX100*100*10*10", 1000, "Q355B", 1])
    sheet.append([None, None, "bad-pipe", "PD60*30", 1000, "Q355B", 1])
    sheet.append(["B1", "C1", "构件小计", None, None, None, 1])
    workbook.save(source)
    workbook.close()
    handbook = FakeHandbook()

    outcome = run_auto_pipeline(source, output, handbook_repository=handbook)

    assert output.is_file()
    rows = _organized(output)
    pipe = next(row for row in rows if row["零件号"] == "bad-pipe")
    assert pipe["比重"] is None
    assert pipe["理单重(kg)"] is None
    assert outcome.warning_count >= 1
    assert not handbook.requests
    report = load_workbook(output, data_only=True, read_only=True)
    try:
        report_rows = list(report["处理报告"].iter_rows(min_row=2, values_only=True))
        assert any(
            row[1] == "圆管规格无效"
            and row[4] == "bad-pipe"
            and "D>2t" in str(row[6])
            for row in report_rows
        )
    finally:
        report.close()


def test_handbook_conflict_is_not_downgraded_to_not_found_or_plate_fallback(
    tmp_path: Path,
) -> None:
    source = tmp_path / "conflicting-handbook.xlsx"
    output = tmp_path / "conflicting-handbook-output.xlsx"
    _tekla_workbook(source)

    run_auto_pipeline(
        source,
        output,
        handbook_repository=ConflictHandbook(),
    )

    rows = _organized(output)
    by_part = {str(row["零件号"]): row for row in rows}
    assert by_part["p-bare"]["比重"] == "冲突"
    assert by_part["p-bare"]["规格"] == "9*91"
    assert by_part["p-miss"]["比重"] == "冲突"

    workbook = load_workbook(output, data_only=True, read_only=False)
    try:
        organized_sheet = workbook["整理表"]
        organized_headers = [cell.value for cell in organized_sheet[1]]
        organized_columns = {
            header: index + 1 for index, header in enumerate(organized_headers)
        }
        bare_row = next(
            row
            for row in range(2, organized_sheet.max_row + 1)
            if organized_sheet.cell(
                row, organized_columns["零件号"]
            ).value == "p-bare"
        )
        density_font = organized_sheet.cell(
            bare_row,
            organized_columns["比重"],
        ).font.color
        assert density_font is not None
        assert density_font.rgb.endswith("FF0000")
        report_rows = list(
            workbook["处理报告"].iter_rows(min_row=2, values_only=True)
        )
        conflict_rows = [
            row for row in report_rows if row[1] == "五金手册数据冲突"
        ]
        assert {row[4] for row in conflict_rows} == {"p-bare", "p-miss"}
        assert all("不得自动选取" in str(row[6]) for row in conflict_rows)
    finally:
        workbook.close()


def test_part_projection_keeps_main_component_and_globally_merges_other_parts(
    tmp_path: Path,
) -> None:
    source = tmp_path / "two-components.xlsx"
    output = tmp_path / "two-components-output.xlsx"
    _two_component_shared_part_workbook(source)

    run_auto_pipeline(source, output, handbook_repository=FakeHandbook())

    workbook = load_workbook(output, data_only=True, read_only=True)
    try:
        headers = [cell.value for cell in workbook["part"][1]]
        rows = [
            dict(zip(headers, values, strict=True))
            for values in workbook["part"].iter_rows(min_row=2, values_only=True)
        ]
        main_rows = [row for row in rows if row["导入构件编号"] is not None]
        global_rows = [row for row in rows if row["导入构件编号"] is None]

        assert len(main_rows) == 4
        assert {row["导入构件编号"] for row in main_rows} == {"C1", "C2"}
        assert len(global_rows) == 1
        assert global_rows[0]["导入构件编号"] is None
        assert global_rows[0]["汇总"] == 5
    finally:
        workbook.close()


def test_invalid_confirmed_split_is_reported_without_dropping_source_row(tmp_path: Path) -> None:
    source = tmp_path / "invalid-box.xlsx"
    output = tmp_path / "invalid-box-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原表"
    sheet.append([
        "批次", "构件编号", "零件号", "规格", "长度(mm)", "材质", "数量",
    ])
    sheet.append(["B1", "C1", None, "BOX50*100*10*30", 1000, "Q355B", 1])
    sheet.append([None, None, "bad-box", "BOX50*100*10*30", 1000, "Q355B", 1])
    workbook.save(source)
    workbook.close()

    outcome = run_auto_pipeline(source, output, handbook_repository=FakeHandbook())
    rows = _organized(output)

    assert len(rows) == 1
    assert rows[0]["零件号"] == "bad-box"
    assert rows[0]["截面型材"] == "BOX50*100*10*30"
    assert outcome.quality_status == "severe_warning"
    result = load_workbook(output, read_only=True, data_only=True)
    try:
        assert result["part"].max_row == 1
        report = list(result["处理报告"].iter_rows(min_row=2, values_only=True))
        assert any(row[1] == "拆板几何异常" and row[4] == "bad-box" for row in report)
    finally:
        result.close()


def test_box_three_dimension_shorthand_is_consistent_end_to_end(tmp_path: Path) -> None:
    source = tmp_path / "box-short.xlsx"
    output = tmp_path / "box-short-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原表"
    sheet.append([
        "构件编号", "零件号", "规格", "长度(mm)", "材质", "数量",
        "单净重(kg)", "总净重(kg)", "单毛重(kg)", "总毛重(kg)",
    ])
    sheet.append(["C1", None, "BOX100*80*10", 1000, "Q355B", 1])
    sheet.append([
        None, "box-short", "BOX100*80*10", 1000, "Q355B", 1,
        25.00, 25.00, 25.12, 25.12,
    ])
    workbook.save(source)
    workbook.close()

    outcome = run_auto_pipeline(source, output, handbook_repository=FakeHandbook())

    assert outcome.quality_status == "ok"
    values = load_workbook(output, read_only=True, data_only=True)
    formulas = load_workbook(output, read_only=True, data_only=False)
    try:
        organized_headers = [cell.value for cell in values["整理表"][1]]
        organized_rows = [
            dict(zip(organized_headers, row, strict=True))
            for row in values["整理表"].iter_rows(min_row=2, values_only=True)
        ]
        assert len(organized_rows) == 2
        assert [row["理单重(kg)"] for row in organized_rows] == [
            pytest.approx(6.28),
            pytest.approx(6.28),
        ]
        assert sum(row["理总重(kg)"] for row in organized_rows) == pytest.approx(25.12)
        part_rows = list(values["part"].iter_rows(min_row=2, values_only=True))
        assert len(part_rows) == 2
        formula_headers = [cell.value for cell in formulas["整理表"][1]]
        theory_column = formula_headers.index("理单重(kg)") + 1
        theory_formulas = [
            formulas["整理表"].cell(row, theory_column).value
            for row in range(2, formulas["整理表"].max_row + 1)
        ]
        assert len(theory_formulas) == 2
        assert all(
            isinstance(formula, str)
            and "3200*" not in formula
            and "*L" in formula
            for formula in theory_formulas
        )
    finally:
        values.close()
        formulas.close()


def test_missing_required_fields_are_preserved_audited_and_excluded_from_part(
    tmp_path: Path,
) -> None:
    source = tmp_path / "missing-fields.xlsx"
    output = tmp_path / "missing-fields-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原表"
    sheet.append(["批次", "构件编号", "零件号", "规格", "长度(mm)", "材质", "数量"])
    sheet.append(["B1", "C1", None, "PL10*100", 1000, "Q355B", 1])
    sheet.append([None, None, None, "PL10*100", 1000, "Q355B", 1])
    sheet.append([None, None, "missing-spec", None, 1000, "Q355B", 1])
    sheet.append([None, None, "missing-length", "PL10*100", None, "Q355B", 1])
    sheet.append([None, None, "missing-material", "PL10*100", 1000, None, 1])
    sheet.append([None, None, "missing-qty", "PL10*100", 1000, "Q355B", None])
    workbook.save(source)
    workbook.close()
    handbook = FakeHandbook()

    outcome = run_auto_pipeline(source, output, handbook_repository=handbook)

    assert outcome.quality_status == "severe_warning"
    assert outcome.severe_warning_count == 2
    assert handbook.requests == []
    result = load_workbook(output, data_only=True)
    try:
        assert result["清洗表"].max_row == 6
        assert result["整理表"].max_row == 6
        assert result["part"].max_row == 1
        report = list(result["处理报告"].iter_rows(min_row=2, values_only=True))
        report_fields = {
            field
            for row in report
            for field in str(row[5]).split("；")
        }
        assert report_fields == {"零件号", "规格", "长度", "材质", "数量"}
        assert all(row[0] == "严重" and row[1] == "关键字段缺失" for row in report)
        assert any("影响 4 行" in str(row[6]) for row in report)
        assert _cell_by_header(result["整理表"], 2, "零件号").value is None
        assert _cell_by_header(result["整理表"], 4, "长度(mm)").value is None
        assert _cell_by_header(result["整理表"], 4, "下料长度(mm)").value is None
        assert _cell_by_header(result["整理表"], 2, "零件号").fill.fill_type == "solid"
        assert _cell_by_header(result["整理表"], 4, "长度(mm)").fill.fill_type == "solid"
    finally:
        result.close()

    formulas = load_workbook(output, data_only=False, read_only=True)
    try:
        assert _cell_by_header(formulas["整理表"], 4, "下料长度(mm)").value is None
    finally:
        formulas.close()


def test_nonpositive_dimensions_counts_and_negative_source_values_are_isolated(
    tmp_path: Path,
) -> None:
    source = tmp_path / "invalid-physics.xlsx"
    output = tmp_path / "invalid-physics-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原表"
    sheet.append([
        "批次", "构件编号", "零件号", "规格", "长度(mm)", "材质", "数量",
        "单净重(kg)", "总净重(kg)", "单毛重(kg)", "总毛重(kg)",
        "单表面积(㎡)", "总表面积(㎡)",
    ])
    sheet.append(["B1", "C1", None, "PL10*100", 1000, "Q355B", 1])
    sheet.append([None, None, "negative-length", "PL10*100", -1, "Q355B", 1])
    sheet.append([None, None, "zero-qty", "PL10*100", 1000, "Q355B", 0])
    sheet.append([None, None, "negative-weight", "NUT24", 1000, "Q355B", 1,
                  -1, -1, -1, -1, -1, -1])
    workbook.save(source)
    workbook.close()

    outcome = run_auto_pipeline(source, output, handbook_repository=FakeHandbook())

    assert outcome.quality_status == "severe_warning"
    result = load_workbook(output, data_only=True, read_only=True)
    try:
        assert result["清洗表"].max_row == 4
        assert result["整理表"].max_row == 4
        assert result["part"].max_row == 1
        report = list(result["处理报告"].iter_rows(min_row=2, values_only=True))
        fields = {
            field
            for row in report
            if row[0] == "严重"
            for field in str(row[5]).split("；")
        }
        assert {
            "长度", "数量", "单净重", "总净重", "单毛重", "总毛重",
            "单表面积", "总表面积",
        }.issubset(fields)
    finally:
        result.close()


def test_initial_table_missing_length_uses_the_same_audited_isolation(
    tmp_path: Path,
) -> None:
    source = tmp_path / "initial-missing.xlsx"
    output = tmp_path / "initial-missing-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "初始表"
    sheet.append(["C1材  料  表构件数量：1构件总重：0"])
    sheet.append(["零件号", "截面型材", "长度", "材质", "数量", "单重", "总重", "总面积", "备注"])
    sheet.append(["P1", "PL10*100", None, "Q355B", 1])
    workbook.save(source)
    workbook.close()

    outcome = run_auto_pipeline(source, output, handbook_repository=FakeHandbook())

    assert outcome.quality_status == "severe_warning"
    result = load_workbook(output, data_only=True, read_only=True)
    try:
        assert result["清洗表"]["J2"].value is None
        assert _cell_by_header(result["整理表"], 2, "长度(mm)").value is None
        assert _cell_by_header(result["整理表"], 2, "下料长度(mm)").value is None
        assert result["part"].max_row == 1
        report = list(result["处理报告"].iter_rows(min_row=2, values_only=True))
        assert len(report) == 1
        assert report[0][1] == "关键字段缺失"
        assert report[0][5] == "长度"
    finally:
        result.close()

    formulas = load_workbook(output, data_only=False, read_only=True)
    try:
        assert _cell_by_header(formulas["整理表"], 2, "下料长度(mm)").value is None
    finally:
        formulas.close()


def test_blank_spec_m_series_part_is_an_explicit_skipped_bolt(
    tmp_path: Path,
) -> None:
    source = tmp_path / "blank-spec-bolt.xlsx"
    output = tmp_path / "blank-spec-bolt-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原表"
    sheet.append([
        "构件编号", "零件号", "规格", "长度(mm)", "材质", "数量",
        "单净重(kg)", "总净重(kg)",
    ])
    sheet.append(["C1", None, "H100", 1000, "Q355B", 1])
    sheet.append([None, "M22", None, 60, "TS10.9", 10, Decimal("0.3"), Decimal("3.0")])
    workbook.save(source)
    workbook.close()

    handbook = FakeHandbook()
    outcome = run_auto_pipeline(source, output, handbook_repository=handbook)

    assert outcome.quality_status == "ok"
    assert not handbook.requests
    values = load_workbook(output, data_only=True, read_only=True)
    try:
        row = _organized(output)[0]
        assert row["截面型材"] is None
        assert row["规格"] == "M22"
        assert row["比重"] is None
        assert row["理单重(kg)"] is None
        assert values["part"].max_row == 1
        assert values["处理报告"]["A2"].value == "无"
    finally:
        values.close()


def test_stacked_studs_with_blank_spec_and_length_remain_blank_without_report_noise(
    tmp_path: Path,
) -> None:
    source = tmp_path / "stacked-studs.xlsx"
    output = tmp_path / "stacked-studs-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "导出"
    sheet.append([
        "构件编号", "零件号", "规格", "长度(mm)", "材质", "数量",
        "单净重(kg)", "总净重(kg)",
    ])
    sheet.append(["C1", None, "H100", 1000, "Q355B", 1])
    sheet.append([
        None,
        "M19X100\nM19X100",
        None,
        None,
        "STUD\nSTUD",
        "42\n350",
    ])
    workbook.save(source)
    workbook.close()

    handbook = FakeHandbook()
    outcome = run_auto_pipeline(source, output, handbook_repository=handbook)

    assert outcome.quality_status == "ok"
    assert not handbook.requests
    rows = _organized(output)
    assert [row["零件号"] for row in rows] == ["M19X100", "M19X100"]
    assert [row["原数量"] for row in rows] == [42, 350]
    assert all(row["截面型材"] is None for row in rows)
    assert all(row["长度(mm)"] is None for row in rows)
    assert all(row["下料长度(mm)"] is None for row in rows)
    assert all(row["比重"] is None for row in rows)
    assert all(row["理单重(kg)"] is None for row in rows)

    values = load_workbook(output, data_only=True, read_only=True)
    try:
        assert values["part"].max_row == 1
        assert values["处理报告"]["A2"].value == "无"
    finally:
        values.close()


def test_multi_sheet_workbook_processes_first_sheet_and_discards_ignored_sheets(
    tmp_path: Path,
) -> None:
    source = tmp_path / "multi-sheet.xlsx"
    output = tmp_path / "multi-sheet-output.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "导出"
    first.append(["构件编号", "零件号", "规格", "长度(mm)", "材质", "数量"])
    first.append(["C1", None, "H100", 1000, "Q355B", 1])
    first.append([None, "P1", "PL10*100", 100, "Q355B", 2])
    ignored = workbook.create_sheet("历史part")
    ignored["A1"] = "不得进入规范结果"
    workbook.save(source)
    workbook.close()

    outcome = run_auto_pipeline(source, output, handbook_repository=FakeHandbook())

    assert outcome.quality_status == "warning"
    result = load_workbook(output, data_only=True, read_only=True)
    try:
        assert result.sheetnames == [
            "原表",
            "清洗表",
            "构件表",
            "整理表",
            "part",
            "处理报告",
        ]
        assert result["原表"]["A1"].value == "构件编号"
        report = list(result["处理报告"].iter_rows(min_row=2, values_only=True))
        multi_sheet_rows = [row for row in report if row[1] == "多工作表输入"]
        assert len(multi_sheet_rows) == 1
        assert "历史part" in multi_sheet_rows[0][6]
    finally:
        result.close()


def test_conflicting_component_identity_blocks_flat_steel_from_part(
    tmp_path: Path,
) -> None:
    source = tmp_path / "conflicting-flat.xlsx"
    output = tmp_path / "conflicting-flat-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原表"
    sheet.append(["批次", "构件编号", "零件号", "规格", "长度(mm)", "材质", "数量"])
    sheet.append(["B1", "C1", None, "H100", 1000, "Q355B", 1])
    sheet.append([None, None, "flat", "PL6*30", 1000, "Q355B", 1])
    sheet.append(["B1", "C1", "构件小计", None, None, None, 1])
    sheet.append(["B1", "C1", None, "H100", 1000, "Q355B", 2])
    workbook.save(source)
    workbook.close()

    outcome = run_auto_pipeline(source, output, handbook_repository=FakeHandbook())

    assert outcome.quality_status == "severe_warning"
    result = load_workbook(output, read_only=True, data_only=True)
    try:
        assert result["part"].max_row == 1
        report = list(result["处理报告"].iter_rows(min_row=2, values_only=True))
        assert any(row[1] == "构件编号冲突" for row in report)
    finally:
        result.close()


def test_invalid_component_summary_physics_blocks_every_component_part(
    tmp_path: Path,
) -> None:
    source = tmp_path / "invalid-component-summary.xlsx"
    output = tmp_path / "invalid-component-summary-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原表"
    sheet.append([
        "批次", "构件编号", "零件号", "规格", "长度(mm)", "材质", "数量",
        "单净重(kg)", "总净重(kg)", "单毛重(kg)", "总毛重(kg)",
        "单表面积(㎡)", "总表面积(㎡)", "长度(mm)", "宽度(mm)", "高度(mm)",
    ])
    sheet.append(["B1", "C1", None, "H100", 1000, "Q355B", 1])
    sheet.append([None, None, "flat", "PL6*30", 1000, "Q355B", 1])
    sheet.append([
        "B1", "C1", "构件小计", None, None, None, 1,
        1, -1, 1, 1, 1, 1, 1000, 0, 100,
    ])
    workbook.save(source)
    workbook.close()

    outcome = run_auto_pipeline(source, output, handbook_repository=FakeHandbook())

    assert outcome.quality_status == "severe_warning"
    result = load_workbook(output, read_only=True, data_only=True)
    try:
        assert result["part"].max_row == 1
        report = list(result["处理报告"].iter_rows(min_row=2, values_only=True))
        component_issues = [row for row in report if row[1] == "构件物理量非法"]
        assert {
            field
            for row in component_issues
            for field in str(row[5]).split("；")
        } == {"总净重", "构件宽度"}
        assert {row[2] for row in component_issues} == {"原表!4"}
    finally:
        result.close()


def test_initial_table_missing_component_number_is_audited_and_isolated(
    tmp_path: Path,
) -> None:
    source = tmp_path / "initial-missing-component.xlsx"
    output = tmp_path / "initial-missing-component-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "初始表"
    sheet.append(["材料表构件数量：1构件总重：0"])
    sheet.append(["零件号", "截面型材", "长度", "材质", "数量", "单重", "总重", "总面积", "备注"])
    sheet.append(["P1", "PL10*100", 1000, "Q355B", 1])
    workbook.save(source)
    workbook.close()

    outcome = run_auto_pipeline(source, output, handbook_repository=FakeHandbook())

    assert outcome.quality_status == "severe_warning"
    result = load_workbook(output, read_only=True, data_only=True)
    try:
        assert result["清洗表"].max_row == 2
        assert result["整理表"].max_row == 2
        assert result["part"].max_row == 1
        report = list(result["处理报告"].iter_rows(min_row=2, values_only=True))
        assert any(row[1] == "关键字段缺失" and row[5] == "构件编号" for row in report)
    finally:
        result.close()


def test_documentation_contract_rejects_legacy_production_rules() -> None:
    stage_root = Path(__file__).resolve().parents[1]
    readme = (stage_root / "README.md").read_text(encoding="utf-8")
    process = (stage_root / "PROCESS.md").read_text(encoding="utf-8")

    combined_production_docs = readme + "\n" + process
    forbidden = (
        "标准 25 步",
        "回退到第 6 行",
        "回退到假设第 6 行",
        "BH / HA",
        "I / HI",
        "公式回退计算",
        "NUT_M24",
        "五个子表",
        "整理表_拆板后",
    )
    assert all(term not in combined_production_docs for term in forbidden)
    for required in (
        "固定六表",
        "plate_constant:7.85",
        "flat_steel",
        "round_bar",
        "rebar",
        "PipelineOutcome",
        "处理报告",
        "公式缓存",
        "表净重",
        "表毛重",
    ):
        assert required in combined_production_docs
    assert not (stage_root / "multi_split").exists()
    production_modules = "\n".join(
        path.read_text(encoding="utf-8")
        for path in stage_root.glob("*.py")
    )
    assert "multi_split" not in production_modules


def test_bbh_variable_height_split_lands_in_both_organized_and_part(
    tmp_path: Path,
) -> None:
    source = tmp_path / "bbh.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原表"
    sheet.append(["测试零件清单"])
    sheet.append([
        "批次", "构件编号", "零件号", "规格", "长度(mm)", "材质", "数量",
        "单净重(kg)", "总净重(kg)", "单毛重(kg)", "总毛重(kg)",
        "单表面积(㎡)", "总表面积(㎡)", "长度(mm)", "宽度(mm)",
        "高度(mm)", "版本",
    ])
    sheet.append(["B1", "C1", None, "BBH700~500*300*16*30", 1000, "Q355B", 1])
    sheet.append([None, None, "p-bbh", "BBH700~500*300*16*30", 1000, "Q355B", 1])
    sheet.append(["B1", "C1", "构件小计", None, None, None, 1])
    workbook.save(source)
    workbook.close()

    output = tmp_path / "bbh-output.xlsx"
    outcome = run_auto_pipeline(
        source,
        output,
        handbook_repository=FakeHandbook(),
    )
    assert outcome.severe_warning_count == 0

    organized = {row["零件号"]: row for row in _organized(output) if row["类型"] in {"BBH腹", "BBH翼"}}
    assert set(organized) == {"p-bbh"}
    web = organized["p-bbh"]
    assert organized["p-bbh"]["导入零件号"] in {"p-bbh-BBH腹", "p-bbh-BBH翼"}
    by_type = {row["类型"]: row for row in _organized(output) if row["类型"] in {"BBH腹", "BBH翼"}}
    assert (str(by_type["BBH腹"]["规格"]), int(by_type["BBH腹"]["宽度"])) == ("16", 540)
    assert (str(by_type["BBH翼"]["规格"]), int(by_type["BBH翼"]["宽度"])) == ("30", 300)
    assert int(by_type["BBH翼"]["数量"]) == 2

    parts = load_workbook(output, data_only=True, read_only=True)["part"]
    try:
        part_rows = [
            dict(zip((c.value for c in parts[1]), row, strict=True))
            for row in parts.iter_rows(min_row=2, values_only=True)
        ]
    finally:
        parts.parent.close()
    bbh_types = [
        row["类型"]
        for row in part_rows
        if str(row["导入零件号"]).startswith("p-bbh-BBH")
    ]
    assert set(bbh_types) == {"BBH腹", "BBH翼"}
