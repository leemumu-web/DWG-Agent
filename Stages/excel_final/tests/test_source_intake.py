from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from input_contract import InputContractError
from source_intake import SourceFormat, read_production_source


def _standard_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原表"
    sheet.append(["构件零件清单"])
    sheet.append([
        "构件编号",
        "零件号",
        "规格",
        "长度(mm)",
        "材质",
        "数量",
        "单毛重(kg)",
        "总毛重(kg)",
    ])
    sheet.append(["C1", None, "BH500*300*12*20", 1000, "Q355B", 2])
    sheet.append([None, "P1", "PL10*200", 100, "Q355B", 3, 1.57, 4.71])
    sheet.append(["C1", "构件小计", None, None, None, 2, 1.57, 3.14])
    workbook.save(path)
    workbook.close()
    return path


def _initial_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "初始表"
    sheet.append(["C2材  料  表构件数量：4构件总重：20.0"])
    sheet.append([
        "零件号",
        "截面型材",
        "长度(mm)",
        "材质",
        "数量",
        "单重(kg)",
        "总重(kg)",
        "总面积(m2)",
        "备注",
    ])
    sheet.append(["P2", "PL8*100", 200, "Q355B", 2, 1.26, 2.52, 0.1, None])
    sheet.append(["合计"])
    workbook.save(path)
    workbook.close()
    return path


def test_source_intake_detects_standard_workbook(tmp_path: Path) -> None:
    source = _standard_workbook(tmp_path / "standard.xlsx")

    result = read_production_source(source)

    assert result.source_path == source.resolve()
    assert result.source_format is SourceFormat.STANDARD_WORKBOOK
    assert result.sheet_name == "原表"
    assert len(result.parts) == 1
    assert result.parts[0].component_no == "C1"
    assert len(result.component_rows) == 1
    assert result.component_rows[0].component_qty == Decimal("2")


def test_source_intake_accepts_units_with_non_parenthetical_separators(
    tmp_path: Path,
) -> None:
    source = tmp_path / "unit-suffixed.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原表"
    sheet.append(
        [
            "批次/批号",
            "构件编号(号)",
            "零件号",
            "规格",
            "长度/mm",
            "材质",
            "数量-件",
            "单毛重[kg]",
            "总毛重【kg】",
        ]
    )
    sheet.append(["B1", "C1", None, "PL10*100", 200, "Q355B", 1, None, None])
    sheet.append([None, None, "P1", "PL10*100", 200, "Q355B", 2, 1.5, 3.0])
    workbook.save(source)
    workbook.close()

    result = read_production_source(source)

    assert len(result.parts) == 1
    assert result.parts[0].length == Decimal("200")
    assert result.parts[0].original_qty == Decimal("2")
    assert result.parts[0].source_unit_gross == Decimal("1.5")
    assert result.parts[0].source_total_gross == Decimal("3.0")


def test_multi_sheet_workbook_keeps_first_sheet_and_records_warning(tmp_path: Path) -> None:
    source = _standard_workbook(tmp_path / "multi.xlsx")
    workbook = Workbook()
    first = workbook.active
    first.title = "原表"
    first.append(["构件编号", "零件号", "规格", "长度(mm)", "材质", "数量"])
    first.append(["C1", None, "BH500*300*12*20", 1000, "Q355B", 1])
    first.append([None, "P1", "PL10*200", 100, "Q355B", 1])
    workbook.create_sheet("整理")
    workbook.create_sheet("part")
    workbook.save(source)
    workbook.close()

    result = read_production_source(source)

    assert result.sheet_name == "原表"
    assert result.ignored_sheets == ("整理", "part")
    assert len(result.warnings) == 1
    assert result.issues[0].category == "多工作表输入"
    assert result.issues[0].level.value == "警告"


def test_standard_workbook_accepts_identical_headers_between_component_blocks(
    tmp_path: Path,
) -> None:
    source = tmp_path / "repeated-section-headers.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原表"
    header = [
        "构件号",
        "零件号",
        "截面型材",
        "长度(mm)",
        "材质",
        "数量",
        "单重(kg)",
        "总重(kg)",
    ]
    sheet.append(header)
    sheet.append(["C1", None, "BH500*300*12*20", 1000, "Q355B", 1])
    sheet.append([None, "P1", "PL10*200", 100, "Q355B", 2, 1.57, 3.14])
    sheet.append(header)
    sheet.append(["C2", None, "BOX500*300*12*20", 1200, "Q355B", 1])
    sheet.append([None, "P2", "PL12*200", 120, "Q355B", 3, 2.26, 6.78])
    workbook.save(source)
    workbook.close()

    result = read_production_source(source)

    assert result.source_format is SourceFormat.STANDARD_WORKBOOK
    assert [part.part_no for part in result.parts] == ["P1", "P2"]
    assert [part.component_no for part in result.parts] == ["C1", "C2"]
    assert [row.component_no for row in result.component_rows] == ["C1", "C2"]
    assert result.diagnostics["header_row"] == 1


def test_source_intake_detects_initial_workbook(tmp_path: Path) -> None:
    source = _initial_workbook(tmp_path / "initial.xlsx")

    result = read_production_source(source)

    assert result.source_path == source.resolve()
    assert result.source_format is SourceFormat.INITIAL_WORKBOOK
    assert result.sheet_name == "初始表"
    assert len(result.parts) == 1
    assert result.parts[0].component_no == "C2"
    assert len(result.component_rows) == 1
    assert result.component_rows[0].component_qty == Decimal("4")


def test_initial_metadata_and_header_can_move(tmp_path: Path) -> None:
    source = tmp_path / "moved-initial.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "材料表"
    sheet.append(["项目说明"])
    sheet.append([])
    sheet.append(["C3材  料  表构件数量：3构件总重：30.0"])
    sheet.append([])
    sheet.append([
        "零件号",
        "截面型材",
        "长度(mm)",
        "材质",
        "数量",
        "单重(kg)",
        "总重(kg)",
        "总面积(m2)",
        "备注",
    ])
    sheet.append(["P3", "PL10*100", 300, "Q355B", 2, 2.36, 4.72, 0.2, None])
    sheet.append(["合计"])
    workbook.save(source)
    workbook.close()

    result = read_production_source(source)

    assert result.source_format is SourceFormat.INITIAL_WORKBOOK
    assert result.parts[0].source_row == 6
    assert result.parts[0].component_no == "C3"
    assert result.component_rows[0].source_row == 3
    assert result.component_rows[0].component_qty == Decimal("3")


def test_fixed_width_blank_spec_does_not_shift_bolt_columns(tmp_path: Path) -> None:
    source = tmp_path / "fixed-width.xls"
    source.write_text(
        "构件编号             零件编号       型 材                     构件名称        材   质       长度(mm)     数量     单净重(kg)  总净重(kg)    单毛重(kg)  总毛重(kg)   单面积(m2)  总面积(m2)    备 注\n"
        "SKG-C-WGKL-27                       PL80*1200                albl_Top_f       Q345GJB-      2335        1         4384.47      4384.47      4625.72      4625.72      26.34      26.34\n"
        "                     M22                                                      TS10.9         60         10        0.3          3.2\n",
        encoding="utf-8",
    )

    result = read_production_source(source)

    assert result.source_format is SourceFormat.FIXED_WIDTH_TEKLA_TEXT
    assert len(result.parts) == 1
    part = result.parts[0]
    assert part.part_no == "M22"
    assert part.original_spec == ""
    assert part.material == "TS10.9"
    assert part.length == Decimal("60")
    assert part.original_qty == Decimal("10")


def test_invalid_numeric_cell_reports_sheet_row_field_and_value(tmp_path: Path) -> None:
    source = tmp_path / "invalid-number.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原表"
    sheet.append(["构件编号", "零件号", "规格", "长度(mm)", "材质", "数量"])
    sheet.append(["C1", None, "BH500*300*12*20", 1000, "Q355B", 1])
    sheet.append([None, "P1", "PL10*200", "一百", "Q355B", 1])
    workbook.save(source)
    workbook.close()

    with pytest.raises(InputContractError) as caught:
        read_production_source(source)

    failure = caught.value.failure
    assert failure.code == "EXCEL_INPUT_ROW_VALUE_INVALID"
    assert failure.message == "表格中存在无法读取的数值。"
    assert failure.action == "请检查 原表 第 3 行“零件长度”，填写有效数字。"
    assert failure.issues[0].sheet == "原表"
    assert failure.issues[0].row == 3
    assert failure.issues[0].field == "零件长度"
    assert failure.issues[0].value == "一百"
    assert failure.issues[0].reason == "not_numeric"


def test_standard_workbook_expands_stacked_part_lines_without_inventing_blanks(
    tmp_path: Path,
) -> None:
    source = tmp_path / "stacked-parts.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "导出"
    sheet.append(["构件编号", "零件号", "规格", "长度(mm)", "材质", "数量"])
    sheet.append(["C1", None, "BH500*300*12*20", 1000, "Q355B", 1])
    sheet.append([None, "M19X100\nM19X100", None, None, "STUD\nSTUD", "42\n350"])
    workbook.save(source)
    workbook.close()

    result = read_production_source(source)

    assert [part.part_no for part in result.parts] == ["M19X100", "M19X100"]
    assert [part.material for part in result.parts] == ["STUD", "STUD"]
    assert [part.original_qty for part in result.parts] == [
        Decimal("42"),
        Decimal("350"),
    ]
    assert [part.source_row for part in result.parts] == [3, 3]
    assert [part.source_seq for part in result.parts] == [2, 2]
    assert [part.original_spec for part in result.parts] == ["", ""]
    assert [part.length for part in result.parts] == [Decimal("0"), Decimal("0")]
    assert all(part.invalid_fields == ("规格", "长度") for part in result.parts)


def test_stacked_part_lines_repeat_shared_single_value_fields(tmp_path: Path) -> None:
    source = tmp_path / "stacked-shared-fields.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "导出"
    sheet.append(["构件编号", "零件号", "规格", "长度(mm)", "材质", "数量"])
    sheet.append(["C1", None, "BH500*300*12*20", 1000, "Q355B", 1])
    sheet.append([None, "P1\nP2", "PL10*100", "100\n200", "Q355B", "2\n3"])
    workbook.save(source)
    workbook.close()

    result = read_production_source(source)

    assert [part.original_spec for part in result.parts] == [
        "PL10*100",
        "PL10*100",
    ]
    assert [part.material for part in result.parts] == ["Q355B", "Q355B"]
    assert [part.length for part in result.parts] == [
        Decimal("100"),
        Decimal("200"),
    ]


def test_stacked_part_lines_reject_mismatched_item_counts(tmp_path: Path) -> None:
    source = tmp_path / "stacked-count-mismatch.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "导出"
    sheet.append(["构件编号", "零件号", "规格", "长度(mm)", "材质", "数量"])
    sheet.append(["C1", None, "BH500*300*12*20", 1000, "Q355B", 1])
    sheet.append([None, "P1\nP2", "PL10*100", "100\n200\n300", "Q355B", "2\n3"])
    workbook.save(source)
    workbook.close()

    with pytest.raises(InputContractError) as caught:
        read_production_source(source)

    failure = caught.value.failure
    assert failure.code == "EXCEL_INPUT_MULTILINE_ROW_AMBIGUOUS"
    assert failure.message == "同一行中的多条零件无法一一对应。"
    assert failure.action == (
        "请检查 导出 第 3 行：零件号有 2 条，但零件长度有 3 条；"
        "请让各换行字段的条目数一致，或仅保留一个明确的共用值。"
    )
    assert failure.issues[0].field == "零件长度"
    assert failure.issues[0].reason == "multiline_item_count_mismatch"


def test_multiline_values_without_itemized_part_numbers_are_rejected(
    tmp_path: Path,
) -> None:
    source = tmp_path / "stacked-without-part-numbers.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "导出"
    sheet.append(["构件编号", "零件号", "规格", "长度(mm)", "材质", "数量"])
    sheet.append(["C1", None, "BH500*300*12*20", 1000, "Q355B", 1])
    sheet.append([None, "P1", "PL10*100", "100\n200", "Q355B", "2\n3"])
    workbook.save(source)
    workbook.close()

    with pytest.raises(InputContractError) as caught:
        read_production_source(source)

    failure = caught.value.failure
    assert failure.code == "EXCEL_INPUT_MULTILINE_ROW_AMBIGUOUS"
    assert failure.action == (
        "请检查 导出 第 3 行：零件长度有 2 条，但零件号有 1 条；"
        "请让各换行字段的条目数一致，或仅保留一个明确的共用值。"
    )
    assert failure.issues[0].field == "零件号"


def test_part_before_component_reports_structured_relationship_error(tmp_path: Path) -> None:
    source = tmp_path / "orphan-part.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原表"
    sheet.append(["构件编号", "零件号", "规格", "长度(mm)", "材质", "数量"])
    sheet.append([None, "P1", "PL10*200", 100, "Q355B", 1])
    workbook.save(source)
    workbook.close()

    with pytest.raises(InputContractError) as caught:
        read_production_source(source)

    failure = caught.value.failure
    assert failure.code == "EXCEL_INPUT_PART_WITHOUT_COMPONENT"
    assert failure.message == "零件明细前缺少所属构件行。"
    assert failure.action == "请在 原表 第 2 行零件 P1 前补充对应的构件起始行。"
    assert failure.issues[0].sheet == "原表"
    assert failure.issues[0].row == 2
    assert failure.issues[0].field == "构件编号"
    assert failure.issues[0].value == "P1"
    assert failure.issues[0].reason == "part_without_component"


def test_initial_table_invalid_numeric_cell_is_not_silently_blank(tmp_path: Path) -> None:
    source = tmp_path / "invalid-initial-number.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "初始表"
    sheet.append(["C2材  料  表构件数量：4构件总重：20.0"])
    sheet.append([
        "零件号",
        "截面型材",
        "长度(mm)",
        "材质",
        "数量",
        "单重(kg)",
        "总重(kg)",
        "总面积(m2)",
        "备注",
    ])
    sheet.append(["P2", "PL8*100", "二百", "Q355B", 2, 1.26, 2.52, 0.1, None])
    sheet.append(["合计"])
    workbook.save(source)
    workbook.close()

    with pytest.raises(InputContractError) as caught:
        read_production_source(source)

    failure = caught.value.failure
    assert failure.code == "EXCEL_INPUT_ROW_VALUE_INVALID"
    assert failure.action == "请检查 初始表 第 3 行“长度”，填写有效数字。"
    assert failure.issues[0].sheet == "初始表"
    assert failure.issues[0].row == 3
    assert failure.issues[0].field == "长度"
    assert failure.issues[0].value == "二百"
    assert failure.issues[0].reason == "not_numeric"


def test_fixed_width_text_requires_one_unique_header(tmp_path: Path) -> None:
    source = tmp_path / "duplicate-fixed-header.xls"
    header = (
        "构件编号             零件编号       型 材                     构件名称"
        "        材   质       长度(mm)     数量     单净重(kg)  总净重(kg)"
        "    单毛重(kg)  总毛重(kg)   单面积(m2)  总面积(m2)    备 注\n"
    )
    source.write_text(header + header, encoding="utf-8")

    with pytest.raises(InputContractError) as caught:
        read_production_source(source)

    failure = caught.value.failure
    assert failure.code == "EXCEL_INPUT_HEADER_AMBIGUOUS"
    assert failure.meta["candidate_rows"] == [1, 2]
    assert [issue.row for issue in failure.issues] == [1, 2]


def test_component_only_fixed_width_text_has_specific_action(tmp_path: Path) -> None:
    source = tmp_path / "component-only.xls"
    source.write_text(
        "构件编号  截面型材  材质  长度  数量  单净重  总净重  单毛重  总毛重\n"
        "C1        BH500*300*12*20  Q355B  1000  1  10  10  11  11\n",
        encoding="utf-8",
    )

    with pytest.raises(InputContractError) as caught:
        read_production_source(source)

    failure = caught.value.failure
    assert failure.code == "EXCEL_INPUT_COMPONENT_ONLY"
    assert "包含零件号" in failure.action
