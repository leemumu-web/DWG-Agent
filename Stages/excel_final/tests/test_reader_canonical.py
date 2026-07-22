from __future__ import annotations

import importlib
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook


def _reader():
    try:
        return importlib.import_module("reader")
    except ModuleNotFoundError as exc:
        pytest.fail(f"reader module is missing: {exc}")


def _reader_init():
    try:
        return importlib.import_module("reader_init")
    except ModuleNotFoundError as exc:
        pytest.fail(f"initial-table reader module is missing: {exc}")


def _standard_workbook(path: Path, *, duplicate_conflict: bool = False) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原表"
    sheet.append(["零件清单"])
    sheet.append(
        [
            "批次", "构件编号", "零件号", "规格", "长度(mm)", "材质", "数量",
            "单净重(kg)", "总净重(kg)", "单毛重(kg)", "总毛重(kg)",
            "单表面积(㎡)", "总表面积(㎡)", "长度(mm)", "宽度(mm)",
            "高度(mm)", "版本",
        ]
    )
    sheet.append(
        [" B A　T ", " C 1 ", None, " BOX 700 * 700 * 36 * 36 ", 3704,
         " Q 420 B ", 2, None, None, None, None, None, None, None, None, None, None]
    )
    sheet.append(
        [None, None, " p 1 ", " PL 10 * 20 ", 100, " Q 355 B ", 3,
         1.5, 4.5, 1.6, 4.8, 0.1, 0.3, None, None, None, None]
    )
    sheet.append(
        [" B A　T ", " C 1 ", "构件 小计", None, None, None, 2,
         1.5, 3.0, 1.6, 3.2, 0.1, 0.2, 3704, 700, 700, None]
    )
    if duplicate_conflict:
        sheet.append(
            ["BAT", "C1", None, "BOX700*700*36*36", 3704, "Q420B", 3,
             None, None, None, None, None, None, None, None, None, None]
        )
    sheet.append([None, "合 计：", None, None, None, None, None, None, 3.0, None, 3.2])
    workbook.save(path)
    workbook.close()
    return path


def test_canonical_reader_preserves_raw_and_summarizes_each_component_once(
    tmp_path: Path,
) -> None:
    reader = _reader()
    source = _standard_workbook(tmp_path / "standard.xlsx")

    result = reader.read_canonical_workbook(source)

    assert result.header.row_number == 2
    assert len(result.parts) == 1
    part = result.parts[0]
    assert part.batch == "BAT"
    assert part.component_no == "C1"
    assert part.component_qty == Decimal("2")
    assert part.part_no == "p1"
    assert part.original_spec == "PL10*20"
    assert part.material == "Q355B"
    assert part.original_qty == Decimal("3")
    assert part.source_unit_net == Decimal("1.5")
    assert part.source_total_net == Decimal("4.5")
    assert part.source_unit_gross == Decimal("1.6")
    assert part.source_total_gross == Decimal("4.8")

    assert len(result.component_rows) == 1
    component = result.component_rows[0]
    assert (component.kind.value, component.component_no) == ("summary", "C1")
    assert component.batch == "BAT"
    assert component.component_qty == Decimal("2")
    assert component.original_spec == "BOX700*700*36*36"
    assert component.material == "Q420B"
    assert component.source_total_net == Decimal("3.0")
    assert component.source_total_gross == Decimal("3.2")
    assert component.component_length == Decimal("3704")
    assert component.component_width == Decimal("700")
    assert component.component_height == Decimal("700")
    assert component.subtotal_source_row == 5
    assert result.working_values[2][0] == "BAT"
    assert result.working_values[4][2] == "构件小计"

    reopened = load_workbook(source, read_only=True, data_only=False)
    try:
        assert reopened.active.cell(3, 1).value == " B A　T "
        assert reopened.active.cell(4, 3).value == " p 1 "
    finally:
        reopened.close()


def test_inconsistent_duplicate_component_id_creates_severe_issue(tmp_path: Path) -> None:
    reader = _reader()
    source = _standard_workbook(tmp_path / "duplicate.xlsx", duplicate_conflict=True)

    result = reader.read_canonical_workbook(source)

    conflicts = [issue for issue in result.issues if issue.category == "构件编号冲突"]
    assert len(conflicts) == 1
    assert conflicts[0].level.value == "严重"
    assert conflicts[0].affects_part is True
    assert conflicts[0].component_no == "C1"


def test_initial_table_maps_unit_and_total_weight_to_gross_only(tmp_path: Path) -> None:
    reader_init = _reader_init()
    source = tmp_path / "initial.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "初始表"
    sheet.append(["C1材  料  表构件数量：2构件总重：12.3"])
    sheet.append(["零件号", "截面型材", "长度", "材质", "数量", "单重", "总重", "总面积", "备注"])
    sheet.append([" p 1 ", " PL 10 * 20 ", 100, " Q 355 B ", 2, 1.57, 3.14, 0.1, None])
    sheet.append(["合计"])
    workbook.save(source)
    workbook.close()

    result = reader_init.read_init_canonical(source)

    assert len(result) == 1
    part = result[0]
    assert part.component_no == "C1"
    assert part.component_qty == Decimal("2")
    assert part.part_no == "p1"
    assert part.original_spec == "PL10*20"
    assert part.material == "Q355B"
    assert part.source_unit_net is None
    assert part.source_total_net is None
    assert part.source_unit_gross == Decimal("1.57")
    assert part.source_total_gross == Decimal("3.14")


def test_tekla_text_xls_adapts_to_the_same_canonical_records(tmp_path: Path) -> None:
    reader = _reader()
    source = tmp_path / "tekla.xls"
    source.write_text(
        "零件清单\t\t\t\t\t\t\n"
        "批次\t构件编号\t零件号\t规格\t长度(mm)\t材质\t数量\t单毛重(kg)\t总毛重(kg)\n"
        "BAT\tC1\t\tPL10*20\t100\tQ355B\t2\t\t\n"
        "\t\tp1\tPL10*20\t100\tQ355B\t3\t1.6\t4.8\n"
        "BAT\tC1\t构件小计\t\t\t\t2\t1.6\t3.2\n",
        encoding="utf-8",
    )

    result = reader.read_canonical_source(source)

    assert result.source_path == source.resolve()
    assert result.header.row_number == 2
    assert len(result.parts) == 1
    assert result.parts[0].source_unit_gross == Decimal("1.6")
    assert [row.kind.value for row in result.component_rows] == ["summary"]


def test_tekla_text_never_falls_back_to_incomplete_row_six(tmp_path: Path) -> None:
    reader = _reader()
    source = tmp_path / "bad-tekla.xls"
    source.write_text(
        "说明\t空\t空\n" * 5
        + "批次\t构件编号\t零件号\t规格\t长度(mm)\t材质\n",
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="missing required fields"):
        reader.read_canonical_source(source)
