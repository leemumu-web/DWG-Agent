"""BOX 左右进 -> Excel 第二阶段深化的输入输出流，与 BH 完全同构。

完全模仿 tests/test_stage2_workbook.py 的 BH 用例：Tekla 导出 -> stage1 ->
run_stage2_workbook(BH 空合同 + BOX 合同) -> 验证整理表 BOX腹/BOX翼 行、
进退公式、数量倍率（×2/×1）。
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from bh_stage2 import parse_bh_measurement_contract
from box_stage2 import parse_box_measurement_contract
from pipeline import run_auto_pipeline
from stage2_workbook import run_stage2_workbook


class _NoHandbookLookup:
    def lookup(self, *_args, **_kwargs):
        raise AssertionError("BOX stage2 workbook must not query the handbook")

    def log_stats(self) -> None:
        return None


def _box_stage1_workbook(tmp_path: Path) -> Path:
    source = tmp_path / "box-source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Tekla导出"
    sheet.append(["测试BOX零件清单"])
    sheet.append([
        "批次", "构件编号", "零件号", "规格", "长度(mm)", "材质", "数量",
    ])
    sheet.append([
        "B1", "C1", None, "BOX600*600*22*22", 6745, "Q355C", 1,
    ])
    sheet.append([
        None, None, "w4e-cb-10", "BOX600*600*22*22", 6745, "Q355C", 1,
    ])
    sheet.append(["B1", "C1", "构件小计", None, None, None, 1])
    workbook.save(source)
    workbook.close()
    stage1 = tmp_path / "stage1-box.xlsx"
    run_auto_pipeline(
        source,
        stage1,
        handbook_repository=_NoHandbookLookup(),
    )
    return stage1


def _box_complete_contract():
    return parse_box_measurement_contract({
        "schema": "box_setback_measurements/v1",
        "items": [{
            "source_file_id": 201,
            "file_name": "w4e-cb-10.dxf",
            "part_number": "w4e-cb-10",
            "classification_spec": "BOX600*600*22*22",
            "reader_spec": "BOX600*600*22*22",
            "status": "OK",
            "warnings": [],
            "measurements": [
                {"role": "翼", "left_safe": 30, "right_safe": 40},
                {"role": "腹", "left_safe": 50, "right_safe": 60},
            ],
        }],
    })


def _box_split_contract():
    return parse_box_measurement_contract({
        "schema": "box_setback_measurements/v1",
        "items": [{
            "source_file_id": 202,
            "file_name": "w4e-cb-10.dxf",
            "part_number": "w4e-cb-10",
            "classification_spec": "BOX600*600*22*22",
            "reader_spec": "BOX600*600*22*22",
            "status": "OK",
            "warnings": [],
            "measurements": [
                {"role": "翼", "left_safe": 0, "right_safe": 0},
                {"role": "上腹", "left_safe": 20, "right_safe": 0},
                {"role": "下腹", "left_safe": 0, "right_safe": 0},
            ],
        }],
    })


def test_box_stage2_rebuilds_from_the_formal_stage1_then_applies_box_setbacks(
    tmp_path: Path,
) -> None:
    stage1 = _box_stage1_workbook(tmp_path)
    stage2 = tmp_path / "stage2-box.xlsx"
    internal = tmp_path / "stage2-box-internal.xlsx"

    outcome = run_stage2_workbook(
        stage1,
        stage2,
        measurements=parse_bh_measurement_contract({
            "schema": "bh_setback_measurements/v1",
            "items": [],
        }),
        box_measurements=_box_complete_contract(),
        handbook=_NoHandbookLookup(),
        internal_output_path=internal,
    )

    assert outcome.status == "complete"
    assert outcome.output_path == stage2.resolve()
    assert outcome.internal_output_path == internal.resolve()
    assert outcome.matched_occurrence_count == 1
    assert outcome.missing_drawing_count == 0
    assert outcome.manual_occurrence_count == 0
    assert stage2.is_file()
    assert internal.is_file()
    formulas = load_workbook(stage2, data_only=False, read_only=True)
    values = load_workbook(stage2, data_only=True, read_only=True)
    try:
        assert formulas.sheetnames == [
            "原表", "清洗表", "构件表", "整理表", "part", "处理报告",
        ]
        organized_headers = [cell.value for cell in formulas["整理表"][1]]
        organized = {
            header: index + 1 for index, header in enumerate(organized_headers)
        }
        # 两行 BOX 板件：BOX腹（×2）、BOX翼（×2）
        types = [
            values["整理表"].cell(row, organized["类型"]).value
            for row in (2, 3)
        ]
        assert types == ["BOX腹", "BOX翼"]
        # 进退：腹 50/60 -> 下料 6745-50-60 = 6635；翼 30/40 -> 6675
        assert values["整理表"].cell(2, organized["下料长度(mm)"]).value == 6635
        assert values["整理表"].cell(3, organized["下料长度(mm)"]).value == 6675
        # 数量：BOX腹 ×2、BOX翼 ×2
        assert values["整理表"].cell(2, organized["数量"]).value == 2
        assert values["整理表"].cell(3, organized["数量"]).value == 2
        assert formulas["整理表"].cell(2, organized["下料长度(mm)"]).value == (
            "=M2-N2-O2"
        )
    finally:
        formulas.close()
        values.close()


def test_box_stage2_split_upper_lower_webs_keeps_1x_multipliers(
    tmp_path: Path,
) -> None:
    stage1 = _box_stage1_workbook(tmp_path)
    stage2 = tmp_path / "stage2-box-split.xlsx"
    internal = tmp_path / "stage2-box-split-internal.xlsx"

    outcome = run_stage2_workbook(
        stage1,
        stage2,
        measurements=parse_bh_measurement_contract({
            "schema": "bh_setback_measurements/v1",
            "items": [],
        }),
        box_measurements=_box_split_contract(),
        handbook=_NoHandbookLookup(),
        internal_output_path=internal,
    )

    assert outcome.status == "complete"
    values = load_workbook(stage2, data_only=True, read_only=True)
    try:
        organized_headers = [cell.value for cell in values["整理表"][1]]
        organized = {
            header: index + 1 for index, header in enumerate(organized_headers)
        }
        rows = []
        for row in range(2, 5):
            rows.append((
                values["整理表"].cell(row, organized["类型"]).value,
                values["整理表"].cell(row, organized["导入零件号"]).value,
                values["整理表"].cell(row, organized["左进(mm)"]).value,
                values["整理表"].cell(row, organized["数量"]).value,
            ))
        # 上腹 ×1（20 左进）、下腹 ×1、翼 ×2
        assert ("BOX腹", "w4e-cb-10-BOX上腹", 20, 1) in rows
        assert ("BOX腹", "w4e-cb-10-BOX下腹", 0, 1) in rows
        assert ("BOX翼", "w4e-cb-10-BOX翼", 0, 2) in rows
    finally:
        values.close()
