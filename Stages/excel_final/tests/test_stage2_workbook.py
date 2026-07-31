from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook, load_workbook

from canonical_pipeline import build_canonical_projection, write_canonical_projection
from bh_stage2 import parse_bh_measurement_contract
from domain import SourcePart
from pipeline import run_auto_pipeline, run_stage2_pipeline
from stage2_workbook import (
    Stage2BaselineError,
    _formal_source_sheet_by_row,
    read_canonical_baseline_signature,
    run_stage2_workbook,
    verify_canonical_baseline,
)


class _NoHandbookLookup:
    def lookup(self, *_args, **_kwargs):
        raise AssertionError("plate projection must not query the handbook")

    def log_stats(self) -> None:
        return None


def test_formal_source_sheet_map_streams_read_only_rows(monkeypatch, tmp_path: Path) -> None:
    """The Stage 2 source map must stay linear for large read-only workbooks."""

    class _Sheet:
        max_row = 4

        class _HeaderCell:
            def __init__(self, value):
                self.value = value

        def __getitem__(self, key):
            assert key == 1
            return (
                self._HeaderCell("来源sheet"),
                self._HeaderCell("来源行"),
            )

        def iter_rows(self, *, min_row, values_only):
            assert min_row == 2
            assert values_only is True
            return iter((
                ("原表", 3),
                ("原表", 4),
                (None, None),
            ))

        def cell(self, *_args, **_kwargs):
            raise AssertionError("read-only cell() causes quadratic rescans")

    class _Workbook:
        def __init__(self):
            self.sheet = _Sheet()

        def __getitem__(self, name):
            assert name == "清洗表"
            return self.sheet

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return None

        def close(self):
            return None

    monkeypatch.setattr(
        "stage2_workbook.load_workbook",
        lambda *_args, **_kwargs: _Workbook(),
    )

    mapping, default = _formal_source_sheet_by_row(tmp_path / "stage1.xlsx")

    assert mapping == {3: "原表", 4: "原表"}
    assert default == "原表"


def _stage1_workbook(tmp_path: Path) -> Path:
    source_path = tmp_path / "source.xlsx"
    workbook = Workbook()
    workbook.active.title = "上传原表"
    workbook.active["A1"] = "原始数据"
    workbook.save(source_path)
    workbook.close()
    part = SourcePart(
        source_sheet="上传原表",
        source_row=3,
        source_seq=1,
        batch="B1",
        component_no="C1",
        component_qty=Decimal("2"),
        part_no="P1",
        original_spec="PL10*100",
        material="Q355B",
        length=Decimal("1000"),
        original_qty=Decimal("2"),
        source_unit_net=Decimal("7.85"),
        source_total_net=Decimal("15.7"),
        source_unit_gross=Decimal("7.85"),
        source_total_gross=Decimal("15.7"),
        source_unit_area=Decimal("0.22"),
        source_total_area=Decimal("0.44"),
        classification=None,
    )
    projection = build_canonical_projection(
        parts=(part,),
        component_rows=(),
        reader_issues=(),
        handbook=_NoHandbookLookup(),
    )
    output = tmp_path / "stage1.xlsx"
    write_canonical_projection(source_path, output, projection=projection)
    return output


def test_stage1_signature_requires_six_visible_sheets_headers_and_formula_caches(
    tmp_path: Path,
) -> None:
    stage1 = _stage1_workbook(tmp_path)

    signature = read_canonical_baseline_signature(stage1)

    assert signature.sheet_names == (
        "原表",
        "清洗表",
        "构件表",
        "整理表",
        "part",
        "处理报告",
    )
    assert signature.formula_cell_counts["整理表"] > 0
    assert signature.formula_cell_counts["part"] == 1
    assert set(signature.sheet_hashes) == set(signature.sheet_names)

    formulas = load_workbook(stage1, data_only=False)
    try:
        assert all(sheet.sheet_state == "visible" for sheet in formulas.worksheets)
        formulas.remove(formulas["处理报告"])
        malformed = tmp_path / "missing-report.xlsx"
        formulas.save(malformed)
    finally:
        formulas.close()

    with pytest.raises(Stage2BaselineError) as caught:
        read_canonical_baseline_signature(malformed)
    assert caught.value.code == "EXCEL_STAGE2_BASELINE_DRIFT"


def _copy_xlsx_with_replacement(
    source: Path,
    target: Path,
    *,
    entry_name: str,
    old: bytes,
    new: bytes,
) -> None:
    replaced = False
    with ZipFile(source, "r") as original, ZipFile(
        target,
        "w",
        compression=ZIP_DEFLATED,
    ) as changed:
        for info in original.infolist():
            payload = original.read(info.filename)
            if info.filename == entry_name:
                assert old in payload
                payload = payload.replace(old, new, 1)
                replaced = True
            changed.writestr(info, payload)
    assert replaced


def test_baseline_comparison_ignores_metadata_but_detects_business_drift(
    tmp_path: Path,
) -> None:
    stage1 = _stage1_workbook(tmp_path)
    metadata_only = tmp_path / "metadata-only.xlsx"
    _copy_xlsx_with_replacement(
        stage1,
        metadata_only,
        entry_name="docProps/core.xml",
        old=b"2026",
        new=b"2025",
    )

    verified = verify_canonical_baseline(stage1, metadata_only)

    assert verified == read_canonical_baseline_signature(stage1)

    business_drift = tmp_path / "business-drift.xlsx"
    _copy_xlsx_with_replacement(
        stage1,
        business_drift,
        entry_name="xl/worksheets/sheet4.xml",
        old=b"<t>P1</t>",
        new=b"<t>P-DRIFT</t>",
    )

    with pytest.raises(Stage2BaselineError) as caught:
        verify_canonical_baseline(stage1, business_drift)
    assert caught.value.code == "EXCEL_STAGE2_BASELINE_DRIFT"
    assert caught.value.changed_sheets == ("整理表",)


def test_baseline_comparison_allows_stage_specific_report_rebuild(
    tmp_path: Path,
) -> None:
    stage1 = _stage1_workbook(tmp_path)
    report_changed = tmp_path / "report-changed.xlsx"
    _copy_xlsx_with_replacement(
        stage1,
        report_changed,
        entry_name="xl/worksheets/sheet6.xml",
        old=b"<t>\xe6\x97\xa0</t>",
        new=b"<t>\xe8\xad\xa6\xe5\x91\x8a</t>",
    )

    assert verify_canonical_baseline(stage1, report_changed).sheet_names


def test_stage1_signature_rejects_a_formula_with_a_missing_cached_value(
    tmp_path: Path,
) -> None:
    stage1 = _stage1_workbook(tmp_path)
    missing_cache = tmp_path / "missing-formula-cache.xlsx"
    _copy_xlsx_with_replacement(
        stage1,
        missing_cache,
        entry_name="xl/worksheets/sheet4.xml",
        old=b"<f>M2-N2-O2</f><v>1000</v>",
        new=b"<f>M2-N2-O2</f><v></v>",
    )

    with pytest.raises(Stage2BaselineError, match="公式缓存缺失") as caught:
        read_canonical_baseline_signature(missing_cache)

    assert caught.value.code == "EXCEL_STAGE2_BASELINE_DRIFT"


def _bh_stage1_workbook(tmp_path: Path) -> Path:
    source = tmp_path / "bh-source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Tekla导出"
    sheet.append(["测试BH零件清单"])
    sheet.append([
        "批次", "构件编号", "零件号", "规格", "长度(mm)", "材质", "数量",
    ])
    sheet.append([
        "B1", "C1", None, "BH500*200*10*16", 1000, "Q355B", 2,
    ])
    sheet.append([
        None, None, "BH-P1", "BH500*200*10*16", 1000, "Q355B", 1,
    ])
    sheet.append(["B1", "C1", "构件小计", None, None, None, 2])
    workbook.save(source)
    workbook.close()
    stage1 = tmp_path / "stage1-bh.xlsx"
    run_auto_pipeline(
        source,
        stage1,
        handbook_repository=_NoHandbookLookup(),
    )
    return stage1


def _complete_contract():
    return parse_bh_measurement_contract({
        "schema": "bh_setback_measurements/v1",
        "items": [{
            "source_file_id": 101,
            "file_name": "BH-P1.dxf",
            "part_number": "BH-P1",
            "classification_spec": "BH500*200*10*16",
            "reader_spec": "BH500*200*10*16",
            "status": "OK",
            "warnings": [],
            "measurements": [
                {"role": "腹", "left_safe": 10, "right_safe": 20},
                {"role": "翼", "left_safe": 100, "right_safe": 200},
            ],
        }],
    })


def test_stage2_rebuilds_from_the_formal_stage1_then_applies_bh_setbacks(
    tmp_path: Path,
) -> None:
    stage1 = _bh_stage1_workbook(tmp_path)
    stage2 = tmp_path / "stage2-bh.xlsx"
    internal = tmp_path / "stage2-bh-internal.xlsx"

    outcome = run_stage2_workbook(
        stage1,
        stage2,
        measurements=_complete_contract(),
        handbook=_NoHandbookLookup(),
        internal_output_path=internal,
    )

    assert outcome.status == "complete"
    assert outcome.output_path == stage2.resolve()
    assert outcome.internal_output_path == internal.resolve()
    assert outcome.matched_occurrence_count == 1
    assert outcome.missing_drawing_count == 0
    assert outcome.manual_occurrence_count == 0
    assert stage2.is_file()
    assert internal.is_file()
    formulas = load_workbook(stage2, data_only=False, read_only=True)
    values = load_workbook(stage2, data_only=True, read_only=True)
    try:
        assert formulas.sheetnames == [
            "原表", "清洗表", "构件表", "整理表", "part", "处理报告",
        ]
        organized_headers = [cell.value for cell in formulas["整理表"][1]]
        organized = {
            header: index + 1 for index, header in enumerate(organized_headers)
        }
        assert formulas["整理表"].cell(2, organized["下料长度(mm)"]).value == (
            "=M2-N2-O2"
        )
        assert values["整理表"].cell(2, organized["下料长度(mm)"]).value == 970
        assert values["整理表"].cell(3, organized["下料长度(mm)"]).value == 700
        assert values["整理表"].cell(2, organized["数量"]).value == 1
        assert values["整理表"].cell(3, organized["数量"]).value == 2
        assert formulas["整理表"].cell(2, organized["理单重(kg)"]).value == (
            "=ROUND(K2*L2*P2*V2/1000000,3)"
        )
        assert values["整理表"].cell(2, organized["理单重(kg)"]).value == pytest.approx(
            round(
                values["整理表"].cell(2, organized["规格"]).value
                * values["整理表"].cell(2, organized["宽度"]).value
                * values["整理表"].cell(2, organized["下料长度(mm)"]).value
                * values["整理表"].cell(2, organized["比重"]).value
                / 1000000,
                3,
            )
        )
        assert formulas["整理表"].cell(2, organized["总长(mm)"]).value == "=P2*T2"
        assert formulas["part"]["G2"].value == "=SUM('整理表'!T2)"
        assert formulas["part"]["G3"].value == "=SUM('整理表'!T3)"
    finally:
        formulas.close()
        values.close()


def test_stage2_baseline_drift_names_the_sheet_and_publishes_nothing(
    tmp_path: Path,
) -> None:
    stage1 = _bh_stage1_workbook(tmp_path)
    drifted = tmp_path / "stage1-drifted.xlsx"
    _copy_xlsx_with_replacement(
        stage1,
        drifted,
        entry_name="xl/worksheets/sheet4.xml",
        old=b"<t>BH-P1</t>",
        new=b"<t>BH-P1-DRIFT</t>",
    )
    stage2 = tmp_path / "must-not-exist.xlsx"

    with pytest.raises(Stage2BaselineError) as caught:
        run_stage2_workbook(
            drifted,
            stage2,
            measurements=_complete_contract(),
            handbook=_NoHandbookLookup(),
        )

    assert caught.value.code == "EXCEL_STAGE2_BASELINE_DRIFT"
    assert caught.value.changed_sheets == ("整理表",)
    assert not stage2.exists()


def _non_bh_stage1_workbook(tmp_path: Path) -> Path:
    source = tmp_path / "plate-source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "人工命名源表"
    sheet.append([
        "批次", "构件编号", "零件号", "规格", "长度(mm)", "材质", "数量",
        "单净重(kg)", "总净重(kg)", "单毛重(kg)", "总毛重(kg)",
    ])
    sheet.append(["B1", "C1", None, "H100", 1000, "Q355B", 1])
    sheet.append([
        None, None, "PL-P1", "PL10*100", 1000, "Q355B", 1,
        7.85, 7.85, 7.85, 7.85,
    ])
    workbook.save(source)
    workbook.close()
    stage1 = tmp_path / "stage1-plate.xlsx"
    run_auto_pipeline(
        source,
        stage1,
        handbook_repository=_NoHandbookLookup(),
    )
    return stage1


def test_stage2_without_bh_is_a_business_equivalent_noop(
    tmp_path: Path,
) -> None:
    stage1 = _non_bh_stage1_workbook(tmp_path)
    stage2 = tmp_path / "stage2-noop.xlsx"
    internal = tmp_path / "stage2-noop-internal.xlsx"
    empty_contract = parse_bh_measurement_contract({
        "schema": "bh_setback_measurements/v1",
        "items": [],
    })

    outcome = run_stage2_workbook(
        stage1,
        stage2,
        measurements=empty_contract,
        handbook=_NoHandbookLookup(),
        internal_output_path=internal,
    )

    assert outcome.status == "noop"
    assert outcome.matched_occurrence_count == 0
    assert outcome.missing_drawing_count == 0
    assert outcome.unmatched_drawing_count == 0
    assert outcome.manual_occurrence_count == 0
    assert read_canonical_baseline_signature(stage2) == (
        read_canonical_baseline_signature(stage1)
    )
    workbook = load_workbook(stage2, data_only=True, read_only=True)
    try:
        assert workbook["处理报告"]["A2"].value == "无"
    finally:
        workbook.close()
    assert internal.is_file()


def test_stage2_reader_failure_publishes_partial_manual_formulas(
    tmp_path: Path,
) -> None:
    stage1 = _bh_stage1_workbook(tmp_path)
    stage2 = tmp_path / "stage2-partial.xlsx"
    failed_contract = parse_bh_measurement_contract({
        "schema": "bh_setback_measurements/v1",
        "items": [{
            "source_file_id": 102,
            "file_name": "BH-P1-broken.dxf",
            "part_number": "BH-P1",
            "classification_spec": "BH500*200*10*16",
            "reader_spec": "",
            "status": "ERROR_UNHANDLED",
            "warnings": ["DXF无法读取"],
            "measurements": [],
        }],
    })

    outcome = run_stage2_workbook(
        stage1,
        stage2,
        measurements=failed_contract,
        handbook=_NoHandbookLookup(),
    )

    assert outcome.status == "partial"
    assert outcome.matched_occurrence_count == 1
    assert outcome.manual_occurrence_count == 1
    assert outcome.pipeline_outcome.warning_count >= 1
    formulas = load_workbook(stage2, data_only=False, read_only=True)
    values = load_workbook(stage2, data_only=True, read_only=True)
    try:
        assert formulas["整理表"]["P2"].value == (
            '=IF(OR(N2="",O2=""),"",M2-N2-O2)'
        )
        assert formulas["整理表"]["P3"].value == (
            '=IF(OR(N3="",O3=""),"",M3-N3-O3)'
        )
        assert values["整理表"]["P2"].value is None
        assert values["整理表"]["P3"].value is None
        assert formulas["整理表"]["P2"].fill.fill_type == "solid"
        assert formulas["part"]["E2"].value == (
            '=IF(\'整理表\'!P2="","",\'整理表\'!P2)'
        )
        assert formulas["part"]["E3"].value == (
            '=IF(\'整理表\'!P3="","",\'整理表\'!P3)'
        )
        assert values["part"]["E2"].value is None
        assert values["part"]["E3"].value is None
        assert formulas["part"].max_row == 3
        report_categories = {
            row[1]
            for row in values["处理报告"].iter_rows(
                min_row=2,
                values_only=True,
            )
        }
        assert "BH读取失败需补录" in report_categories
    finally:
        formulas.close()
        values.close()


def test_stage2_pipeline_reuses_an_injected_handbook_repository(
    tmp_path: Path,
) -> None:
    stage1 = _bh_stage1_workbook(tmp_path)
    stage2 = tmp_path / "stage2-through-pipeline.xlsx"
    handbook = _NoHandbookLookup()

    outcome = run_stage2_pipeline(
        stage1,
        stage2,
        measurements=_complete_contract(),
        handbook_repository=handbook,
    )

    assert outcome.status == "complete"
    assert outcome.output_path == stage2.resolve()
