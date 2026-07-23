from __future__ import annotations

import hashlib
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


STAGE_ROOT = Path(__file__).resolve().parents[1]
SOURCE = STAGE_ROOT / "data/20260320-首都体育学院B7#地下部分-构件零件清单(毛净重)去gyb(3).xlsx"
PREPROCESSED = (
    STAGE_ROOT
    / "data/preprocessed/20260320-首都体育学院B7#地下部分-构件零件清单(毛净重)去gyb(3)_原表.xlsx"
)
BASELINE_PATH = Path(__file__).parent / "fixtures/ground_truth_baseline.json"
COMPONENT_SCOPED_TYPES = {
    "BH腹", "BH翼", "BOX腹", "BOX翼", "BT腹", "BT翼",
}


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _rows_by_headers(sheet) -> list[dict[str, object]]:
    headers = [cell.value for cell in sheet[1]]
    return [
        dict(zip(headers, values, strict=True))
        for values in sheet.iter_rows(min_row=2, values_only=True)
    ]


@pytest.mark.live_data
def test_real_ground_truth_invariants_with_live_mysql(tmp_path: Path) -> None:
    if not SOURCE.is_file() or not PREPROCESSED.is_file():
        pytest.skip("real ground-truth source or reviewed single-sheet input is absent")
    from app.modules.excel_processing.stage_adapter import run_excel_final_pipeline

    baseline = json.loads(BASELINE_PATH.read_text(encoding="utf-8"))
    assert _sha256(SOURCE) == baseline["sha256"]
    output = tmp_path / "ground-truth-canonical.xlsx"

    result = run_excel_final_pipeline(PREPROCESSED, output, source_format="canonical")

    assert result.protocol_version == 1
    assert result.output_path == output.resolve()
    assert result.quality_status == "ok"
    assert result.warning_count == 0
    assert result.severe_warning_count == 0
    assert result.report_summary["info_count"] == 0
    assert _sha256(SOURCE) == baseline["sha256"]
    formulas = load_workbook(output, data_only=False)
    values = load_workbook(output, read_only=True, data_only=True)
    try:
        assert formulas.sheetnames == ["原表", "清洗表", "构件表", "整理表", "part", "处理报告"]
        cleaned = _rows_by_headers(values["清洗表"])
        components = _rows_by_headers(values["构件表"])
        organized = _rows_by_headers(values["整理表"])
        part = _rows_by_headers(values["part"])
        assert len(cleaned) == baseline["parent_parts"]
        assert len(components) == baseline["components"]
        assert len({row["构件编号"] for row in components}) == baseline["components"]
        assert len(organized) == baseline["organized_rows"]

        source_types: dict[str, int] = {"PL": 0, "BOX": 0, "TT": 0, "D": 0, "NUT": 0}
        for row in cleaned:
            spec = str(row["原规格"])
            for prefix in source_types:
                if spec.startswith(prefix):
                    source_types[prefix] += 1
                    break
        assert source_types == {
            "PL": baseline["pl"],
            "BOX": baseline["box"],
            "TT": baseline["tt"],
            "D": baseline["d"],
            "NUT": baseline["nut"],
        }

        component_scoped = [
            row for row in part if row["类型"] in COMPONENT_SCOPED_TYPES
        ]
        global_scoped = [
            row for row in part if row["类型"] not in COMPONENT_SCOPED_TYPES
        ]
        assert len(part) == baseline["part_rows"]
        assert len(component_scoped) == baseline["part_component_scoped"]
        assert len(global_scoped) == baseline["part_global_scoped"]
        assert all(row["导入构件编号"] for row in component_scoped)
        assert all(row["导入构件编号"] is None for row in global_scoped)
        assert sum(row["汇总"] for row in global_scoped) == baseline["part_global_summary"]
        assert all(row["文件"] is None for row in part)
        assert values["处理报告"]["A2"].value == "无"
        assert values["处理报告"].max_row == 2
        for sheet_name in ("清洗表", "构件表", "整理表", "part", "处理报告"):
            worksheet = formulas[sheet_name]
            for column in range(1, worksheet.max_column + 1):
                letter = get_column_letter(column)
                width = worksheet.column_dimensions[letter].width
                if sheet_name == "处理报告" and column in (7, 8):
                    assert 16 <= width <= 48
                else:
                    assert 8 <= width <= 32
        for coordinate in ("G2", "H2"):
            assert formulas["处理报告"][coordinate].alignment.wrap_text is True
            assert formulas["处理报告"][coordinate].alignment.vertical == "top"
        assert formulas["构件表"].auto_filter.ref == "A1:O1"
        assert formulas["整理表"].auto_filter.ref == "A1:AF1"
        for sheet_name, removed_headers in (
            ("整理表", ("比重来源", "净材利用率", "重量核验")),
            ("构件表", ("来源sheet", "行类型", "小计来源行")),
        ):
            worksheet = formulas[sheet_name]
            header_values = {cell.value for cell in worksheet[1]}
            assert not (set(removed_headers) & header_values)

        d_rows = [row for row in organized if str(row["截面型材"]).startswith("D")]
        assert len(d_rows) == baseline["d"]
        assert {row["规格"] for row in d_rows} == {24, 30}
        assert {row["比重"] for row in d_rows} == {3.55, 5.55}
        assert all(row["理单重(kg)"] is not None for row in d_rows)

        skipped = [
            row for row in organized
            if str(row["截面型材"]).startswith(("NUT", "TT"))
        ]
        assert len(skipped) == baseline["nut"] + baseline["tt"]
        assert all(row["比重"] is None for row in skipped)
        assert all(row["理单重(kg)"] is None and row["理总重(kg)"] is None for row in skipped)

        box_rows: dict[object, list[dict[str, object]]] = {}
        for row in organized:
            if str(row["截面型材"]).startswith("BOX"):
                box_rows.setdefault(row["序号"], []).append(row)
        assert len(box_rows) == baseline["box"]
        assert all(len(rows) == 2 for rows in box_rows.values())
        assert all(
            sum(row["单毛重(kg)"] is not None for row in rows) == 1
            and sum(row["理总重(kg)"] is not None for row in rows) == 1
            for rows in box_rows.values()
        )

        assert formulas["整理表"]["P2"].value == "=M2-N2-O2"
        assert values["整理表"]["P2"].value is not None
    finally:
        formulas.close()
        values.close()
