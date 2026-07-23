from __future__ import annotations

import importlib
from decimal import Decimal

from openpyxl import Workbook, load_workbook


def test_ooxml_formula_patch_preserves_formula_and_adds_numeric_cache(tmp_path) -> None:
    cache = importlib.import_module("ooxml_formula")
    path = tmp_path / "formula.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "整理表"
    sheet["A1"] = 1000
    sheet["B1"] = 10
    sheet["C1"] = 5
    sheet["D1"] = "=A1-B1-C1"
    workbook.save(path)
    workbook.close()

    cache.patch_formula_caches(
        path,
        "整理表",
        {"D1": cache.FormulaCache("=A1-B1-C1", Decimal("985"))},
    )

    formulas = load_workbook(path, data_only=False, read_only=True)
    values = load_workbook(path, data_only=True, read_only=True)
    try:
        assert formulas["整理表"]["D1"].value == "=A1-B1-C1"
        assert values["整理表"]["D1"].value == 985
    finally:
        formulas.close()
        values.close()
