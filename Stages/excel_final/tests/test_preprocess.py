from __future__ import annotations

import hashlib
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

import preprocess
from conftest import STAGE_ROOT


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _run_preprocessor(source: Path, output: Path) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [sys.executable, str(STAGE_ROOT / "preprocess.py"), str(source), str(output)],
        cwd=STAGE_ROOT,
        capture_output=True,
        text=True,
        check=False,
    )


def test_preprocessor_extracts_only_equivalent_raw_sheet(
    styled_multisheet_workbook: Path,
    tmp_path: Path,
) -> None:
    source = styled_multisheet_workbook
    output = tmp_path / "preprocessed" / "raw-only.xlsx"
    source_hash_before = _sha256(source)

    completed = _run_preprocessor(source, output)

    assert completed.returncode == 0, completed.stderr
    assert output.is_file()
    assert _sha256(source) == source_hash_before

    source_book = load_workbook(source, data_only=False)
    output_book = load_workbook(output, data_only=False)
    try:
        assert output_book.sheetnames == ["原表"]
        source_sheet = source_book.worksheets[0]
        output_sheet = output_book["原表"]
        assert output_sheet.max_row == source_sheet.max_row
        assert output_sheet.max_column == source_sheet.max_column
        assert output_sheet.calculate_dimension() == source_sheet.calculate_dimension()
        assert set(output_sheet.merged_cells.ranges) == set(source_sheet.merged_cells.ranges)
        assert output_sheet.freeze_panes == source_sheet.freeze_panes
        assert output_sheet.sheet_view.showGridLines == source_sheet.sheet_view.showGridLines

        for source_row, output_row in zip(source_sheet.iter_rows(), output_sheet.iter_rows()):
            for source_cell, output_cell in zip(source_row, output_row):
                assert output_cell.value == source_cell.value
                assert output_cell.data_type == source_cell.data_type
                assert output_cell.style_id == source_cell.style_id

        assert output_sheet.column_dimensions["A"].width == source_sheet.column_dimensions["A"].width
        assert output_sheet.column_dimensions["C"].hidden is True
        assert output_sheet.row_dimensions[1].height == source_sheet.row_dimensions[1].height
    finally:
        source_book.close()
        output_book.close()


def test_preprocessor_rejects_overwriting_source(
    styled_multisheet_workbook: Path,
) -> None:
    source = styled_multisheet_workbook
    source_hash_before = _sha256(source)

    completed = _run_preprocessor(source, source)

    assert completed.returncode != 0
    assert "different" in completed.stderr.lower() or "不同" in completed.stderr
    assert _sha256(source) == source_hash_before


def test_preprocessor_returns_verified_summary(
    styled_multisheet_workbook: Path,
    tmp_path: Path,
) -> None:
    source = styled_multisheet_workbook
    output = tmp_path / "verified.xlsx"

    result = preprocess.preprocess_ground_truth(source, output)

    assert result.source_path == source.resolve()
    assert result.output_path == output.resolve()
    assert result.source_sha256 == _sha256(source)
    assert result.output_sha256 == _sha256(output)
    assert result.source_sheet_name == "原始清单"
    assert result.output_sheet_name == "原表"
    assert result.dimension == "A1:C3"


def test_preprocessor_compares_many_merge_ranges_without_order_dependence(
    tmp_path: Path,
) -> None:
    source = tmp_path / "many-merges.xlsx"
    output = tmp_path / "many-merges-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原始清单"
    for block in range(35):
        start = block * 5 + 1
        sheet.cell(start, 1, f"构件-{block}")
        sheet.merge_cells(start_row=start, start_column=1, end_row=start + 2, end_column=1)
        sheet.cell(start + 3, 2, f"小计-{block}")
        sheet.merge_cells(
            start_row=start + 3,
            start_column=2,
            end_row=start + 3,
            end_column=4,
        )
    workbook.save(source)
    workbook.close()

    preprocess.preprocess_ground_truth(source, output)

    source_book = load_workbook(source, data_only=False)
    output_book = load_workbook(output, data_only=False)
    try:
        source_merges = {str(item) for item in source_book.worksheets[0].merged_cells.ranges}
        output_merges = {str(item) for item in output_book["原表"].merged_cells.ranges}
        assert output_merges == source_merges
    finally:
        source_book.close()
        output_book.close()
