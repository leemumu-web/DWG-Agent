from __future__ import annotations

from dataclasses import FrozenInstanceError, replace
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

import canonical_pipeline
from domain import ComponentRowKind, ComponentSourceRow, SourcePart
from writer_parts import FormulaLengthBasis


class NoLookupHandbook:
    def lookup(self, *_args, **_kwargs):
        raise AssertionError("plate projection must not query the handbook")


def _source_part() -> SourcePart:
    return SourcePart(
        source_sheet="原表",
        source_row=3,
        source_seq=1,
        batch="B1",
        component_no="C1",
        component_qty=Decimal("2"),
        part_no="P1",
        original_spec="PL10*100",
        material="Q355B",
        length=Decimal("1000"),
        original_qty=Decimal("3"),
        source_unit_net=Decimal("7.8"),
        source_total_net=Decimal("23.4"),
        source_unit_gross=Decimal("7.85"),
        source_total_gross=Decimal("23.55"),
        source_unit_area=Decimal("0.22"),
        source_total_area=Decimal("0.66"),
        classification=None,
    )


def _component_row() -> ComponentSourceRow:
    return ComponentSourceRow(
        source_sheet="原表",
        source_row=2,
        kind=ComponentRowKind.START,
        batch="B1",
        component_no="C1",
        component_qty=Decimal("2"),
        original_spec="BH500*200*10*16",
        material="Q355B",
        source_unit_net=None,
        source_total_net=None,
        source_unit_gross=None,
        source_total_gross=None,
        source_unit_area=None,
        source_total_area=None,
        component_length=Decimal("1000"),
        component_width=Decimal("200"),
        component_height=Decimal("500"),
    )


def test_build_canonical_projection_returns_immutable_rows_without_writing() -> None:
    source = _source_part()
    component = _component_row()
    assert hasattr(canonical_pipeline, "build_canonical_projection")

    projection = canonical_pipeline.build_canonical_projection(
        parts=iter((source,)),
        component_rows=iter((component,)),
        reader_issues=(),
        handbook=NoLookupHandbook(),
    )

    assert projection.cleaned_parts == (replace(source, classification="板材"),)
    assert projection.component_rows == (component,)
    assert len(projection.organized_rows) == 1
    assert projection.organized_rows[0]["导入零件号"] == "P1"
    assert projection.organized_rows[0]["下料长度(mm)"] == Decimal("1000")
    assert len(projection.part_candidates) == 1
    assert projection.part_candidates[0].import_part_no == "P1"
    assert isinstance(projection.issues, tuple)

    with pytest.raises(FrozenInstanceError):
        projection.cleaned_parts = ()


def _source_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "原表"
    workbook.active["A1"] = "原始数据"
    workbook.save(path)
    workbook.close()


def test_write_canonical_projection_emits_the_existing_six_sheet_contract(
    tmp_path: Path,
) -> None:
    assert hasattr(canonical_pipeline, "write_canonical_projection")
    source_path = tmp_path / "source.xlsx"
    output_path = tmp_path / "output.xlsx"
    _source_workbook(source_path)
    projection = canonical_pipeline.build_canonical_projection(
        parts=(_source_part(),),
        component_rows=(_component_row(),),
        reader_issues=(),
        handbook=NoLookupHandbook(),
    )

    outcome = canonical_pipeline.write_canonical_projection(
        source_path,
        output_path,
        projection=projection,
    )

    assert outcome.output_path == output_path.resolve()
    workbook = load_workbook(output_path, data_only=False, read_only=True)
    try:
        assert workbook.sheetnames == [
            "原表",
            "清洗表",
            "构件表",
            "整理表",
            "part",
            "处理报告",
        ]
        assert workbook["part"]["G2"].value == "=SUM('整理表'!T2)"
    finally:
        workbook.close()


def test_write_canonical_projection_forwards_the_formula_length_policy(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source_path = tmp_path / "source.xlsx"
    output_path = tmp_path / "output.xlsx"
    _source_workbook(source_path)
    projection = canonical_pipeline.build_canonical_projection(
        parts=(_source_part(),),
        component_rows=(_component_row(),),
        reader_issues=(),
        handbook=NoLookupHandbook(),
    )
    captured: dict[str, object] = {}
    expected_outcome = object()

    def capture_writer(*_args, **kwargs):
        captured.update(kwargs)
        return expected_outcome

    monkeypatch.setattr(canonical_pipeline, "write_canonical_workbook", capture_writer)

    outcome = canonical_pipeline.write_canonical_projection(
        source_path,
        output_path,
        projection=projection,
        formula_length_basis=FormulaLengthBasis.CUT_LENGTH,
    )

    assert outcome is expected_outcome
    assert captured["formula_length_basis"] is FormulaLengthBasis.CUT_LENGTH
