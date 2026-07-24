from __future__ import annotations

import importlib
import logging
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from domain import ComponentRowKind, ComponentSourceRow, SourcePart
from part_builder import PartRow
from quality import IssueLevel, QualityIssue


def _writer():
    return importlib.import_module("writer_parts")


def _source(path: Path) -> tuple[SourcePart, tuple[ComponentSourceRow, ...]]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原表"
    sheet["A1"] = "原始标题"
    sheet["A1"].font = Font(bold=True, color="123456")
    sheet["B2"] = " 保 留 空 格 "
    sheet.column_dimensions["A"].width = 27.5
    sheet.column_dimensions["B"].width = 13.25
    workbook.save(path)
    workbook.close()

    part = SourcePart(
        source_sheet="原表",
        source_row=8,
        source_seq=7,
        batch="B1",
        component_no="C1",
        component_qty=Decimal("2"),
        part_no="p1",
        original_spec="PL10*100",
        material="Q355B",
        length=Decimal("1000"),
        original_qty=Decimal("3"),
        source_unit_net=Decimal("7.8"),
        source_total_net=Decimal("23.4"),
        source_unit_gross=Decimal("7.85"),
        source_total_gross=Decimal("23.55"),
        source_unit_area=Decimal("0.22"),
        source_total_area=Decimal("0.66"),
        classification="板材",
    )
    rows = (
        ComponentSourceRow(
            source_sheet="原表",
            source_row=7,
            kind=ComponentRowKind.SUMMARY,
            batch="B1",
            component_no="C1",
            component_qty=Decimal("2"),
            original_spec="BOX700*700*36*36",
            material="Q355B",
            source_unit_net=Decimal("7.8"),
            source_total_net=Decimal("15.6"),
            source_unit_gross=Decimal("7.85"),
            source_total_gross=Decimal("15.7"),
            source_unit_area=Decimal("0.22"),
            source_total_area=Decimal("0.44"),
            component_length=Decimal("1000"),
            component_width=Decimal("100"),
            component_height=Decimal("10"),
            subtotal_source_row=9,
        ),
    )
    return part, rows


def _organized_row(**overrides: object) -> dict[str, object]:
    row: dict[str, object] = {
        "序号": 7,
        "构件编号": "C1",
        "导入构件编号": "C1",
        "构件数": Decimal("2"),
        "类型": "板材",
        "班组": "",
        "批次": "B1",
        "零件号": "p1",
        "导入零件号": "p1",
        "截面型材": "PL10*100",
        "规格": Decimal("10"),
        "宽度": Decimal("100"),
        "长度(mm)": Decimal("1000"),
        "左进(mm)": None,
        "右进(mm)": None,
        "下料长度(mm)": Decimal("1000"),
        "材质": "Q355B",
        "原数量": Decimal("3"),
        "数量": Decimal("3"),
        "总数": Decimal("6"),
        "总长(mm)": Decimal("6000"),
        "比重": "查无",
        "比重来源": "flat_steel:not_found",
        "理单重(kg)": Decimal("7.85"),
        "理总重(kg)": Decimal("47.10"),
        "单净重(kg)": Decimal("7.8"),
        "总净重(kg)": Decimal("23.4"),
        "表净重(kg)": Decimal("46.8"),
        "单毛重(kg)": Decimal("7.85"),
        "总毛重(kg)": Decimal("23.55"),
        "表毛重(kg)": Decimal("47.10"),
        "净材利用率": Decimal("0.993630573"),
        "重量核验": "严重",
        "单表面积(㎡)": Decimal("0.22"),
        "总表面积(㎡)": Decimal("0.66"),
        "_source_row": 8,
    }
    row.update(overrides)
    return row


def _issues() -> tuple[QualityIssue, ...]:
    common = {
        "source_sheet": "原表",
        "source_row": 8,
        "component_no": "C1",
        "part_no": "p1",
        "spec": "PL10*100",
        "absolute_error": None,
        "relative_error": None,
        "density_source": "flat_steel:not_found",
    }
    return (
        QualityIssue(
            level=IssueLevel.WARNING,
            category="五金手册查无",
            field="比重",
            actual_value="查无",
            expected_value="手册命中",
            affects_part=False,
            description="扁钢规格查无",
            **common,
        ),
        QualityIssue(
            level=IssueLevel.SEVERE,
            category="源重量链异常",
            field="单毛重",
            actual_value=Decimal("8.5"),
            expected_value=Decimal("7.85"),
            affects_part=True,
            description="单毛重与总毛重链异常",
            **common,
        ),
    )


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (None, 0),
        ("ABCD", 4),
        ("构件编号", 8),
        ("ＡB", 3),
        ("a\n中文", 4),
    ],
)
def test_display_width_counts_east_asian_characters_and_longest_line(
    value: object,
    expected: int,
) -> None:
    assert _writer()._display_width(value) == expected


def test_canonical_writer_emits_six_sheets_with_adaptive_widths_and_audited_styles(
    tmp_path: Path,
) -> None:
    writer = _writer()
    source = tmp_path / "source.xlsx"
    part, component_rows = _source(source)
    output = tmp_path / "output.xlsx"
    internal_output = tmp_path / "internal-output.xlsx"
    organized = [
        _organized_row(**{"重量核验": "通过"}),
        _organized_row(
            类型="BOX翼",
            序号=7,
            导入零件号="p1-BOX翼",
            截面型材="PL" + "1" * 80,
        ),
    ]
    part_rows = (
        PartRow(
            "",
            "p1",
            Decimal("10"),
            Decimal("100"),
            Decimal("1000"),
            "Q355B",
            Decimal("6"),
            "",
            "",
            "板材",
        ),
    )

    outcome = writer.write_canonical_workbook(
        source,
        output,
        cleaned_parts=(part,),
        component_rows=component_rows,
        organized_rows=organized,
        part_rows=part_rows,
        issues=_issues(),
        internal_output_path=internal_output,
    )

    workbook = load_workbook(output, data_only=False)
    try:
        assert workbook.sheetnames == ["原表", "清洗表", "构件表", "整理表", "part", "处理报告"]
        assert workbook["原表"]["A1"].value == "原始标题"
        assert workbook["原表"]["A1"].font.bold is True
        assert workbook["原表"]["B2"].value == " 保 留 空 格 "
        assert workbook["原表"].column_dimensions["A"].width == 27.5
        assert workbook["原表"].column_dimensions["B"].width == 13.25
        assert [cell.value for cell in workbook["整理表"][1]] == [
            header
            for header in writer.ORGANIZED_HEADERS
            if header not in {"比重来源", "净材利用率", "重量核验"}
        ]
        assert [cell.value for cell in workbook["构件表"][1]] == [
            header
            for header in writer.COMPONENT_HEADERS
            if header not in {"来源sheet", "行类型", "小计来源行"}
        ]
        assert [cell.value for cell in workbook["part"][1]] == [
            *writer.PART_HEADERS[:9],
            "备注",
            "文件",
            "类型",
        ]
        assert workbook["part"]["J1"].value == "备注"
        assert workbook["part"]["K1"].value == "文件"
        assert writer.REPORT_HEADERS == [
            "级别",
            "类别",
            "来源位置",
            "构件编号",
            "零件号",
            "涉及字段",
            "说明",
            "建议操作",
        ]
        assert [cell.value for cell in workbook["处理报告"][1]] == writer.REPORT_HEADERS
        assert workbook["处理报告"]["H2"].value
        assert workbook["处理报告"]["H3"].value
        assert workbook["整理表"]["A2"].value == workbook["整理表"]["A3"].value == 7
        assert workbook["整理表"]["P2"].value == "=M2-N2-O2"
        organized_headers = [cell.value for cell in workbook["整理表"][1]]
        theo_unit_column = get_column_letter(organized_headers.index("理单重(kg)") + 1)
        gross_unit_column = get_column_letter(organized_headers.index("单毛重(kg)") + 1)
        density_column = get_column_letter(organized_headers.index("比重") + 1)
        assert workbook["整理表"][f"{theo_unit_column}2"].number_format == "0.000"
        assert workbook["整理表"][f"{density_column}2"].font.color.rgb.endswith("FF0000")
        assert workbook["整理表"][f"{gross_unit_column}2"].fill.fill_type == "solid"
        assert workbook["part"].max_row == 2
        assert workbook["part"]["H2"].value is None
        assert workbook["part"]["I2"].value is None
        assert workbook["part"]["J2"].value is None
        assert workbook["part"]["K2"].value is None
        assert workbook["part"]["L2"].value is None
        assert workbook["处理报告"].max_row == 3
        assert workbook["构件表"].max_row == 2
        assert workbook["构件表"]["A2"].value == 7
        assert workbook["构件表"]["B2"].value == "B1"
        assert workbook["构件表"]["C2"].value == "C1"
        assert workbook["构件表"].auto_filter.ref == "A1:O1"
        assert workbook["整理表"].auto_filter.ref == "A1:AF1"
        assert workbook["整理表"].column_dimensions["B"].width == 10
        assert workbook["整理表"].column_dimensions["J"].width == 32
        assert workbook["part"].column_dimensions["A"].width == 14
        assert 16 <= workbook["处理报告"].column_dimensions["G"].width <= 48
        assert 16 <= workbook["处理报告"].column_dimensions["H"].width <= 48
        for coordinate in ("G2", "H2", "G3", "H3"):
            assert workbook["处理报告"][coordinate].alignment.wrap_text is True
            assert workbook["处理报告"][coordinate].alignment.vertical == "top"
        assert not ({"比重来源", "净材利用率", "重量核验"} & set(organized_headers))
        assert workbook["整理表"]["E2"].value is None
        assert workbook["整理表"]["E3"].value == "BOX翼"
        component_headers = [cell.value for cell in workbook["构件表"][1]]
        assert not ({"来源sheet", "行类型", "小计来源行"} & set(component_headers))
        assert workbook["part"]["L1"].value == "类型"
    finally:
        workbook.close()

    internal_workbook = load_workbook(internal_output, data_only=False)
    try:
        assert [cell.value for cell in internal_workbook["整理表"][1]] == (writer.ORGANIZED_HEADERS)
        assert [cell.value for cell in internal_workbook["构件表"][1]] == (writer.COMPONENT_HEADERS)
        assert internal_workbook["整理表"]["P2"].value == "=M2-N2-O2"
    finally:
        internal_workbook.close()

    assert outcome.output_path == output.resolve()
    assert outcome.warning_count == 1
    assert outcome.severe_warning_count == 1
    assert outcome.quality_status == "severe_warning"


def test_final_workbook_exposes_only_split_child_types_and_keeps_file_in_k(
    tmp_path: Path,
) -> None:
    writer = _writer()
    source = tmp_path / "source.xlsx"
    source_part, component_rows = _source(source)
    output = tmp_path / "output.xlsx"
    internal_output = tmp_path / "internal-output.xlsx"
    split_types = ("BH腹", "BH翼", "BOX腹", "BOX翼", "BT腹", "BT翼")
    row_types = ("板材", *split_types)
    organized_rows = [
        _organized_row(
            **{
                "类型": part_type,
                "导入零件号": f"p-{index}",
                "比重": Decimal("7.85"),
                "比重来源": "plate_constant:7.85",
                "重量核验": "通过",
            }
        )
        for index, part_type in enumerate(row_types)
    ]
    part_rows = tuple(
        PartRow(
            "C1" if part_type in split_types else "",
            f"p-{index}",
            Decimal("10"),
            Decimal("100"),
            Decimal("1000"),
            "Q355B",
            Decimal("6"),
            "",
            "",
            part_type,
        )
        for index, part_type in enumerate(row_types)
    )

    writer.write_canonical_workbook(
        source,
        output,
        cleaned_parts=(source_part,),
        component_rows=component_rows,
        organized_rows=organized_rows,
        part_rows=part_rows,
        issues=(),
        internal_output_path=internal_output,
    )

    workbook = load_workbook(output, data_only=False)
    try:
        organized = workbook["整理表"]
        organized_headers = [cell.value for cell in organized[1]]
        organized_type_column = organized_headers.index("类型") + 1
        organized_types = [
            organized.cell(row, organized_type_column).value
            for row in range(2, organized.max_row + 1)
        ]
        assert organized_types[0] is None
        assert tuple(organized_types[1:]) == split_types

        part = workbook["part"]
        part_headers = [cell.value for cell in part[1]]
        assert part_headers[9:12] == ["备注", "文件", "类型"]
        assert part["K1"].value == "文件"
        part_type_column = part_headers.index("类型") + 1
        part_types = [
            part.cell(row, part_type_column).value
            for row in range(2, part.max_row + 1)
        ]
        assert part_types[0] is None
        assert tuple(part_types[1:]) == split_types
    finally:
        workbook.close()

    internal = load_workbook(internal_output, data_only=False)
    try:
        assert internal["整理表"]["E2"].value == "板材"
        assert internal["part"]["J2"].value == "板材"
    finally:
        internal.close()


def test_writer_formula_cache_is_immediately_readable_data_only(tmp_path: Path) -> None:
    writer = _writer()
    source = tmp_path / "source.xlsx"
    part, component_rows = _source(source)
    output = tmp_path / "output.xlsx"

    writer.write_canonical_workbook(
        source,
        output,
        cleaned_parts=(part,),
        component_rows=component_rows,
        organized_rows=[
            _organized_row(
                **{
                    "左进(mm)": Decimal("10"),
                    "右进(mm)": Decimal("5"),
                    "下料长度(mm)": Decimal("985"),
                    "比重": Decimal("7.85"),
                    "比重来源": "plate_constant:7.85",
                    "重量核验": "通过",
                }
            )
        ],
        part_rows=(
            PartRow(
                "",
                "p1",
                Decimal("10"),
                Decimal("100"),
                Decimal("985"),
                "Q355B",
                Decimal("6"),
                "",
                "",
                "板材",
            ),
        ),
        issues=(),
    )

    formulas = load_workbook(output, data_only=False, read_only=True)
    values = load_workbook(output, data_only=True, read_only=True)
    try:
        assert formulas["整理表"]["P2"].value == "=M2-N2-O2"
        assert formulas["整理表"]["T2"].value == "=D2*S2"
        assert formulas["整理表"]["U2"].value == "=M2*T2"
        assert formulas["整理表"]["W2"].value == "=ROUND(K2*L2*M2*V2/1000000,3)"
        assert formulas["整理表"]["X2"].value == ("=ROUND(K2*L2*M2*V2/1000000*T2,3)")
        assert formulas["整理表"]["AA2"].value == "=ROUND(Z2*D2,3)"
        assert formulas["整理表"]["AD2"].value == "=ROUND(AC2*D2,3)"
        assert formulas["part"]["G2"].value == "=SUM('整理表'!T2)"
        assert values["整理表"]["P2"].value == 985
        assert values["整理表"]["T2"].value == 6
        assert values["整理表"]["U2"].value == 6000
        assert values["整理表"]["W2"].value == 7.85
        assert values["整理表"]["X2"].value == 47.1
        assert values["整理表"]["AA2"].value == 46.8
        assert values["整理表"]["AD2"].value == 47.1
        assert values["part"]["G2"].value == 6
        assert formulas["处理报告"]["A2"].value == "无"
        assert formulas["处理报告"].max_row == 2
    finally:
        formulas.close()
        values.close()


def test_canonical_writer_rejects_non_xlsx_output(tmp_path: Path) -> None:
    writer = _writer()
    source = tmp_path / "source.xlsx"
    part, component_rows = _source(source)
    output = tmp_path / "output.xlsm"

    with pytest.raises(ValueError, match=r"\.xlsx"):
        writer.write_canonical_workbook(
            source,
            output,
            cleaned_parts=(part,),
            component_rows=component_rows,
            organized_rows=[_organized_row()],
            part_rows=(),
            issues=(),
        )

    assert not output.exists()


def test_canonical_writer_log_does_not_disclose_absolute_path(
    tmp_path: Path, caplog: pytest.LogCaptureFixture
) -> None:
    writer = _writer()
    source = tmp_path / "private-source.xlsx"
    part, component_rows = _source(source)
    output = tmp_path / "private-output.xlsx"

    with caplog.at_level(logging.INFO, logger="writer_parts"):
        writer.write_canonical_workbook(
            source,
            output,
            cleaned_parts=(part,),
            component_rows=component_rows,
            organized_rows=[_organized_row()],
            part_rows=(),
            issues=(),
        )

    assert output.name in caplog.text
    assert str(tmp_path) not in caplog.text
