"""Read 初始表 (initial table) format — the 9-column flat format from DWG extraction.

Row 1: component info text → ComponentInfo
Row 2: column headers (零件号/截面型材/长度/材质/数量/单重/总重/总面积/备注)
Row 3+: part data rows, terminated by 合计 row
"""

from __future__ import annotations

import math
import re
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from types import MappingProxyType
from typing import Any, Mapping

import openpyxl

from domain import SourcePart
from input_contract import InputContractError, InputKind, inspect_production_input
from input_errors import ExcelInputIssue, input_failure
from legacy_xls import open_legacy_workbook
from utils import safe_float, safe_str

# ── Dataclasses ──────────────────────────────────────────────────


@dataclass
class ComponentInfo:
    """Parsed component metadata from Row 1."""
    component_no: str       # e.g. "B7-4FD-ZL-19"
    component_qty: int      # e.g. 1
    total_weight: float     # e.g. 1739.26
    raw_text: str           # full Row 1 text
    source_row: int = 1


@dataclass
class PartRow:
    """A single part data row from the initial table."""
    part_no: str            # 零件号
    spec: str               # 截面型材 (original spec string)
    length: float | None    # 长度(mm)
    material: str           # 材质
    qty: float | None       # 数量
    unit_weight: float | None   # 单重(kg)
    total_weight: float | None  # 总重(kg)
    surface_area: float | None  # 总面积(m2)
    note: str               # 备注
    original_seq: int       # 1-based position in data rows
    source_row: int


@dataclass(frozen=True, slots=True)
class InitialLayout:
    metadata_row: int
    header_row: int
    columns: Mapping[str, int]


_INITIAL_HEADER_ALIASES = {
    "part_no": frozenset({"零件号", "零件编号"}),
    "spec": frozenset({"截面型材", "型材", "规格"}),
    "length": frozenset({"长度"}),
    "material": frozenset({"材质"}),
    "qty": frozenset({"数量"}),
    "unit_weight": frozenset({"单重", "单毛重"}),
    "total_weight": frozenset({"总重", "总毛重"}),
    "surface_area": frozenset({"总面积", "总表面积"}),
    "note": frozenset({"备注"}),
}
_INITIAL_ALIAS_TO_FIELD = {
    alias: field
    for field, aliases in _INITIAL_HEADER_ALIASES.items()
    for alias in aliases
}
_INITIAL_REQUIRED_FIELDS = frozenset({"part_no", "spec", "length", "material", "qty"})
_INITIAL_FIELD_LABELS = {
    "length": "长度",
    "qty": "数量",
    "unit_weight": "单重",
    "total_weight": "总重",
    "surface_area": "总面积",
}

# ── Public API ───────────────────────────────────────────────────


def read_init_table(filepath: str | Path) -> tuple[ComponentInfo, list[PartRow]]:
    """Read an 初始表-format .xlsx file.

    Returns (ComponentInfo, list of PartRow).
    """
    filepath = Path(filepath)
    inspected = inspect_production_input(filepath)
    if inspected.sheet_name is None:
        raise ValueError("initial-table reader requires a named worksheet")
    wb = (
        open_legacy_workbook(inspected.path)
        if inspected.kind is InputKind.LEGACY_WORKBOOK
        else openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    )

    # Locate the 初始表 sheet
    if "初始表" in wb.sheetnames:
        ws = wb["初始表"]
    else:
        ws = wb.worksheets[0]

    layout = detect_initial_layout(ws)
    metadata_text = _join_row(ws, layout.metadata_row)
    parsed = parse_component_info(metadata_text)
    comp_info = ComponentInfo(
        component_no=parsed.component_no,
        component_qty=parsed.component_qty,
        total_weight=parsed.total_weight,
        raw_text=parsed.raw_text,
        source_row=layout.metadata_row,
    )

    # Read data rows after the detected header until 合计.
    parts: list[PartRow] = []
    seq = 0
    for row_idx in range(layout.header_row + 1, ws.max_row + 1):
        values = {
            field: _cell_str(ws, row_idx, column)
            for field, column in layout.columns.items()
        }
        part_no = values.get("part_no", "")
        spec = values.get("spec", "")

        # Stop at 合计 row
        if "合计" in part_no or "合计" in spec:
            break

        # Skip empty rows
        if not any(values.values()):
            continue

        seq += 1
        parts.append(PartRow(
            part_no=part_no,
            spec=spec,
            length=_layout_float(ws, row_idx, layout, "length"),
            material=values.get("material", ""),
            qty=_layout_float(ws, row_idx, layout, "qty"),
            unit_weight=_layout_float(ws, row_idx, layout, "unit_weight"),
            total_weight=_layout_float(ws, row_idx, layout, "total_weight"),
            surface_area=_layout_float(ws, row_idx, layout, "surface_area"),
            note=values.get("note", ""),
            original_seq=seq,
            source_row=row_idx,
        ))

    wb.close()
    return comp_info, parts


def _compact_working_text(value: str) -> str:
    return value.replace(" ", "").replace("　", "")


def _decimal_or_none(value: float | None) -> Decimal | None:
    if value is None:
        return None
    result = Decimal(str(value))
    if not result.is_finite():
        raise ValueError("initial-table numeric value must be finite")
    return result


def read_init_canonical(filepath: str | Path) -> tuple[SourcePart, ...]:
    """Adapt a one-sheet initial table to canonical parts without changing source data."""
    inspected = inspect_production_input(Path(filepath))
    if inspected.kind not in {InputKind.WORKBOOK, InputKind.LEGACY_WORKBOOK}:
        raise ValueError("initial-table canonical reader requires a one-sheet workbook source")
    if inspected.sheet_name is None:
        raise ValueError("initial-table canonical reader requires a named worksheet")

    component, rows = read_init_table(inspected.path)
    component_no = _compact_working_text(component.component_no)
    component_qty = Decimal(str(component.component_qty))
    result: list[SourcePart] = []
    for row in rows:
        length = _decimal_or_none(row.length)
        quantity = _decimal_or_none(row.qty)
        part_no = _compact_working_text(row.part_no)
        original_spec = _compact_working_text(row.spec)
        material = _compact_working_text(row.material)
        invalid_fields = tuple(
            field
            for field, missing in (
                ("构件编号", not component_no),
                ("零件号", not part_no),
                ("规格", not original_spec),
                ("长度", length is None),
                ("材质", not material),
                ("数量", quantity is None),
                ("构件数", component_qty == 0),
            )
            if missing
        )
        result.append(SourcePart(
            source_sheet=inspected.sheet_name,
            source_row=row.source_row,
            source_seq=row.original_seq,
            batch=None,
            component_no=component_no,
            component_qty=component_qty,
            part_no=part_no,
            original_spec=original_spec,
            material=material,
            length=length or Decimal("0"),
            original_qty=quantity or Decimal("0"),
            source_unit_net=None,
            source_total_net=None,
            source_unit_gross=_decimal_or_none(row.unit_weight),
            source_total_gross=_decimal_or_none(row.total_weight),
            source_unit_area=None,
            source_total_area=_decimal_or_none(row.surface_area),
            classification=None,
            invalid_fields=invalid_fields,
        ))
    return tuple(result)


def _normalized_initial_header(value: Any) -> str:
    compact = "".join(str(value or "").split())
    return re.sub(r"[（(][^）)]*[）)]", "", compact)


def _initial_columns(values: tuple[Any, ...]) -> tuple[dict[str, int], tuple[str, ...]]:
    matches: dict[str, list[int]] = {}
    for index, value in enumerate(values, start=1):
        field = _INITIAL_ALIAS_TO_FIELD.get(_normalized_initial_header(value))
        if field is not None:
            matches.setdefault(field, []).append(index)
    conflicts = tuple(
        f"{field} columns={indexes}"
        for field, indexes in matches.items()
        if len(indexes) > 1
    )
    return {field: indexes[0] for field, indexes in matches.items()}, conflicts


def detect_initial_layout(worksheet: Any) -> InitialLayout:
    candidates: list[tuple[int, dict[str, int], tuple[str, ...]]] = []
    for row_number, values in enumerate(
        worksheet.iter_rows(
            min_row=1,
            max_row=min(worksheet.max_row, 100),
            values_only=True,
        ),
        start=1,
    ):
        columns, conflicts = _initial_columns(values)
        if _INITIAL_REQUIRED_FIELDS <= columns.keys():
            candidates.append((row_number, columns, conflicts))
    valid = [candidate for candidate in candidates if not candidate[2]]
    if not valid:
        if candidates:
            failure = input_failure(
                "EXCEL_INPUT_DUPLICATE_COLUMNS",
                "初始材料表中同一业务字段对应了多个标题列。",
                "请删除或改名重复标题，使每个业务字段只对应一列。",
                issues=tuple(
                    ExcelInputIssue.create(
                        row=candidates[0][0],
                        field=conflict.split(" columns=", 1)[0],
                        reason="duplicate_column",
                    )
                    for conflict in candidates[0][2]
                ),
                meta={"conflicts": list(candidates[0][2])},
            )
            raise InputContractError(
                failure,
                diagnostic=(
                    "conflicting initial-table header aliases: "
                    f"{list(candidates[0][2])}"
                ),
            )
        failure = input_failure(
            "EXCEL_INPUT_REQUIRED_COLUMNS_MISSING",
            "初始材料表缺少必需列。",
            "请确认表中包含零件号、截面型材、长度、材质和数量列。",
        )
        raise InputContractError(
            failure,
            diagnostic="initial-table header is missing required fields",
        )
    best_score = max(len(columns) for _, columns, _ in valid)
    winners = [
        (row_number, columns)
        for row_number, columns, _ in valid
        if len(columns) == best_score
    ]
    if len(winners) != 1:
        rows = [row for row, _ in winners]
        failure = input_failure(
            "EXCEL_INPUT_HEADER_AMBIGUOUS",
            "初始材料表中检测到多个同等有效的标题行。",
            "请只保留一行正式列标题，并删除重复标题行。",
            issues=tuple(
                ExcelInputIssue.create(row=row, reason="ambiguous_header")
                for row in rows
            ),
            meta={"candidate_rows": rows},
        )
        raise InputContractError(
            failure,
            diagnostic=f"ambiguous initial-table header at rows {rows}",
        )
    header_row, columns = winners[0]
    metadata_rows = [
        row_number
        for row_number in range(1, header_row)
        if "构件数量" in _join_row(worksheet, row_number)
    ]
    if len(metadata_rows) != 1:
        failure = input_failure(
            "EXCEL_INPUT_SCHEMA_AMBIGUOUS",
            "初始材料表的构件信息行缺失或不唯一。",
            "请只保留一行包含“构件数量”的构件信息。",
            issues=tuple(
                ExcelInputIssue.create(row=row, reason="component_metadata_candidate")
                for row in metadata_rows
            ),
            meta={"candidate_rows": metadata_rows},
        )
        raise InputContractError(
            failure,
            diagnostic=(
                f"initial-table component metadata is not unique: rows={metadata_rows}"
            ),
        )
    return InitialLayout(
        metadata_row=metadata_rows[0],
        header_row=header_row,
        columns=MappingProxyType(columns),
    )


def parse_component_info(text: str) -> ComponentInfo:
    """Parse component metadata from Row 1 text.

    Example input:
        "B7-4FD-ZL-19材  料  表构件数量：1构件总重：1739.26"

    Returns ComponentInfo with extracted fields.
    """
    text = text.strip()

    # 构件号: everything before the first "材"
    component_no = ""
    m = re.match(r"^(.*?)材", text)
    if m:
        component_no = m.group(1).strip()

    # 构件数量
    component_qty = 0
    m = re.search(r"构件数量[：:]\s*(\d+)", text)
    if m:
        component_qty = int(m.group(1))

    # 构件总重
    total_weight = 0.0
    m = re.search(r"构件总重[：:]\s*([\d.]+)", text)
    if m:
        total_weight = float(m.group(1))

    return ComponentInfo(
        component_no=component_no,
        component_qty=component_qty,
        total_weight=total_weight,
        raw_text=text,
    )


# ── Internal helpers ─────────────────────────────────────────────


def _cell_str(ws, row: int, col: int) -> str:
    """Get a cell value as a cleaned string."""
    return safe_str(ws.cell(row=row, column=col).value)


def _layout_float(
    worksheet: Any,
    row: int,
    layout: InitialLayout,
    field: str,
) -> float | None:
    column = layout.columns.get(field)
    if column is None:
        return None
    value = worksheet.cell(row=row, column=column).value
    if value is None or (isinstance(value, str) and value.strip() in {"", "-"}):
        return None
    result = safe_float(value)
    if result is not None and math.isfinite(result):
        return result
    display_field = _INITIAL_FIELD_LABELS[field]
    failure = input_failure(
        "EXCEL_INPUT_ROW_VALUE_INVALID",
        "表格中存在无法读取的数值。",
        f"请检查 {worksheet.title} 第 {row} 行“{display_field}”，填写有效数字。",
        issues=(
            ExcelInputIssue.create(
                sheet=worksheet.title,
                row=row,
                field=display_field,
                value=value,
                reason="not_numeric" if result is None else "not_finite",
            ),
        ),
    )
    raise InputContractError(
        failure,
        diagnostic=(
            f"initial-table row {row} field {display_field} is not a finite number: "
            f"{value!r}"
        ),
    )


def _join_row(ws, row: int) -> str:
    """Join all cells in a row into a single string."""
    parts = []
    for c in range(1, ws.max_column + 1):
        v = safe_str(ws.cell(row=row, column=c).value)
        if v:
            parts.append(v)
    return " ".join(parts)
