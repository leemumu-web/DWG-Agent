"""Single production-source intake boundary for Excel Final."""

from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from enum import StrEnum
from pathlib import Path
from types import MappingProxyType
from typing import Any, Mapping

import openpyxl

from domain import ComponentRowKind, ComponentSourceRow, SourcePart
from input_contract import (
    InputContractError,
    InputKind,
    detect_canonical_header,
    inspect_production_input,
)
from input_errors import input_failure
from legacy_xls import open_legacy_workbook
from quality import IssueLevel, QualityIssue
from reader import CanonicalWorkbookRead, read_canonical_source, read_canonical_workbook
from reader_init import detect_initial_layout, read_init_canonical, read_init_table


class SourceFormat(StrEnum):
    STANDARD_WORKBOOK = "standard_workbook"
    LEGACY_WORKBOOK = "legacy_workbook"
    INITIAL_WORKBOOK = "initial_workbook"
    LEGACY_INITIAL_WORKBOOK = "legacy_initial_workbook"
    DELIMITED_TEKLA_TEXT = "delimited_tekla_text"
    FIXED_WIDTH_TEKLA_TEXT = "fixed_width_tekla_text"


@dataclass(frozen=True, slots=True)
class SourceIntakeResult:
    source_path: Path
    source_format: SourceFormat
    sheet_name: str
    working_values: tuple[tuple[Any, ...], ...]
    parts: tuple[SourcePart, ...]
    component_rows: tuple[ComponentSourceRow, ...]
    issues: tuple[QualityIssue, ...]
    diagnostics: Mapping[str, object]
    warnings: tuple[str, ...] = ()
    ignored_sheets: tuple[str, ...] = ()


def _workbook_values(worksheet: Any) -> tuple[tuple[Any, ...], ...]:
    return tuple(tuple(value for value in row) for row in worksheet.iter_rows(values_only=True))


def _selection_issues(
    sheet_name: str,
    warnings: tuple[str, ...],
    ignored_sheets: tuple[str, ...],
) -> tuple[QualityIssue, ...]:
    return tuple(
        QualityIssue(
            level=IssueLevel.WARNING,
            category="多工作表输入",
            source_sheet=sheet_name,
            source_row=1,
            component_no=None,
            part_no=None,
            spec=None,
            field="工作表",
            actual_value=ignored_sheets,
            expected_value=sheet_name,
            absolute_error=None,
            relative_error=None,
            affects_part=False,
            density_source=None,
            description=warning,
        )
        for warning in warnings
    )


def _from_canonical(
    read: CanonicalWorkbookRead,
    source_format: SourceFormat,
    *,
    warnings: tuple[str, ...] = (),
    ignored_sheets: tuple[str, ...] = (),
) -> SourceIntakeResult:
    selection_issues = _selection_issues(read.sheet_name, warnings, ignored_sheets)
    return SourceIntakeResult(
        source_path=read.source_path,
        source_format=source_format,
        sheet_name=read.sheet_name,
        working_values=read.working_values,
        parts=read.parts,
        component_rows=read.component_rows,
        issues=(*selection_issues, *read.issues),
        diagnostics=MappingProxyType({
            "header_row": read.header.row_number,
            "repeated_header_rows": read.header.repeated_rows,
            "part_count": len(read.parts),
            "component_count": len(read.component_rows),
            "warnings": warnings,
            "ignored_sheets": ignored_sheets,
        }),
        warnings=warnings,
        ignored_sheets=ignored_sheets,
    )


def _initial_component_row(path: Path, sheet_name: str) -> ComponentSourceRow:
    component, _ = read_init_table(path)
    component_no = component.component_no.replace(" ", "").replace("　", "")
    return ComponentSourceRow(
        source_sheet=sheet_name,
        source_row=component.source_row,
        kind=ComponentRowKind.SUMMARY,
        batch=None,
        component_no=component_no,
        component_qty=Decimal(str(component.component_qty)),
        original_spec=None,
        material=None,
        source_unit_net=None,
        source_total_net=None,
        source_unit_gross=None,
        source_total_gross=None,
        source_unit_area=None,
        source_total_area=None,
        component_length=None,
        component_width=None,
        component_height=None,
    )


def _read_initial(path: Path, sheet_name: str, header_row: int) -> SourceIntakeResult:
    inspected = inspect_production_input(path)
    workbook = (
        open_legacy_workbook(path)
        if inspected.kind is InputKind.LEGACY_WORKBOOK
        else openpyxl.load_workbook(path, read_only=True, data_only=False)
    )
    try:
        values = _workbook_values(workbook[sheet_name])
    finally:
        workbook.close()
    parts = read_init_canonical(path)
    component_rows = (_initial_component_row(path, sheet_name),)
    return SourceIntakeResult(
        source_path=path,
        source_format=(
            SourceFormat.LEGACY_INITIAL_WORKBOOK
            if inspected.kind is InputKind.LEGACY_WORKBOOK
            else SourceFormat.INITIAL_WORKBOOK
        ),
        sheet_name=sheet_name,
        working_values=values,
        parts=parts,
        component_rows=component_rows,
        issues=_selection_issues(sheet_name, inspected.warnings, inspected.ignored_sheets),
        diagnostics=MappingProxyType({
            "header_row": header_row,
            "part_count": len(parts),
            "component_count": len(component_rows),
            "warnings": inspected.warnings,
            "ignored_sheets": inspected.ignored_sheets,
        }),
        warnings=inspected.warnings,
        ignored_sheets=inspected.ignored_sheets,
    )


def _read_workbook(path: Path, sheet_name: str) -> SourceIntakeResult:
    inspected = inspect_production_input(path)
    workbook = (
        open_legacy_workbook(path)
        if inspected.kind is InputKind.LEGACY_WORKBOOK
        else openpyxl.load_workbook(path, read_only=True, data_only=False)
    )
    try:
        worksheet = workbook[sheet_name]
        try:
            initial_layout = detect_initial_layout(worksheet)
        except InputContractError:
            initial_layout = None
        try:
            detect_canonical_header(worksheet)
        except InputContractError as canonical_error:
            if initial_layout is None:
                raise canonical_error
        else:
            if initial_layout is not None:
                failure = input_failure(
                    "EXCEL_INPUT_SCHEMA_AMBIGUOUS",
                    "工作表同时匹配两种输入表结构。",
                    "请只保留一种表头结构，不要在同一工作表混合 Tekla 清单和初始材料表。",
                    sheets=(sheet_name,),
                )
                raise InputContractError(
                    failure,
                    diagnostic="workbook matches both canonical and initial-table schemas",
                )
            return _from_canonical(
                read_canonical_workbook(path),
                (
                    SourceFormat.LEGACY_WORKBOOK
                    if inspected.kind is InputKind.LEGACY_WORKBOOK
                    else SourceFormat.STANDARD_WORKBOOK
                ),
                warnings=inspected.warnings,
                ignored_sheets=inspected.ignored_sheets,
            )
    finally:
        workbook.close()
    if initial_layout is None:
        failure = input_failure(
            "EXCEL_INPUT_SCHEMA_AMBIGUOUS",
            "无法确定工作表属于哪一种受支持的输入结构。",
            "请上传 Tekla 构件零件明细清单，或标准初始材料表。",
            sheets=(sheet_name,),
        )
        raise InputContractError(
            failure,
            diagnostic="workbook format could not be detected",
        )
    return _read_initial(path, sheet_name, initial_layout.header_row)


def _text_format(path: Path) -> SourceFormat:
    for encoding in ("utf-8-sig", "gb18030", "gbk", "gb2312"):
        try:
            text = path.read_text(encoding=encoding)
        except UnicodeError:
            continue
        return (
            SourceFormat.DELIMITED_TEKLA_TEXT
            if "\t" in text
            else SourceFormat.FIXED_WIDTH_TEKLA_TEXT
        )
    failure = input_failure(
        "EXCEL_INPUT_TEXT_ENCODING_UNSUPPORTED",
        "无法识别 Tekla 文本文件的字符编码。",
        "请使用 UTF-8 或 GB18030 编码重新导出 Tekla 文本文件。",
    )
    raise InputContractError(
        failure,
        diagnostic="Tekla text encoding could not be detected",
    )


def read_production_source(path: str | Path) -> SourceIntakeResult:
    """Detect and adapt one supported production source into canonical records."""
    inspected = inspect_production_input(Path(path))
    if inspected.kind in {InputKind.WORKBOOK, InputKind.LEGACY_WORKBOOK}:
        if inspected.sheet_name is None:
            failure = input_failure(
                "EXCEL_INPUT_NO_WORKSHEET",
                "Excel 文件中没有可读取的工作表。",
                "请添加一张 Tekla 原始零件明细工作表后重新上传。",
            )
            raise InputContractError(
                failure,
                diagnostic="validated workbook has no worksheet",
            )
        return _read_workbook(inspected.path, inspected.sheet_name)

    source_format = _text_format(inspected.path)
    return _from_canonical(
        read_canonical_source(inspected.path),
        source_format,
    )
