"""Baseline verification and workbook orchestration for Excel Final Stage 2."""

from __future__ import annotations

from dataclasses import dataclass, replace
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
import hashlib
import json
import os
from pathlib import Path
import tempfile
from types import MappingProxyType
from typing import Mapping

from openpyxl import load_workbook

from bh_stage2 import BhMeasurementContract, enhance_bh_projection
from box_stage2 import BoxMeasurementContract, enhance_box_projection
from canonical_pipeline import (
    HandbookReader,
    build_canonical_projection,
    write_canonical_projection,
)
from domain import PipelineOutcome
from source_intake import read_production_source
from writer_parts import (
    CANONICAL_SHEET_NAMES,
    CLEAN_HEADERS,
    COMPONENT_HEADERS,
    ORGANIZED_HEADERS,
    PART_HEADERS,
    REPORT_HEADERS,
    FormulaLengthBasis,
)


BASELINE_DRIFT_CODE = "EXCEL_STAGE2_BASELINE_DRIFT"
_GENERATED_SHEETS = tuple(CANONICAL_SHEET_NAMES[1:])
_REMOVED_HEADERS = {
    "构件表": frozenset({"来源sheet", "行类型", "小计来源行"}),
    "整理表": frozenset({"比重来源", "净材利用率", "重量核验"}),
}
_VISIBLE_HEADERS = {
    "清洗表": tuple(CLEAN_HEADERS),
    "构件表": tuple(
        header for header in COMPONENT_HEADERS
        if header not in _REMOVED_HEADERS["构件表"]
    ),
    "整理表": tuple(
        header for header in ORGANIZED_HEADERS
        if header not in _REMOVED_HEADERS["整理表"]
    ),
    "part": tuple((*PART_HEADERS[:9], "备注", "文件", "类型")),
    "处理报告": tuple(REPORT_HEADERS),
}
_BASELINE_BUSINESS_SHEETS = (
    "原表",
    "清洗表",
    "构件表",
    "整理表",
    "part",
)


class Stage2BaselineError(ValueError):
    code = BASELINE_DRIFT_CODE

    def __init__(
        self,
        message: str,
        *,
        changed_sheets: tuple[str, ...] = (),
    ) -> None:
        self.changed_sheets = changed_sheets
        super().__init__(message)


@dataclass(frozen=True, slots=True)
class CanonicalBaselineSignature:
    sheet_names: tuple[str, ...]
    sheet_hashes: Mapping[str, str]
    formula_cell_counts: Mapping[str, int]

    def __post_init__(self) -> None:
        object.__setattr__(
            self,
            "sheet_hashes",
            MappingProxyType(dict(self.sheet_hashes)),
        )
        object.__setattr__(
            self,
            "formula_cell_counts",
            MappingProxyType(dict(self.formula_cell_counts)),
        )


@dataclass(frozen=True, slots=True)
class Stage2WorkbookOutcome:
    output_path: Path
    internal_output_path: Path | None
    status: str
    matched_occurrence_count: int
    missing_drawing_count: int
    unmatched_drawing_count: int
    manual_occurrence_count: int
    pipeline_outcome: PipelineOutcome

    def __post_init__(self) -> None:
        object.__setattr__(self, "output_path", self.output_path.resolve())
        if self.internal_output_path is not None:
            object.__setattr__(
                self,
                "internal_output_path",
                self.internal_output_path.resolve(),
            )


def _normalized_number(value: object) -> str | None:
    if isinstance(value, bool) or not isinstance(value, (int, float, Decimal)):
        return None
    try:
        number = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise Stage2BaselineError(f"工作簿包含非法数值: {value!r}") from exc
    if not number.is_finite():
        raise Stage2BaselineError(f"工作簿包含非有限数值: {value!r}")
    if number == 0:
        return "0"
    rendered = format(number.normalize(), "f")
    return rendered.rstrip("0").rstrip(".") if "." in rendered else rendered


def _value_token(value: object) -> tuple[str, object]:
    if value is None:
        return "blank", ""
    if isinstance(value, bool):
        return "bool", value
    number = _normalized_number(value)
    if number is not None:
        return "number", number
    if isinstance(value, (datetime, date, time)):
        return "datetime", value.isoformat()
    return "text", str(value)


def _validate_generated_sheet_layout(workbook) -> None:
    if tuple(workbook.sheetnames) != tuple(CANONICAL_SHEET_NAMES):
        raise Stage2BaselineError(
            "Stage 1 必须恰好包含六张规范工作表，且顺序不得改变"
        )
    for sheet in workbook.worksheets:
        if sheet.sheet_state != "visible":
            raise Stage2BaselineError(f"Stage 1 工作表不可隐藏: {sheet.title}")
    for sheet_name in _GENERATED_SHEETS:
        sheet = workbook[sheet_name]
        expected_headers = _VISIBLE_HEADERS[sheet_name]
        actual_headers = tuple(
            sheet.cell(row=1, column=column).value
            for column in range(1, len(expected_headers) + 1)
        )
        if actual_headers != expected_headers or sheet.max_column != len(expected_headers):
            raise Stage2BaselineError(
                f"Stage 1 {sheet_name} 表头与当前规范不一致"
            )
        hidden_rows = [
            index for index, dimension in sheet.row_dimensions.items()
            if dimension.hidden
        ]
        hidden_columns = [
            name for name, dimension in sheet.column_dimensions.items()
            if dimension.hidden
        ]
        if hidden_rows or hidden_columns:
            raise Stage2BaselineError(
                f"Stage 1 {sheet_name} 不得保留隐藏行列"
            )


def _sheet_digest(formula_sheet, value_sheet) -> tuple[str, int]:
    digest = hashlib.sha256()
    formula_count = 0
    digest.update(
        json.dumps(
            {
                "title": formula_sheet.title,
                "max_row": formula_sheet.max_row,
                "max_column": formula_sheet.max_column,
                "merged_ranges": sorted(
                    str(value) for value in formula_sheet.merged_cells.ranges
                ),
            },
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
    )
    for row in formula_sheet.iter_rows():
        for cell in row:
            formula_value = cell.value
            cached_value = value_sheet[cell.coordinate].value
            is_formula = (
                isinstance(formula_value, str)
                and formula_value.startswith("=")
            )
            if is_formula:
                formula_count += 1
                if formula_sheet.title in _GENERATED_SHEETS and cached_value is None:
                    raise Stage2BaselineError(
                        f"Stage 1 公式缓存缺失: {formula_sheet.title}!{cell.coordinate}"
                    )
            token = (
                cell.coordinate,
                _value_token(formula_value),
                _value_token(cached_value),
            )
            digest.update(
                json.dumps(
                    token,
                    ensure_ascii=False,
                    separators=(",", ":"),
                ).encode("utf-8")
            )
    return digest.hexdigest(), formula_count


def read_canonical_baseline_signature(
    workbook_path: str | Path,
) -> CanonicalBaselineSignature:
    """Validate and hash one formal Stage 1 workbook's business contract."""
    path = Path(workbook_path).resolve()
    if not path.is_file():
        raise Stage2BaselineError(f"Stage 1 Excel 不存在: {path}")
    formulas = None
    values = None
    try:
        formulas = load_workbook(path, data_only=False, read_only=False)
        values = load_workbook(path, data_only=True, read_only=False)
    except Exception as exc:
        if formulas is not None:
            formulas.close()
        raise Stage2BaselineError("Stage 1 Excel 无法读取") from exc
    try:
        _validate_generated_sheet_layout(formulas)
        if tuple(values.sheetnames) != tuple(CANONICAL_SHEET_NAMES):
            raise Stage2BaselineError("Stage 1 公式缓存工作簿结构不一致")
        sheet_hashes: dict[str, str] = {}
        formula_counts: dict[str, int] = {}
        for sheet_name in CANONICAL_SHEET_NAMES:
            digest, count = _sheet_digest(
                formulas[sheet_name],
                values[sheet_name],
            )
            sheet_hashes[sheet_name] = digest
            formula_counts[sheet_name] = count
        return CanonicalBaselineSignature(
            sheet_names=tuple(CANONICAL_SHEET_NAMES),
            sheet_hashes=sheet_hashes,
            formula_cell_counts=formula_counts,
        )
    finally:
        formulas.close()
        values.close()


def verify_canonical_baseline(
    formal_stage1_path: str | Path,
    rebuilt_stage1_path: str | Path,
) -> CanonicalBaselineSignature:
    """Fail closed unless the rebuilt and formal Stage 1 business data agree."""
    formal = read_canonical_baseline_signature(formal_stage1_path)
    rebuilt = read_canonical_baseline_signature(rebuilt_stage1_path)
    changed_sheets = tuple(
        sheet_name
        for sheet_name in _BASELINE_BUSINESS_SHEETS
        if (
            formal.sheet_hashes[sheet_name] != rebuilt.sheet_hashes[sheet_name]
            or formal.formula_cell_counts[sheet_name]
            != rebuilt.formula_cell_counts[sheet_name]
        )
    )
    if changed_sheets:
        raise Stage2BaselineError(
            "Stage 1 正式结果与当前代码重建基线不一致: "
            + "、".join(changed_sheets),
            changed_sheets=changed_sheets,
        )
    return formal


def _temporary_xlsx(parent: Path, prefix: str) -> Path:
    parent.mkdir(parents=True, exist_ok=True)
    descriptor, name = tempfile.mkstemp(
        prefix=prefix,
        suffix=".xlsx",
        dir=parent,
    )
    os.close(descriptor)
    path = Path(name)
    path.unlink()
    return path


def _materialize_formal_source(stage1_path: Path, target: Path) -> None:
    workbook = load_workbook(stage1_path, data_only=False, read_only=False)
    try:
        for generated_sheet in tuple(workbook.worksheets[1:]):
            workbook.remove(generated_sheet)
        workbook.save(target)
    finally:
        workbook.close()


def _formal_source_sheet_by_row(stage1_path: Path) -> tuple[dict[int, str], str]:
    workbook = load_workbook(stage1_path, data_only=True, read_only=True)
    try:
        sheet = workbook["清洗表"]
        headers = [cell.value for cell in sheet[1]]
        try:
            source_sheet_column = headers.index("来源sheet") + 1
            source_row_column = headers.index("来源行") + 1
        except ValueError as exc:
            raise Stage2BaselineError("Stage 1 清洗表缺少来源身份列") from exc
        source_sheet_by_row: dict[int, str] = {}
        source_names: list[str] = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            source_name = (
                values[source_sheet_column - 1]
                if len(values) >= source_sheet_column
                else None
            )
            source_row = (
                values[source_row_column - 1]
                if len(values) >= source_row_column
                else None
            )
            if source_name in (None, "") or source_row in (None, ""):
                continue
            try:
                normalized_row = int(source_row)
            except (TypeError, ValueError) as exc:
                raise Stage2BaselineError(
                    f"Stage 1 清洗表来源行无效: {source_row!r}"
                ) from exc
            normalized_name = str(source_name)
            previous = source_sheet_by_row.setdefault(
                normalized_row,
                normalized_name,
            )
            if previous != normalized_name:
                raise Stage2BaselineError(
                    f"Stage 1 清洗表来源行 {normalized_row} 对应多个sheet"
                )
            if normalized_name not in source_names:
                source_names.append(normalized_name)
        if len(source_names) > 1:
            raise Stage2BaselineError("Stage 1 清洗表包含多个源工作表身份")
        return source_sheet_by_row, source_names[0] if source_names else "原表"
    finally:
        workbook.close()


def _rebuild_projection(
    stage1_path: Path,
    raw_source_path: Path,
    *,
    handbook: HandbookReader,
):
    _materialize_formal_source(stage1_path, raw_source_path)
    intake = read_production_source(raw_source_path)
    source_sheet_by_row, default_source_sheet = _formal_source_sheet_by_row(
        stage1_path
    )

    def source_sheet(source_row: int) -> str:
        return source_sheet_by_row.get(source_row, default_source_sheet)

    parts = tuple(
        replace(part, source_sheet=source_sheet(part.source_row))
        for part in intake.parts
    )
    components = tuple(
        replace(component, source_sheet=source_sheet(component.source_row))
        for component in intake.component_rows
    )
    issues = tuple(
        replace(issue, source_sheet=source_sheet(issue.source_row))
        for issue in intake.issues
    )
    return build_canonical_projection(
        parts=parts,
        component_rows=components,
        reader_issues=issues,
        handbook=handbook,
    )


def _published_outcome(
    candidate_outcome: PipelineOutcome,
    output_path: Path,
) -> PipelineOutcome:
    return replace(candidate_outcome, output_path=output_path)


def run_stage2_workbook(
    formal_stage1_path: str | Path,
    output_path: str | Path,
    *,
    measurements: BhMeasurementContract,
    box_measurements: BoxMeasurementContract | None = None,
    handbook: HandbookReader,
    internal_output_path: str | Path | None = None,
) -> Stage2WorkbookOutcome:
    """Validate Stage 1, enhance BH/BOX rows, and publish a six-sheet Stage 2."""
    stage1 = Path(formal_stage1_path).resolve()
    output = Path(output_path).resolve()
    internal_output = (
        Path(internal_output_path).resolve()
        if internal_output_path is not None
        else None
    )
    if output.suffix.lower() != ".xlsx":
        raise ValueError("Excel Stage 2 output must use the .xlsx extension")
    if internal_output is not None and internal_output.suffix.lower() != ".xlsx":
        raise ValueError("Excel Stage 2 internal output must use the .xlsx extension")
    if output == stage1 or internal_output in {stage1, output}:
        raise ValueError("Excel Stage 2 input and output paths must be distinct")

    read_canonical_baseline_signature(stage1)
    raw_source = _temporary_xlsx(output.parent, f".{output.stem}.source.")
    baseline_candidate = _temporary_xlsx(
        output.parent,
        f".{output.stem}.baseline.",
    )
    baseline_internal = (
        _temporary_xlsx(
            internal_output.parent,
            f".{internal_output.stem}.baseline.",
        )
        if internal_output is not None
        else None
    )
    stage2_candidate: Path | None = None
    stage2_internal: Path | None = None
    try:
        projection = _rebuild_projection(
            stage1,
            raw_source,
            handbook=handbook,
        )
        baseline_outcome = write_canonical_projection(
            stage1,
            baseline_candidate,
            projection=projection,
            internal_output_path=baseline_internal,
            formula_length_basis=FormulaLengthBasis.MODEL_LENGTH,
        )
        verify_canonical_baseline(stage1, baseline_candidate)
        enhanced = enhance_bh_projection(projection, measurements)
        if box_measurements is not None:
            # BH 与 BOX 是不同构件（零件号不同），链式增强互不干扰；
            # 状态取"更不完整"的一侧（noop < complete < partial）。
            box_enhanced = enhance_box_projection(
                enhanced.projection,
                box_measurements,
            )
            if box_enhanced.status != "noop":
                combined = enhanced.status
                if combined == "noop" or box_enhanced.status == "partial":
                    combined = box_enhanced.status
                enhanced = replace(
                    enhanced,
                    projection=box_enhanced.projection,
                    status=combined,
                    matched_occurrence_count=(
                        enhanced.matched_occurrence_count
                        + box_enhanced.matched_occurrence_count
                    ),
                    missing_drawing_count=(
                        enhanced.missing_drawing_count
                        + box_enhanced.missing_drawing_count
                    ),
                    unmatched_drawing_count=(
                        enhanced.unmatched_drawing_count
                        + box_enhanced.unmatched_drawing_count
                    ),
                    manual_occurrence_count=(
                        enhanced.manual_occurrence_count
                        + box_enhanced.manual_occurrence_count
                    ),
                )

        if enhanced.status == "noop":
            formal_candidate = baseline_candidate
            internal_candidate = baseline_internal
            candidate_outcome = baseline_outcome
        else:
            stage2_candidate = _temporary_xlsx(
                output.parent,
                f".{output.stem}.stage2.",
            )
            stage2_internal = (
                _temporary_xlsx(
                    internal_output.parent,
                    f".{internal_output.stem}.stage2.",
                )
                if internal_output is not None
                else None
            )
            candidate_outcome = write_canonical_projection(
                stage1,
                stage2_candidate,
                projection=enhanced.projection,
                internal_output_path=stage2_internal,
                formula_length_basis=FormulaLengthBasis.CUT_LENGTH,
            )
            formal_candidate = stage2_candidate
            internal_candidate = stage2_internal

        published = _published_outcome(candidate_outcome, output)
        result = Stage2WorkbookOutcome(
            output_path=output,
            internal_output_path=internal_output,
            status=enhanced.status,
            matched_occurrence_count=enhanced.matched_occurrence_count,
            missing_drawing_count=enhanced.missing_drawing_count,
            unmatched_drawing_count=enhanced.unmatched_drawing_count,
            manual_occurrence_count=enhanced.manual_occurrence_count,
            pipeline_outcome=published,
        )
        if internal_output is not None:
            if internal_candidate is None or not internal_candidate.is_file():
                raise RuntimeError("Excel Stage 2 internal candidate was not created")
            internal_output.parent.mkdir(parents=True, exist_ok=True)
            os.replace(internal_candidate, internal_output)
        output.parent.mkdir(parents=True, exist_ok=True)
        os.replace(formal_candidate, output)
        return result
    finally:
        for temporary in (
            raw_source,
            baseline_candidate,
            baseline_internal,
            stage2_candidate,
            stage2_internal,
        ):
            if temporary is not None:
                temporary.unlink(missing_ok=True)
