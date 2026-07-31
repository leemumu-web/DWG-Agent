"""Strict production-input boundary for Excel Final."""

from __future__ import annotations

import re
from dataclasses import dataclass
from enum import StrEnum
from pathlib import Path
from types import MappingProxyType
from typing import Any, Mapping

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from input_errors import (
    ExcelInputIssue,
    InputContractError,
    input_failure,
)
from header_normalization import normalize_header
from legacy_xls import open_legacy_workbook


class InputKind(StrEnum):
    WORKBOOK = "workbook"
    LEGACY_WORKBOOK = "legacy_workbook"
    TEKLA_TEXT = "tekla_text"


@dataclass(frozen=True, slots=True)
class ProductionInput:
    path: Path
    kind: InputKind
    sheet_name: str | None
    warnings: tuple[str, ...] = ()
    ignored_sheets: tuple[str, ...] = ()


def _first_sheet_selection(
    sheetnames: tuple[str, ...],
) -> tuple[str, tuple[str, ...], tuple[str, ...]]:
    first = sheetnames[0]
    ignored = sheetnames[1:]
    visible = ignored[:10]
    suffix = "" if len(visible) == len(ignored) else "、另有更多"
    warning = (
        f"检测到 {len(sheetnames)} 张工作表，仅处理第一张“{first}”；"
        f"其余工作表已忽略：{'、'.join(visible)}{suffix}。"
    )
    return first, (warning,), ignored[:10]


@dataclass(frozen=True, slots=True)
class HeaderDetection:
    row_number: int
    columns: Mapping[str, int]
    candidate_scores: tuple[HeaderCandidateScore, ...]
    repeated_rows: tuple[int, ...] = ()


@dataclass(frozen=True, slots=True)
class HeaderCandidateScore:
    row_number: int
    score: int
    missing: tuple[str, ...]
    conflicts: tuple[str, ...] = ()


_HEADER_ALIASES = {
    "批次": frozenset({"批次"}),
    "构件编号": frozenset({"构件编号", "构件号"}),
    "零件号": frozenset({"零件号", "零件编号"}),
    "规格": frozenset({"规格", "型材", "截面型材"}),
    "材质": frozenset({"材质"}),
    "数量": frozenset({"数量"}),
    "单净重": frozenset({"单净重"}),
    "总净重": frozenset({"总净重"}),
    "单毛重": frozenset({"单毛重", "单重"}),
    "总毛重": frozenset({"总毛重", "总重"}),
    "单表面积": frozenset({"单表面积", "单面积", "单涂装面积"}),
    "总表面积": frozenset({"总表面积", "总面积", "总涂装面积"}),
    "版本": frozenset({"版本"}),
}
_ALIAS_TO_HEADER = {
    alias: canonical
    for canonical, aliases in _HEADER_ALIASES.items()
    for alias in aliases
}

_REQUIRED_HEADERS = (
    "构件编号",
    "零件号",
    "规格",
    "零件长度",
    "材质",
    "数量",
)


def _normalized_header(value: Any) -> str:
    return normalize_header(value)


def _columns_for_header_row(
    values: tuple[Any, ...],
) -> tuple[dict[str, int], tuple[str, ...]]:
    normalized = [_normalized_header(value) for value in values]
    matches: dict[str, list[int]] = {}
    for index, value in enumerate(normalized, start=1):
        canonical = _ALIAS_TO_HEADER.get(value)
        if canonical is not None:
            matches.setdefault(canonical, []).append(index)

    # A standard Tekla export may abbreviate both net/gross pairs as two
    # occurrences of 单重 and 总重.  Only resolve the four-column pattern as a
    # whole; a lone duplicate remains ambiguous instead of inventing meaning.
    unit_indexes = matches.get("单毛重", [])
    total_indexes = matches.get("总毛重", [])
    if (
        len(unit_indexes) == 2
        and len(total_indexes) == 2
        and not matches.get("单净重")
        and not matches.get("总净重")
    ):
        matches["单净重"] = [unit_indexes[0]]
        matches["总净重"] = [total_indexes[0]]
        matches["单毛重"] = [unit_indexes[1]]
        matches["总毛重"] = [total_indexes[1]]

    columns = {
        canonical: indexes[0]
        for canonical, indexes in matches.items()
    }
    conflicts = tuple(
        f"{canonical} columns={indexes}"
        for canonical, indexes in matches.items()
        if len(indexes) > 1
    )

    length_columns = [index for index, value in enumerate(normalized, start=1) if value == "长度"]
    if length_columns:
        columns["零件长度"] = length_columns[0]
        for index in length_columns[1:]:
            if normalized[index:index + 2] == ["宽度", "高度"]:
                columns["构件长度"] = index
                columns["构件宽度"] = index + 1
                columns["构件高度"] = index + 2
                break
    return columns, conflicts


def _candidate_diagnostics(candidates: list[tuple[HeaderCandidateScore, dict[str, int]]]) -> str:
    details = [
        (
            f"row={candidate.row_number} score={candidate.score} "
            f"missing={list(candidate.missing)} conflicts={list(candidate.conflicts)}"
        )
        for candidate, _ in candidates[:15]
    ]
    return "first 15 candidate scores: " + "; ".join(details)


def _duplicate_fields(conflicts: tuple[str, ...]) -> dict[str, list[int]]:
    duplicates: dict[str, list[int]] = {}
    for conflict in conflicts:
        match = re.fullmatch(r"(.+?) columns=\[([0-9, ]+)\]", conflict)
        if match is None:
            continue
        duplicates[match.group(1)] = [
            int(value.strip()) for value in match.group(2).split(",")
        ]
    return duplicates


def _section_has_payload(
    worksheet: Any,
    *,
    header_row: int,
    next_header_row: int,
    columns: Mapping[str, int],
) -> bool:
    identity_columns = tuple(
        columns[field]
        for field in ("构件编号", "零件号")
        if field in columns
    )
    return any(
        any(
            column <= len(values) and values[column - 1] not in (None, "")
            for column in identity_columns
        )
        for values in worksheet.iter_rows(
            min_row=header_row + 1,
            max_row=next_header_row - 1,
            values_only=True,
        )
    )


def is_repeated_canonical_header(
    values: tuple[Any, ...],
    columns: Mapping[str, int],
) -> bool:
    """Return whether a later row repeats the selected canonical column layout."""
    repeated_columns, conflicts = _columns_for_header_row(values)
    return not conflicts and repeated_columns == dict(columns)


def detect_canonical_header(worksheet: Any) -> HeaderDetection:
    """Locate the strongest canonical header without assuming a fixed row."""
    candidates: list[tuple[HeaderCandidateScore, dict[str, int]]] = []
    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=1, max_row=min(worksheet.max_row, 100), values_only=True),
        start=1,
    ):
        columns, conflicts = _columns_for_header_row(values)
        missing = tuple(field for field in _REQUIRED_HEADERS if field not in columns)
        candidates.append((
            HeaderCandidateScore(row_number, len(columns), missing, conflicts),
            columns,
        ))
    if not candidates:
        failure = input_failure(
            "EXCEL_INPUT_HEADER_NOT_FOUND",
            "未检测到可用的列标题。",
            "请确认工作表中包含构件编号、零件号、规格、长度、材质和数量列。",
        )
        raise InputContractError(
            failure,
            diagnostic="worksheet does not contain a detectable header",
        )

    valid = [
        (candidate, columns)
        for candidate, columns in candidates
        if not candidate.missing and not candidate.conflicts
    ]
    diagnostics = _candidate_diagnostics(candidates)
    if not valid:
        best = max(candidates, key=lambda item: item[0].score)[0]
        if best.conflicts:
            duplicate_fields = _duplicate_fields(best.conflicts)
            issues = tuple(
                ExcelInputIssue.create(
                    row=best.row_number,
                    column=get_column_letter(column),
                    field=field,
                    reason="duplicate_column",
                )
                for field, columns in duplicate_fields.items()
                for column in columns
            )
            failure = input_failure(
                "EXCEL_INPUT_DUPLICATE_COLUMNS",
                "同一业务字段对应了多个标题列。",
                "请删除或改名重复标题，使每个业务字段只对应一列。",
                issues=issues,
                meta={"duplicate_fields": duplicate_fields},
            )
            raise InputContractError(
                failure,
                diagnostic=(
                    f"conflicting header aliases: {list(best.conflicts)}; {diagnostics}"
                ),
            )
        if best.missing == ("零件号",):
            failure = input_failure(
                "EXCEL_INPUT_COMPONENT_ONLY",
                "输入只有构件汇总，没有零件明细。",
                "请从 Tekla 导出包含零件号的构件零件明细清单后重新上传。",
                issues=(
                    ExcelInputIssue.create(
                        row=best.row_number,
                        field="零件号",
                        reason="required_column_missing",
                    ),
                ),
                meta={"missing_fields": ["零件号"]},
            )
            raise InputContractError(
                failure,
                diagnostic="输入只有构件汇总，没有零件明细，不能生成 Excel Final part",
            )
        if best.score == 0:
            failure = input_failure(
                "EXCEL_INPUT_HEADER_NOT_FOUND",
                "未检测到可用的列标题。",
                "请确认工作表中包含构件编号、零件号、规格、长度、材质和数量列。",
                meta={"candidate_row": best.row_number},
            )
            raise InputContractError(failure, diagnostic=diagnostics)
        issues = tuple(
            ExcelInputIssue.create(
                row=best.row_number,
                field=field,
                reason="required_column_missing",
            )
            for field in best.missing
        )
        missing_text = "、".join(best.missing)
        failure = input_failure(
            "EXCEL_INPUT_REQUIRED_COLUMNS_MISSING",
            "表格缺少 Excel 第一阶段所需列。",
            f"请在正式标题行中补充：{missing_text}。",
            issues=issues,
            meta={"missing_fields": list(best.missing)},
        )
        raise InputContractError(
            failure,
            diagnostic=f"missing required fields: {list(best.missing)}; {diagnostics}",
        )

    best_score = max(candidate.score for candidate, _ in valid)
    winners = [(candidate, columns) for candidate, columns in valid if candidate.score == best_score]
    if len(winners) != 1:
        rows = [candidate.row_number for candidate, _ in winners]
        layouts = {
            tuple(sorted(columns.items()))
            for _, columns in winners
        }
        repeated_sections = len(layouts) == 1 and all(
            _section_has_payload(
                worksheet,
                header_row=candidate.row_number,
                next_header_row=(
                    winners[index + 1][0].row_number
                    if index + 1 < len(winners)
                    else worksheet.max_row + 1
                ),
                columns=columns,
            )
            for index, (candidate, columns) in enumerate(winners)
        )
        if not repeated_sections:
            failure = input_failure(
                "EXCEL_INPUT_HEADER_AMBIGUOUS",
                "表格中检测到多个同等有效的标题行。",
                "请只保留一行正式列标题，并删除重复标题行。",
                issues=tuple(
                    ExcelInputIssue.create(row=row, reason="ambiguous_header")
                    for row in rows
                ),
                meta={"candidate_rows": rows},
            )
            raise InputContractError(
                failure,
                diagnostic=f"ambiguous canonical header at rows {rows}; {diagnostics}",
            )

    winner, columns = winners[0]
    scores = tuple(candidate for candidate, _ in candidates)
    return HeaderDetection(
        winner.row_number,
        MappingProxyType(columns),
        scores,
        tuple(candidate.row_number for candidate, _ in winners[1:]),
    )


def inspect_production_input(path: Path) -> ProductionInput:
    resolved = path.resolve()
    if not resolved.is_file():
        failure = input_failure(
            "EXCEL_INPUT_UNREADABLE",
            "无法读取上传的 Excel 文件。",
            "请重新选择文件并上传；如果问题持续，请确认文件没有被移动或删除。",
        )
        raise InputContractError(
            failure,
            diagnostic=f"production input does not exist: {resolved}",
        )
    if resolved.stat().st_size == 0:
        failure = input_failure(
            "EXCEL_INPUT_EMPTY",
            "上传的文件为空。",
            "请选择包含 Tekla 零件明细的 Excel 文件后重新上传。",
        )
        raise InputContractError(failure, diagnostic="production input is empty")

    suffix = resolved.suffix.lower()
    if suffix == ".xls":
        with resolved.open("rb") as source:
            signature = source.read(8)
        if signature == bytes.fromhex("D0CF11E0A1B11AE1"):
            try:
                workbook = open_legacy_workbook(resolved)
            except Exception as exc:
                failure = input_failure(
                    "EXCEL_INPUT_UNREADABLE",
                    "Excel 文件损坏、加密或与扩展名不一致。",
                    "请使用 Excel 打开并另存为未加密的 XLSX 文件后重新上传。",
                )
                raise InputContractError(
                    failure,
                    diagnostic=f"legacy binary workbook is unreadable: {exc.__class__.__name__}",
                ) from exc
            try:
                if not workbook.sheetnames:
                    failure = input_failure(
                        "EXCEL_INPUT_NO_WORKSHEET",
                        "Excel 文件中没有可读取的工作表。",
                        "请添加一张 Tekla 原始零件明细工作表后重新上传。",
                    )
                    raise InputContractError(
                        failure,
                        diagnostic="legacy binary workbook has no worksheet",
                    )
                sheetnames = tuple(workbook.sheetnames)
                if len(sheetnames) > 1:
                    first, warnings, ignored = _first_sheet_selection(sheetnames)
                    return ProductionInput(
                        resolved,
                        InputKind.LEGACY_WORKBOOK,
                        first,
                        warnings,
                        ignored,
                    )
                return ProductionInput(
                    resolved,
                    InputKind.LEGACY_WORKBOOK,
                    sheetnames[0],
                )
            finally:
                workbook.close()
        return ProductionInput(resolved, InputKind.TEKLA_TEXT, None)
    if suffix not in {".xlsx", ".xlsm"}:
        failure = input_failure(
            "EXCEL_INPUT_UNSUPPORTED_EXTENSION",
            "文件格式不受支持。",
            "请上传 .xlsx、.xlsm，或 Tekla 文本格式的 .xls 文件。",
            meta={"extension": suffix or None},
        )
        raise InputContractError(
            failure,
            diagnostic=f"unsupported production input extension: {suffix or '<none>'}",
        )

    try:
        workbook = load_workbook(resolved, read_only=True, data_only=False)
    except Exception as exc:
        failure = input_failure(
            "EXCEL_INPUT_UNREADABLE",
            "Excel 文件损坏、加密或与扩展名不一致。",
            "请使用 Excel 打开并另存为未加密的 XLSX 文件后重新上传。",
        )
        raise InputContractError(
            failure,
            diagnostic=f"production workbook is unreadable: {exc.__class__.__name__}",
        ) from exc
    try:
        if not workbook.sheetnames:
            failure = input_failure(
                "EXCEL_INPUT_NO_WORKSHEET",
                "Excel 文件中没有可读取的工作表。",
                "请添加一张 Tekla 原始零件明细工作表后重新上传。",
            )
            raise InputContractError(failure, diagnostic="production workbook has no worksheet")
        sheetnames = tuple(workbook.sheetnames)
        if len(sheetnames) > 1:
            first, warnings, ignored = _first_sheet_selection(sheetnames)
            return ProductionInput(
                resolved,
                InputKind.WORKBOOK,
                first,
                warnings,
                ignored,
            )
        sheet_name = sheetnames[0]
    finally:
        workbook.close()
    return ProductionInput(resolved, InputKind.WORKBOOK, sheet_name)
