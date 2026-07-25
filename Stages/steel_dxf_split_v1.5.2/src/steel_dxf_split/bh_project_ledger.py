from __future__ import annotations

import os
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from .artifact_io import fsync_directory
from .pipeline import SplitResult

BH_PROJECT_LEDGER_FILENAME = "BH拆板信息表.xlsx"
BH_PROJECT_LEDGER_SHEET = "BH拆板信息"
BH_PROJECT_LEDGER_HEADERS = ("零件号", "BH尺寸", "上下翼板是否相同")


@dataclass(frozen=True, slots=True)
class BHProjectLedgerRow:
    part_number: str
    section_spec: str
    upper_lower_flanges_same: bool

    def to_excel_row(self) -> tuple[str, str, str]:
        return (
            self.part_number,
            self.section_spec,
            "是" if self.upper_lower_flanges_same else "否",
        )


def _required_text(value: object, *, field: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise ValueError(f"BH 拆板信息缺少有效的{field}。")
    return value.strip()


def _flanges_are_same(manufacturing_ir: dict[str, object]) -> bool:
    plates = manufacturing_ir.get("plates")
    if not isinstance(plates, list):
        raise ValueError("BH 制造 IR 缺少板件清单。")
    flanges = [
        plate
        for plate in plates
        if isinstance(plate, dict)
        and plate.get("role") in {"upper_flange", "lower_flange"}
    ]
    if {plate.get("role") for plate in flanges} != {
        "upper_flange",
        "lower_flange",
    } or len(flanges) != 2:
        raise ValueError("BH 制造 IR 必须包含唯一的上翼板和下翼板。")
    merge_groups = {plate.get("merge_group_id") for plate in flanges}
    return (
        all(plate.get("merge_authorized") is True for plate in flanges)
        and len(merge_groups) == 1
        and None not in merge_groups
        and "" not in merge_groups
    )


def ledger_row_from_result(result: SplitResult) -> BHProjectLedgerRow | None:
    if result.family != "BH":
        return None
    manufacturing_ir = result.report.get("manufacturing_ir")
    if not isinstance(manufacturing_ir, dict):
        raise ValueError("BH 拆板报告缺少制造 IR，无法生成项目拆板信息表。")
    return BHProjectLedgerRow(
        part_number=_required_text(
            manufacturing_ir.get("part_number"),
            field="零件号",
        ),
        section_spec=_required_text(
            manufacturing_ir.get("profile"),
            field="BH尺寸",
        ),
        upper_lower_flanges_same=_flanges_are_same(manufacturing_ir),
    )


def collect_bh_project_ledger_rows(
    results: Iterable[SplitResult],
) -> tuple[BHProjectLedgerRow, ...]:
    rows = tuple(
        row
        for result in results
        if (row := ledger_row_from_result(result)) is not None
    )
    seen: set[tuple[str, str]] = set()
    for row in rows:
        key = (row.part_number.casefold(), row.section_spec.casefold())
        if key in seen:
            raise ValueError(
                "BH 拆板信息表中“零件号 + BH尺寸”必须唯一："
                f"{row.part_number} / {row.section_spec}"
            )
        seen.add(key)
    return tuple(
        sorted(
            rows,
            key=lambda row: (
                row.part_number.casefold(),
                row.section_spec.casefold(),
            ),
        )
    )


def _build_workbook(rows: tuple[BHProjectLedgerRow, ...]) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = BH_PROJECT_LEDGER_SHEET
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    worksheet.append(BH_PROJECT_LEDGER_HEADERS)
    for row in rows:
        worksheet.append(row.to_excel_row())

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Microsoft YaHei", color="FFFFFF", bold=True)
    body_font = Font(name="Microsoft YaHei", color="1F2937")
    thin_gray = Side(style="thin", color="D9E2F3")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="medium", color="9EADBA"))
    for row in worksheet.iter_rows(min_row=2, max_col=3):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(
                horizontal="center" if cell.column == 3 else "left",
                vertical="center",
            )
            cell.border = Border(bottom=thin_gray)

    worksheet.row_dimensions[1].height = 26
    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 24
    worksheet.column_dimensions["C"].width = 22
    worksheet.auto_filter.ref = f"A1:C{max(1, len(rows) + 1)}"
    if rows:
        table = Table(
            displayName="BHProjectSplitLedger",
            ref=f"A1:C{len(rows) + 1}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    workbook.properties.title = "BH拆板信息表"
    workbook.properties.subject = "供 Excel 第二阶段定位 BH 零件并判断上下翼板是否相同"
    return workbook


def write_bh_project_ledger(
    rows: Iterable[BHProjectLedgerRow],
    output_dir: Path,
) -> Path:
    frozen_rows = tuple(rows)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    destination = output_dir / BH_PROJECT_LEDGER_FILENAME
    pending = output_dir / f".{destination.name}.{uuid4().hex}.pending.xlsx"
    workbook = _build_workbook(frozen_rows)
    try:
        workbook.save(pending)
        workbook.close()
        with pending.open("rb+") as stream:
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(pending, destination)
        fsync_directory(output_dir)
    finally:
        workbook.close()
        pending.unlink(missing_ok=True)
    return destination


def publish_bh_project_ledger(
    results: Iterable[SplitResult],
    output_dir: Path,
) -> Path:
    rows = collect_bh_project_ledger_rows(results)
    return write_bh_project_ledger(rows, output_dir)
