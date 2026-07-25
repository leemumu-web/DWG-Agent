from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from steel_dxf_split.bh_project_ledger import (
    BH_PROJECT_LEDGER_FILENAME,
    BH_PROJECT_LEDGER_HEADERS,
    BHProjectLedgerRow,
    collect_bh_project_ledger_rows,
    ledger_row_from_result,
    publish_bh_project_ledger,
    write_bh_project_ledger,
)


def _result(
    *,
    family: str = "BH",
    part_number: str = "BH-001",
    profile: str = "BH200*200*20*30",
    same: bool = True,
) -> SimpleNamespace:
    group = "merge:BH-001:flanges" if same else None
    return SimpleNamespace(
        family=family,
        report={
            "manufacturing_ir": {
                "part_number": part_number,
                "profile": profile,
                "plates": [
                    {
                        "role": "web",
                        "merge_authorized": False,
                        "merge_group_id": None,
                    },
                    {
                        "role": "upper_flange",
                        "merge_authorized": same,
                        "merge_group_id": group,
                    },
                    {
                        "role": "lower_flange",
                        "merge_authorized": same,
                        "merge_group_id": group,
                    },
                ],
            }
        },
    )


@pytest.mark.parametrize(
    ("same", "expected"),
    [(True, True), (False, False)],
)
def test_ledger_row_uses_authorized_flange_merge_as_yes_no_authority(
    same: bool,
    expected: bool,
) -> None:
    row = ledger_row_from_result(_result(same=same))

    assert row == BHProjectLedgerRow(
        part_number="BH-001",
        section_spec="BH200*200*20*30",
        upper_lower_flanges_same=expected,
    )
    assert row.to_excel_row()[2] == ("是" if expected else "否")


def test_project_ledger_contains_only_bh_and_requires_unique_lookup_key() -> None:
    rows = collect_bh_project_ledger_rows(
        [
            _result(part_number="BH-002", same=False),
            _result(family="BOX", part_number="BOX-001"),
            _result(part_number="BH-001", same=True),
        ]
    )

    assert [row.part_number for row in rows] == ["BH-001", "BH-002"]
    with pytest.raises(ValueError, match="零件号 \\+ BH尺寸.*唯一"):
        collect_bh_project_ledger_rows([_result(), _result()])


def test_project_ledger_writes_exact_three_column_excel_contract(
    tmp_path: Path,
) -> None:
    destination = write_bh_project_ledger(
        [
            BHProjectLedgerRow("BH-001", "BH200*200*20*30", True),
            BHProjectLedgerRow("BH-002", "BH300*250*16*20", False),
        ],
        tmp_path,
    )

    assert destination == tmp_path / BH_PROJECT_LEDGER_FILENAME
    workbook = load_workbook(destination, read_only=False, data_only=False)
    try:
        worksheet = workbook["BH拆板信息"]
        assert tuple(cell.value for cell in worksheet[1]) == BH_PROJECT_LEDGER_HEADERS
        assert list(worksheet.values)[1:] == [
            ("BH-001", "BH200*200*20*30", "是"),
            ("BH-002", "BH300*250*16*20", "否"),
        ]
        assert worksheet.freeze_panes == "A2"
        assert tuple(worksheet.tables) == ("BHProjectSplitLedger",)
    finally:
        workbook.close()


def test_successful_project_without_bh_replaces_stale_ledger_with_empty_table(
    tmp_path: Path,
) -> None:
    write_bh_project_ledger(
        [BHProjectLedgerRow("OLD", "BH1*1*1*1", True)],
        tmp_path,
    )

    destination = publish_bh_project_ledger([_result(family="BOX")], tmp_path)

    workbook = load_workbook(destination, read_only=True, data_only=True)
    try:
        assert list(workbook["BH拆板信息"].values) == [
            BH_PROJECT_LEDGER_HEADERS
        ]
    finally:
        workbook.close()
