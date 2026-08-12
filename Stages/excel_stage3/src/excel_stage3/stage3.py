"""Excel 第三阶段核心逻辑 — 异孔折判断对接与回填。"""

from __future__ import annotations

import logging
import os
import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from yikongzhe.classifier import classify_directory

logger = logging.getLogger(__name__)

# 待人工单元格红色背景
_RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# DXF 文件名中需要剥离的后缀
_DXF_STRIP_SUFFIXES = [
    "_拆板后", "_拆分后", "_拆板", "_拆分",
    "_处理后", "_处理", "_正常拆板", "_正常拆分",
    "_切边后", "_切断后", "_切割后",
]

# DXF 文件名前缀模式：BYSJ@...@ → 提取后面的板号
# 例如 "BYSJ@板零件图@5b1-cb-14_正常拆板.dxf" → "5b1-cb-14_正常拆板"
_DXF_PREFIX_RE = re.compile(r"^[A-Z]+@[^@]+@")

# 零件名中标识 BH/BOX 类型的前缀模式（捕获完整后缀）
_PART_TYPE_RE = re.compile(r"-(BH|BOX)(.+)")

# 分类结果 part_name 中的前缀
_CLASSIFICATION_NAME_PREFIX = "p="

# 匹配中文字符（用于在 DXF part_name 中定位后缀起始位置）
_CHINESE_CHAR_RE = re.compile(r"[\u4e00-\u9fff]")


def _normalize_part_type(suffix: str) -> str:
    """将零件名后缀归一化为"腹"或"翼"。

    "上翼" / "下翼" / "翼" → "翼"
    "腹" / "腹板1"        → "腹"
    """
    if "翼" in suffix:
        return "翼"
    if "腹" in suffix:
        return "腹"
    return suffix


def _extract_base_part_number(part_name: str) -> str | None:
    """从零件名中提取基础板号。

    "3b1-cb-133-BH腹"  → "3b1-cb-133"
    "3b1-cb-4-BH上翼"  → "3b1-cb-4"
    "3b1-s-19"         → None（非 BH/BOX 零件）
    """
    m = _PART_TYPE_RE.search(part_name)
    if not m:
        return None
    return part_name[:m.start()]


def _extract_part_category(part_name: str) -> str | None:
    """从零件名中提取归一化的板件部位类别。

    "3b1-cb-133-BH腹"  → "腹"
    "3b1-cb-4-BH上翼"  → "翼"
    "3b1-cb-4-BH腹板1" → "腹"
    "3b1-s-19"         → None
    """
    m = _PART_TYPE_RE.search(part_name)
    if not m:
        return None
    return _normalize_part_type(m.group(2))


def _strip_dxf_suffix(filename: str) -> str:
    """去除 DXF 文件名中的前缀和后缀，返回干净的 stem。

    "BYSJ@板零件图@5b1-cb-14_正常拆板.dxf" → "5b1-cb-14"
    "b4-1-cb-15_拆板后.dxf"               → "b4-1-cb-15"
    "b4-1-cb-15.dxf"                      → "b4-1-cb-15"
    """
    stem = Path(filename).stem
    # 先剥离 BYSJ@...@ 前缀
    stem = _DXF_PREFIX_RE.sub("", stem)
    for suffix in _DXF_STRIP_SUFFIXES:
        if stem.endswith(suffix):
            stem = stem[:-len(suffix)]
            break
    return stem


def _build_dxf_index(dxf_dir: str) -> dict[str, str]:
    """构建 DXF 文件名索引。

    遍历目录下所有 .dxf/.dwg 文件，建立
    {干净stem → 完整文件名} 的映射。

    Returns:
        {"b4-1-cb-15": "b4-1-cb-15_拆板后.dxf", ...}
    """
    index: dict[str, str] = {}
    base = Path(dxf_dir)
    if not base.is_dir():
        return index

    for f in base.iterdir():
        if f.is_file() and f.suffix.lower() in (".dxf", ".dwg"):
            clean = _strip_dxf_suffix(f.name)
            index[clean] = f.name
    return index


def _extract_raw_part_suffix(part_name: str) -> str:
    """从 DXF 分类结果的 part_name 中提取原始后缀，不做归一化。

    "p=b4-1-cb-15上翼" → "上翼"
    "p=b4-1-cb-15下翼" → "下翼"
    "p=b4-1-cb-15腹"   → "腹"
    "p=b4-1-cb-15腹板1" → "腹板1"
    """
    clean = part_name
    if clean.startswith(_CLASSIFICATION_NAME_PREFIX):
        clean = clean[len(_CLASSIFICATION_NAME_PREFIX):]
    # 找到第一个中文字符的位置，从那里开始即为后缀
    m = _CHINESE_CHAR_RE.search(clean)
    if m:
        return clean[m.start():]
    return clean


def _classify_part_name_to_type(name: str) -> str | None:
    """从分类结果的 part_name 中提取归一化的板件部位类别。

    "p=b4-1-cb-15腹" → "腹"
    "p=b4-1-cb-15翼" → "翼"
    """
    clean = name
    if clean.startswith(_CLASSIFICATION_NAME_PREFIX):
        clean = clean[len(_CLASSIFICATION_NAME_PREFIX):]
    if "翼" in clean:
        return "翼"
    if "腹" in clean:
        return "腹"
    return None


class Stage3Runner:
    """第三阶段处理运行器。

    1. 读取第二阶段 Excel 的 part 表
    2. 提取 BH/BOX 零件，匹配拆板后 DXF
    3. 调用 yikongzhe 分类
    4. 回填"图形"列，标红"待人工"
    5. 输出两个 Excel 文件
    """

    def __init__(
        self,
        stage2_excel_path: str,
        dxf_dir: str,
        output_dir: str,
        encoding: str = "utf-8",
    ):
        self.stage2_excel_path = stage2_excel_path
        self.dxf_dir = dxf_dir
        self.output_dir = output_dir
        self.encoding = encoding
        # 列索引（1-based），在 _read_part_rows 中检测
        self._col_part_name: int | None = None
        self._col_component: int | None = None
        self._col_graphics: int | None = None
        self._col_summary: int | None = None

    def run(self) -> dict[str, object]:
        """执行完整流水线，返回统计信息。"""
        os.makedirs(self.output_dir, exist_ok=True)

        # 1. 读取 part 表
        logger.info("读取 Stage2 Excel: %s", self.stage2_excel_path)
        wb = openpyxl.load_workbook(self.stage2_excel_path)
        if "part" not in wb.sheetnames:
            raise ValueError(f"Excel 文件中未找到 'part' 工作表，现有表: {wb.sheetnames}")
        ws = wb["part"]

        part_rows = self._read_part_rows(ws)
        logger.info("part 表共 %d 行数据", len(part_rows))

        # 2. 构建 DXF 索引
        dxf_index = _build_dxf_index(self.dxf_dir)
        logger.info("DXF 目录中共 %d 个文件", len(dxf_index))

        # 3. 匹配 part → DXF
        part_dxf_map, unmatched = self._match_parts_to_dxfs(part_rows, dxf_index)
        bh_box_count = sum(
            1 for r in part_rows
            if _extract_base_part_number(str(r.get("_part_name", "") or ""))
        )
        logger.info(
            "BH/BOX 零件共 %d 行, 匹配到 DXF: %d 行, 未匹配: %d 行",
            bh_box_count, len(part_dxf_map), len(unmatched),
        )

        # 4. 收集需要分类的唯一 DXF
        unique_dxfs = set(part_dxf_map.values())
        logger.info("待分类的唯一 DXF: %d 个", len(unique_dxfs))

        # 5. 运行分类（获取原始部件名 + 归一化映射）
        classification_map, raw_parts_map = self._run_classification(unique_dxfs)
        logger.info("分类完成: %d 个 DXF", len(classification_map))

        # 6. [NEW] 检测需要拆分的通用行
        split_groups = self._find_split_groups(part_rows, part_dxf_map, raw_parts_map)
        if split_groups:
            logger.info("检测到 %d 行需要拆分", len(split_groups))

            # 读取汇总列缓存值（公式 → 数值）
            summary_values = self._read_cached_summaries(split_groups.keys())

            # 在 worksheet 中拆分行
            new_rows = self._split_rows_in_part_sheet(ws, split_groups, summary_values)
            logger.info("已拆分: 新增 %d 行", len(new_rows))

            # 重新读取 part 行（行号已变）
            part_rows = self._read_part_rows(ws)
            logger.info("拆分后 part 表共 %d 行数据", len(part_rows))

            # 重新匹配 part → DXF
            part_dxf_map, unmatched = self._match_parts_to_dxfs(part_rows, dxf_index)
            logger.info("拆分后重新匹配: %d 行 → DXF, %d 行未匹配",
                        len(part_dxf_map), len(unmatched))
        else:
            logger.info("无需拆分行")

        # 7. 回填"图形"列
        filled, manual = self._fill_graphics_column(ws, part_rows, part_dxf_map, classification_map)

        # 8. 输出文件
        output_paths = self._write_outputs(wb, part_rows, part_dxf_map, classification_map)

        wb.close()
        logger.info("所有输出已保存到: %s", self.output_dir)

        return {
            "classification_excel": output_paths.get("classification", ""),
            "deepened_excel": output_paths.get("deepened", ""),
            "bh_box_count": bh_box_count,
            "matched_count": len(part_dxf_map),
            "unmatched_count": len(unmatched),
            "classified_dxf_count": len(classification_map),
            "filled_count": filled,
            "manual_count": manual,
        }

    # ------------------------------------------------------------------
    # 内部方法
    # ------------------------------------------------------------------

    @staticmethod
    def _find_col(headers: list, keyword: str) -> int | None:
        """在表头中按关键字查找列索引（1-based）。"""
        for i, h in enumerate(headers):
            if h and keyword in str(h):
                return i + 1
        return None

    def _read_part_rows(self, ws) -> list[dict[str, Any]]:
        """读取 part 表数据行，同时检测关键列索引。

        Returns:
            [{"_row": 2, "_part_name": "...", "_component": "...", "_summary": ..., ...}, ...]
        """
        headers = []
        for c in range(1, ws.max_column + 1):
            headers.append(ws.cell(1, c).value)

        # 按关键字检测列
        self._col_part_name = self._find_col(headers, "零件")
        self._col_component = self._find_col(headers, "构件")
        self._col_graphics = self._find_col(headers, "图形")
        self._col_summary = self._find_col(headers, "汇总")

        logger.debug(
            "列检测: 零件名=col%d, 构件编号=col%d, 图形=col%d, 汇总=col%d",
            self._col_part_name or -1,
            self._col_component or -1,
            self._col_graphics or -1,
            self._col_summary or -1,
        )

        rows = []
        for r in range(2, ws.max_row + 1):
            row_data: dict[str, Any] = {"_row": r}
            # 读取关键列
            if self._col_part_name:
                row_data["_part_name"] = ws.cell(r, self._col_part_name).value
            if self._col_component:
                row_data["_component"] = ws.cell(r, self._col_component).value
            if self._col_summary:
                row_data["_summary"] = ws.cell(r, self._col_summary).value
            # 跳过完全空行
            if any(v is not None for k, v in row_data.items() if k != "_row"):
                rows.append(row_data)
        return rows

    def _match_parts_to_dxfs(
        self, part_rows: list[dict], dxf_index: dict[str, str]
    ) -> tuple[dict[int, str], list[int]]:
        """建立 part 行 → DXF 文件名 的映射。

        匹配策略：精确匹配 → 大小写不敏感匹配 → 包含匹配。
        包含匹配会检查零件基础板号是否出现在 DXF 干净 stem 中，
        以兼容 DXF 文件名中未剥离干净的前后缀。

        Returns:
            ({row_num → dxf_filename}, [unmatched_row_nums])
        """
        part_dxf_map: dict[int, str] = {}
        unmatched: list[int] = []

        for row in part_rows:
            part_name = str(row.get("_part_name", "") or "")
            base = _extract_base_part_number(part_name)
            if base is None:
                continue  # 非 BH/BOX 零件

            if base in dxf_index:
                part_dxf_map[row["_row"]] = dxf_index[base]
                continue

            # 尝试不区分大小写精确匹配
            base_lower = base.lower()
            found = None
            for key, val in dxf_index.items():
                if key.lower() == base_lower:
                    found = val
                    break
            if found:
                part_dxf_map[row["_row"]] = found
                continue

            # 尝试包含匹配：零件基础板号是否出现在 DXF stem 中
            for key, val in dxf_index.items():
                if base_lower in key.lower():
                    found = val
                    break
            if found:
                part_dxf_map[row["_row"]] = found
                continue

            unmatched.append(row["_row"])
            logger.debug("未匹配到 DXF: %s (base=%s)", part_name, base)

        return part_dxf_map, unmatched

    def _run_classification(
        self, dxf_filenames: set[str]
    ) -> tuple[dict[str, dict[str, str]], dict[str, list[object]]]:
        """对指定 DXF 文件所在目录运行分类。

        classify_directory 会处理整个目录，我们只提取需要的 DXF 结果。

        以原始后缀（上翼/下翼/腹/腹板1）作为 part_map 的 key，
        保留具体子类型信息，避免归一化导致覆盖。

        Returns:
            (
                {dxf_filename: {"上翼": "方孔折", "下翼": "方", "腹": "方孔", ...}},
                {dxf_filename: [PartClassification, ...]},
            )
        """
        if not dxf_filenames:
            return {}, {}

        all_results = classify_directory(self.dxf_dir, encoding=self.encoding)
        classification_map: dict[str, dict[str, str]] = {}
        raw_parts_map: dict[str, list[object]] = {}

        for result in all_results:
            if result.dxf_file not in dxf_filenames:
                continue
            part_map: dict[str, str] = {}
            raw_parts_map[result.dxf_file] = list(result.parts)
            for pc in result.parts:
                raw_suffix = _extract_raw_part_suffix(pc.part_name)
                if raw_suffix:
                    part_map[raw_suffix] = pc.category
                else:
                    # 无法提取后缀 → fallback 用归一化类型
                    ptype = _classify_part_name_to_type(pc.part_name)
                    if ptype:
                        part_map[ptype] = pc.category
                    else:
                        clean = pc.part_name
                        if clean.startswith(_CLASSIFICATION_NAME_PREFIX):
                            clean = clean[len(_CLASSIFICATION_NAME_PREFIX):]
                        part_map[clean] = pc.category
            classification_map[result.dxf_file] = part_map

        return classification_map, raw_parts_map

    @staticmethod
    def _find_raw_suffix(part_name: str) -> str | None:
        """从 Excel 零件名中提取原始后缀（不做归一化）。

        "3b1-cb-4-BH上翼"  → "上翼"
        "3b1-cb-4-BH翼"    → "翼"
        "3b1-cb-4-BH腹板1"  → "腹板1"
        "3b1-s-19"          → None
        """
        m = _PART_TYPE_RE.search(part_name)
        if not m:
            return None
        return m.group(2)

    def _find_split_groups(
        self,
        part_rows: list[dict],
        part_dxf_map: dict[int, str],
        raw_parts_map: dict[str, list],
    ) -> dict[int, list[str]]:
        """检测需要拆分的 part 行。

        对每个 DXF 分类结果，将部件按归一化类型分组。
        如果某归一化类型有 >1 个具体后缀，且 Excel 中仅有通用行，
        则标记该通用行需要拆分。

        Returns:
            {row_num: ["3b1-cb-4-BH上翼", "3b1-cb-4-BH下翼"], ...}
        """
        split_groups: dict[int, list[str]] = {}

        # 先按基础板号分组 part_rows，便于查找
        by_base: dict[str, list[dict]] = {}
        for row in part_rows:
            pn = str(row.get("_part_name", "") or "")
            base = _extract_base_part_number(pn)
            if base:
                by_base.setdefault(base, []).append(row)

        for row in part_rows:
            row_num = row["_row"]
            if row_num not in part_dxf_map:
                continue
            dxf_filename = part_dxf_map[row_num]
            if dxf_filename not in raw_parts_map:
                continue

            part_name = str(row.get("_part_name", "") or "")
            base = _extract_base_part_number(part_name)
            m = _PART_TYPE_RE.search(part_name)
            if base is None or not m:
                continue

            profile = m.group(1)   # "BH" or "BOX"
            suffix = m.group(2)    # current suffix: "翼", "上翼", etc.
            normalized = _normalize_part_type(suffix)

            # 只拆分通用行（当前后缀与归一化相同，即"翼"/"腹"无修饰）
            if suffix != normalized:
                continue

            # 从 DXF 原始部件中收集属于该归一化类型的具体后缀
            raw_parts = raw_parts_map[dxf_filename]
            specific_suffixes: list[str] = []
            for pc in raw_parts:
                raw_suffix = _extract_raw_part_suffix(pc.part_name)
                n = _normalize_part_type(raw_suffix)
                if n == normalized and raw_suffix != normalized:
                    specific_suffixes.append(raw_suffix)

            # 去重并保持顺序
            seen: set[str] = set()
            unique_suffixes: list[str] = []
            for s in specific_suffixes:
                if s not in seen:
                    seen.add(s)
                    unique_suffixes.append(s)

            if len(unique_suffixes) <= 1:
                continue  # 只有一个或没有具体后缀，无需拆分

            # 检查是否已有其他行覆盖了这些具体后缀
            existing_suffixes: set[str] = set()
            for other in by_base.get(base, []):
                if other["_row"] == row_num:
                    continue
                other_suffix = self._find_raw_suffix(
                    str(other.get("_part_name", "") or "")
                )
                if other_suffix:
                    existing_suffixes.add(other_suffix)

            if existing_suffixes >= set(unique_suffixes):
                continue  # 已有足够的行覆盖各个子类型

            # 构造新零件名
            prefix = part_name[:m.start()]  # e.g. "3b1-cb-4"
            new_names = [
                f"{prefix}-{profile}{suffix}" for suffix in unique_suffixes
            ]
            split_groups[row_num] = new_names

        return split_groups

    def _read_cached_summaries(
        self, row_nums: set[int]
    ) -> dict[int, float]:
        """通过 data_only 模式读取汇总列的缓存数值。

        Stage 2 通过 patch_formula_caches 写入了缓存值，
        读取这些值即可获得公式的计算结果。

        Returns:
            {row_num: cached_value, ...}
        """
        result: dict[int, float] = {}
        if self._col_summary is None:
            return result

        try:
            data_wb = openpyxl.load_workbook(
                self.stage2_excel_path, data_only=True,
            )
        except Exception:
            logger.warning("无法以 data_only 模式打开 Excel，汇总值拆分可能不准确")
            return result

        try:
            if "part" not in data_wb.sheetnames:
                return result
            data_ws = data_wb["part"]
            for row_num in row_nums:
                val = data_ws.cell(row_num, self._col_summary).value
                if val is not None:
                    try:
                        result[row_num] = float(val)
                    except (ValueError, TypeError):
                        pass
        finally:
            data_wb.close()

        return result

    def _split_rows_in_part_sheet(
        self,
        ws,
        split_groups: dict[int, list[str]],
        summary_values: dict[int, float],
    ) -> list[int]:
        """在 part 表 worksheet 中拆分行。

        从高行号向低行号处理，避免插入行导致行号错位。
        拆分后每行的"导入零件号"更新为具体名称，
        "汇总"列除以拆分数 N。

        Returns:
            新增行的行号列表
        """
        if not split_groups:
            return []

        max_col = ws.max_column
        new_row_nums: list[int] = []

        for row_num in sorted(split_groups.keys(), reverse=True):
            new_names = split_groups[row_num]
            N = len(new_names)
            if N <= 1:
                continue

            # 收集原行所有单元格数据
            original_cells: dict[int, dict] = {}
            for c in range(1, max_col + 1):
                src = ws.cell(row_num, c)
                original_cells[c] = {
                    "value": src.value,
                    "font": src.font.copy(),
                    "fill": src.fill.copy(),
                    "border": src.border.copy(),
                    "alignment": src.alignment.copy(),
                    "number_format": src.number_format,
                }

            # 插入 N-1 行（在原始行下方）
            if N > 1:
                ws.insert_rows(row_num + 1, N - 1)

            original_summary = summary_values.get(row_num)

            # 写入各拆分行
            for i in range(N):
                target_row = row_num + i
                if i > 0:
                    # 复制原行数据
                    for c in range(1, max_col + 1):
                        cell = ws.cell(target_row, c)
                        orig = original_cells[c]
                        cell.value = orig["value"]
                        cell.font = orig["font"]
                        cell.fill = orig["fill"]
                        cell.border = orig["border"]
                        cell.alignment = orig["alignment"]
                        cell.number_format = orig["number_format"]
                    new_row_nums.append(target_row)

                # 更新导入零件号
                if self._col_part_name:
                    ws.cell(target_row, self._col_part_name).value = new_names[i]

                # 清空图形列（待后续分类回填）
                if self._col_graphics:
                    ws.cell(target_row, self._col_graphics).value = None

                # 汇总列 ÷ N
                if (
                    self._col_summary
                    and original_summary is not None
                    and original_summary != 0
                ):
                    ws.cell(target_row, self._col_summary).value = original_summary / N

        return new_row_nums

    def _fill_graphics_column(
        self,
        ws,
        part_rows: list[dict],
        part_dxf_map: dict[int, str],
        classification_map: dict[str, dict[str, str]],
    ) -> tuple[int, int]:
        """回填"图形"列并标红"待人工"。

        优先用原始后缀（上翼/下翼/腹板1）精确匹配 DXF 分类结果，
        精确匹配不到再降级为归一化匹配或模糊匹配。

        Returns:
            (filled_count, manual_count)
        """
        if self._col_graphics is None:
            logger.warning("未检测到'图形'列，跳过回填")
            return 0, 0

        filled = 0
        manual = 0

        for row in part_rows:
            row_num = row["_row"]
            if row_num not in part_dxf_map:
                continue

            dxf_filename = part_dxf_map[row_num]
            if dxf_filename not in classification_map:
                continue

            part_maps = classification_map[dxf_filename]
            part_name = str(row.get("_part_name", "") or "")
            raw_suffix = self._find_raw_suffix(part_name)
            normalized = _extract_part_category(part_name)

            # 确定分类结果：优先级
            # 1. 原始后缀精确匹配（如"上翼" → "上翼"）
            # 2. 归一化匹配（如"翼" → "翼"）
            # 3. DXF 只有一个结果 → 直接使用
            # 4. 模糊匹配（key 出现在 part_name 中）
            category: str | None = None
            if raw_suffix and raw_suffix in part_maps:
                category = part_maps[raw_suffix]
            elif normalized and normalized in part_maps:
                category = part_maps[normalized]
            elif part_maps:
                if len(part_maps) == 1:
                    category = next(iter(part_maps.values()))
                else:
                    for key, val in part_maps.items():
                        if key and key in part_name:
                            category = val
                            break

            if category is None:
                logger.debug("无法确定分类: row=%d, part=%s", row_num, part_name)
                continue

            cell = ws.cell(row_num, self._col_graphics)
            cell.value = category
            filled += 1

            if category == "待人工":
                cell.fill = _RED_FILL
                manual += 1

        logger.info("回填完成: %d 行, 其中待人工 %d 行", filled, manual)
        return filled, manual

    def _write_outputs(
        self,
        wb,
        part_rows: list[dict],
        part_dxf_map: dict[int, str],
        classification_map: dict[str, dict[str, str]],
    ) -> dict[str, str]:
        """输出两个 Excel 文件，返回路径映射。"""
        classification_path = self._write_classification_excel(
            part_rows, part_dxf_map, classification_map,
        )

        deepened_path = os.path.join(self.output_dir, "stage2_with_graphics.xlsx")
        wb.save(deepened_path)
        logger.info("深化后 Excel 已保存: %s", deepened_path)

        return {"classification": classification_path, "deepened": deepened_path}

    def _write_classification_excel(
        self,
        part_rows: list[dict],
        part_dxf_map: dict[int, str],
        classification_map: dict[str, dict[str, str]],
    ) -> str:
        """生成分类结果 Excel，返回输出路径。"""
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = "分类结果"

        out_headers = ["零件名称", "构件编号", "基础板号", "部位", "图形类别", "来源DXF"]
        for c, h in enumerate(out_headers, 1):
            ws_out.cell(1, c, h)

        row_out = 2
        for row in part_rows:
            row_num = row["_row"]
            part_name = str(row.get("_part_name", "") or "")
            component = str(row.get("_component", "") or "")
            base = _extract_base_part_number(part_name)
            ptype = _extract_part_category(part_name)
            raw_suffix = self._find_raw_suffix(part_name)

            if row_num not in part_dxf_map:
                continue

            dxf_filename = part_dxf_map[row_num]
            if dxf_filename not in classification_map:
                continue

            part_maps = classification_map[dxf_filename]
            category: str | None = None
            if raw_suffix and raw_suffix in part_maps:
                category = part_maps[raw_suffix]
            elif ptype and ptype in part_maps:
                category = part_maps[ptype]
            elif len(part_maps) == 1:
                category = next(iter(part_maps.values()))
            else:
                for key, val in part_maps.items():
                    if key and key in part_name:
                        category = val
                        break

            if category is None:
                continue

            # 部位列显示原始后缀（更具体）
            display_suffix = raw_suffix or ptype or ""
            ws_out.cell(row_out, 1, part_name)
            ws_out.cell(row_out, 2, component)
            ws_out.cell(row_out, 3, base or "")
            ws_out.cell(row_out, 4, display_suffix)
            ws_out.cell(row_out, 5, category)
            ws_out.cell(row_out, 6, dxf_filename)

            if category == "待人工":
                for c in range(1, len(out_headers) + 1):
                    ws_out.cell(row_out, c).fill = _RED_FILL

            row_out += 1

        output_path = os.path.join(self.output_dir, "分类结果.xlsx")
        wb_out.save(output_path)
        logger.info("分类结果 Excel 已保存: %s", output_path)
        return output_path
