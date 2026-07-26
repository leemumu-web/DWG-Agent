"""Excel output writer — produces a 4-sheet .xlsx workbook.

Sheets:
  1. all_rows       — standardized detail rows
  2. raw_like_original — original grid layout preservation
  3. table_summary  — per-table statistics
  4. warnings       — quality anomalies
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import TableResult, WarningInfo

# Styling constants
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11)
DATA_FONT = Font(size=10)
WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

ALL_ROWS_HEADERS = [
    "source_file", "drawing_type", "row_subtype",
    "component_no", "part_no", "spec", "length_mm",
    "material", "quantity", "unit_weight_kg", "total_weight_kg",
    "area_m2", "remark", "confidence", "row_index",
]

SUMMARY_HEADERS = [
    "source_file", "table_index", "source_block",
    "bbox_x1", "bbox_y1", "bbox_x2", "bbox_y2",
    "num_rows", "num_cols", "data_cols",
    "text_count", "line_count",
    "candidate_score", "grid_regularity", "fill_rate", "grid_score",
]

WARN_HEADERS = [
    "source_file", "table_index", "row_index",
    "warning_code", "message", "raw_value",
]


def write_excel(
    output_path: Path,
    table_results: list[TableResult],
    warnings: list[WarningInfo],
) -> None:
    """Write the complete 4-sheet Excel workbook.

    Args:
        output_path: Destination .xlsx path.
        table_results: All extracted table results.
        warnings: All quality warnings.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    _write_all_rows(wb, table_results)
    _write_raw_like_original(wb, table_results)
    _write_table_summary(wb, table_results)
    _write_warnings(wb, warnings)

    wb.save(str(output_path))


# ---- Sheet 1: all_rows ----

def _write_all_rows(wb: openpyxl.Workbook, results: list[TableResult]) -> None:
    ws = wb.create_sheet("all_rows")

    # Flatten all data rows
    flat: list[dict] = []
    for table in results:
        for row in table.data_rows:
            flat.append({
                "source_file": row.source_file,
                "drawing_type": row.drawing_type.value,
                "row_subtype": row.row_subtype,
                "component_no": row.component_no or "",
                "part_no": row.part_no or "",
                "spec": row.spec or "",
                "length_mm": row.length_mm,
                "material": row.material or "",
                "quantity": row.quantity,
                "unit_weight_kg": row.unit_weight_kg,
                "total_weight_kg": row.total_weight_kg,
                "area_m2": row.area_m2,
                "remark": row.remark or "",
                "confidence": row.confidence,
                "row_index": row.row_index,
            })

    if not flat:
        _write_header_row(ws, ALL_ROWS_HEADERS)
        return

    df = pd.DataFrame(flat)
    # Reorder columns
    df = df[ALL_ROWS_HEADERS]

    # Write via pandas for speed, then style
    _write_df_to_sheet(ws, df, ALL_ROWS_HEADERS)

    # Style: bold header, auto-filter
    _style_header(ws, len(ALL_ROWS_HEADERS))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(ALL_ROWS_HEADERS))}{len(flat) + 1}"

    # Conditional: highlight low-confidence rows yellow
    conf_col = ALL_ROWS_HEADERS.index("confidence") + 1  # 1-based
    for row_idx in range(2, len(flat) + 2):
        cell = ws.cell(row=row_idx, column=conf_col)
        if cell.value is not None and isinstance(cell.value, (int, float)) and cell.value < 0.8:
            for col in range(1, len(ALL_ROWS_HEADERS) + 1):
                ws.cell(row=row_idx, column=col).fill = WARN_FILL


# ---- Sheet 2: raw_like_original ----

def _write_raw_like_original(wb: openpyxl.Workbook, results: list[TableResult]) -> None:
    ws = wb.create_sheet("raw_like_original")
    current_row = 1

    for ti, table in enumerate(results):
        # Section header
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=table.num_cols,
        )
        cell = ws.cell(row=current_row, column=1)
        cell.value = f"{table.source_file}  |  block={table.source_block}  |  {table.num_rows}×{table.num_cols}"
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        current_row += 1

        # Write grid rows
        for grid_row in table.grid_rows:
            for j, cell in enumerate(grid_row.cells):
                ws_cell = ws.cell(row=current_row, column=j + 1)
                ws_cell.value = cell.merged_text if cell.merged_text else (
                    cell.texts[0].text if cell.texts else ""
                )
                ws_cell.font = DATA_FONT
                ws_cell.border = THIN_BORDER
                ws_cell.alignment = Alignment(wrap_text=True, vertical="center")

                # Style header/total rows
                if grid_row.row_type.value in ("header", "subheader"):
                    ws_cell.font = Font(bold=True, size=10)
                    ws_cell.fill = HEADER_FILL
                elif grid_row.row_type.value in ("total", "summary"):
                    ws_cell.font = Font(bold=True, size=10)
            current_row += 1

        current_row += 2  # gap between tables

    # Auto-fit column widths (approximate)
    for col_idx in range(1, 10):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16


# ---- Sheet 3: table_summary ----

def _write_table_summary(wb: openpyxl.Workbook, results: list[TableResult]) -> None:
    ws = wb.create_sheet("table_summary")

    rows = []
    for ti, table in enumerate(results):
        rows.append({
            "source_file": table.source_file,
            "table_index": ti,
            "source_block": table.source_block,
            "bbox_x1": table.bbox_x1,
            "bbox_y1": table.bbox_y1,
            "bbox_x2": table.bbox_x2,
            "bbox_y2": table.bbox_y2,
            "num_rows": table.num_rows,
            "num_cols": table.num_cols,
            "data_cols": table.data_cols,
            "text_count": table.text_count,
            "line_count": table.line_count,
            "candidate_score": table.candidate_score,
            "grid_regularity": table.grid_regularity,
            "fill_rate": table.fill_rate,
            "grid_score": table.grid_score,
        })

    if rows:
        df = pd.DataFrame(rows)
        df = df[SUMMARY_HEADERS]
        _write_df_to_sheet(ws, df, SUMMARY_HEADERS)
    else:
        _write_header_row(ws, SUMMARY_HEADERS)

    _style_header(ws, len(SUMMARY_HEADERS))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(SUMMARY_HEADERS))}{max(len(rows), 1) + 1}"


# ---- Sheet 4: warnings ----

def _write_warnings(wb: openpyxl.Workbook, warnings: list[WarningInfo]) -> None:
    ws = wb.create_sheet("warnings")

    rows = [
        {
            "source_file": w.source_file,
            "table_index": w.table_index,
            "row_index": w.row_index,
            "warning_code": w.warning_code,
            "message": w.message,
            "raw_value": w.raw_value,
        }
        for w in warnings
    ]

    if rows:
        df = pd.DataFrame(rows)
        df = df[WARN_HEADERS]
        _write_df_to_sheet(ws, df, WARN_HEADERS)
    else:
        _write_header_row(ws, WARN_HEADERS)
        ws.cell(row=2, column=1).value = "(no warnings)"

    _style_header(ws, len(WARN_HEADERS))


# ---- Helpers ----

def _write_df_to_sheet(ws, df: pd.DataFrame, headers: list[str]) -> None:
    """Write a pandas DataFrame to an openpyxl worksheet."""
    # Write header
    for j, h in enumerate(headers):
        ws.cell(row=1, column=j + 1).value = h

    # Write data rows
    for i, (_, row) in enumerate(df.iterrows()):
        for j, h in enumerate(headers):
            val = row.get(h)
            # Convert numpy/pandas types to Python native
            if hasattr(val, "item"):
                val = val.item()
            elif hasattr(val, "iloc"):
                val = str(val)
            ws.cell(row=i + 2, column=j + 1).value = val


def _write_header_row(ws, headers: list[str]) -> None:
    """Write just the header row."""
    for j, h in enumerate(headers):
        ws.cell(row=1, column=j + 1).value = h


def _style_header(ws, n_cols: int) -> None:
    """Apply header styling to row 1."""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
